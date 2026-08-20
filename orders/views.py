from __future__ import annotations
from functools import wraps
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import decimal
from datetime import datetime, timedelta, time
from django.template.loader import render_to_string
from urllib.parse import quote
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest, HttpResponseForbidden, Http404
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_GET, require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_exempt
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import (
    Count,
    Q,
    Sum,
    F,
    Value,
    Max,
    DecimalField,
    ExpressionWrapper,
    Case,
    When,
)
from django.db import transaction
from wallets.models import WalletTransaction
from wallets.services import (
    get_or_create_wallet_for_customer,
    get_or_create_wallet_for_delivery_partner,
    credit_wallet,
    debit_wallet,
    distribute_order_revenues,
)
from django.db.models.functions import Coalesce, Cast, TruncDate
from django.views.decorators.http import require_POST, require_http_methods
from django.contrib.admin.views.decorators import staff_member_required
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.encoding import smart_str

from orders.utils.pricing import compute_order_amounts
from orders.utils import build_order_canonical_snapshot
from orders.presenters import build_order_display_summary, build_order_finance_summary
from orders.pricing_engine import compute_order_pricing
from orders.utils.address_rules import clean_address_or_empty, is_probably_valid_address

from orders.utils.settings_loader import get_pricing_settings
from orders.utils.geocoding import ensure_order_geocoded
from orders.utils.geo import resolve_pickup_coords, resolve_delivery_coords, resolve_provider_coords
from .config_models import InvoiceSettings
from .models import (
    Order,
    Customer,
    OrderItem,
    OrderItemPhoto,
    ServiceCategory,
    ServiceItem,
    DeliveryLeg,
    OrderStatusHistory,
    OrderUpsell,
    haversine_distance_km,
    LogisticsConfig,
)
from partners.models import LaundryPartner, DeliveryPartner
from mlm.services import generate_mlm_commissions_for_order
from wallets.models import Wallet
from mlm.models import ReferralLink, ReferralCommission
from .assignment import assign_best_driver, assign_best_laundry
from .utils import build_order_display_context
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from orders.utils.assign import get_active_drivers
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except Exception:
    WEASYPRINT_AVAILABLE = False
    HTML = None
import uuid
import csv
import base64
import json
import re
import logging


DEC12 = DecimalField(max_digits=12, decimal_places=2)
DEC = DecimalField(max_digits=12, decimal_places=2)
DECIMAL_ZERO = Decimal("0")


logger = logging.getLogger(__name__)


# =========================
# FAGNI MONETISATION ENGINE
# =========================
def apply_fagni_monetization(order):
    from decimal import Decimal
    from wallets.services import get_or_create_wallet_for_customer, credit_wallet
    from mlm.models import ReferralLink

    if not order or not getattr(order, "customer", None):
        return

    customer = order.customer

    try:
        total = Decimal(str(getattr(order, "total_client_ttc", 0) or 0))
        total = total - Decimal(str(getattr(order, "coupon_discount_applied", 0) or 0))
        if total < 0:
            total = Decimal("0")
    except Exception:
        total = Decimal("0")

    # ---- CASHBACK 2% ----
    cashback = (total * Decimal("0.02")).quantize(Decimal("0.01"))

    try:
        wallet = get_or_create_wallet_for_customer(customer)
        if cashback > 0:
            credit_wallet(
                wallet=wallet,
                amount=cashback,
                description=f"Cashback FAGNI commande {order.code}",
                order=order,
            )
    except Exception:
        import logging
        logging.getLogger("fagni.views.wallet").exception("Echec silencieux: credit_wallet cashback commande | order_id=%s", getattr(order, "id", None) if "order" in dir() else None)

    # ---- REFERRAL ----
    # Supprime le 19 juillet 2026 : bloc mort depuis toujours (ReferralLink
    # n'a jamais eu de champs 'referred_customer'/'referrer' - vrais noms
    # 'customer'/'sponsor' - chaque appel levait FieldError, capturee
    # silencieusement, donc ni le credit parrain (500) ni le bonus
    # bienvenue (300) n'ont jamais ete verses via ce chemin. Le vrai
    # systeme actif est handle_referral_reward() (orders/models.py, sur
    # transition payment_status vers 'paid'), avec le bon montant
    # (1000 FCFA, Pilot Growth Plan du 9 juillet 2026). Le bonus
    # bienvenue (300 FCFA) n'a pas d'equivalent actif - a recreer
    # separement si ce comportement est encore voulu.




def referral_dashboard_data():
    """
    KPI business du moteur de parrainage FAGNI.
    """
    from django.db.models import Sum, Q, Count
    from wallets.models import WalletTransaction
    from orders.models import Order
    from decimal import Decimal

    referred_orders = (
        Order.objects
        .exclude(referral_code__isnull=True)
        .exclude(referral_code__exact="")
    )

    total_orders = referred_orders.count()

    paid_orders_qs = referred_orders.filter(
        Q(payment_status="paid") | Q(status__in=["paid", "completed", "done"])
    )
    paid_orders = paid_orders_qs.count()

    total_ca = paid_orders_qs.aggregate(
        total=Sum("total_client_ttc")
    ).get("total") or Decimal("0")

    reward_qs = WalletTransaction.objects.filter(
        type="mlm_commission",
        direction="in",
    )

    total_rewards = reward_qs.aggregate(
        total=Sum("amount")
    ).get("total") or Decimal("0")

    rewarded_users = (
        reward_qs.values("wallet__customer_id")
        .exclude(wallet__customer_id__isnull=True)
        .distinct()
        .count()
    )

    activated_wallet_users = (
        WalletTransaction.objects
        .filter(
            wallet__owner_type="customer",
            direction="out",
            amount__gt=0,
            wallet__customer_id__in=reward_qs.values("wallet__customer_id"),
        )
        .values("wallet__customer_id")
        .distinct()
        .count()
    )

    conversion = Decimal("0")
    if total_orders > 0:
        conversion = (Decimal(str(paid_orders)) / Decimal(str(total_orders))) * Decimal("100")

    wallet_activation_rate = Decimal("0")
    if rewarded_users > 0:
        wallet_activation_rate = (
            Decimal(str(activated_wallet_users)) / Decimal(str(rewarded_users))
        ) * Decimal("100")

    roi_ratio = Decimal("0")
    if total_rewards > 0:
        roi_ratio = Decimal(str(total_ca)) / Decimal(str(total_rewards))

    top_sponsors = (
        reward_qs
        .filter(wallet__owner_type="customer")
        .values(
            "wallet__customer__id",
            "wallet__customer__name",
            "wallet__customer__phone",
        )
        .annotate(
            total_commission=Sum("amount"),
            rewards_count=Count("id"),
        )
        .order_by("-total_commission")[:10]
    )

    return {
        "total_orders": total_orders,
        "paid_orders": paid_orders,
        "conversion": conversion.quantize(Decimal("0.1")) if total_orders else Decimal("0.0"),
        "total_ca": total_ca,
        "total_rewards": total_rewards,
        "rewarded_users": rewarded_users,
        "activated_wallet_users": activated_wallet_users,
        "wallet_activation_rate": wallet_activation_rate.quantize(Decimal("0.1")) if rewarded_users else Decimal("0.0"),
        "roi_ratio": roi_ratio.quantize(Decimal("0.1")) if total_rewards else Decimal("0.0"),
        "top_sponsors": top_sponsors,
    }


@staff_member_required
def admin_referral_dashboard(request):
    ctx = referral_dashboard_data()
    return render(request, "orders/admin_referral_dashboard.html", ctx)

# =========================
#  Helpers AJAX / Client auth
# =========================


def _build_client_display_items(order):
    rows = []
    try:
        qs = order.items.all()
    except Exception:
        return rows

    for it in qs:
        label = (
            getattr(it, "designation", None)
            or getattr(getattr(it, "service", None), "name", None)
            or getattr(it, "service_type", None)
            or "Article"
        )
        qty = getattr(it, "quantity", None) or 1
        price = (
            getattr(it, "total_price", None)
            or getattr(it, "unit_price", None)
            or 0
        )
        rows.append({
            "label": label,
            "qty": qty,
            "price": price,
        })
    return rows



def _get_order_upsell_total(order) -> Decimal:
    try:
        upsell = getattr(order, "upsell", None)
        if upsell:
            return upsell.total
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=302")
    return Decimal("0.00")


CLIENT_PHONE_COOKIE = "client_phone"
CLIENT_REF_COOKIE = "client_ref_code"


def _is_json_request(request) -> bool:
    # AJAX (fetch) + Accept JSON
    xrw = (request.headers.get("X-Requested-With") or "").lower()
    accept = (request.headers.get("Accept") or "").lower()
    return xrw == "xmlhttprequest" or "application/json" in accept


def _client_phone(request) -> str | None:
    """
    Retourne le téléphone client "authentifié" côté app client.
    Stratégie (simple) : cookie HTTPOnly.
    """
    phone = request.COOKIES.get(CLIENT_PHONE_COOKIE)
    if phone:
        phone = phone.strip()
    return phone or None


def _client_ref_code(request) -> str | None:
    ref = request.GET.get("ref") or request.COOKIES.get(CLIENT_REF_COOKIE)
    ref = (ref or "").strip().upper()
    return ref or None


def client_login_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not _client_phone(request):
            # ✅ IMPORTANT: pour les appels JS, JSON 401 (pas redirect HTML)
            if _is_json_request(request):
                return JsonResponse({"ok": False, "error": "not_authenticated"}, status=401)


            return redirect("orders:client_login")
        return view_func(request, *args, **kwargs)
    return _wrapped
# Aliases compat (anciens noms)
client_required = client_login_required
_client_required = client_login_required


# =========================
#  Login "client" (par téléphone)
# =========================
@require_http_methods(["GET", "POST"])
def client_login(request):
    """
    Login client minimal: on saisit un téléphone => pose cookie client_phone.
    Capture aussi un éventuel code de parrainage entrant (?ref=CODE).
    """
    incoming_ref = (request.GET.get("ref") or request.COOKIES.get(CLIENT_REF_COOKIE) or "").strip().upper()

    if request.method == "GET":
        resp = render(request, "orders/client_login.html", {"ok": True, "incoming_ref": incoming_ref})
        if incoming_ref:
            resp.set_cookie(
                CLIENT_REF_COOKIE,
                incoming_ref,
                max_age=30 * 24 * 3600,
                httponly=True,
                samesite="Lax",
            )
        return resp

    phone = (request.POST.get("phone") or "").strip()
    incoming_ref = (request.POST.get("incoming_ref") or incoming_ref or "").strip().upper()

    if not phone:
        if _is_json_request(request):
            return JsonResponse({"ok": False, "error": "missing_phone"}, status=400)
        return render(
            request,
            "orders/client_login.html",
            {"ok": False, "error": "missing_phone", "incoming_ref": incoming_ref},
        )

    resp = redirect("orders:client_home")
    resp.set_cookie(
        CLIENT_PHONE_COOKIE,
        phone,
        max_age=7 * 24 * 3600,
        httponly=True,
        samesite="Lax",
    )

    if incoming_ref:
        resp.set_cookie(
            CLIENT_REF_COOKIE,
            incoming_ref,
            max_age=30 * 24 * 3600,
            httponly=True,
            samesite="Lax",
        )

    return resp


@require_http_methods(["POST"])
def client_logout(request):
    resp = redirect("orders:client_login")
    resp.delete_cookie(CLIENT_PHONE_COOKIE)
    return resp


@client_login_required
def client_wallet(request):
    phone = _client_phone(request)
    customer = Customer.objects.filter(phone=phone).order_by("-id").first()
    if not customer:
        messages.error(request, "Session client invalide. Reconnecte-toi.")
        return redirect("orders:client_login")

    wallet = (
        Wallet.objects
        .filter(owner_type="customer", customer=customer)
        .order_by("-id")
        .first()
    )

    if wallet:
        transactions = (
            WalletTransaction.objects
            .filter(wallet=wallet)
            .select_related("order")
            .order_by("-created_at")[:50]
        )
        balance = getattr(wallet, "balance", Decimal("0.00")) or Decimal("0.00")
    else:
        transactions = []
        balance = Decimal("0.00")

    ctx = {
        "phone": phone,
        "customer": customer,
        "wallet": wallet,
        "balance": balance,
        "transactions": transactions,
    }
    resp = render(request, "orders/client_wallet.html", ctx)
    resp["Cache-Control"] = "no-store"
    return resp


@client_login_required
def client_referrals(request):
    phone = _client_phone(request)
    customer = Customer.objects.filter(phone=phone).order_by("-id").first()
    if not customer:
        messages.error(request, "Session client invalide. Reconnecte-toi.")
        return redirect("orders:client_login")

    def _clean_phone_for_code(value: str) -> str:
        raw = re.sub(r"\D+", "", str(value or ""))
        if raw:
            return raw[-8:]
        return f"{customer.id:04d}"

    def _build_referral_code() -> str:
        base = f"FAGNI-{_clean_phone_for_code(getattr(customer, 'phone', ''))}"
        code = base
        i = 2
        while ReferralLink.objects.exclude(customer=customer).filter(referral_code=code).exists():
            code = f"{base}-{i}"
            i += 1
        return code

    activate_now = (request.GET.get("activate_ref") or "").strip() == "1"

    profile = (
        ReferralLink.objects
        .select_related("customer", "sponsor")
        .filter(customer=customer)
        .order_by("-id")
        .first()
    )

    if not profile and activate_now:
        profile = ReferralLink.objects.create(
            customer=customer,
            referral_code=_build_referral_code(),
            actor_type="client",
        )
        messages.success(request, "Ton code de parrainage FAGNI est maintenant activé.")

    if profile:
        direct_referrals = list(
            profile.direct_referrals
            .select_related("customer")
            .order_by("-created_at")[:50]
        )

        commissions = list(
            ReferralCommission.objects
            .filter(beneficiary_profile=profile)
            .select_related("order", "order__customer")
            .order_by("-created_at")[:50]
        )

        total_commissions = (
            ReferralCommission.objects
            .filter(beneficiary_profile=profile)
            .aggregate(total=Sum("commission_amount"))
            .get("total")
            or 0
        )

        referral_url = f"https://fagni.app/invite/{profile.referral_code}"
        referral_whatsapp_text = (
            f"💰 {customer.name if customer and getattr(customer, 'name', None) else 'Je'} viens de découvrir une astuce simple pour gagner de l’argent avec FAGNI\n\n"
            "Tu peux commander ton linge et gagner des récompenses en invitant 1 proche 👌\n\n"
            "👉 Teste ici :\n"
            f"{referral_url}\n\n"
            "Franchement, ça vaut le coup d’essayer 🔥"
        )
        referral_whatsapp_url = (
            "https://wa.me/?text=" + quote(referral_whatsapp_text)
        )
    else:
        direct_referrals = []
        commissions = []
        total_commissions = 0
        referral_url = ""
        referral_whatsapp_text = ""
        referral_whatsapp_url = ""

    gains = compute_referral_gains(profile) if profile else {}

    projected_gain_low = 5000
    projected_gain_high = 20000

    earning_scenarios = [
        {"label": "1 proche actif", "amount": 2000},
        {"label": "3 proches actifs", "amount": 5000},
        {"label": "5 proches actifs", "amount": 10000},
    ]

    wallet_activation_message = ""
    wallet_reminder_text = ""
    wallet_reminder_whatsapp_url = ""
    child_referral_discount = 500

    if gains and gains.get("total", 0):
        wallet_activation_message = (
            f"Tu as déjà jusqu’à {gains.get('total', 0)} FCFA de gains potentiels dans ton réseau FAGNI."
        )
        wallet_reminder_text = (
            f"Bonjour, j’ai déjà jusqu’à {gains.get('total', 0)} FCFA de gains potentiels dans mon espace FAGNI. "
            "Je veux activer mes gains sur ma prochaine commande. "
            "Peux-tu m’aider à finaliser ma commande maintenant ?"
        )
        wallet_reminder_whatsapp_url = "https://wa.me/?text=" + quote(wallet_reminder_text)

    ctx = {
        "phone": phone,
        "customer": customer,
        "profile": profile,
        "direct_referrals": direct_referrals,
        "commissions": commissions,
        "total_commissions": total_commissions,
        "referral_url": referral_url,
        "referral_whatsapp_text": referral_whatsapp_text,
        "referral_whatsapp_url": referral_whatsapp_url,
        "projected_gain_low": projected_gain_low,
        "projected_gain_high": projected_gain_high,
        "earning_scenarios": earning_scenarios,
        "gains": gains,
        "wallet_activation_message": wallet_activation_message,
        "wallet_reminder_text": wallet_reminder_text,
        "wallet_reminder_whatsapp_url": wallet_reminder_whatsapp_url,
        "child_referral_discount": child_referral_discount,
    }
    resp = render(request, "orders/client_referrals.html", ctx)
    resp["Cache-Control"] = "no-store"
    return resp


# =========================
#  LIVE endpoint (exemple)
# =========================
@client_login_required
def client_order_live_status(request, order_id: int):
    """
    Endpoint JSON live pour le suivi commande client.
    Règle d'accès: le cookie client_phone doit matcher le customer.phone (ou autre logique).
    """
    from orders.models import OrderItem, Order
    from collections import Counter

    phone = _client_phone(request)

    o = (
        Order.objects
        .select_related("customer")
        .prefetch_related("items", "legs")
        .filter(pk=order_id)
        .first()
    )
    if not o:
        return JsonResponse({"ok": False, "error": "not_found"}, status=404)

    customer_phone = _normalize_phone(getattr(o.customer, "phone", "") or "")
    phone = _normalize_phone(phone or "")

    if not customer_phone or not phone or customer_phone != phone:
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    items = []
    for it in o.items.all():
        qty = float(it.quantity or 0)
        price = float(it.unit_price or 0)
        items.append({
            "id": it.id,
            "designation": it.designation,
            "quantity": qty,
            "unit_price": price,
            "total": float(getattr(it, "line_total", qty * price)),
        })

    legs = []
    legs_by_id = {}
    for leg in o.legs.all().order_by("id"):
        row = {
            "id": leg.id,
            "leg_type": leg.leg_type,
            "status": leg.status,
            "driver_amount": float(leg.driver_amount or 0),
            "driver_id": leg.driver_id,
        }
        legs.append(row)
        legs_by_id[leg.id] = row

    # ✅ Source unique de vérité pour les montants
    pricing = _compute_order_pricing(o)

    # (optionnel mais recommandé) sync sans effet DB pour refléter Payments

    # ✅ Paiement CANONIQUE (APP CLIENT)
    # IMPORTANT: ne pas appeler sync_payment_status_from_payments() ici,
    # car même save=False peut MUTER o.amount_paid / o.payment_status en mémoire.
    from decimal import Decimal

    total_ttc_dec = Decimal(str(pricing.get("total_client", 0) or 0))

    # ✅ On lit la DB (Order.amount_paid) comme source de vérité
    paid_raw = getattr(o, "amount_paid", None)
    try:
        paid_dec = Decimal(str(paid_raw or 0))
    except Exception:
        paid_dec = Decimal("0")
    if paid_dec < 0:
        paid_dec = Decimal("0")

    # ✅ Clamp UI : surpaiement => on affiche max = total
    paid_ui_dec = paid_dec
    if paid_ui_dec > total_ttc_dec:
        paid_ui_dec = total_ttc_dec

    remaining_dec = total_ttc_dec - paid_ui_dec
    if remaining_dec < 0:
        remaining_dec = Decimal("0")

    # ✅ statut paiement canonique
    if total_ttc_dec <= 0:
        payment_ui = "waiting_amount"
        payment_status_ui = "unpaid"
    else:
        if paid_ui_dec <= 0:
            payment_ui = "unpaid"
            payment_status_ui = "unpaid"
        elif paid_ui_dec >= total_ttc_dec:
            payment_ui = "paid"
            payment_status_ui = "paid"
        else:
            payment_ui = "partial"
            payment_status_ui = "partial"

    overpaid_dec = paid_dec - total_ttc_dec
    if overpaid_dec < 0:
        overpaid_dec = Decimal("0")

    amounts = {
        "prestation_total": float(pricing["items_total"]),
        "service_fee": float(pricing["service_fee"]),
        "delivery_fee": float(pricing["delivery_fee"]),
        "express_extra_fee": float(pricing.get("express_extra_fee", 0)),
        "vat_fagni": float(pricing["vat_fagni"]),
        "total_ttc": float(pricing["total_client"]),
        "amount_paid": float(paid_ui_dec),
        "amount_paid_raw": float(paid_dec),
        "overpaid": float(overpaid_dec),
        "amount_remaining": float(remaining_dec),
    }

    # =========================
    # ✅ PREUVES / PHOTOS (OrderEvidencePhoto)
    # =========================
    evidence_photos = []
    evidence_counts = {}
    try:
        from .models import OrderEvidencePhoto

        # mapping label (KIND_CHOICES)
        kind_labels = {k: lbl for (k, lbl) in getattr(OrderEvidencePhoto, "KIND_CHOICES", [])}

        qs = (
            OrderEvidencePhoto.objects
            .filter(order=o)
            .select_related("leg")
            .order_by("-created_at")[:50]
        )

        kinds = []
        for p in qs:
            kind = getattr(p, "kind", "") or ""
            kinds.append(kind)

            url = ""
            try:
                if getattr(p, "image", None) and getattr(p.image, "url", None):
                    url = p.image.url
            except Exception:
                url = ""

            leg_id = getattr(p, "leg_id", None)
            leg_info = None
            if leg_id and leg_id in legs_by_id:
                leg_info = {
                    "id": leg_id,
                    "leg_type": legs_by_id[leg_id].get("leg_type"),
                    "status": legs_by_id[leg_id].get("status"),
                }

            evidence_photos.append({
                "id": p.id,
                "kind": kind,
                "kind_label": kind_labels.get(kind, kind or "Photo"),
                "caption": getattr(p, "caption", "") or "",
                "created_at": p.created_at.isoformat() if getattr(p, "created_at", None) else None,
                "url": url,
                "leg_id": leg_id,
                "leg": leg_info,
            })

        evidence_counts = dict(Counter(kinds))
    except Exception:
        # safe: endpoint continue même si modèle indispo
        evidence_photos = []
        evidence_counts = {}

    # =========================
    # ✅ RATING (OrderRating)
    # =========================
    rating = None
    try:
        from .models import OrderRating
        r = OrderRating.objects.filter(order=o).first()
        if r:
            rating = {
                "score": int(getattr(r, "score", 0) or 0),
                "comment": (getattr(r, "comment", "") or "").strip(),
                "created_at": r.created_at.isoformat() if getattr(r, "created_at", None) else None,
            }
        else:
            rating = None
    except Exception:
        rating = None

    return JsonResponse({
        "ok": True,
        "order": {
            "id": o.id,
            "code": o.code,
            "status": o.status,
            "payment_status": payment_status_ui,
            "amount_paid": float(paid_dec),
            "items": items,
            "pickup_time": o.pickup_time.isoformat() if o.pickup_time else None,
            "wash_complete_time": o.wash_complete_time.isoformat() if o.wash_complete_time else None,
            "delivered_time": o.delivered_time.isoformat() if o.delivered_time else None,
            "express_extra_fee": float(pricing.get("express_extra_fee", 0)),
        },
        "amounts": amounts,
        "legs": legs,

        # ✅ AJOUTS MVP PILOTE
        "evidence_photos": evidence_photos,
        "evidence_counts": evidence_counts,
        "rating": rating,
    })


@require_http_methods(["POST"])
@client_login_required
def client_order_rating(request, order_id: int):
    """
    Enregistre (ou met à jour) l'évaluation client pour une commande.
    Règles:
    - Accès: cookie client_phone doit matcher le customer.phone
    - Autorisé seulement si commande terminée (status=done) OU delivered_time non null
    - Idempotent: upsert (update si existe)
    - Réponse JSON si AJAX, sinon redirect vers détail commande
    """
    from orders.models import Order, OrderRating

    phone = _normalize_phone((_client_phone(request) or ""))
    o = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id)
        .first()
    )
    if not o:
        return JsonResponse({"ok": False, "error": "not_found"}, status=404)

    customer_phone = _normalize_phone(getattr(o.customer, "phone", "") or "")
    if not customer_phone or not phone or customer_phone != phone:
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    # Autoriser seulement si terminé/livré
    if (getattr(o, "status", "") != "done") and (getattr(o, "delivered_time", None) is None):
        return JsonResponse({"ok": False, "error": "not_allowed"}, status=400)

    # Payload
    score_raw = (request.POST.get("score") or request.POST.get("rating") or "").strip()
    comment = (request.POST.get("comment") or request.POST.get("review") or "").strip()

    try:
        score = int(score_raw)
    except Exception:
        score = 0

    if score < 1 or score > 5:
        return JsonResponse({"ok": False, "error": "invalid_score"}, status=400)

    if len(comment) > 500:
        comment = comment[:500]

    r, created = OrderRating.objects.get_or_create(order=o, defaults={"score": score, "comment": comment})
    if not created:
        r.score = score
        r.comment = comment
        r.save(update_fields=["score", "comment", "updated_at"])

    payload = {
        "ok": True,
        "rating": {
            "score": int(r.score or 0),
            "comment": (r.comment or "").strip(),
            "created_at": r.created_at.isoformat() if getattr(r, "created_at", None) else None,
        }
    }

    if _is_json_request(request):
        return JsonResponse(payload)

    return redirect("orders:client_order_detail", order_id=o.id)


# =========================
# ✅ CLIENT — Evidence upload (preuves photos)
# =========================
@require_http_methods(["POST"])
@client_login_required
def client_order_evidence_upload(request, order_id: int):
    """
    ✅ CLIENT — Evidence upload (preuves photos)
    - Auth: cookie client_phone
    - Règle: phone doit matcher order.customer.phone
    - Input:
        kind (pickup|laundry|dropoff) + files[] (ou photos[])
        step (collecte|lavage|livraison) accepté aussi (mapping vers kind)
    - Output JSON:
        {"ok": True, "created": n, "rejected": n, "rejected_reasons": {...},
         "count_total": n, "evidence_photos": [...], "evidence_counts": {...}, "last_added_ids":[...]}
    """
    from collections import Counter
    from datetime import timedelta
    from django.http import JsonResponse
    from django.utils import timezone
    from django.conf import settings
    from django.views.decorators.http import require_http_methods
    from orders.models import Order, OrderEvidencePhoto

    # ------- helpers -------
    def _is_json(req):
        try:
            if req.headers.get("x-requested-with") == "XMLHttpRequest":
                return True
            acc = (req.headers.get("accept") or "").lower()
            return "application/json" in acc
        except Exception:
            return True

    def _norm_phone(x):
        try:
            x = (x or "").strip()
            x = x.replace(" ", "").replace("-", "")
            if x.startswith("+225"):
                x = x[4:]
            if x.startswith("00225"):
                x = x[5:]
            return x
        except Exception:
            return (x or "").strip()

    # ------- load order FIRST -------
    o = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id)
        .first()
    )
    if not o:
        return JsonResponse({"ok": False, "error": "not_found"}, status=404)

    # ------- auth check -------
    # utilise ta fonction existante si elle existe, sinon fallback cookie direct
    phone = ""
    try:
        phone = _normalize_phone(_client_phone(request) or "")
    except Exception:
        phone = _norm_phone(request.COOKIES.get("client_phone") or "")

    customer_phone = ""
    try:
        customer_phone = _normalize_phone(getattr(o.customer, "phone", "") or "")
    except Exception:
        customer_phone = _norm_phone(getattr(o.customer, "phone", "") or "")

    if not customer_phone or not phone or customer_phone != phone:
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    # ------- business rules -------
    status_raw = (getattr(o, "status", "") or "").lower()
    if status_raw == "canceled":
        return JsonResponse({"ok": False, "error": "order_canceled"}, status=403)

    delivered_at = getattr(o, "delivered_at", None)
    is_delivered = False

    if delivered_at:
        is_delivered = True
    else:
        if status_raw in {"done", "delivered"}:
            is_delivered = True
            delivered_at = getattr(o, "updated_at", None) or getattr(o, "created_at", None)

    if is_delivered:
        delivered_at = delivered_at or timezone.now()
        if timezone.now() - delivered_at > timedelta(hours=24):
            return JsonResponse({"ok": False, "error": "evidence_window_closed"}, status=403)

    # ------- input (kind/step) -------
    kind_raw = (request.POST.get("kind") or "").strip().lower()
    step_raw = (request.POST.get("step") or "").strip().lower()

    step_map = {
        "collecte": "pickup",
        "pickup": "pickup",
        "ramassage": "pickup",
        "lavage": "laundry",
        "laundry": "laundry",
        "pressing": "laundry",
        "livraison": "dropoff",
        "dropoff": "dropoff",
        "delivery": "dropoff",
    }

    kind = "pickup"
    if kind_raw in {"pickup", "laundry", "dropoff"}:
        kind = kind_raw
    elif step_raw:
        kind = step_map.get(step_raw, "pickup")

    # Si livré, seul dropoff autorisé
    if is_delivered and kind != "dropoff":
        return JsonResponse({"ok": False, "error": "only_dropoff_allowed_after_delivery"}, status=403)

    # ------- files -------
    files = []
    files += list(request.FILES.getlist("files"))
    files += list(request.FILES.getlist("photos"))
    f1 = request.FILES.get("file")
    if f1:
        files.append(f1)

    files = [f for f in files if f]
    if not files:
        return JsonResponse({"ok": False, "error": "no_files"}, status=400)

    allowed_mimes = {"image/jpeg", "image/png", "image/webp"}
    max_mb = int(getattr(settings, "FAGNI_EVIDENCE_MAX_MB", 5) or 5)
    max_bytes = max_mb * 1024 * 1024

    # détecter le champ FileField du modèle
    file_field = None
    try:
        for cand in ("file", "photo", "image", "img", "picture"):
            if cand in [f.name for f in OrderEvidencePhoto._meta.fields]:
                file_field = cand
                break
        if not file_field:
            # fallback: premier FileField/ImageField
            for f in OrderEvidencePhoto._meta.fields:
                if f.get_internal_type() in ("FileField", "ImageField"):
                    file_field = f.name
                    break
    except Exception:
        file_field = "file"

    created = 0
    rejected = 0
    reasons = Counter()
    last_ids = []

    for f in files:
        try:
            ctype = (getattr(f, "content_type", "") or "").lower()
            if ctype and ctype not in allowed_mimes:
                rejected += 1
                reasons["bad_type"] += 1
                continue
            if getattr(f, "size", 0) and f.size > max_bytes:
                rejected += 1
                reasons["too_large"] += 1
                continue

            obj = OrderEvidencePhoto(order=o, kind=kind)
            # set file field
            try:
                setattr(obj, file_field, f)
            except Exception:
                # si field différent, tente 'file'
                try:
                    setattr(obj, "file", f)
                except Exception:
                    import logging
                    logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=1046")

            obj.save()
            created += 1
            last_ids.append(getattr(obj, "id", None))
        except Exception:
            rejected += 1
            reasons["error"] += 1

    # refresh list
    qs = OrderEvidencePhoto.objects.filter(order=o).order_by("-created_at", "-id")
    photos = []
    counts = Counter()

    for ph in qs[:40]:
        k = getattr(ph, "kind", "") or ""
        counts[k] += 1

        url = ""
        try:
            ff = getattr(ph, file_field, None)
            url = getattr(ff, "url", "") or ""
        except Exception:
            url = ""

        photos.append({
            "id": getattr(ph, "id", None),
            "kind": k,
            "kind_label": k,
            "url": url,
            "created_at": getattr(ph, "created_at", None).isoformat() if getattr(ph, "created_at", None) else None,
        })

    payload = {
        "ok": True,
        "created": created,
        "rejected": rejected,
        "rejected_reasons": {
            "too_large": int(reasons.get("too_large", 0)),
            "bad_type": int(reasons.get("bad_type", 0)),
            "error": int(reasons.get("error", 0)),
        },
        "count_total": int(qs.count()),
        "evidence_photos": photos,
        "evidence_counts": dict(counts),
        "last_added_ids": [x for x in last_ids if x is not None],
    }
    return JsonResponse(payload)
def _wallet_net_expr():
    """
    Net = somme(direction=in) - somme(direction=out)
    Utilisé pour revenus livreurs basés sur WalletTransaction.
    """
    return Sum(
        Case(
            When(direction="in", then=F("amount")),
            When(direction="out", then=-F("amount")),
            default=Value(0),
            output_field=DEC,
        )
    )


@login_required
def driver_app_alias(request):
    qs = request.META.get("QUERY_STRING", "")
    url = reverse("orders:driver_app")  # => /orders/driver-app/
    return redirect(f"{url}?{qs}" if qs else url)

def _strip_django_comment_blocks(txt: str) -> str:
    if not txt:
        return ""
    cleaned = re.sub(r"{#.*?#}", "", str(txt), flags=re.DOTALL)
    return "\n".join([line.rstrip() for line in cleaned.splitlines() if line.strip()])


def get_invoice_settings_clean():
    """
    Retourne InvoiceSettings (singleton pk=1) avec header/footer nettoyés
    pour éviter l'affichage de {# ... #} dans les PDF/tickets.
    """
    try:
        invoice_settings, _ = InvoiceSettings.objects.get_or_create(pk=1)

        # Nettoyage pour affichage seulement (pas de save)
        invoice_settings.header_note = _strip_django_comment_blocks(
            getattr(invoice_settings, "header_note", "")
        )
        invoice_settings.footer_text = _strip_django_comment_blocks(
            getattr(invoice_settings, "footer_text", "")
        )
        return invoice_settings
    except Exception:
        return None


def _qr_png_base64(data: str) -> str:
    """
    Génère un QR code PNG et renvoie le base64 (sans préfixe data:).
    """
    try:
        import qrcode
    except Exception:
        return ""

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=6,
        border=2,
    )
    qr.add_data(data or "")
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB").convert("RGB")

    buf = BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def generate_order_code():
    return uuid.uuid4().hex[:10].upper()

# Champ décimal générique pour les expressions
DEC = DecimalField(max_digits=12, decimal_places=2)
DECIMAL_ZERO = Decimal("0")

# -------------------------------------------------

# Helpers FAGNI : parsing des lignes + calcul frais
# -------------------------------------------------
def fagni_parse_items_from_post(request):
    """
    Lit les tableaux envoyés par create.html :
    service_id[], designation[], quantity[], unit_price[]
    et renvoie (items, total_ht)
    """
    service_ids   = request.POST.getlist('service_id[]')
    designations  = request.POST.getlist('designation[]')
    quantities    = request.POST.getlist('quantity[]')
    unit_prices   = request.POST.getlist('unit_price[]')

    items = []
    total_ht = Decimal('0')

    for i in range(len(service_ids)):
        sid = service_ids[i].strip() if i < len(service_ids) else ""
        if not sid:
            continue

        designation = designations[i].strip() if i < len(designations) else ""
        q_raw = quantities[i] if i < len(quantities) else "0"
        pu_raw = unit_prices[i] if i < len(unit_prices) else "0"

        try:
            qty = int(q_raw)
        except (TypeError, ValueError):
            qty = 0

        try:
            pu = Decimal(str(pu_raw))
        except (TypeError, ValueError, ArithmeticError):
            pu = Decimal('0')

        if qty <= 0 or pu <= 0:
            continue

        line_total = pu * qty
        total_ht += line_total

        items.append({
            "service_id": sid,
            "designation": designation or "",
            "quantity": qty,
            "unit_price": pu,
            "total": line_total,
        })

    return items, total_ht


# ============================================================
#  HELPER : PRICING FAGNI
# ============================================================
def apply_fagni_pricing(order):
    """
    Recalcule tous les montants financiers d'une commande FAGNI selon les règles métier.

    Rappel des règles :
    - service_fee = max(5% du sous-total prestations, 500 FCFA) si prestation_total > 0, sinon 0
    - revenu FAGNI HT = commission_laundry_ht + commission_delivery_ht + logistic_margin + service_fee
    - TVA FAGNI (18%) = 18% du revenu FAGNI HT
    - revenu FAGNI TTC = revenu FAGNI HT + TVA FAGNI
    - total_client_ttc = prestation_total + service_fee + delivery_fee + TVA FAGNI
    """

    DEC = Decimal
    ZERO = DEC("0")

    # 1) Récup des bases
    prestation_total = order.prestation_total or ZERO
    delivery_fee = order.delivery_fee or ZERO

    commission_laundry = getattr(order, "commission_laundry_ht", ZERO) or ZERO
    commission_delivery = getattr(order, "commission_delivery_ht", ZERO) or ZERO
    logistic_margin = order.logistic_margin or ZERO

    # 2) Service FAGNI (HT)
    if prestation_total > ZERO:
        service_fee = (prestation_total * DEC("0.05"))
        if service_fee < DEC("500"):
            service_fee = DEC("500")
    else:
        service_fee = ZERO

    service_fee = service_fee.quantize(DEC("1."), rounding=ROUND_HALF_UP)
    order.service_fee = service_fee

    # 3) Revenu FAGNI HT
    fagni_ht = commission_laundry + commission_delivery + logistic_margin + service_fee
    fagni_ht = fagni_ht.quantize(DEC("1."), rounding=ROUND_HALF_UP)
    order.fagni_revenue_ht = fagni_ht

    # 4) TVA FAGNI (18%)
    vat_fagni = (fagni_ht * DEC("0.18")).quantize(DEC("1."), rounding=ROUND_HALF_UP)
    order.vat_fagni = vat_fagni

    # 5) Revenu FAGNI TTC
    order.fagni_revenue_ttc = fagni_ht + vat_fagni

    # 6) Total TTC facturé au client
    total_client_ttc = prestation_total + service_fee + delivery_fee + vat_fagni
    total_client_ttc = total_client_ttc.quantize(DEC("1."), rounding=ROUND_HALF_UP)
    order.total_client_ttc = total_client_ttc


@login_required
def portal_dashboard(request):
    """
    Dashboard central FAGNI (PORTAL) :
    - accès rapide aux stats commandes
    - liens vers OPS, finance, livreurs, etc.
    - tu pourras le rendre ultra premium visuellement dans orders/portal_dashboard.html
    """
    return render(request, "orders/portal_dashboard.html", {})


def fagni_compute_fees(total_ht: Decimal):
    """
    Règle FAGNI : frais de service et livraison.
    NB : on reste simple ici, la vraie logique distance/km
    pourra venir plus tard dans un module dédié.
    """
    if total_ht is None:
        total_ht = Decimal('0')

    # 5% du HT, min 500 si total > 0
    SERVICE_RATE = Decimal('0.05')
    SERVICE_MIN  = Decimal('500')

    if total_ht > 0:
        service_fee = total_ht * SERVICE_RATE
        if service_fee < SERVICE_MIN:
            service_fee = SERVICE_MIN
    else:
        service_fee = Decimal('0')

    # Frais de livraison : pour l'instant min fixe si total > 0
    default_delivery_min = getattr(settings, 'FAGNI_DELIVERY_MIN_FEE', '0')
    try:
        DELIVERY_MIN = Decimal(str(default_delivery_min))
    except Exception:
        DELIVERY_MIN = Decimal('0')

    delivery_fee = DELIVERY_MIN if total_ht > 0 else Decimal('0')

    grand_total = total_ht + service_fee + delivery_fee

    return {
        "total_ht": total_ht,
        "service_fee": service_fee,
        "delivery_fee": delivery_fee,
        "grand_total": grand_total,
    }


# ============================================================
#  UTILITAIRE INTERNE : ANNOTATION DES TOTAUX
# ============================================================
def _annotate_totals(qs):
    """
    Ajoute :
      - items_total = somme (quantity * unit_price) des lignes
      - total_display = total (DB) OU items_total si total est null
    """
    items_sum = Coalesce(
        Sum(Cast(F("items__quantity"), DEC) * Cast(F("items__unit_price"), DEC)),
        Value(0, output_field=DEC),
    )

    return (
        qs.select_related("customer")
        .annotate(items_total=items_sum)
        .annotate(total_display=Coalesce(F("total"), F("items_total")))
    )


# ============================================================
#  LISTE DES COMMANDES
# ============================================================
def orders_list(request):
    """
    Liste des commandes FAGNI avec :
    - filtre par statut (all / pending / in_progress / done / canceled)
    - recherche plein texte (code, client, téléphone)
    - filtre par date de création (du / au)
    - pagination
    """
    qs = (
        Order.objects
        .select_related("customer", "laundry_partner", "delivery_partner")
        .order_by("-created_at")
    )

    # --- Statut ---
    current_status = request.GET.get("status", "all")
    valid_statuses = ("pending", "in_progress", "done", "canceled")

    if current_status in valid_statuses:
        qs = qs.filter(status=current_status)

    # --- Recherche plein texte ---
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(code__icontains=q)
            | Q(customer__name__icontains=q)
            | Q(customer__phone__icontains=q)
        )

    # --- Filtre date (création) ---
    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""

    if date_from:
        df = parse_date(date_from)
        if df:
            qs = qs.filter(created_at__date__gte=df)

    if date_to:
        dt = parse_date(date_to)
        if dt:
            qs = qs.filter(created_at__date__lte=dt)

    # --- Stats globales (tous statuts confondus) ---
    stats_qs = Order.objects.all()
    stats = stats_qs.aggregate(
        total_count=Count("id"),
        done_total=Sum("total_client_ttc", filter=Q(status="done")),
        pending_count=Count("id", filter=Q(status="pending")),
        in_progress_count=Count("id", filter=Q(status="in_progress")),
        done_count=Count("id", filter=Q(status="done")),
        canceled_count=Count("id", filter=Q(status="canceled")),
    )

    # --- Pagination ---
    page = request.GET.get("page") or 1
    paginator = Paginator(qs, 25)  # 25 commandes par page

    try:
        orders_page = paginator.page(page)
    except PageNotAnInteger:
        orders_page = paginator.page(1)
    except EmptyPage:
        orders_page = paginator.page(paginator.num_pages)

    context = {
        "orders": orders_page,
        "page_obj": orders_page,
        "paginator": paginator,
        "is_paginated": orders_page.has_other_pages(),

        "total_count": stats["total_count"] or 0,
        "done_total": stats["done_total"] or 0,
        "pending_count": stats["pending_count"] or 0,
        "in_progress_count": stats["in_progress_count"] or 0,
        "done_count": stats["done_count"] or 0,
        "canceled_count": stats["canceled_count"] or 0,
        "pending_payment_count": Order.objects.filter(payment_status="declared").count(),
        "current_status": current_status or "all",
        # pour que les champs filtres restent pré-remplis
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
    }
    return render(request, "orders/orders_list.html", context)


# ============================================================
#  TABLEAU OPS FAGNI (COLLECTE / LAVAGE / LIVRAISON)
# ============================================================
@login_required
def ops_dashboard(request):
    """
    Tableau de bord opérationnel FAGNI :
    - En attente
    - En cours
    - Terminées

    Important :
    - on affiche le TOTAL CLIENT (total_client_ttc)
    - on recalcule à l'affichage (sans save) pour éviter les incohérences
    - pagination indépendante par colonne (pending / in_progress / done)

    Lot 4.9:
    - Alertes SLA (commandes bloquées)
    - Alertes livreurs (inactive/offline)
    - drivers_json inclut updated_at + server_time

    Lot 4.9.2:
    - filtre OPS par livreur via ?driver_id=<id>
    """
    # ============================================================
    #  DRIVER FILTER (Lot 4.9.2)
    # ============================================================
    raw_driver_id = request.GET.get("driver_id")
    selected_driver = None
    selected_driver_id = None
    try:
        selected_driver_id = int(raw_driver_id) if raw_driver_id else None
    except (TypeError, ValueError):
        selected_driver_id = None

    if selected_driver_id:
        try:
            selected_driver = DeliveryPartner.objects.filter(is_active=True).get(pk=selected_driver_id)
        except DeliveryPartner.DoesNotExist:
            selected_driver = None
            selected_driver_id = None

    # ============================================================
    #  BASE QS (filtrable)
    # ============================================================
    base_qs = (
        Order.objects
        .select_related("customer", "laundry_partner", "delivery_partner")
        .prefetch_related("items__photos")
        .order_by("-created_at")
    )

    # applique le filtre driver si présent
    if selected_driver_id:
        base_qs = base_qs.filter(legs__driver_id=selected_driver_id).distinct()

    pending_qs = base_qs.filter(status="pending")
    in_progress_qs = base_qs.filter(status="in_progress")
    done_qs = base_qs.filter(status="done")

    # --------- Pagination (3 colonnes indépendantes) ----------
    per_page = int(request.GET.get("per_page", 12))

    pending_p = Paginator(pending_qs, per_page)
    progress_p = Paginator(in_progress_qs, per_page)
    done_p = Paginator(done_qs, per_page)

    pending_page_num = request.GET.get("pending_page", 1)
    progress_page_num = request.GET.get("progress_page", 1)
    done_page_num = request.GET.get("done_page", 1)

    try:
        pending_page_obj = pending_p.page(pending_page_num)
    except (PageNotAnInteger, EmptyPage):
        pending_page_obj = pending_p.page(1)

    try:
        progress_page_obj = progress_p.page(progress_page_num)
    except (PageNotAnInteger, EmptyPage):
        progress_page_obj = progress_p.page(1)

    try:
        done_page_obj = done_p.page(done_page_num)
    except (PageNotAnInteger, EmptyPage):
        done_page_obj = done_p.page(1)

    # --- Recompute display amounts (no save) for consistency ---
    def _refresh_amounts(page_obj):
        for o in page_obj.object_list:
            try:
                o.compute_totals(save=False)
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=1536")

            # Auto-heal legs (évite commandes "cassées" sans legs)
            try:
                if o.delivery_partner_id and (o.delivery_fee or 0) > 0 and not o.legs.exists():
                    from orders.models import sync_delivery_legs_for_order
                    sync_delivery_legs_for_order(o)
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=1544")

    _refresh_amounts(pending_page_obj)
    _refresh_amounts(progress_page_obj)
    _refresh_amounts(done_page_obj)

    # Total livré (global) — ici "global" = selon le filtre driver si actif
    try:
        DEC = DecimalField(max_digits=12, decimal_places=2)
        agg = done_qs.aggregate(
            prestation_sum=Coalesce(Sum(Cast("prestation_total", DEC)), Value(0, output_field=DEC)),
            total_sum=Coalesce(Sum(Cast("total", DEC)), Value(0, output_field=DEC)),
        )
        done_total = (agg["prestation_sum"] or Decimal("0"))
        if done_total <= 0:
            done_total = agg["total_sum"] or Decimal("0")
    except Exception:
        done_total = sum((o.prestation_total or o.total or Decimal("0")) for o in done_qs[:500])

    # ============================================================
    #  DRIVERS MAP (initial JSON) — on garde la carte globale
    # ============================================================
    today = timezone.localdate()
    start_week = today - timezone.timedelta(days=today.weekday())

    drivers_qs = DeliveryPartner.objects.filter(
        is_active=True,
        latitude__isnull=False,
        longitude__isnull=False,
    ).order_by("name")

    # Stats semaine (1 requête)
    stats_qs = (
        Order.objects.filter(
            delivery_partner__in=drivers_qs,
            created_at__date__gte=start_week,
            created_at__date__lte=today,
        )
        .values("delivery_partner_id")
        .annotate(
            week_orders=Count("id"),
            week_earnings=Coalesce(Sum("amount_driver_partner"), Decimal("0")),
        )
    )
    stats_map = {
        row["delivery_partner_id"]: {
            "week_orders": int(row["week_orders"] or 0),
            "week_earnings": int(row["week_earnings"] or 0),
        }
        for row in stats_qs
    }

    drivers = []
    for d in drivers_qs:
        try:
            lat = float(d.latitude)
            lng = float(d.longitude)
        except (TypeError, ValueError):
            continue

        st = stats_map.get(d.id, {"week_orders": 0, "week_earnings": 0})
        drivers.append({
            "id": d.id,
            "name": d.name,
            "city": getattr(d, "city", "") or "",
            "latitude": lat,
            "longitude": lng,
            "week_orders": st["week_orders"],
            "week_earnings": st["week_earnings"],
            "updated_at": d.updated_at.isoformat() if getattr(d, "updated_at", None) else None,
        })

    payload = {
        "ok": True,
        "count": len(drivers),
        "drivers": drivers,
        "server_time": timezone.now().isoformat(),
    }
    drivers_json = json.dumps(payload, ensure_ascii=False)

    # --------- helper querystring : conserver les params GET ----------
    def _qs_without(param_name: str) -> str:
        qd = request.GET.copy()
        if param_name in qd:
            qd.pop(param_name)
        s = qd.urlencode()
        return ("&" + s) if s else ""

    # reset url (retire driver_id + focus + highlight, garde le reste)
    qd_reset = request.GET.copy()
    for k in ["driver_id", "focus", "highlight", "pending_page", "progress_page", "done_page"]:
        if k in qd_reset:
            qd_reset.pop(k)
    reset_qs = qd_reset.urlencode()
    reset_url = reverse("orders:ops_dashboard")
    if reset_qs:
        reset_url = f"{reset_url}?{reset_qs}"

    highlight = request.GET.get("highlight")
    try:
        highlight_order_id = int(highlight) if highlight else None
    except (TypeError, ValueError):
        highlight_order_id = None

    # ============================================================
    #  Lot 4.9 — ALERTES (SLA) + LITIGES PESÉE (filtrées si driver actif)
    # ============================================================

    from .models import OrderWeighing, OrderEvidencePhoto

    SLA_PICKUP_H = 2        # pending -> pickup
    SLA_DROPOFF_H = 3       # pickup -> dropoff
    SLA_WASH_H = 48         # dropoff -> wash_done
    SLA_RETURN_H = 3        # wash_done -> return
    SLA_DELIVERED_H = 6     # return -> delivered

    alerts_disputes = []   # 🧷 litiges / pesées contestées
    alerts_sla = []        # 🚨 retards SLA

    def _hours(dt_from, dt_to):
        if not dt_from or not dt_to:
            return None
        return (dt_to - dt_from).total_seconds() / 3600.0

    now = timezone.now()

    # -----------------------------
    # A) LITIGES : PESÉE CONTESTÉE
    # -----------------------------
    weigh_qs = (
        OrderWeighing.objects
        .select_related(
            "order",
            "order__customer",
            "order__laundry_partner",
            "order__delivery_partner",
        )
        .filter(status="disputed")
        .order_by("-confirmed_at", "-updated_at", "-id")
    )

    if selected_driver_id:
        weigh_qs = weigh_qs.filter(order__legs__driver_id=selected_driver_id).distinct()

    weigh_qs = weigh_qs[:200]

    seen_order_ids = set()

    for ow in weigh_qs:
        o = ow.order
        if not o:
            continue

        seen_order_ids.add(o.id)

        latest_issue = (
            OrderEvidencePhoto.objects
            .filter(order=o, kind="issue")
            .order_by("-created_at", "-id")
            .first()
        )

        latest_scale = (
            OrderEvidencePhoto.objects
            .filter(order=o, kind__in=["weighing_scale", "issue"])
            .order_by("-created_at", "-id")
            .first()
        )

        has_scale_photo = bool(latest_scale and getattr(latest_scale, "image", None))
        scale_photo_url = ""
        if has_scale_photo:
            try:
                scale_photo_url = latest_scale.image.url
            except Exception:
                scale_photo_url = ""

        issue_txt = ""
        if latest_issue and getattr(latest_issue, "caption", ""):
            issue_txt = str(latest_issue.caption).strip()

        age_h = _hours(getattr(ow, "confirmed_at", None) or getattr(ow, "updated_at", None), now) or 0

        reason = f"⚠️ Pesée contestée : poids retenu {ow.weight_kg} kg"
        if issue_txt:
            reason += f" — {issue_txt}"

        alerts_disputes.append({
            "order": o,
            "reason": reason,
            "age_h": round(age_h, 1),
            "next_step": "weighing_disputed",
              "resolve_url": reverse("orders:ops_weighing_resolve", args=[o.id]),
            "has_scale_photo": has_scale_photo,
            "scale_photo_url": scale_photo_url,
        })

    # -----------------------------
    # B) ALERTES SLA (pending / in_progress)
    # -----------------------------
    scan_qs = (
        Order.objects
        .select_related("customer", "laundry_partner", "delivery_partner")
        .filter(status__in=["pending", "in_progress"])
        .order_by("-created_at")
    )
    if selected_driver_id:
        scan_qs = scan_qs.filter(legs__driver_id=selected_driver_id).distinct()

    scan_qs = scan_qs[:250]

    for o in scan_qs:
        # si déjà litige pesée => pas de doublon
        if o.id in seen_order_ids:
            continue

        reason = None
        age_h = None
        next_step = None

        if o.status == "pending" and not o.pickup_time:
            age_h = _hours(o.created_at, now) or 0
            if age_h >= SLA_PICKUP_H:
                reason = f"Collecte en retard (>{SLA_PICKUP_H}h)"
                next_step = "pickup"

        if o.status == "in_progress":
            if o.pickup_time and not o.dropoff_time:
                age_h = _hours(o.pickup_time, now) or 0
                if age_h >= SLA_DROPOFF_H:
                    reason = f"Dépôt blanchisserie en retard (>{SLA_DROPOFF_H}h après collecte)"
                    next_step = "dropoff"

            elif o.dropoff_time and not o.wash_complete_time:
                age_h = _hours(o.dropoff_time, now) or 0
                if age_h >= SLA_WASH_H:
                    reason = f"Lavage trop long (>{SLA_WASH_H}h après dépôt)"
                    next_step = "wash_done"

            elif o.wash_complete_time and not o.return_time:
                age_h = _hours(o.wash_complete_time, now) or 0
                if age_h >= SLA_RETURN_H:
                    reason = f"Reprise livreur en retard (>{SLA_RETURN_H}h après lavage)"
                    next_step = "return"

            elif o.return_time and not o.delivered_time:
                age_h = _hours(o.return_time, now) or 0
                if age_h >= SLA_DELIVERED_H:
                    reason = f"Livraison client en retard (>{SLA_DELIVERED_H}h après reprise)"
                    next_step = "delivered"

        if reason:
            alerts_sla.append({
                "order": o,
                "reason": reason,
                "age_h": round(age_h or 0, 1),
                "next_step": next_step,
            })


    # Alertes drivers (offline > 10min) — global
    ACTIVE_MS = 10 * 60  # 10 minutes en secondes
    drivers_offline = []
    for d in drivers_qs:
        if not getattr(d, "updated_at", None):
            drivers_offline.append({"driver": d, "age_min": None})
            continue
        delta = (timezone.now() - d.updated_at).total_seconds()
        if delta > ACTIVE_MS:
            drivers_offline.append({"driver": d, "age_min": int(delta // 60)})

    # ============================================================
    #  DRIVERS LIST (dropdown filtre) — tous les drivers actifs
    # ============================================================
    drivers_list = DeliveryPartner.objects.filter(is_active=True).order_by("name")

    # Total affiché (selon filtre)
    displayed_total_count = base_qs.count()

    context = {
        # counts (selon filtre driver si actif)
        "pending_count": pending_qs.count(),
        "in_progress_count": in_progress_qs.count(),
        "done_count": done_qs.count(),

        "pending_page_obj": pending_page_obj,
        "progress_page_obj": progress_page_obj,
        "done_page_obj": done_page_obj,

        "pending_qs": _qs_without("pending_page"),
        "progress_qs": _qs_without("progress_page"),
        "done_qs": _qs_without("done_page"),

        "per_page": per_page,
        "done_total": done_total,

        # MAP
        "drivers_json": drivers_json,
        "drivers": drivers,

        "highlight_id": highlight_order_id,

        # FILTER UI
        "selected_driver": selected_driver,
        "selected_driver_id": selected_driver_id,
        "ops_reset_url": reset_url,

        # ALERTES
        "alerts_sla_orders": alerts_sla[:20],
        "alerts_sla_count": len(alerts_sla),

        "alerts_disputes_orders": alerts_disputes[:20],
        "alerts_disputes_count": len(alerts_disputes),

        # (optionnel) rétro-compat si tu veux garder l'ancien nom ailleurs
        "alerts_orders": (alerts_disputes + alerts_sla)[:20],
        "alerts_orders_count": (len(alerts_disputes) + len(alerts_sla)), 

        "drivers_offline": drivers_offline[:20],
        "drivers_offline_count": len(drivers_offline),

        # FILTRE DRIVER (Lot 4.9.2/4.9.3)
        "drivers_list": drivers_list,
        "reset_url": reset_url,
        "displayed_total_count": displayed_total_count,
        "pending_payment_count": Order.objects.filter(payment_status="declared").count(),
    }
    return render(request, "orders/ops_dashboard.html", context)


@staff_member_required
def ops_weighing_resolve(request, order_id: int):
    from decimal import Decimal
    """
    OPS — Trancher un litige de pesée (OrderWeighing.status = disputed)
    - choisit un poids final (final_weight_kg)
    - marque la pesée comme "resolved"
    """
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect, render
    from django.urls import reverse
    from django.utils import timezone

    from .models import Order, OrderWeighing, OrderEvidencePhoto

    order = get_object_or_404(Order, pk=order_id)



    ow, _ = OrderWeighing.objects.get_or_create(order=order)
    # Garde-fou : resolution seulement pour les pesees "contestees" (GET et POST)
    if ow.status != "disputed":
        messages.info(request, "Aucun litige de pesée à trancher pour cette commande.")
        return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}&focus=order")

    # Evidence (dernier "issue" et dernier "weighing_scale" si dispo)
    latest_issue = (OrderEvidencePhoto.objects
                    .filter(order=order, kind="issue")
                    .order_by("-created_at", "-id")
                    .first())
    latest_scale = (OrderEvidencePhoto.objects
                    .filter(order=order, kind="weighing_scale")
                    .order_by("-created_at", "-id")
                    .first())

    issue_url = ""
    scale_url = ""
    if latest_issue and getattr(latest_issue, "image", None):
        try:
            issue_url = latest_issue.image.url
        except Exception:
            issue_url = ""
    if latest_scale and getattr(latest_scale, "image", None):
        try:
            scale_url = latest_scale.image.url
        except Exception:
            scale_url = ""

    if request.method == "POST":
        w_raw = (request.POST.get("final_weight_kg") or "").strip().replace(",", ".")
        note = (request.POST.get("resolution_notes") or "").strip()

        try:
            w = Decimal(w_raw)
        except Exception:
            messages.error(request, "Poids final invalide.")
            return redirect(reverse("orders:ops_weighing_resolve", args=[order.id]))

        if w <= 0:
            messages.error(request, "Le poids final doit être > 0.")
            return redirect(reverse("orders:ops_weighing_resolve", args=[order.id]))

        ow.final_weight_kg = w
        ow.status = "resolved"
        ow.resolved_by = getattr(request, "user", None)
        ow.resolved_at = timezone.now()
        ow.resolution_notes = note
        ow.save(update_fields=["final_weight_kg", "status", "resolved_by", "resolved_at", "resolution_notes", "updated_at"])

        messages.success(request, f"Litige pesée résolu : poids final retenu = {w} kg.")
        return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}&focus=order")

    ctx = {
        "order": order,
        "weighing": ow,
        "issue_url": issue_url,
        "scale_url": scale_url,
        "latest_issue": latest_issue,
        "latest_scale": latest_scale,
    }
    return render(request, "orders/ops_weighing_resolve.html", ctx)


# ============================================================
#  📅 OPS PLANNING (Next actions + échéances)
# ============================================================
@login_required
def ops_planning(request):
    """
    Planning OPS = liste des prochaines actions à faire par commande,
    triées et groupées par date d'échéance (SLA).
    - pending -> pickup (SLA_PICKUP_H)
    - pickup -> dropoff (SLA_DROPOFF_H)
    - dropoff -> wash_done (SLA_WASH_H)
    - wash_done -> return (SLA_RETURN_H)
    - return -> delivered (SLA_DELIVERED_H)

    Filtrable par livreur via ?driver_id=<id>
    Fenêtre : ?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD (par défaut J à J+2)
    """
    # mêmes SLA que ton dashboard (cohérence)
    SLA_PICKUP_H = 2
    SLA_DROPOFF_H = 3
    SLA_WASH_H = 48
    SLA_RETURN_H = 3
    SLA_DELIVERED_H = 6

    # ---- filtres dates
    today = timezone.localdate()
    date_from = request.GET.get("date_from") or str(today)
    date_to = request.GET.get("date_to") or str(today + timedelta(days=2))

    df = parse_date(date_from) or today
    dt = parse_date(date_to) or (today + timedelta(days=2))
    if dt < df:
        dt = df

    start_dt = timezone.make_aware(datetime.combine(df, time.min))
    end_dt = timezone.make_aware(datetime.combine(dt, time.max))

    # ---- filtre driver
    raw_driver_id = request.GET.get("driver_id")
    selected_driver = None
    selected_driver_id = None
    try:
        selected_driver_id = int(raw_driver_id) if raw_driver_id else None
    except (TypeError, ValueError):
        selected_driver_id = None

    if selected_driver_id:
        try:
            selected_driver = DeliveryPartner.objects.filter(is_active=True).get(pk=selected_driver_id)
        except DeliveryPartner.DoesNotExist:
            selected_driver = None
            selected_driver_id = None

    qs = (
        Order.objects
        .select_related("customer", "laundry_partner", "delivery_partner")
        .prefetch_related("items")
        .filter(status__in=["pending", "in_progress"])
        .order_by("-created_at")
    )
    if selected_driver_id:
        qs = qs.filter(legs__driver_id=selected_driver_id).distinct()

    now = timezone.now()

    def _next_action(o):
        """
        Retourne dict {action, label, due_at, stage}
        ou None si rien à planifier.
        """
        if o.status == "pending" and not o.pickup_time:
            due = o.created_at + timedelta(hours=SLA_PICKUP_H)
            return {"action": "pickup", "label": "Collecte", "due_at": due, "stage": "pending"}

        if o.pickup_time and not o.dropoff_time:
            due = o.pickup_time + timedelta(hours=SLA_DROPOFF_H)
            return {"action": "dropoff", "label": "Dépôt blanchisserie", "due_at": due, "stage": "in_progress"}

        if o.dropoff_time and not o.wash_complete_time:
            due = o.dropoff_time + timedelta(hours=SLA_WASH_H)
            return {"action": "wash_done", "label": "Lavage terminé", "due_at": due, "stage": "in_progress"}

        if o.wash_complete_time and not o.return_time:
            due = o.wash_complete_time + timedelta(hours=SLA_RETURN_H)
            return {"action": "return", "label": "Reprise livreur", "due_at": due, "stage": "in_progress"}

        if o.return_time and not o.delivered_time:
            due = o.return_time + timedelta(hours=SLA_DELIVERED_H)
            return {"action": "delivered", "label": "Livrée au client", "due_at": due, "stage": "in_progress"}

        return None

    tasks = []
    overdue_count = 0

    # fenêtre large mais maîtrisée
    for o in qs[:500]:
        # refresh affichage sans casser DB
        try:
            o.compute_totals(save=False)
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=2071")

        info = _next_action(o)
        if not info:
            continue

        due_at = info["due_at"]
        if due_at < start_dt or due_at > end_dt:
            continue

        is_overdue = bool(due_at < now)
        if is_overdue:
            overdue_count += 1

        # montants affichés (total client canon)
        total_disp = getattr(o, "total_client_ttc", None)
        if total_disp is None:
            total_disp = getattr(o, "total", None)
        total_disp = total_disp or Decimal("0")

        tasks.append({
            "order": o,
            "action": info["action"],
            "label": info["label"],
            "due_at": due_at,
            "due_date": timezone.localtime(due_at).date(),
            "overdue": is_overdue,
            "total_disp": total_disp,
        })

    # tri par échéance
    tasks.sort(key=lambda x: x["due_at"])

    # group by date
    by_date = {}
    for t in tasks:
        by_date.setdefault(t["due_date"], []).append(t)

    drivers_list = DeliveryPartner.objects.filter(is_active=True).order_by("name")

    context = {
        "date_from": df,
        "date_to": dt,
        "tasks": tasks,
        "by_date": sorted(by_date.items(), key=lambda kv: kv[0]),
        "overdue_count": overdue_count,
        "selected_driver": selected_driver,
        "selected_driver_id": selected_driver_id,
        "drivers_list": drivers_list,
    }
    return render(request, "orders/ops_planning.html", context)


@require_POST
@login_required
def ops_update_step(request, order_id, action):
    """
    Met à jour les timestamps opérationnels :
    - pickup, dropoff, wash_done, return, delivered
    + bascule éventuellement le statut et les legs.

    Sécurité :
    - empêche les étapes dans le mauvais ordre
    - empêche toute étape OPS si la commande n'est pas chiffrée
      (au moins 1 item + total_client_ttc > 0, et si livreur => delivery_fee > 0)
    """
    order = get_object_or_404(Order, pk=order_id)

    # Frontière d'autorité V2 :
    # cette ancienne route OPS pilote des timestamps et des DeliveryLeg
    # selon le workflow V1. Dès qu'une ServiceExecution canonique existe,
    # elle ne doit plus pouvoir piloter l'état opérationnel de la commande.
    from services.services import order_uses_canonical_service_executions
    if order_uses_canonical_service_executions(order=order):
        return JsonResponse({
            "error": "autorite_v2",
            "message": (
                "Cette commande est pilotée par le moteur "
                "ServiceExecution."
            ),
        }, status=409)

    
    # POST only (évite clic direct / appels GET)
    if request.method != "POST":
        messages.error(request, "Action refusée (méthode invalide).")
        return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}")

    allowed = {"pickup", "dropoff", "wash_done", "return", "delivered"}
    if action not in allowed:
        messages.error(request, "Action OPS invalide.")
        return redirect("orders:ops_dashboard")

    mapping = {
        "pickup": ("pickup_time", "Collecte validée"),
        "dropoff": ("dropoff_time", "Dépôt blanchisserie validé"),
        "wash_done": ("wash_complete_time", "Lavage terminé validé"),
        "return": ("return_time", "Reprise livreur validée"),
        "delivered": ("delivered_time", "Livraison client validée"),
    }
    field_name, label = mapping[action]

    prerequisites = {
        "pickup": [],
        "dropoff": ["pickup_time"],
        "wash_done": ["dropoff_time"],
        "return": ["wash_complete_time"],
        "delivered": ["return_time"],
    }

    # Déjà fait ?
    already_done = bool(getattr(order, field_name, None))
    if already_done and action != "wash_done":
        messages.warning(request, f"Déjà fait : {label}.")
        return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}")
    # 🔒 Garde-fou métier : les étapes liées au pressing exigent un pressing assigné
    if action in {"dropoff", "wash_done", "return"} and not getattr(order, "laundry_partner_id", None):
        messages.error(request, "Impossible : aucun pressing n’est assigné à cette commande.")
        return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}")

    # Ordre des étapes
    missing = [f for f in prerequisites[action] if not getattr(order, f, None)]
    if missing:
        messages.error(request, "Impossible : étape précédente non validée.")
        return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}")

    skip_pricing_guard = (action == "wash_done")

    # ------------------------------------------------------------
    # 1) 🔁 SYNC FINANCIER (DB doit être la source de vérité)
    # ------------------------------------------------------------
    try:
        order.compute_totals(save=True)  # update_financials(save=True)
    except Exception:
        # On ne bloque pas si le recalcul plante, mais on traitera comme non chiffré
        pass

    # ------------------------------------------------------------
    # 2) 🔒 GARDE-FOU : commande doit être chiffrée
    # ------------------------------------------------------------
    has_items = False
    try:
        has_items = order.items.exists()
    except Exception:
        has_items = False

    total_client = getattr(order, "total_client_ttc", None)
    if total_client is None:
        total_client = getattr(order, "total", None)
    total_client = total_client or Decimal("0")
    if not isinstance(total_client, Decimal):
        try:
            total_client = Decimal(str(total_client))
        except Exception:
            total_client = Decimal("0")

    delivery_fee = getattr(order, "delivery_fee", None) or Decimal("0")
    if not isinstance(delivery_fee, Decimal):
        try:
            delivery_fee = Decimal(str(delivery_fee))
        except Exception:
            delivery_fee = Decimal("0")

    has_driver = bool(order.delivery_partner_id)

    if (not skip_pricing_guard) and (not has_items):
        messages.error(
            request,
            "Commande non chiffrée (aucun article / prestation). Ajoute au moins 1 prestation avant de valider l'OPS."
        )
        return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}")

    if (not skip_pricing_guard) and (total_client <= 0):
        messages.error(
            request,
            "Commande non chiffrée (total client = 0). Ajoute au moins 1 prestation / recalcule les montants."
        )
        return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}")

    if (not skip_pricing_guard) and has_driver and delivery_fee <= 0:
        messages.error(
            request,
            "Commande non chiffrée (livreur assigné mais frais de livraison = 0). Recalcule/ajuste la livraison."
        )
        return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}")

    if (not skip_pricing_guard) and (not has_driver) and delivery_fee > 0:
        messages.error(
            request,
            "Commande incohérente (frais de livraison > 0 sans livreur). Assigne un livreur ou remets la livraison à 0."
        )
        return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}")

    # ------------------------------------------------------------
    # 3) (optionnel) sync legs si besoin (après recalcul)
    # ------------------------------------------------------------
    try:
        from orders.models import DeliveryLeg, sync_delivery_legs_for_order
        if not DeliveryLeg.objects.filter(order=order).exclude(status="canceled").exists():
            sync_delivery_legs_for_order(order)
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=2258")

    # ------------------------------------------------------------
    # 4) ✅ VALIDER L'ÉTAPE
    # ------------------------------------------------------------
    # Ne pas écraser wash_complete_time si déjà renseigné (cas wash_done)
    if (action != "wash_done") or (not already_done):
        setattr(order, field_name, timezone.now())
    # Status + legs
    if action == "pickup":
        if order.status == "pending":
            order.status = "in_progress"

        for leg in order.legs.filter(leg_type="pickup"):
            update_leg_status(leg, "start", user=request.user)

    elif action == "dropoff":
        for leg in order.legs.filter(leg_type="pickup"):
            update_leg_status(leg, "finish", user=request.user)

    elif action == "wash_done":
        # OPS: wash_done => activer / réactiver la jambe return (pending), sans auto-start
        try:
            from orders.models import DeliveryLeg
    
            # 1) s'assurer qu'un return existe
            DeliveryLeg.objects.get_or_create(
                order=order,
                leg_type="return",
                defaults={
                    "status": "pending",
                    "driver": order.delivery_partner if getattr(order, "delivery_partner_id", None) else None,
                },
            )
    
            # 2) assigner driver si besoin
            if getattr(order, "delivery_partner_id", None):
                DeliveryLeg.objects.filter(order=order, leg_type="return", driver__isnull=True).update(driver=order.delivery_partner)
    
            # 3) protéger legs payés (si wallets dispo)
            paid_leg_ids = set()
            try:
                from wallets.models import WalletTransaction
                paid_leg_ids = set(
                    WalletTransaction.objects.filter(
                        order_id=order.id,
                        wallet__owner_type="driver",
                        type="payout",
                        direction="in",
                    )
                    .exclude(leg_id__isnull=True)
                    .values_list("leg_id", flat=True)
                )
            except Exception:
                paid_leg_ids = set()
    
            # 4) forcer return => pending (sauf done, sauf payé)
            qs = DeliveryLeg.objects.filter(order=order, leg_type="return").exclude(id__in=paid_leg_ids).exclude(status="done")
            n = qs.update(status="pending")
            logger.info("OPS wash_done: forced return pending n=%s order=%s", n, getattr(order, "id", None))
        except Exception:
            logger.exception("OPS wash_done failed order=%s", getattr(order, "id", None))

    elif action == "return":
        for leg in order.legs.filter(leg_type="return"):
            update_leg_status(leg, "start", user=request.user)

    elif action == "delivered":
        # 1) Tenter de finir toutes les jambes non terminées
        for leg in order.legs.exclude(status__in=("done", "canceled")):
            update_leg_status(leg, "finish", user=request.user)

        # 2) Statut commande: source de vérité = sync_order_status_from_legs (pickup+return)
        try:
            from orders.models import sync_order_status_from_legs
            sync_order_status_from_legs(order, save=False)
        except Exception:
            # fallback safe: ne pas forcer DONE ici
            if order.status == "pending":
                order.status = "in_progress"

    update_fields = {field_name, "status", "updated_at"}
    order.save(update_fields=list(update_fields))
    if action == "wash_done" and already_done:
        messages.success(request, "✅ Return réactivé (pending).")
    else:
        messages.success(request, label)

    return redirect(f"{reverse('orders:ops_dashboard')}?highlight={order.id}")


def _ensure_delivery_legs_for_order(order):
    """
    Crée les legs pickup/return si la commande est en mode livraison et n'a aucun leg.
    Idempotent.

    - Crée 2 legs (pickup + return) en status="pending"
    - Remplit client_fee_share / driver_amount / fagni_margin (NOT NULL)
    - Pas de payout rétro (on ne met jamais done)
    """
    from decimal import Decimal, ROUND_HALF_UP
    from orders.models import DeliveryLeg

    def _round_fcfa(v):
        if v is None:
            v = Decimal("0")
        if not isinstance(v, Decimal):
            v = Decimal(str(v))
        # FCFA arrondi à l'entier (stocké en decimal(10,2))
        return v.quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    delivery_fee = Decimal(str(getattr(order, "delivery_fee", 0) or 0))
    delivery_mode = getattr(order, "delivery_mode", None)
    amount_driver = Decimal(str(getattr(order, "amount_driver_partner", 0) or 0))
    margin_total = Decimal(str(getattr(order, "logistic_margin", 0) or 0))

    is_delivery = (
        bool(getattr(order, "delivery_partner_id", None))
        or delivery_fee > 0
        or (delivery_mode not in (None, "", "none"))
        or amount_driver > 0
    )
    if not is_delivery:
        return 0

    # idempotent
    if DeliveryLeg.objects.filter(order=order).exists():
        return 0

    driver = getattr(order, "delivery_partner", None)

    client_1 = _round_fcfa(delivery_fee / 2)
    client_2 = _round_fcfa(delivery_fee - client_1)

    driver_1 = _round_fcfa(amount_driver / 2)
    driver_2 = _round_fcfa(amount_driver - driver_1)

    margin_1 = _round_fcfa(margin_total / 2)
    margin_2 = _round_fcfa(margin_total - margin_1)

    DeliveryLeg.objects.create(
        order=order,
        leg_type="pickup",
        status="pending",
        driver=driver,
        client_fee_share=client_1,
        driver_amount=driver_1,
        fagni_margin=margin_1,
    )
    DeliveryLeg.objects.create(
        order=order,
        leg_type="return",
        status="pending",
        driver=driver,
        client_fee_share=client_2,
        driver_amount=driver_2,
        fagni_margin=margin_2,
    )
    return 2


@login_required
@require_POST
def order_mark_paid(request, order_id):
    """
    ✅ Encaisser (Back-office)
    - Marque la commande payée
    - Aligne amount_paid sur la source unique (_build_invoice_context)
    - Verrouille (via status paiement + UI) ; évite double encaissement
    - Distribue revenus : blanchisserie + wallet interne (distribute_order_revenues)
    - Déclenche payouts livreur UNIQUEMENT pour les legs déjà done (anti-doublon par leg)
    - Met FNE en pending (si activée / pas déjà envoyée)
    """

    if not request.user.is_staff:
        return HttpResponseForbidden("Accès refusé.")

    order = get_object_or_404(
        Order.objects.select_related("customer", "laundry_partner", "delivery_partner"),
        pk=order_id
    )

    # Déjà payée => rien à faire
    if getattr(order, "payment_status", None) == "paid":
        messages.info(request, "Commande déjà payée.")
        return redirect("orders:detail", order_id=order.id)

    # Source unique montants (canon)
    ctx = _build_invoice_context(order)
    total_ttc_client = ctx.get("total_ttc_client") or 0

    from decimal import Decimal

    try:
        total_ctx = Decimal(str(total_ttc_client or 0))
    except Exception:
        total_ctx = Decimal("0")

    total_db = Decimal(str(getattr(order, "total_client_ttc", 0) or 0))
    total_eff = total_ctx if total_ctx > 0 else total_db

    if total_eff <= 0:
        messages.error(request, "Impossible d'encaisser : total client = 0 (commande non chiffrée). Ajoute les prestations / calcule les montants avant.")
        return redirect("orders:detail", order_id=order.id)

    # Saisie optionnelle
    method = (
        request.POST.get("payment_method") or ""
    ).strip().lower() or None
    ref = (request.POST.get("payment_reference") or "").strip() or None

    # Sécurité pilote :
    # cette route générique ne permet que l'encaissement CASH réellement
    # constaté par un opérateur habilité.
    #
    # Wave, wallet et futurs PSP ont leurs workflows dédiés.
    if method != "cash":
        messages.error(
            request,
            (
                "Encaissement générique refusé. "
                "Cette action est réservée au CASH réellement encaissé. "
                "Wave doit passer par le workflow sécurisé de vérification."
            ),
        )
        return redirect("orders:detail", order_id=order.id)

    from decimal import Decimal

    # ✅ SAFETY: si la commande est en mode livraison et qu'elle n'a pas de legs, on les crée (idempotent)
    try:
        created = _ensure_delivery_legs_for_order(order)
        if created:
            messages.info(request, f"Legs logistiques créés automatiquement ({created}) pour cette commande.")
    except Exception as e:
        messages.warning(request, f"Attention: impossible d'initialiser les legs logistiques: {e}")


    # 1) Paiement via source-of-truth (Payment + sync)
    try:
        # Total canon (Option A)
        total_ttc = Decimal(str(getattr(order, "total_client_ttc", 0) or 0))

        # Si ctx fournit mieux, on garde en fallback
        try:
            total_ctx = Decimal(str(total_ttc_client or 0))
            if total_ctx > 0:
                total_ttc = total_ctx
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=2489")

        already_paid = Decimal(str(getattr(order, "amount_paid", 0) or 0))
        remaining = total_ttc - already_paid
        if remaining < 0:
            remaining = Decimal("0")

        # ✅ Montant saisi (optionnel) : acompte / solde
        raw_amt = (request.POST.get("payment_amount") or "").strip()
        pay_amt = None
        if raw_amt:
            try:
                pay_amt = Decimal(str(raw_amt).replace(" ", "").replace("\u202f", ""))
            except Exception:
                pay_amt = None


        # Référence idempotente si l'utilisateur ne saisit rien
        ref_eff = ref or f"BO-PAID-{order.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}"

        if remaining > 0:
            # channel = method (ex: cash, om, moov, card…), sinon "manual"
            order.add_payment(
                amount=(min(pay_amt, remaining) if pay_amt is not None else remaining),
                channel=(method or "manual"),
                reference=ref_eff,
                source="backoffice",
                confirmed_by=request.user,
                save=True,
            )
        else:
            # Rien à ajouter, mais on resynchronise au cas où
            order.sync_payment_status_from_payments(save=True)

    except Exception as e:
        messages.warning(request, f"Paiement: erreur de sync via payments: {e}")

    # (Optionnel) garder la trace du moyen / référence sur la commande
    try:
        if method is not None:
            order.payment_method = method
        if ref is not None:
            order.payment_reference = ref
        order.save(update_fields=["payment_method", "payment_reference", "updated_at"])
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=2534")

    # 🔒 Stop si la commande n'est pas devenue "paid" ou "partial" après sync
    if getattr(order, "payment_status", "unpaid") not in ("paid", "partial"):
        messages.error(request, "Paiement non pris en compte (statut inchangé).")
        return redirect("orders:detail", order_id=order.id)

    # 2) Facture / numéros / financials (si ton update_financials gère invoice_number)
    try:
        order.update_financials(save=True)
    except Exception:
        # au pire, on sauvegarde quand même le paiement
        order.save()

    # 3) FNE : passer en pending si activée (et pas déjà "sent/accepted/rejected/error")
    try:
        current_fne = getattr(order, "fne_status", None) or "disabled"
        if current_fne not in ("disabled", "sent", "accepted", "rejected", "error"):
            order.fne_status = "pending"
            order.save(update_fields=["fne_status", "updated_at"])
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=2555")

    # ✅ Refresh order (assure état paiement à jour avant distribution/payout)
    try:
        order.refresh_from_db()
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=2561")

    # ✅ Ne distribuer / payer le livreur QUE si la commande est soldée
    if getattr(order, "payment_status", None) != "paid":
        messages.info(request, "Paiement enregistré. Commande non soldée (payout et distribution reportés).")
        return redirect("orders:detail", order_id=order.id)

    # 4) Distribution des revenus (blanchisserie + wallet interne)
    #    (anti-doublon via order.wallets_distributed + contraintes uniques côté tx)
    try:
        distribute_order_revenues(order, recompute=True, force=False)
    except Exception as e:
        # On ne bloque pas l'encaissement si distribution échoue, mais on le signale
        messages.warning(request, f"Paiement OK, mais distribution wallets incomplète : {e}")
    # 5) ✅ Payout livreur
    #    Géré automatiquement par les signals :
    #    - DeliveryLeg -> done
    #    - Order -> paid (si legs done avant paiement)
    messages.success(request, "Paiement OK. Payout livreur géré automatiquement.")
    return redirect("orders:detail", order_id=order.id)


@login_required
def ops_drivers_live(request):
    """
    Lot 4.9 — JSON LIVE pour la carte Leaflet du OPS Dashboard.
    Retourne la liste des DeliveryPartner actifs avec latitude/longitude + stats semaine.
    + updated_at + server_time
    """
    today = timezone.localdate()
    start_week = today - timedelta(days=today.weekday())

    drivers_qs = DeliveryPartner.objects.filter(
        is_active=True,
        latitude__isnull=False,
        longitude__isnull=False,
    ).order_by("name")

    stats_qs = (
        Order.objects.filter(
            delivery_partner__in=drivers_qs,
            created_at__date__gte=start_week,
            created_at__date__lte=today,
        )
        .values("delivery_partner_id")
        .annotate(
            week_orders=Count("id"),
            week_earnings=Coalesce(Sum("amount_driver_partner"), Decimal("0")),
        )
    )

    stats_map = {
        row["delivery_partner_id"]: {
            "week_orders": int(row["week_orders"] or 0),
            "week_earnings": int(row["week_earnings"] or 0),
        }
        for row in stats_qs
    }

    payload = []
    for d in drivers_qs:
        st = stats_map.get(d.id, {"week_orders": 0, "week_earnings": 0})

        updated_at = None
        if getattr(d, "updated_at", None):
            try:
                updated_at = timezone.localtime(d.updated_at).isoformat()
            except Exception:
                updated_at = None

        payload.append(
            {
                "id": d.id,
                "name": getattr(d, "name", "") or "",
                "phone": getattr(d, "phone", "") or "",
                "lat": float(d.latitude),
                "lng": float(d.longitude),
                "is_active": bool(getattr(d, "is_active", True)),
                "updated_at": updated_at,
                "week_orders": st["week_orders"],
                "week_earnings": st["week_earnings"],
            }
        )

    return JsonResponse(
        {
            "drivers": payload,
            "count": len(payload),
            "server_time": timezone.localtime(timezone.now()).isoformat(),
        }
    )


# ============================================================
#  DASHBOARD GLOBAL DES COMMANDES
# ============================================================
def orders_dashboard(request):
    base_qs = _annotate_totals(Order.objects.all())

    total_count = base_qs.count()
    pending_count = base_qs.filter(status="pending").count()
    in_progress_count = base_qs.filter(status="in_progress").count()
    done_count = base_qs.filter(status="done").count()
    canceled_count = base_qs.filter(status="canceled").count()

    done_total = (
        base_qs.filter(status="done").aggregate(total=Sum("total"))["total"]
        or Decimal("0")
    )

    if total_count > 0:
        completion_rate = (done_count / total_count) * 100
        cancel_rate = (canceled_count / total_count) * 100
    else:
        completion_rate = 0
        cancel_rate = 0

    if done_count > 0:
        avg_ticket_done = done_total / done_count
    else:
        avg_ticket_done = Decimal("0")

    if total_count > 0:
        avg_ticket_all = done_total / total_count
    else:
        avg_ticket_all = Decimal("0")

    today = timezone.localdate()
    today_qs = base_qs.filter(created_at__date=today)
    today_count = today_qs.count()
    today_done_total = (
        today_qs.filter(status="done").aggregate(total=Sum("total"))["total"]
        or Decimal("0")
    )

    start_week = today - timedelta(days=today.weekday())
    start_month = today.replace(day=1)

    week_qs = base_qs.filter(created_at__date__gte=start_week)
    week_count = week_qs.count()
    week_done_total = (
        week_qs.filter(status="done").aggregate(total=Sum("total"))["total"]
        or Decimal("0")
    )

    month_qs = base_qs.filter(created_at__date__gte=start_month)
    month_count = month_qs.count()
    month_done_total = (
        month_qs.filter(status="done").aggregate(total=Sum("total"))["total"]
        or Decimal("0")
    )

    days_param = request.GET.get("days", "7")
    try:
        days_int = int(days_param)
    except (TypeError, ValueError):
        days_int = 7
    if days_int not in (7, 14, 30):
        days_int = 7

    chart_labels = []
    chart_orders = []
    chart_totals = []

    for i in range(days_int - 1, -1, -1):
        day = today - timedelta(days=i)
        day_qs = base_qs.filter(created_at__date=day)

        chart_labels.append(day.strftime("%d/%m"))
        chart_orders.append(day_qs.count())

        day_total = (
            day_qs.filter(status="done").aggregate(total=Sum("total"))["total"]
            or Decimal("0")
        )
        chart_totals.append(float(day_total))

    top_clients = (
        Order.objects
        .filter(status="done", customer__isnull=False)
        .values("customer__name", "customer__phone")
        .annotate(
            total_spent=Sum("total"),
            orders_count=Count("id"),
        )
        .order_by("-total_spent")[:5]
    )

    top_items = (
        OrderItem.objects
        .filter(order__status="done")
        .values("designation")
        .annotate(
            total_qty=Sum("quantity"),
            total_amount=Sum("total"),
        )
        .order_by("-total_qty")[:5]
    )

    context = {
        "total_count": total_count,
        "pending_count": pending_count,
        "in_progress_count": in_progress_count,
        "done_count": done_count,
        "canceled_count": canceled_count,
        "done_total": done_total,
        "today": today,
        "today_count": today_count,
        "today_done_total": today_done_total,
        "start_week": start_week,
        "start_month": start_month,
        "week_count": week_count,
        "week_done_total": week_done_total,
        "month_count": month_count,
        "month_done_total": month_done_total,
        "current_days": days_int,
        "chart_labels": chart_labels,
        "chart_orders": chart_orders,
        "chart_totals": chart_totals,
        "top_clients": top_clients,
        "top_items": top_items,
        "completion_rate": completion_rate,
        "cancel_rate": cancel_rate,
        "avg_ticket_done": avg_ticket_done,
        "avg_ticket_all": avg_ticket_all,
    }
    return render(request, "orders/dashboard.html", context)


@login_required
def export_finance_xlsx(request):
    """
    Export Excel du dashboard financier – SOURCE DE VÉRITÉ :
    compute_order_amounts() + TVA sur fagni_revenue_ht.
    - Onglet 1 : Synthèse
    - Onglet 2 : Détail (mêmes colonnes que dashboard)
    """
    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""
    status_filter = request.GET.get("status") or "all"   # paid / partial / unpaid / all
    min_amount_input = request.GET.get("min_amount") or ""

    qs = (
        Order.objects
        .select_related("customer", "laundry_partner", "delivery_partner")
        .order_by("-created_at")
    )

    # Période
    if date_from:
        df = parse_date(date_from)
        if df:
            qs = qs.filter(created_at__date__gte=df)

    if date_to:
        dt = parse_date(date_to)
        if dt:
            qs = qs.filter(created_at__date__lte=dt)

    raw_orders = list(qs[:500])

    cfg = get_pricing_settings()
    vat_rate = _get_vat_rate_percent(cfg)

    def q(x):
        try:
            return _q(x)
        except Exception:
            try:
                return Decimal(str(x))
            except Exception:
                return Decimal("0")

    # Filtre montant mini
    try:
        min_amount = Decimal(min_amount_input) if min_amount_input else Decimal("0")
    except Exception:
        min_amount = Decimal("0")

    rows = []
    total_ca = Decimal("0")
    total_prestations = Decimal("0")
    total_service = Decimal("0")
    total_delivery = Decimal("0")
    total_express = Decimal("0")
    total_vat_fagni = Decimal("0")
    total_fagni_ttc = Decimal("0")
    total_logistic_margin = Decimal("0")
    total_laundry = Decimal("0")
    total_driver = Decimal("0")

    for o in raw_orders:
        amounts = compute_order_amounts(o)

        subtotal = q(amounts.get("subtotal", 0))
        service_fee_ht = q(amounts.get("service_fee_ht", 0))
        delivery_fee_client = q(amounts.get("delivery_fee_client", 0))
        express = q(
            amounts.get("express_for_client", 0)
            or amounts.get("express_extra_fee_client", 0)
            or amounts.get("express_fee_client", 0)
            or getattr(o, "express_extra_fee", 0)
            or 0
        )

        fagni_revenue_ht = q(amounts.get("fagni_revenue_ht", 0))
        vat_fagni = q((fagni_revenue_ht * q(vat_rate)) / Decimal("100"))

        total_ht_client = q(subtotal + service_fee_ht + delivery_fee_client + express)
        total_ttc_client = q(total_ht_client + vat_fagni)

        logistic_margin = q(
            amounts.get("logistic_margin", 0)
            or getattr(o, "logistic_margin", 0)
            or 0
        )
        fagni_ttc = q(fagni_revenue_ht + vat_fagni)

        amount_laundry = q(
            amounts.get("amount_laundry_partner", 0)
            or getattr(o, "amount_laundry_partner", 0)
            or 0
        )
        amount_driver = q(
            amounts.get("amount_driver_partner", 0)
            or getattr(o, "amount_driver_partner", 0)
            or amounts.get("driver_income", 0)
            or getattr(o, "driver_logistic_cost", 0)
            or 0
        )

        # filtres statut paiement "export"
        paid = q(getattr(o, "amount_paid", 0))
        due = q(getattr(o, "amount_due", 0))
        is_fully_paid = (due <= 0)

        if status_filter == "paid" and not is_fully_paid:
            continue
        if status_filter == "partial" and not (paid > 0 and due > 0):
            continue
        if status_filter == "unpaid" and not (paid == 0 and due > 0):
            continue

        if min_amount > 0 and total_ttc_client < min_amount:
            continue

        rows.append({
            "date": o.created_at,
            "code": o.code or str(o.id),
            "client": (o.customer.name if o.customer else ""),
            "laundry": (o.laundry_partner.name if o.laundry_partner else ""),
            "payment_status": o.payment_status or "",
            "total_client": total_ttc_client,
            "prestations": subtotal,
            "service_fee": service_fee_ht,
            "delivery": delivery_fee_client,
            "express": express,
            "vat_fagni": vat_fagni,
            "fagni_ttc": fagni_ttc,
            "logistic_margin": logistic_margin,
            "amount_laundry": amount_laundry,
            "amount_driver": amount_driver,
            "partners_total": q(amount_laundry + amount_driver),
            "paid": paid,
            "due": due,
        })

        total_ca += total_ttc_client
        total_prestations += subtotal
        total_service += service_fee_ht
        total_delivery += delivery_fee_client
        total_express += express
        total_vat_fagni += vat_fagni
        total_fagni_ttc += fagni_ttc
        total_logistic_margin += logistic_margin
        total_laundry += amount_laundry
        total_driver += amount_driver

    # ---------- Excel ----------
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Synthèse"

    title_font = Font(size=16, bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="0056B3")
    section_fill = PatternFill("solid", fgColor="FF7A00")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Titre
    ws1["A1"] = "FAGNI – Dashboard financier (export)"
    ws1["A1"].font = title_font
    ws1["A1"].fill = header_fill
    ws1.merge_cells("A1:D1")

    ws1["A2"] = "Période"
    ws1["A2"].font = label_font
    ws1["B2"] = f"{date_from or '—'} → {date_to or '—'}"

    ws1["A4"] = "KPI"
    ws1["A4"].font = header_font
    ws1["A4"].fill = section_fill
    ws1.merge_cells("A4:D4")

    kpis = [
        ("Nb commandes", len(rows)),
        ("CA total client (TTC)", total_ca),
        ("Prestations (HT)", total_prestations),
        ("Service fee (HT)", total_service),
        ("Livraison client", total_delivery),
        ("Express", total_express),
        ("TVA FAGNI", total_vat_fagni),
        ("Revenu FAGNI (TTC)", total_fagni_ttc),
        ("Marge logistique", total_logistic_margin),
        ("Montant blanchisserie", total_laundry),
        ("Montant livreur", total_driver),
        ("Total partenaires", total_laundry + total_driver),
    ]

    r = 5
    for label, val in kpis:
        ws1[f"A{r}"] = label
        ws1[f"A{r}"].font = label_font
        ws1[f"B{r}"] = float(val) if isinstance(val, Decimal) else val
        ws1[f"B{r}"].alignment = right
        r += 1

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 22

    # Onglet détail
    ws2 = wb.create_sheet("Commandes")
    headers = [
        "Date", "Code", "Client", "Blanchisserie", "Paiement",
        "Total client TTC", "Prestations HT", "Service fee HT", "Livraison", "Express",
        "TVA FAGNI", "FAGNI TTC", "Marge logistique",
        "Montant blanchisserie", "Montant livreur", "Total partenaires",
        "Payé", "Reste dû",
    ]
    ws2.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row in rows:
        ws2.append([
            row["date"].strftime("%d/%m/%Y %H:%M") if row["date"] else "",
            row["code"],
            row["client"],
            row["laundry"],
            row["payment_status"],
            float(row["total_client"]),
            float(row["prestations"]),
            float(row["service_fee"]),
            float(row["delivery"]),
            float(row["express"]),
            float(row["vat_fagni"]),
            float(row["fagni_ttc"]),
            float(row["logistic_margin"]),
            float(row["amount_laundry"]),
            float(row["amount_driver"]),
            float(row["partners_total"]),
            float(row["paid"]),
            float(row["due"]),
        ])

    # tailles
    widths = [18, 14, 22, 22, 12, 16, 14, 14, 12, 12, 12, 12, 14, 18, 14, 16, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = "FAGNI_finance_dashboard.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _safe_dec(value):
    """
    Convertit en Decimal en gérant None / types bizarres.
    Renvoie Decimal('0') si la valeur n'est pas convertible.
    """
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")


def _order_effective_total(order):
    """
    Montant réellement facturé au client pour une commande.
    On essaie dans l'ordre :
    - total_client_ttc (champ dédié si présent)
    - grand_total / total_ttc / total_ht / total
    - sinon on reconstruit : items + service + livraison
    """
    # 1) Champs agrégés
    for field in ["total_client_ttc", "grand_total", "total_ttc", "total_ht", "total"]:
        if hasattr(order, field):
            d = _safe_dec(getattr(order, field))
            if d > 0:
                return d

    # 2) Fallback : recomposer à partir des items + frais
    items_sum = Decimal("0")
    try:
        for it in order.items.all():
            line_total = getattr(it, "line_total", None)
            if line_total is not None:
                items_sum += _safe_dec(line_total)
            else:
                q = _safe_dec(getattr(it, "quantity", 0))
                p = _safe_dec(getattr(it, "unit_price", 0))
                items_sum += q * p
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=3092")

    service_fee = _safe_dec(getattr(order, "service_fee", None))
    delivery_fee = _safe_dec(getattr(order, "delivery_fee", None))
    return items_sum + service_fee + delivery_fee


DEC_ZERO = Decimal("0")


def _dec_or_zero(val):
    """
    Convertit une valeur en Decimal de façon robuste.
    Retourne Decimal('0') si la conversion échoue.
    """
    if val is None:
        return DEC_ZERO
    if isinstance(val, Decimal):
        return val
    try:
        return Decimal(str(val))
    except (InvalidOperation, TypeError, ValueError):
        return DEC_ZERO


def _compute_order_pricing(order):
    """
    Bridge legacy -> pricing engine canonique.
    CORRECTIF 6 juillet 2026 : compute_order_pricing(order) etait appele avec l'objet
    commande entier au lieu d'un nombre d'articles (TypeError systematique), provoquant
    un 500 sur la page de paiement client /pay/wave/. Lecture directe des montants deja
    verrouilles sur la commande (ADR-001) au lieu d'un recalcul via un moteur incompatible.
    """
    total_locked = _safe_dec(getattr(order, "total_client_ttc", 0))
    service_fee = _safe_dec(getattr(order, "service_fee", 0))
    delivery_fee = _safe_dec(getattr(order, "delivery_fee", 0))
    amount_driver = _safe_dec(getattr(order, "amount_driver_partner", 0))
    amount_laundry = _safe_dec(getattr(order, "amount_laundry_partner", 0))
    fagni_revenue = _safe_dec(getattr(order, "fagni_revenue_ht", 0))
    prestation_total = total_locked - delivery_fee - service_fee

    child_referral_discount = get_child_referral_discount_amount(order)
    coupon_discount = _safe_dec(getattr(order, "coupon_discount_applied", 0))
    try:
        total_client_adjusted = Decimal(str(total_locked or 0)) - Decimal(str(child_referral_discount or 0)) - Decimal(str(coupon_discount or 0))
    except Exception:
        total_client_adjusted = Decimal(str(total_locked or 0))

    if total_client_adjusted < 0:
        total_client_adjusted = Decimal("0")

    total_client_adjusted = total_client_adjusted.quantize(Decimal("1"))

    return {
        "items_total": prestation_total,
        "service_fee": service_fee,
        "delivery_fee": delivery_fee,
        "total_client": total_client_adjusted,
        "total_client_ttc": total_client_adjusted,
        "child_referral_discount": child_referral_discount,
        "coupon_discount": coupon_discount,
        "driver_income": amount_driver,
        "vat_fagni": Decimal("0"),
        "express_extra_fee": Decimal("0"),
        "fagni_revenue_ht": fagni_revenue,
        "fagni_revenue_ttc": fagni_revenue,
        "laundry_amount": amount_laundry,
        "logistic_margin": Decimal("0"),
    }

def _guess_delivery_fee(order, total_client=None, items_total=None, service_fee=None):
    """
    Essaie de deviner le montant de la livraison pour une commande.

    Logique :
    livraison ≈ total_client_ttc - prestations - service FAGNI - TVA FAGNI

    - total_client : TTC client (total_client_ttc, total ou annotation)
    - items_total : sous-total prestations (annotation ou recalcul)
    - service_fee : service FAGNI (order.service_fee)
    Utilise aussi order.vat_fagni si disponible.
    """
    # 1) TTC client
    total_client = _safe_dec(
        total_client
        or getattr(order, "total_client_ttc", None)
        or getattr(order, "total", None)
    )

    # 2) Sous-total prestations
    if items_total is None:
        items_total = _safe_dec(getattr(order, "items_total", None))
    else:
        items_total = _safe_dec(items_total)

    # 3) Service FAGNI et TVA FAGNI
    vat_fagni = _safe_dec(getattr(order, "vat_fagni", None))
    if service_fee is None:
        service_fee = _safe_dec(getattr(order, "service_fee", None))
    else:
        service_fee = _safe_dec(service_fee)

    if total_client <= 0:
        return Decimal("0")

    # livraison = TTC - prestations - service - TVA
    delivery_guess = total_client - items_total - service_fee - vat_fagni
    if delivery_guess > 0:
        return delivery_guess

    return Decimal("0")


def _order_driver_income(order, delivery_fee=None):
    """
    Calcule le revenu du livreur pour une commande.

    Priorité :
    1) order.amount_driver_partner si > 0
    2) order.driver_logistic_cost si > 0
    3) somme des DeliveryLeg.driver_amount si > 0
    4) delivery_fee - logistic_margin (après avoir deviné la livraison si nécessaire)
    Sinon, 0.
    """
    # 1) Montant partenaire livreur direct
    if hasattr(order, "amount_driver_partner") and order.amount_driver_partner is not None:
        d = _safe_dec(order.amount_driver_partner)
        if d > 0:
            return d

    # 2) Coût logistique payé au livreur
    if hasattr(order, "driver_logistic_cost") and order.driver_logistic_cost is not None:
        d = _safe_dec(order.driver_logistic_cost)
        if d > 0:
            return d

    # 3) Fallback : legs
    legs_rel = getattr(order, "legs", None)
    if legs_rel is not None:
        try:
            total_legs = Decimal("0")
            for leg in legs_rel.all():
                da = _safe_dec(getattr(leg, "driver_amount", None))
                total_legs += da
            if total_legs > 0:
                return total_legs
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=3227")

    # 4) Gestion livraison
    if delivery_fee is None:
        delivery_fee = _safe_dec(getattr(order, "delivery_fee", None))
    else:
        delivery_fee = _safe_dec(delivery_fee)

    if delivery_fee <= 0:
        delivery_fee = _guess_delivery_fee(order)

    logistic_margin = _safe_dec(getattr(order, "logistic_margin", None))

    if delivery_fee > 0 and logistic_margin >= 0:
        candidate = delivery_fee - logistic_margin
        if candidate > 0:
            return candidate

    return Decimal("0")


@login_required
def finance_dashboard(request):
    """
    Dashboard financier FAGNI – Source de vérité :
    - compute_order_amounts() + TVA sur fagni_revenue_ht
    - évite les agrégations DB incohérentes (total_client_ttc / total / etc.)
    """
    payment_filter = request.GET.get("payment", "all")

    qs = (
        Order.objects
        .select_related("customer", "laundry_partner", "delivery_partner")
        .order_by("-created_at")
    )

    if payment_filter == "paid":
        qs = qs.filter(payment_status="paid")
    elif payment_filter == "unpaid":
        qs = qs.exclude(payment_status="paid")

    # on limite l'affichage, mais les KPI doivent être calculés sur un volume raisonnable
    # (tu peux monter à 500 si besoin)
    orders = list(qs[:500])

    cfg = get_pricing_settings()
    vat_rate = _get_vat_rate_percent(cfg)

    def q(x):
        try:
            return _q(x)
        except Exception:
            try:
                return Decimal(str(x))
            except Exception:
                return Decimal("0")

    ca_client_total = Decimal("0")
    partners_laundry_total = Decimal("0")
    partners_driver_total = Decimal("0")
    fagni_revenue_total = Decimal("0")      # TTC
    logistic_margin_total = Decimal("0")
    service_fee_total = Decimal("0")

    enriched = []

    for o in orders:
        amounts = compute_order_amounts(o)

        subtotal = q(amounts.get("subtotal", 0))
        service_fee_ht = q(amounts.get("service_fee_ht", 0))
        delivery_fee_client = q(amounts.get("delivery_fee_client", 0))

        express = q(
            amounts.get("express_for_client", 0)
            or amounts.get("express_extra_fee_client", 0)
            or amounts.get("express_fee_client", 0)
            or getattr(o, "express_extra_fee", 0)
            or 0
        )

        fagni_revenue_ht = q(amounts.get("fagni_revenue_ht", 0))
        vat_fagni = q((fagni_revenue_ht * q(vat_rate)) / Decimal("100"))

        total_ht_client = q(subtotal + service_fee_ht + delivery_fee_client + express)
        total_ttc_client = q(total_ht_client + vat_fagni)

        amount_laundry = q(
            amounts.get("amount_laundry_partner", 0)
            or getattr(o, "amount_laundry_partner", 0)
            or 0
        )
        amount_driver = q(
            amounts.get("amount_driver_partner", 0)
            or getattr(o, "amount_driver_partner", 0)
            or amounts.get("driver_income", 0)
            or getattr(o, "driver_logistic_cost", 0)
            or 0
        )

        logistic_margin = q(
            amounts.get("logistic_margin", 0)
            or getattr(o, "logistic_margin", 0)
            or 0
        )

        fagni_revenue_ttc = q(fagni_revenue_ht + vat_fagni)
        partners_total = q(amount_laundry + amount_driver)

        # stocke sur l'objet pour le template
        o.fin_total_client = total_ttc_client
        o.fin_subtotal = subtotal
        o.fin_service_fee = service_fee_ht
        o.fin_delivery_fee = delivery_fee_client
        o.fin_express = express
        o.fin_vat_fagni = vat_fagni
        o.fin_fagni_ttc = fagni_revenue_ttc
        o.fin_amount_laundry = amount_laundry
        o.fin_amount_driver = amount_driver
        o.fin_partners_total = partners_total
        o.fin_logistic_margin = logistic_margin

        ca_client_total += total_ttc_client
        partners_laundry_total += amount_laundry
        partners_driver_total += amount_driver
        fagni_revenue_total += fagni_revenue_ttc
        logistic_margin_total += logistic_margin
        service_fee_total += service_fee_ht

        enriched.append(o)

    order_count = len(enriched)
    avg_ticket = (ca_client_total / Decimal(order_count)).quantize(Decimal("0.01")) if order_count else Decimal("0.00")
    partners_total = partners_laundry_total + partners_driver_total

    context = {
        "payment_filter": payment_filter,
        "orders": enriched[:50],  # affichage seulement

        "ca_client_total": ca_client_total,
        "order_count": order_count,
        "avg_ticket": avg_ticket,

        "partners_total": partners_total,
        "partners_laundry_total": partners_laundry_total,
        "partners_driver_total": partners_driver_total,

        "fagni_revenue_total": fagni_revenue_total,
        "logistic_margin_total": logistic_margin_total,
        "service_fee_total": service_fee_total,

        "vat_rate": vat_rate,  # utile si tu veux l'afficher
    }

    return render(request, "orders/finance_dashboard.html", context)


# ============================================================
#  EXPORT CSV COMMANDES
# ============================================================
def export_orders_csv(request):
    status = request.GET.get("status")

    qs = Order.objects.select_related("customer").all()

    valid_statuses = {"pending", "in_progress", "done", "canceled"}
    if status in valid_statuses:
        qs = qs.filter(status=status)

    response = HttpResponse(content_type="text/csv")
    filename = f"commandes_fagni_{timezone.now().strftime('%Y%m%d_%H%M')}.csv"
    response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'

    writer = csv.writer(response, delimiter=";")

    writer.writerow(
        [
            "ID",
            "Code",
            "Date création",
            "Statut",
            "Client",
            "Téléphone",
            "Adresse",
            "Total (DB)",
            "Total HT",
            "TVA",
            "Total TTC",
            "Montant payé",
            "Montant dû",
        ]
    )

    for order in qs.order_by("-created_at"):
        customer = getattr(order, "customer", None)
        customer_name = getattr(customer, "name", "") if customer else ""
        customer_phone = getattr(customer, "phone", "") if customer else ""
        customer_address = getattr(customer, "address", "") if customer else ""

        try:
            total_ht = order.total_ht
        except Exception:
            total_ht = ""

        try:
            tva_amount = order.tva_amount
        except Exception:
            tva_amount = ""

        try:
            total_ttc = order.total_ttc
        except Exception:
            total_ttc = ""

        try:
            amount_paid = order.amount_paid
        except Exception:
            amount_paid = ""

        try:
            amount_due = order.amount_due
        except Exception:
            amount_due = ""

        writer.writerow(
            [
                order.id,
                order.code or "",
                order.created_at.strftime("%d/%m/%Y %H:%M") if order.created_at else "",
                order.get_status_display(),
                customer_name,
                customer_phone,
                customer_address,
                getattr(order, "total_client_display", None) or getattr(order, "total_client_ttc", None) or "",
                total_ht,
                tva_amount,
                total_ttc,
                amount_paid,
                amount_due,
            ]
        )

    return response


def export_orders_xlsx(request):
    """
    Export Excel des commandes FAGNI :
    - Reprend les filtres de la liste (status, q, date_from, date_to)
    - Onglet 1 : Synthèse
    - Onglet 2 : Détail des commandes
    """
    status = request.GET.get("status", "all")
    q = (request.GET.get("q") or "").strip()
    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""

    qs = (
        Order.objects
        .select_related("customer", "laundry_partner", "delivery_partner")
        .order_by("-created_at")
    )

    valid_statuses = ("pending", "in_progress", "done", "canceled")
    if status in valid_statuses:
        qs = qs.filter(status=status)

    if q:
        qs = qs.filter(
            Q(code__icontains=q)
            | Q(customer__name__icontains=q)
            | Q(customer__phone__icontains=q)
        )

    if date_from:
        df = parse_date(date_from)
        if df:
            qs = qs.filter(created_at__date__gte=df)

    if date_to:
        dt = parse_date(date_to)
        if dt:
            qs = qs.filter(created_at__date__lte=dt)

    orders = list(qs[:1000])  # limite de sécurité

    # Helper décimal
    def d(val):
        if isinstance(val, Decimal):
            return val
        if val in (None, "", 0):
            return Decimal("0")
        try:
            return Decimal(str(val))
        except Exception:
            return Decimal("0")

    # Stats synthèse
    total_count = len(orders)
    pending_count = sum(1 for o in orders if o.status == "pending")
    in_progress_count = sum(1 for o in orders if o.status == "in_progress")
    done_count = sum(1 for o in orders if o.status == "done")
    canceled_count = sum(1 for o in orders if o.status == "canceled")
    done_total = sum(d(getattr(o, "total", None)) for o in orders if o.status == "done")

    # Styles Excel
    wb = Workbook()
    title_font = Font(size=16, bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0056B3")
    section_fill = PatternFill("solid", fgColor="FF7A00")
    label_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    # ---------- Onglet 1 : Synthèse ----------
    ws1 = wb.active
    ws1.title = "Synthèse"

    ws1.merge_cells("A1:D1")
    cell_title = ws1["A1"]
    cell_title.value = "FAGNI – Export commandes"
    cell_title.font = title_font
    cell_title.fill = header_fill
    cell_title.alignment = center

    ws1["A3"] = "Généré le"
    ws1["A3"].font = label_font
    ws1["B3"] = timezone.localtime().strftime("%d/%m/%Y %H:%M")

    ws1["A4"] = "Statut"
    ws1["A4"].font = label_font
    ws1["B4"] = status if status in valid_statuses else "Tous"

    ws1["A5"] = "Recherche"
    ws1["A5"].font = label_font
    ws1["B5"] = q or "-"

    ws1["A6"] = "Période"
    ws1["A6"].font = label_font
    if date_from or date_to:
        per = ""
        if date_from:
            per += f"du {date_from} "
        if date_to:
            per += f"au {date_to}"
        ws1["B6"] = per
    else:
        ws1["B6"] = "Toutes les dates"

    ws1.merge_cells("A8:D8")
    sec = ws1["A8"]
    sec.value = "Synthèse des commandes"
    sec.font = Font(bold=True, color="FFFFFF")
    sec.fill = section_fill
    sec.alignment = left

    data_rows = [
        ("Nombre de commandes", total_count),
        ("En attente", pending_count),
        ("En cours", in_progress_count),
        ("Terminées", done_count),
        ("Annulées", canceled_count),
        ("Total commandes terminées (Total DB)", f"{done_total} FCFA"),
    ]

    start_row = 10
    for i, (label, value) in enumerate(data_rows):
        r = start_row + i
        ws1[f"A{r}"] = label
        ws1[f"A{r}"].font = label_font
        ws1[f"B{r}"] = value

    ws1.column_dimensions["A"].width = 40
    ws1.column_dimensions["B"].width = 35

    # ---------- Onglet 2 : Commandes ----------
    ws2 = wb.create_sheet(title="Commandes")

    headers = [
        "Code",
        "Date création",
        "Statut",
        "Client",
        "Téléphone",
        "Adresse",
        "Total (DB)",
        "Total HT",
        "TVA",
        "Total TTC",
        "Montant payé",
        "Montant dû",
        "Service FAGNI",
        "Livraison",
        "Blanchisserie",
        "Livreur",
    ]

    for col_idx, head in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col_idx, value=head)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    row_idx = 2

    for o in orders:
        customer = getattr(o, "customer", None)

        # champs financiers optionnels (comme dans export_orders_csv)
        try:
            total_ht = o.total_ht
        except Exception:
            total_ht = ""

        try:
            tva_amount = o.tva_amount
        except Exception:
            tva_amount = ""

        try:
            total_ttc = o.total_ttc
        except Exception:
            total_ttc = ""

        try:
            amount_paid = o.amount_paid
        except Exception:
            amount_paid = ""

        try:
            amount_due = o.amount_due
        except Exception:
            amount_due = ""

        row_vals = [
            o.code or "",
            o.created_at.strftime("%d/%m/%Y %H:%M") if o.created_at else "",
            o.get_status_display(),
            customer.name if customer else "",
            customer.phone if customer else "",
            customer.address if customer else "",
            d(getattr(o, "total", None)),
            d(total_ht) if total_ht not in ("", None) else "",
            d(tva_amount) if tva_amount not in ("", None) else "",
            d(total_ttc) if total_ttc not in ("", None) else "",
            d(amount_paid) if amount_paid not in ("", None) else "",
            d(amount_due) if amount_due not in ("", None) else "",
            d(getattr(o, "service_fee", None)),
            d(getattr(o, "delivery_fee", None)),
            o.laundry_partner.name if o.laundry_partner else "",
            o.delivery_partner.name if o.delivery_partner else "",
        ]

        for col_idx, val in enumerate(row_vals, start=1):
            c = ws2.cell(row=row_idx, column=col_idx, value=float(val) if isinstance(val, Decimal) else val)
            c.border = thin_border
            if isinstance(val, Decimal):
                c.alignment = right
                c.number_format = "#,##0"
            elif col_idx in (4, 5, 6):
                c.alignment = wrap
            else:
                c.alignment = left

        row_idx += 1

    # Largeurs colonnes
    widths = [14, 18, 16, 20, 14, 30, 14, 14, 12, 14, 14, 14, 14, 14, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    ws2.auto_filter.ref = ws2.dimensions

    # Réponse HTTP
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"fagni_commandes_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ============================================================
#  LISTE DES COMMANDES D'UN CLIENT (FICHE CLIENT)
# ============================================================
def orders_by_customer(request, customer_id):
    """
    Fiche client FAGNI :
    - Infos client
    - Stats globales (nb commandes, CA, service, livraison)
    - Historique des commandes
    """
    customer = get_object_or_404(Customer, pk=customer_id)

    qs = (
        Order.objects
        .filter(customer=customer)
        .select_related("customer", "laundry_partner", "delivery_partner")
        .order_by("-created_at")
    )

    agg = qs.aggregate(
        total_orders=Count("id"),
        done_orders=Count("id", filter=Q(status="done")),
        total_amount=Coalesce(Sum("total"), Decimal("0.00")),
        total_service_fee=Coalesce(Sum("service_fee"), Decimal("0.00")),
        total_delivery_fee=Coalesce(Sum("delivery_fee"), Decimal("0.00")),
    )

    context = {
        "customer": customer,
        "orders": qs,
        "total_orders": agg["total_orders"] or 0,
        "total_done": agg["done_orders"] or 0,
        "total_amount": agg["total_amount"] or Decimal("0.00"),
        "total_service_fee": agg["total_service_fee"] or Decimal("0.00"),
        "total_delivery_fee": agg["total_delivery_fee"] or Decimal("0.00"),
    }
    return render(request, "orders/orders_by_customer.html", context)


# ============================================================
#  LISTE CLIENTS – MINI CRM
# ============================================================
def customers_list(request):
    """
    Liste des clients FAGNI (mini CRM) :
    - Recherche (nom, téléphone, adresse)
    - Filtre min_orders (nb min de commandes)
    - Stats : nb commandes, montant total, service FAGNI, livraison
    Basé directement sur la relation Customer -> orders.
    """

    q = (request.GET.get("q") or "").strip()
    min_orders = request.GET.get("min_orders") or ""

    # 1) Base : tous les clients avec agrégats sur la relation "orders"
    qs = (
        Customer.objects
        .annotate(
            total_orders=Count("orders", distinct=True),
            total_amount=Coalesce(Sum("orders__total"), DECIMAL_ZERO),
            total_service_fee=Coalesce(Sum("orders__service_fee"), DECIMAL_ZERO),
            total_delivery_fee=Coalesce(Sum("orders__delivery_fee"), DECIMAL_ZERO),
            last_order_date=Max("orders__created_at"),
        )
    )

    # 2) Recherche plein texte
    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(phone__icontains=q)
            | Q(address__icontains=q)
        )

    # 3) Filtre min_orders
    if min_orders:
        try:
            min_o = int(min_orders)
        except (TypeError, ValueError):
            min_o = 0
        if min_o > 0:
            qs = qs.filter(total_orders__gte=min_o)

    # 4) Tri : dernière commande d'abord, puis nom
    qs = qs.order_by(
        F("last_order_date").desc(nulls_last=True),
        "name",
    )

    # 5) Totaux de synthèse globaux (pour la barre en haut)
    total_customers = qs.count()
    total_with_orders = qs.filter(total_orders__gt=0).count()

    agg = qs.aggregate(
        total_amount_global=Coalesce(Sum("total_amount"), DECIMAL_ZERO),
        total_service_global=Coalesce(Sum("total_service_fee"), DECIMAL_ZERO),
        total_delivery_global=Coalesce(Sum("total_delivery_fee"), DECIMAL_ZERO),
    )

    context = {
        "customers": qs,
        "total_customers": total_customers,
        "total_with_orders": total_with_orders,
        "q": q,
        "min_orders": min_orders,
        # synthèse globale
        "total_amount": agg["total_amount_global"],
        "total_service": agg["total_service_global"],
        "total_delivery": agg["total_delivery_global"],
    }
    return render(request, "orders/customers_list.html", context)


# ============================================================
#  EXPORT CSV CLIENTS – MINI CRM
# ============================================================
def export_customers_csv(request):
    """
    Export CSV de la liste des clients avec stats agrégées.
    Les mêmes filtres (q, min_orders) que la liste HTML sont appliqués.
    """
    q = (request.GET.get("q") or "").strip()
    min_orders = request.GET.get("min_orders") or ""

    qs = (
        Customer.objects
        .annotate(
            total_orders=Count("orders", distinct=True),
            total_amount=Coalesce(Sum("orders__total"), Decimal("0.00")),
            total_service_fee=Coalesce(Sum("orders__service_fee"), Decimal("0.00")),
            total_delivery_fee=Coalesce(Sum("orders__delivery_fee"), Decimal("0.00")),
            last_order_date=Max("orders__created_at"),
        )
    )

    # mêmes filtres que la liste
    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(phone__icontains=q)
            | Q(address__icontains=q)
        )

    if min_orders:
        try:
            min_o = int(min_orders)
        except (ValueError, TypeError):
            min_o = 0
        if min_o > 0:
            qs = qs.filter(total_orders__gte=min_o)

    # Réponse CSV
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    filename = f"clients_fagni_{timezone.now().strftime('%Y%m%d_%H%M')}.csv"
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'

    # En-tête
    resp.write(
        "Nom;Téléphone;Adresse;Nb commandes;Montant total;Service FAGNI;Livraison;Dernière commande\n"
    )

    for c in qs:
        last_date_str = (
            c.last_order_date.strftime("%d/%m/%Y %H:%M")
            if c.last_order_date else ""
        )

        line = ";".join([
            (c.name or "").replace(";", ","),
            (c.phone or "").replace(";", ","),
            (c.address or "").replace(";", ","),
            str(c.total_orders or 0),
            str(c.total_amount or 0),
            str(c.total_service_fee or 0),
            str(c.total_delivery_fee or 0),
            last_date_str,
        ])
        resp.write(line + "\n")

    return resp


def export_customers_xlsx(request):
    """
    Export Excel de la liste des clients avec stats agrégées.
    Reprend les mêmes filtres (q, min_orders) que la liste HTML.
    """
    q = (request.GET.get("q") or "").strip()
    min_orders = request.GET.get("min_orders") or ""

    qs = (
        Customer.objects
        .annotate(
            total_orders=Count("orders", distinct=True),
            total_amount=Coalesce(Sum("orders__total"), Decimal("0.00")),
            total_service_fee=Coalesce(Sum("orders__service_fee"), Decimal("0.00")),
            total_delivery_fee=Coalesce(Sum("orders__delivery_fee"), Decimal("0.00")),
            last_order_date=Max("orders__created_at"),
        )
    )

    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(phone__icontains=q)
            | Q(address__icontains=q)
        )

    if min_orders:
        try:
            min_o = int(min_orders)
        except (ValueError, TypeError):
            min_o = 0
        if min_o > 0:
            qs = qs.filter(total_orders__gte=min_o)

    qs = qs.order_by(
        F("last_order_date").desc(nulls_last=True),
        "name",
    )

    customers = list(qs)

    # Helper décimal
    def d(val):
        if isinstance(val, Decimal):
            return val
        if val in (None, "", 0):
            return Decimal("0")
        try:
            return Decimal(str(val))
        except Exception:
            return Decimal("0")

    total_customers = len(customers)
    total_with_orders = sum(1 for c in customers if (c.total_orders or 0) > 0)
    total_amount = sum(d(c.total_amount) for c in customers)
    total_service = sum(d(c.total_service_fee) for c in customers)
    total_delivery = sum(d(c.total_delivery_fee) for c in customers)

    # Styles
    wb = Workbook()
    title_font = Font(size=16, bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0056B3")
    section_fill = PatternFill("solid", fgColor="FF7A00")
    label_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    # --- Synthèse ---
    ws1 = wb.active
    ws1.title = "Synthèse"

    ws1.merge_cells("A1:D1")
    t = ws1["A1"]
    t.value = "FAGNI – Export clients"
    t.font = title_font
    t.fill = header_fill
    t.alignment = center

    ws1["A3"] = "Généré le"
    ws1["A3"].font = label_font
    ws1["B3"] = timezone.localtime().strftime("%d/%m/%Y %H:%M")

    ws1["A4"] = "Recherche"
    ws1["A4"].font = label_font
    ws1["B4"] = q or "-"

    ws1["A5"] = "Min. commandes"
    ws1["A5"].font = label_font
    ws1["B5"] = min_orders or "0"

    ws1.merge_cells("A7:D7")
    s = ws1["A7"]
    s.value = "Synthèse du portefeuille clients"
    s.font = Font(bold=True, color="FFFFFF")
    s.fill = section_fill
    s.alignment = left

    lines = [
        ("Nombre total de clients", total_customers),
        ("Clients avec au moins 1 commande", total_with_orders),
        ("Montant total commandes", f"{total_amount} FCFA"),
        ("Service FAGNI cumulé", f"{total_service} FCFA"),
        ("Livraison facturée cumulée", f"{total_delivery} FCFA"),
    ]

    start_row = 9
    for i, (label, value) in enumerate(lines):
        r = start_row + i
        ws1[f"A{r}"] = label
        ws1[f"A{r}"].font = label_font
        ws1[f"B{r}"] = value

    ws1.column_dimensions["A"].width = 45
    ws1.column_dimensions["B"].width = 35

    # --- Détail clients ---
    ws2 = wb.create_sheet(title="Clients")

    headers = [
        "Nom",
        "Téléphone",
        "Adresse",
        "Nb commandes",
        "Montant total",
        "Service FAGNI",
        "Livraison",
        "Dernière commande",
    ]

    for col_idx, head in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col_idx, value=head)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    row_idx = 2
    for cst in customers:
        last_date_str = (
            cst.last_order_date.strftime("%d/%m/%Y %H:%M")
            if cst.last_order_date else ""
        )
        row_vals = [
            cst.name or "",
            cst.phone or "",
            cst.address or "",
            int(cst.total_orders or 0),
            float(d(cst.total_amount)),
            float(d(cst.total_service_fee)),
            float(d(cst.total_delivery_fee)),
            last_date_str,
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (4, 5, 6, 7):
                cell.alignment = right
                if col_idx >= 5:
                    cell.number_format = "#,##0"
            elif col_idx == 3:
                cell.alignment = wrap
            else:
                cell.alignment = left
        row_idx += 1

    widths = [24, 16, 30, 14, 16, 16, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    ws2.auto_filter.ref = ws2.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"fagni_clients_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ============================================================
#  LOGISTIQUE : VALIDATION COLLECTE / LIVRAISON (BLOC 1)
# ============================================================

# Créneaux "standards" qu'on utilise côté front (Bloc 3)
DEFAULT_SLOTS = [
    ("08:00-10:00", time(8, 0), time(10, 0)),
    ("10:00-12:00", time(10, 0), time(12, 0)),
    ("12:00-14:00", time(12, 0), time(14, 0)),
    ("14:00-16:00", time(14, 0), time(16, 0)),
    ("16:00-18:00", time(16, 0), time(18, 0)),
    ("18:00-20:00", time(18, 0), time(20, 0)),
]


def _find_slot(slot_value):
    """
    slot_value est censé être une string du type "08:00-10:00".
    On renvoie (start_time, end_time) ou (None, None) si introuvable.
    """
    if not slot_value:
        return None, None

    # On tolère quelques formats, mais on recommande "HH:MM-HH:MM"
    cleaned = slot_value.strip().replace(" ", "").replace("h", ":").replace("–", "-")
    # ex : "08h00 – 10h00" -> "08:00-10:00"

    for key, start_t, end_t in DEFAULT_SLOTS:
        key_clean = key.replace(" ", "")
        if cleaned == key_clean:
            return start_t, end_t

    # Dernier recours : essayer un split basique
    try:
        if "-" in cleaned:
            left, right = cleaned.split("-", 1)
            h1, m1 = left.split(":")
            h2, m2 = right.split(":")
            return time(int(h1), int(m1)), time(int(h2), int(m2))
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=4137")

    return None, None


def _combine_date_and_time(date_str, time_obj):
    """
    Combine une date (YYYY-MM-DD) et un time en datetime aware.
    """
    if not date_str or not time_obj:
        return None

    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return None

    naive_dt = datetime.combine(d, time_obj)
    return timezone.make_aware(naive_dt, timezone.get_current_timezone())


def compute_pickup_and_delivery_datetimes(post_data, config, now=None):
    """
    Calcule et VALIDE la logique collecte / livraison.

    Retourne :
    (pickup_mode, pickup_dt,
     delivery_mode, delivery_dt)

    - pickup_dt : datetime de début du créneau de collecte (ou "maintenant" si immédiate)
    - delivery_dt : datetime de début du créneau de livraison (ou None si non pertinent)
    """

    if now is None:
        now = timezone.localtime()

    # Sécurité si pas de config
    cutoff_hour = config.pickup_cutoff_hour if config else 10
    standard_sla_hours = config.standard_sla_hours if config else 48
    express_sla_hours = config.express_sla_hours if config else 24
    express_enabled = bool(config.express_enabled) if config else True
    scheduled_enabled = bool(config.scheduled_slots_enabled) if config else True

    pickup_mode = post_data.get("pickup_mode") or Order.PICKUP_MODE_NOW
    delivery_mode = post_data.get("delivery_mode") or Order.DELIVERY_MODE_STANDARD

    # -----------------------------
    # 1) COLLECTE
    # -----------------------------
    pickup_dt = None

    if pickup_mode == Order.PICKUP_MODE_NOW:
        # Collecte immédiate : on fixe la référence à maintenant
        pickup_dt = now

    elif pickup_mode == Order.PICKUP_MODE_LATER:
        pickup_date_str = post_data.get("pickup_date") or ""
        pickup_slot_value = post_data.get("pickup_slot") or ""

        if not pickup_date_str or not pickup_slot_value:
            raise ValidationError("Merci de choisir une date et un créneau horaire pour la collecte programmée.")

        start_t, end_t = _find_slot(pickup_slot_value)
        if not start_t or not end_t:
            raise ValidationError("Le créneau de collecte choisi est invalide.")

        pickup_dt = _combine_date_and_time(pickup_date_str, start_t)
        if not pickup_dt:
            raise ValidationError("La date de collecte programmée est invalide.")

        # Interdiction d'un créneau passé
        if pickup_dt <= now:
            raise ValidationError("Le créneau de collecte doit être dans le futur.")

    else:
        # Mode inconnu → on force sur NOW
        pickup_mode = Order.PICKUP_MODE_NOW
        pickup_dt = now

    # -----------------------------
    # 2) LIVRAISON
    # -----------------------------
    delivery_dt = None

    # --- Mode STANDARD (≈ 48h) ---
    if delivery_mode == Order.DELIVERY_MODE_STANDARD:
        # La livraison est calculée automatiquement : collecte + 48h
        delivery_dt = pickup_dt + timedelta(hours=standard_sla_hours)

    # --- Mode EXPRESS (≈ 24h) ---
    elif delivery_mode == Order.DELIVERY_MODE_EXPRESS:
        if not express_enabled:
            raise ValidationError("La livraison express n'est pas disponible pour le moment.")

        # Si la collecte est après l'heure limite, on bloque l'express
        if pickup_dt.hour >= cutoff_hour:
            raise ValidationError(
                f"Après {cutoff_hour}h, la formule express n'est plus garantie. "
                "Merci de choisir Standard ou Programmée."
            )

        delivery_dt = pickup_dt + timedelta(hours=express_sla_hours)

    # --- Mode PROGRAMMÉ ---
    elif delivery_mode == Order.DELIVERY_MODE_SCHEDULED:
        if not scheduled_enabled:
            raise ValidationError("La livraison programmée n'est pas disponible pour le moment.")

        delivery_date_str = post_data.get("delivery_date") or ""
        delivery_slot_value = post_data.get("delivery_slot") or ""

        if not delivery_date_str or not delivery_slot_value:
            raise ValidationError("Merci de choisir une date et un créneau horaire pour la livraison programmée.")

        d_start_t, d_end_t = _find_slot(delivery_slot_value)
        if not d_start_t or not d_end_t:
            raise ValidationError("Le créneau de livraison choisi est invalide.")

        delivery_dt = _combine_date_and_time(delivery_date_str, d_start_t)
        if not delivery_dt:
            raise ValidationError("La date de livraison programmée est invalide.")

        # Interdiction d'un créneau passé
        if delivery_dt <= now:
            raise ValidationError("Le créneau de livraison doit être dans le futur.")

        # Interdiction de livraison avant ou en même temps que la collecte
        if delivery_dt <= pickup_dt:
            raise ValidationError("La livraison ne peut pas être avant ou au même créneau que la collecte.")

    else:
        # Sécurité : si mode inconnu -> on revient à STANDARD
        delivery_mode = Order.DELIVERY_MODE_STANDARD
        delivery_dt = pickup_dt + timedelta(hours=standard_sla_hours)

    return pickup_mode, pickup_dt, delivery_mode, delivery_dt


def validate_delivery_mode(pickup_dt, delivery_dt, mode):
    if not pickup_dt or not delivery_dt:
        return None

    delta_hours = (delivery_dt - pickup_dt).total_seconds() / 3600

    # Express ⇢ ~24h
    if mode == "express":
        if not (12 <= delta_hours <= 36):
            return (
                "Le mode EXPRESS doit correspondre à une livraison environ 24h après la collecte."
            )

    # Standard ⇢ ~48h
    if mode == "standard":
        if not (36 <= delta_hours <= 72):
            return (
                "Le mode STANDARD doit correspondre à une livraison environ 48h après la collecte."
            )

    # Programmée ⇢ ≥ 72h
    if mode == "scheduled":
        if delta_hours < 72:
            return (
                "Le mode PROGRAMMÉ nécessite une livraison au moins 72h après la collecte. "
                "Pour une livraison avant 72h, choisis Standard (48h) ou Express (24h)."
            )

    return None


# ============================================================
#  CRÉATION COMMANDE
# ============================================================
@login_required
def create(request):
    """
    Création d'une commande FAGNI :
    - Client (nom, téléphone, adresse, lat/lng)
    - Lignes de prestations issues du catalogue
    - Photos multiples par ligne (photos_0, photos_1, ...)
    - Assignation automatique d'un livreur si possible
    - Code parrain (referral_code)
    - Logistique : collecte / livraison validées (Bloc 1)
    """

    service_categories = ServiceCategory.objects.all()
    service_items = ServiceItem.objects.select_related("category").all()

    # Paramètres logistiques généraux (tarifs, etc.)
    logi = getattr(settings, "FAGNI_LOGISTICS", {})

    # Configuration logistique (SLA, cutoff, express, programmée…)
    config = LogisticsConfig.current()

    # --- Gestion du code parrain initial (GET ou POST) ---
    if request.method == "POST":
        referral_initial = (request.POST.get("referral_code") or "").strip()
    else:
        referral_initial = (
            (request.GET.get("ref") or "").strip()
            or (request.GET.get("aff") or "").strip()
        )

    # --- Valeurs par défaut / re-remplissage pour la collecte & la livraison ---
    if request.method == "POST":
        pickup_mode_val = request.POST.get("pickup_mode") or Order.PICKUP_MODE_NOW
        pickup_date_val = request.POST.get("pickup_date") or ""
        pickup_slot_val = request.POST.get("pickup_slot") or ""

        delivery_mode_val = request.POST.get("delivery_mode") or Order.DELIVERY_MODE_STANDARD
        delivery_date_val = request.POST.get("delivery_date") or ""
        delivery_slot_val = request.POST.get("delivery_slot") or ""
    else:
        pickup_mode_val = Order.PICKUP_MODE_NOW
        pickup_date_val = ""
        pickup_slot_val = ""

        delivery_mode_val = Order.DELIVERY_MODE_STANDARD
        delivery_date_val = ""
        delivery_slot_val = ""

    # Contexte de base (utilisé aussi en cas d'erreur POST)
    context = {
        "service_categories": service_categories,
        "service_items": service_items,
        "error": None,
        "client_phone": request.POST.get("client_phone", "") if request.method == "POST" else "",
        "client_name": request.POST.get("client_name", "") if request.method == "POST" else "",
        "client_address": request.POST.get("client_address", "") if request.method == "POST" else "",
        "client_lat": request.POST.get("client_lat", "") if request.method == "POST" else "",
        "client_lng": request.POST.get("client_lng", "") if request.method == "POST" else "",
        "delivery_address": request.POST.get("delivery_address", "") if request.method == "POST" else "",
        "referral_code": referral_initial,
        "order_notes": request.POST.get("order_notes", "") if request.method == "POST" else "",
        "GOOGLE_MAPS_API_KEY": getattr(settings, "GOOGLE_MAPS_API_KEY", ""),
        "delivery_min_fee": logi.get("client_min_fee", 1000),
        "delivery_price_per_km": logi.get("client_price_per_km", 150),
        "delivery_fixed_fee": logi.get("client_fixed_fee", 300),
        # Infos config logistique (pour affichage dans le template si besoin)
        "logi_pickup_cutoff_hour": config.pickup_cutoff_hour if config else 10,
        "logi_standard_sla_hours": config.standard_sla_hours if config else 48,
        "logi_express_sla_hours": config.express_sla_hours if config else 24,
        "logi_express_enabled": bool(config.express_enabled) if config else True,
        "logi_express_extra_flat": config.express_extra_flat if config else 0,
        "logi_express_extra_percent": config.express_extra_percent if config else 0,
        "logi_scheduled_enabled": bool(config.scheduled_slots_enabled) if config else True,
        "pickup_mode": pickup_mode_val,
        "pickup_date": pickup_date_val,
        "pickup_slot": pickup_slot_val,
        "delivery_mode": delivery_mode_val,
        "delivery_date": delivery_date_val,
        "delivery_slot": delivery_slot_val,
    }

    context["drivers"] = get_active_drivers()

    # --- LOT E : Blanchisseries géolocalisees (pour estimation front) ---
    laundries_geo = list(
        LaundryPartner.objects.filter(
            is_active=True,
            latitude__isnull=False,
            longitude__isnull=False,
        ).values("id", "name", "latitude", "longitude")
    )
    context["laundries_geo_json"] = json.dumps(laundries_geo, default=str)


    # --- GET : on affiche le formulaire ---
    if request.method != "POST":
        return render(request, "orders/create.html", context)

    # --- POST : on traite la commande ---
    phone = (request.POST.get("client_phone") or "").strip()
    name = (request.POST.get("client_name") or "").strip()
    address = clean_address_or_empty((request.POST.get("client_address") or "").strip())

    lat_raw = (request.POST.get("client_lat") or "").strip()
    lng_raw = (request.POST.get("client_lng") or "").strip()

    delivery_address_input = clean_address_or_empty((request.POST.get("delivery_address") or "").strip())
    referral_code = referral_initial
    notes = (request.POST.get("order_notes") or "").strip()

    # 1) Validation minimale client
    if not phone or not name:
        context["error"] = "Merci de renseigner au moins le nom et le téléphone du client."
        return render(request, "orders/create.html", context)

    if not is_probably_valid_address(address):
        context["error"] = "Merci de renseigner une adresse de collecte plus précise."
        return render(request, "orders/create.html", context)

    # 2) Création / mise à jour du client
    try:
        customer = Customer.objects.filter(phone=phone).order_by("-id").first()
        if not customer:
            customer = Customer.objects.create(phone=phone, name=name, address=address)

    except Customer.MultipleObjectsReturned:
        customer = (
            Customer.objects.filter(phone=phone)
            .order_by("-id")
            .first()
        )

    changed = False
    if name and customer.name != name:
        customer.name = name
        changed = True
    if address and customer.address != address:
        customer.address = address
        changed = True

    # latitude / longitude si présentes
    try:
        if lat_raw:
            customer.latitude = Decimal(lat_raw)
            changed = True
        if lng_raw:
            customer.longitude = Decimal(lng_raw)
            changed = True
    except Exception:
        # On ignore les erreurs de parsing lat/lng
        pass

    if changed:
        customer.save()

    # 3) Création de la commande (vide au début, pour avoir un ID)
    code = uuid.uuid4().hex[:8].upper()
    while Order.objects.filter(code=code).exists():
        code = uuid.uuid4().hex[:8].upper()

    order = Order.objects.create(
        customer=customer,
        status="pending",
        code=code,
        referral_code=referral_code or None,
        notes=notes,
    )

    # 4) Adresses collecte / livraison + géolocalisation
    order.pickup_address = clean_address_or_empty(address or customer.address or "")

    if delivery_address_input:
        order.delivery_address = clean_address_or_empty(delivery_address_input)
    else:
        # Si rien saisi, livraison = collecte
        order.delivery_address = order.pickup_address

    try:
        order.pickup_lat = float(lat_raw) if lat_raw else None
    except (TypeError, ValueError):
        order.pickup_lat = None

    try:
        order.pickup_lng = float(lng_raw) if lng_raw else None
    except (TypeError, ValueError):
        order.pickup_lng = None

    # Pour l'instant, si aucune lat/lng spécifique pour la livraison :
    # - si adresse livraison = collecte → on copie
    delivery_lat_raw = request.POST.get("delivery_lat") or ""
    delivery_lng_raw = request.POST.get("delivery_lng") or ""

    if not delivery_address_input:
        order.delivery_lat = order.pickup_lat
        order.delivery_lng = order.pickup_lng
    else:
        try:
            order.delivery_lat = float(delivery_lat_raw) if delivery_lat_raw else None
        except (TypeError, ValueError):
            order.delivery_lat = None
        try:
            order.delivery_lng = float(delivery_lng_raw) if delivery_lng_raw else None
        except (TypeError, ValueError):
            order.delivery_lng = None

    # 5) Logistique : collecte & livraison (BLOC 1)
    try:
        pickup_mode, pickup_dt, delivery_mode, delivery_dt = compute_pickup_and_delivery_datetimes(
            request.POST, config
        )
    except ValidationError as e:
        # On supprime la commande vide, on remonte l'erreur à l'UI
        order.delete()
        context["error"] = str(e)
        return render(request, "orders/create.html", context)

    order.pickup_mode = pickup_mode
    order.delivery_mode = delivery_mode

    # Si collecte programmée → on stocke date & heure de début
    if pickup_mode == Order.PICKUP_MODE_LATER and pickup_dt:
        order.pickup_scheduled_date = pickup_dt.date()
        order.pickup_scheduled_time = pickup_dt.time()
    else:
        order.pickup_scheduled_date = None
        order.pickup_scheduled_time = None

    # Si livraison programmée → on stocke date & heure de début
    if delivery_mode == Order.DELIVERY_MODE_SCHEDULED and delivery_dt:
        order.delivery_scheduled_date = delivery_dt.date()
        order.delivery_scheduled_time = delivery_dt.time()
    else:
        order.delivery_scheduled_date = None
        order.delivery_scheduled_time = None

    error = validate_delivery_mode(pickup_dt, delivery_dt, delivery_mode)
    if error:
        # IMPORTANT: éviter les "commandes fantômes" en base
        order.delete()
        context["error"] = error
        return render(request, "orders/create.html", context)

    order.save()

    # 1) Géocodage (avant assignation)
    ensure_order_geocoded(order, save=True)

    # 2) Assignation automatique d'une blanchisserie (si autorisée) + raison
    try:
        from orders.assignment import pick_best_laundry
        laundry, laundry_reason = pick_best_laundry(order)
    except Exception:
        laundry, laundry_reason = None, None

    if laundry:
        order.laundry_partner = laundry
        # si tu as un champ reason côté Order un jour, tu pourras le stocker ici
    else:
        # optionnel : si tu as un champ sur Order du genre laundry_partner_unassigned_reason
        # order.laundry_partner_unassigned_reason = laundry_reason
        pass

    # 3) Assignation automatique d'un livreur (si autorisée) + raison
    try:
        from orders.assignment import pick_best_driver
        driver, reason = pick_best_driver(order)
    except Exception:
        driver, reason = None, None

    if driver:
        order.delivery_partner = driver

    # 3.b) Préparation logistique initiale
    # - persiste les partenaires auto-assignés
    # - synchronise les DeliveryLeg standard
    # - normalise les statuts de jambes
    try:
        update_fields = []
        if hasattr(order, "laundry_partner_id"):
            update_fields.append("laundry_partner")
        if hasattr(order, "delivery_partner_id"):
            update_fields.append("delivery_partner")
        if update_fields:
            order.save(update_fields=update_fields)
        else:
            order.save()
    except Exception:
        order.save()

    try:
        from orders.models import sync_delivery_legs_for_order
        sync_delivery_legs_for_order(order)
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=4602")

    try:
        from orders.service_layer.legs import normalize_order_legs
        normalize_order_legs(order, save=True)
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=4608")

    # 3.c) Bootstrap statut global commande
    # Si la commande a déjà une assignation partenaire/logistique,
    # on la sort de "pending" pour refléter un traitement réel en cours.
    try:
        has_laundry = bool(getattr(order, "laundry_partner_id", None))
        has_driver = bool(getattr(order, "delivery_partner_id", None))
        if getattr(order, "status", None) == "pending" and (has_laundry or has_driver):
            order.status = "in_progress"
            order.save(update_fields=["status"])

            try:
                _ensure_pickup_mission_for_order(order)
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=4623")

            # 🔥 RESYNC FINAL (clé du bug)
            try:
                from orders.models import sync_delivery_legs_for_order, DeliveryLeg
                from orders.service_layer.legs import normalize_order_legs

                sync_delivery_legs_for_order(order)
                normalize_order_legs(order, save=True)

                if getattr(order, "delivery_partner_id", None):
                    DeliveryLeg.objects.filter(order=order).exclude(status="canceled").update(
                        driver_id=order.delivery_partner_id
                    )
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=4638")
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=4640")

    # 9) Lignes de commande (PRIX VERROUILLÉS CÔTÉ SERVEUR)
    service_ids = request.POST.getlist("service_id[]")
    designations = request.POST.getlist("designation[]")
    quantities = request.POST.getlist("quantity[]")

    created_any_item = False

    for idx, (sid, desc, qty_str) in enumerate(zip(service_ids, designations, quantities)):
        sid = (sid or "").strip()
        desc = (desc or "").strip()

        # Quantité
        try:
            qty = int(qty_str)
        except Exception:
            qty = 0

        if qty <= 0:
            continue

        # Service (obligatoire pour sécuriser le prix)
        try:
            service_obj = ServiceItem.objects.get(pk=sid)
        except (ServiceItem.DoesNotExist, ValueError, TypeError):
            service_obj = None

        if not service_obj:
            continue

        # Désignation fallback : si vide, on prend le nom catalogue
        if not desc:
            desc = (service_obj.name or "").strip()
            if not desc:
                continue

        # PRIX : on ignore complètement ce que le navigateur a envoyé
        try:
            pu = Decimal(str(service_obj.default_price))
        except Exception:
            pu = Decimal("0")

        if pu <= 0:
            continue

        line_total = (pu * qty).quantize(Decimal("0.01"))

        item = OrderItem.objects.create(
            order=order,
            service=service_obj,
            designation=desc,
            quantity=qty,
            unit_price=pu,
            total=line_total,
        )
        created_any_item = True

        # Photos
        file_field_name = f"photos_{idx}"
        for photo_file in request.FILES.getlist(file_field_name):
            if not photo_file:
                continue
            OrderItemPhoto.objects.create(
                order_item=item,
                image=photo_file,
            )

    if not created_any_item:
        order.delete()
        context["error"] = "Ajoute au moins une ligne de prestation à la commande."
        return render(request, "orders/create.html", context)

    # 10) Recalcul financier complet (moteur unique) + distances + legs
    try:
        # ✅ on persiste les montants calculés (express, TVA, parts partenaires, etc.)
        # === SAVE_DELIVERY_FIELDS_BEFORE_FINANCE ===
        # Render/Prod: s'assurer que delivery_mode + delivery_fee sont bien persistés
        try:
            real_fields = {f.name for f in order._meta.fields}
            uf = []
            for f in ('delivery_mode', 'delivery_fee'):
                if f in real_fields:
                    uf.append(f)
            if uf:
                order.save(update_fields=uf)
            else:
                order.save()
        except Exception:
            order.save()
        order.update_financials(save=True)
    except Exception as e:
        print("Erreur update_financials:", e)

    try:
        # distances (si ta méthode existe)
        order.recompute_distances_from_positions()
    except Exception as e:
        print("Erreur recompute_distances_from_positions:", e)

    try:
        from orders.models import DeliveryLeg, sync_delivery_legs_for_order
        if not DeliveryLeg.objects.filter(order=order).exclude(status="canceled").exists():
            sync_delivery_legs_for_order(order)
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=4745")

    return redirect("orders:detail", order_id=order.id)


def build_order_finance_context(order):
    """
    Contexte financier normalisé alimenté par le pricing engine.
    Garde les clés legacy utiles aux templates/PDFs, mais la source de vérité
    est désormais compute_order_pricing / build_order_finance_summary.
    """
    finance_summary = build_order_finance_summary(order)

    def q(x):
        try:
            return _q(x)
        except Exception:
            try:
                return Decimal(str(x))
            except Exception:
                return Decimal("0")

    prestation_total = q(finance_summary.get("prestation_total", 0))
    service_fee_ht = q(finance_summary.get("service_fee_ht", 0))
    delivery_fee_client = q(finance_summary.get("delivery_fee_client", 0))
    express_surcharge = q(
        finance_summary.get("express_extra_fee_client", 0)
        or finance_summary.get("express_for_client", 0)
        or finance_summary.get("express_surcharge", 0)
        or 0
    )
    vat_fagni = q(finance_summary.get("vat_fagni", 0))
    total_ttc_client = q(finance_summary.get("total_client_ttc", 0))

    total_ht_client = q(
        prestation_total
        + service_fee_ht
        + delivery_fee_client
        + express_surcharge
    )

    amount_laundry = q(finance_summary.get("amount_laundry", 0))
    amount_driver = q(finance_summary.get("amount_driver", 0))
    partners_total = q(amount_laundry + amount_driver)

    fagni_revenue_ht = q(finance_summary.get("fagni_revenue_ht", 0))
    margin_delivery = q(finance_summary.get("margin_delivery", 0))

    return {
        "order": order,

        # client
        "prestation_total": prestation_total,
        "service_fee_ht": service_fee_ht,
        "delivery_fee_client": delivery_fee_client,
        "express_surcharge": express_surcharge,
        "vat_fagni": vat_fagni,
        "total_ttc_client": total_ttc_client,

        # partenaires
        "amount_laundry": amount_laundry,
        "amount_driver": amount_driver,
        "partners_total": partners_total,

        # FAGNI
        "fagni_revenue_ht": fagni_revenue_ht,
        "margin_delivery": margin_delivery,

        # compat
        "total_ht_client": total_ht_client,
        "finance_summary": finance_summary,
    }


# ============================================================
#   DETAIL COMMANDE
# ============================================================
@login_required
def detail(request, order_id):
    """
    Détail d'une commande FAGNI :
    - recalcule si besoin les frais de livraison via le moteur intégré de la commande
    - applique le modèle financier FAGNI via update_financials()
    - affiche les lignes, photos et historique de statut
    - (Back-office) permet d'ajouter un paiement via POST (order.add_payment)
    """
    # --- Récupération de la commande ---
    order = (
        Order.objects
        .select_related("customer", "laundry_partner", "delivery_partner")
        .filter(pk=order_id)
        .first()
    )

    if not order:
        return redirect("orders:list")

    # ====================================================================
    # ✅ POST: Ajouter un paiement (back-office)
    # ====================================================================
    if request.method == "POST" and (request.POST.get("action") or "").strip() == "add_payment":
        from django.contrib import messages
        from django.core.exceptions import ValidationError
        from decimal import Decimal

        try:
            amt_raw = (request.POST.get("amount") or "").strip()
            channel = (
                request.POST.get("channel") or ""
            ).strip().lower() or "cash"
            ref = (request.POST.get("reference") or "").strip()

            # Sécurité pilote :
            # l'ajout générique Back-office est réservé au CASH.
            # Les paiements électroniques passent obligatoirement par
            # leurs workflows de vérification dédiés.
            if channel != "cash":
                raise ValidationError({
                    "channel": (
                        "Seul un paiement CASH réellement encaissé peut "
                        "être enregistré depuis cette action générique."
                    )
                })

            amount = Decimal(amt_raw)
            if amount <= 0:
                raise ValidationError({"amount": "Le montant doit être > 0."})

            # Référence auto si vide (utile pour tracer)
            if not ref:
                ref = f"BO-{order.code}-{int(amount)}"

            # ✅ utilise ton garde-fou + idempotence + cap surpaiement
            order.add_payment(
                amount=int(amount),
                channel=channel,
                reference=ref,
                source="backoffice",
                save=True,
            )

            messages.success(request, f"Paiement enregistré : {int(amount)} FCFA ({channel}).")
        except ValidationError as e:
            # rendre l'erreur lisible
            msg = "Erreur paiement."
            try:
                if hasattr(e, "message_dict") and e.message_dict:
                    flat = []
                    for k, vals in e.message_dict.items():
                        if isinstance(vals, (list, tuple)):
                            for v in vals:
                                flat.append(str(v))
                        else:
                            flat.append(str(vals))
                    if flat:
                        msg = " | ".join(flat)
                elif e.messages:
                    msg = " | ".join([str(x) for x in e.messages])
            except Exception:
                msg = "Erreur paiement."
            messages.error(request, msg)
        except Exception as e:
            messages.error(request, f"Erreur technique paiement : {e}")

        return redirect("orders:detail", order_id=order.id)

    # --- Lignes de commande (prestations) ---
    items = (
        OrderItem.objects
        .filter(order=order)
        .select_related("service", "service__category")
        .prefetch_related("photos")
    )

    # ====================================================================
    # 1) FRAIS DE LIVRAISON : si 0 ou null, on utilise le moteur interne
    #    de la commande (compute_delivery_fee)
    # ====================================================================
    needs_delivery_recompute = (
        (not order.delivery_fee or decimal.Decimal(str(order.delivery_fee)) == decimal.Decimal("0"))
        and order.customer
        and order.laundry_partner
    )

    if needs_delivery_recompute:
        # Modèle pressing pilote FAGNI : livraison aller-retour fixe facturée au client
        delivery_fee = decimal.Decimal("2000")
        order.delivery_fee = delivery_fee
        order.save(update_fields=["delivery_fee"])
    else:
        if order.delivery_fee is not None:
            order.delivery_fee = decimal.Decimal(str(order.delivery_fee))

    # ====================================================================
    # 2) APPLICATION DU MODÈLE FAGNI (update_financials)
    # ====================================================================
    data = order.update_financials(save=True)

    # ====================================================================
    # 3) PHOTOS & HISTORIQUE
    # ====================================================================
    all_photos = OrderItemPhoto.objects.filter(order_item__order=order)

    status_history = (
        OrderStatusHistory.objects
        .filter(order=order)
        .order_by("changed_at")
    )

    ticket_url = reverse("orders:order_ticket_pdf", args=[order.id])
    ticket_thermal_url = reverse("orders:order_ticket_thermal_pdf", args=[order.id])

    finance = build_order_finance_context(order)

    from orders.models import Payment
    payments = Payment.objects.filter(order=order).order_by("-id")

    try:
        total = decimal.Decimal(str(getattr(order, "total_client_ttc", 0) or 0)) - decimal.Decimal(str(getattr(order, "coupon_discount_applied", 0) or 0))
        paid = decimal.Decimal(str(getattr(order, "amount_paid", 0) or 0))
    except Exception:
        total, paid = decimal.Decimal("0"), decimal.Decimal("0")

    remaining = (total - paid)
    if remaining < 0:
        remaining = decimal.Decimal("0")

    context = {
        "order": order,
        "items": items,
        "status_history": status_history,
        "all_photos": all_photos,
        "ticket_url": ticket_url,
        "ticket_thermal_url": ticket_thermal_url,
        "finance": finance,
        "financial_data": data,
        "payments": payments,
        "remaining": remaining,
    }

    ctx_amounts = _build_invoice_context(order)
    context.update(ctx_amounts)
    context["express_client"] = (ctx_amounts.get("amounts") or {}).get("express_for_client", 0)

    return render(request, "orders/detail.html", context)


# -------------------------------------------------------------------
# Helpers Client (session phone)
# -------------------------------------------------------------------

# Clé session canonique (1 seule vérité)
DEFAULT_CLIENT_SESSION_KEY = "fagni_client_phone"

def _client_session_key() -> str:
    # Permet override via settings.py si tu veux
    return getattr(settings, "CLIENT_SESSION_KEY", DEFAULT_CLIENT_SESSION_KEY)

def _normalize_phone(raw: str) -> str:
    """
    Normalisation légère pour éviter les espaces/traits.
    - garde uniquement les chiffres
    - retire 00225 / 225 si présent (CIV)
    """
    s = (raw or "").strip()
    digits = re.sub(r"\D+", "", s)

    # Côte d'Ivoire: +225 / 00225
    if digits.startswith("00225"):
        digits = digits[5:]
    elif digits.startswith("225") and len(digits) > 10:
        digits = digits[3:]

    return digits


@require_GET
@staff_member_required
def order_live_status(request, order_id: int):
    """
    Endpoint JSON "live" pour la page back-office detail.html
    (polling JS, sans websockets).
    """
    o = (
        Order.objects
        .select_related("customer", "delivery_partner", "laundry_partner")
        .filter(pk=order_id)
        .first()
    )
    if not o:
        return JsonResponse({"ok": False, "error": "not_found"}, status=404)

    # --- updated_at (fallback)
    updated_dt = None
    for attr in ("updated_at", "modified_at", "updated", "modified", "last_updated", "created_at"):
        if hasattr(o, attr):
            updated_dt = getattr(o, attr)
            if updated_dt:
                break
    updated_at = updated_dt.isoformat() if updated_dt else None

    # --- status label (comme ton template)
    status = (o.status or "").strip()
    status_label = status
    if status == "pending":
        status_label = "⏳ En attente"
    elif status == "in_progress":
        status_label = "🚦 En cours"
    elif status == "done":
        status_label = "✅ Terminée"
    elif status == "canceled":
        status_label = "❌ Annulée"
    else:
        try:
            status_label = f"📦 {o.get_status_display()}"
        except Exception:
            status_label = status or "—"

    # --- payment label
    payment_label = None
    try:
        if getattr(o, "payment_status", None):
            payment_label = o.get_payment_status_display() or o.payment_status
    except Exception:
        payment_label = getattr(o, "payment_status", None) or None

    # --- driver HTML (même logique que le template)
    if o.delivery_partner:
        driver_html = f"🚴 {o.delivery_partner.name or 'Livreur'}"
        if getattr(o.delivery_partner, "phone", None):
            driver_html += f" – {o.delivery_partner.phone}"
    else:
        reason = getattr(o, "delivery_partner_unassigned_reason", None) or ""
        if reason:
            driver_html = f'<span class="muted">⚠️ Aucun livreur assigné</span><br><span class="muted">{reason}</span>'
        else:
            driver_html = (
                '<span class="muted">⚠️ Aucun livreur assigné</span><br>'
                '<span class="muted">Probablement aucun livreur disponible ou coordonnées client incomplètes.</span>'
            )

    # --- laundry HTML
    if o.laundry_partner:
        laundry_html = f"🧺 {o.laundry_partner.name or 'Blanchisserie partenaire'}"
    else:
        laundry_html = '<span class="muted">Non encore affectée</span>'

    return JsonResponse({
        "ok": True,
        "updated_at": updated_at,
        "status": status,
        "status_label": status_label,
        "payment_label": payment_label,
        "driver_html": driver_html,
        "laundry_html": laundry_html,
    })


# -------------------------------------------------------------------
# Home client (LISTE)  ✅ FIX: on filtre par phone (pas par customer_id)
# + UI: total client + paiement (safe fallbacks)
# -------------------------------------------------------------------

@ensure_csrf_cookie
@client_login_required
def client_home(request):
    """
    Page Accueil Client:
    - Liste des commandes + filtres statut + pagination
    - Totaux robustes (fallback sur items_total)
    - KPI commandes conditionnel selon filtre sélectionné
    """
    from django.core.paginator import Paginator
    from django.db.models import Sum, F, Value, DecimalField, Q
    from django.db.models.functions import Coalesce, Cast

    phone = _client_phone(request)
    if not phone:
        return redirect("orders:client_login")

    customer = Customer.objects.filter(phone=phone).first()

    # Defaults safe
    orders_qs = Order.objects.none()
    orders_count_all = 0
    active_orders_count = 0
    unpaid_orders_count = 0
    wallet_balance = Decimal("0.00")

    if customer:
        items_sum = Coalesce(
            Sum(
                Cast(F("items__quantity"), DecimalField(max_digits=10, decimal_places=2))
                * Cast(F("items__unit_price"), DecimalField(max_digits=10, decimal_places=2)),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=2)),
        )

        base_qs = (
            Order.objects.filter(customer=customer)
            .order_by("-created_at")
            .annotate(items_total=items_sum)
            .annotate(
                total_client_display=Value(
                    0,
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            )
        )

        # KPI globaux (sans filtre)
        # ✅ Pilote: masquer les commandes "vides" (tous montants à 0 / pas d'items)
        nonempty_qs = base_qs.filter(
            Q(total_client_ttc__gt=0)
            | Q(prestation_total__gt=0)
            | Q(items_total__gt=0)
            | Q(service_fee__gt=0)
            | Q(delivery_fee__gt=0)
            | Q(express_extra_fee__gt=0)
            | Q(vat_fagni__gt=0)
        )

        orders_count_all = nonempty_qs.count()
        active_orders_count = nonempty_qs.filter(status__in=["pending", "in_progress"]).count()
        unpaid_orders_count = nonempty_qs.filter(payment_status__in=["pending", "unpaid", "failed"]).count()

        orders_qs = nonempty_qs

        try:
            wallet_obj = get_or_create_wallet_for_customer(customer)
            if wallet_obj and getattr(wallet_obj, "balance", None) is not None:
                wallet_balance = Decimal(str(wallet_obj.balance or 0))
        except Exception:
            wallet_balance = Decimal("0.00")

    # ---------------------------
    # Filtre statut
    # ---------------------------
    current_status = (request.GET.get("status") or "all").strip().lower()
    valid = {"all", "pending", "in_progress", "done", "canceled"}
    if current_status not in valid:
        current_status = "all"

    filtered_qs = orders_qs
    if current_status != "all":
        filtered_qs = orders_qs.filter(status=current_status)

    kpi_orders_count = orders_count_all if current_status == "all" else filtered_qs.count()

    # ---------------------------
    # Pagination
    # ---------------------------
    per_page = 4
    page_number = request.GET.get("page") or 1

    paginator = Paginator(filtered_qs, per_page)
    page_obj = paginator.get_page(page_number)
    is_paginated = paginator.num_pages > 1

    current = page_obj.number
    last = paginator.num_pages
    window = 2
    start = max(2, current - window)
    end = min(last - 1, current + window)
    pages = list(range(start, end + 1)) if last >= 3 else []
    show_left_ellipsis = start > 2
    show_right_ellipsis = end < (last - 1)

    # --- post-traitement : total_client_display canonique pour la page ---
    orders_page = list(page_obj.object_list)
    for o in orders_page:
        pricing = _compute_order_pricing(o)
        o.total_client_display = pricing["total_client"]
        o.express_extra_fee = pricing.get("express_extra_fee", Decimal("0"))

    # LOT 17B — adresse courte pour affichage home client
    def _short_client_address(v):
        raw = str(v or "").strip()
        if not raw:
            return "—"

        parts = [x.strip() for x in raw.split(",") if x.strip()]

        if len(parts) >= 4:
            return ", ".join(parts[-4:-2])

        if len(parts) >= 2:
            return ", ".join(parts[-2:])

        return raw if raw else "Adresse à confirmer"

    for _o in orders_page:
        full_addr = (
            getattr(_o, "pickup_address", None)
            or getattr(getattr(_o, "customer", None), "address", None)
            or ""
        )
        try:
            _o.pickup_address_short = _short_client_address(full_addr)
        except Exception:
            _o.pickup_address_short = "Adresse à confirmer"

    ctx = {
        "phone": phone,
        "customer": customer,

        # liste paginée
        "orders": page_obj,

        # pagination
        "page_obj": page_obj,
        "is_paginated": is_paginated,
        "pages": pages,
        "show_left_ellipsis": show_left_ellipsis,
        "show_right_ellipsis": show_right_ellipsis,

        # filtre
        "current_status": current_status,

        # KPI
        "orders_count_all": orders_count_all,
        "kpi_orders_count": kpi_orders_count,
        "active_orders_count": active_orders_count,
        "unpaid_orders_count": unpaid_orders_count,

        # wallet
        "wallet_balance": wallet_balance,
    }

    return render(request, "orders/client_home.html", ctx)



# -------------------------------------------------------------------
# Nouvelle commande client
# -------------------------------------------------------------------

@require_http_methods(["GET", "POST"])
@client_required
def client_new_order(request):
    """
    Création Client V1 (rapide) :
    - téléphone (session)
    - nom + adresse + notes
    - optionnel : coordonnées GPS + place_id + source géoloc
    -> crée une commande "pending"
    """
    phone = _client_phone(request)
    customer = Customer.objects.filter(phone=phone).order_by("-id").first()

    error = None
    name_init = customer.name if customer else ""
    address_init = getattr(customer, "address", "") if customer else ""

    def _clean_str(v):
        return (v or "").strip()

    def _clean_float(v):
        s = _clean_str(v)
        if not s:
            return None
        try:
            return float(s.replace(",", "."))
        except Exception:
            return None

    customer_field_names = {
        f.name for f in Customer._meta.get_fields()
        if getattr(f, "concrete", False)
    }
    order_field_names = {
        f.name for f in Order._meta.get_fields()
        if getattr(f, "concrete", False)
    }

    if request.method == "POST":
        name = _clean_str(request.POST.get("name"))
        address = clean_address_or_empty(_clean_str(request.POST.get("address")))
        notes = _clean_str(request.POST.get("notes"))

        pickup_lat = _clean_float(request.POST.get("pickup_lat"))
        pickup_lng = _clean_float(request.POST.get("pickup_lng"))
        pickup_place_id = _clean_str(request.POST.get("pickup_place_id"))
        pickup_geo_source = _clean_str(request.POST.get("pickup_geo_source"))
        pickup_geo_label = _clean_str(request.POST.get("pickup_geo_label"))

        if not name:
            error = "Merci de renseigner ton nom."
        elif not is_probably_valid_address(address):
            error = "Merci de renseigner une adresse plus précise."
        else:
            customer_defaults = {
                "name": name,
                "address": address,
            }

            # Compat : on ne remplit que les champs qui existent réellement
            if "lat" in customer_field_names and pickup_lat is not None:
                customer_defaults["lat"] = pickup_lat
            if "lng" in customer_field_names and pickup_lng is not None:
                customer_defaults["lng"] = pickup_lng
            if "latitude" in customer_field_names and pickup_lat is not None:
                customer_defaults["latitude"] = pickup_lat
            if "longitude" in customer_field_names and pickup_lng is not None:
                customer_defaults["longitude"] = pickup_lng
            if "place_id" in customer_field_names and pickup_place_id:
                customer_defaults["place_id"] = pickup_place_id
            if "geo_source" in customer_field_names and pickup_geo_source:
                customer_defaults["geo_source"] = pickup_geo_source
            if "geo_label" in customer_field_names and pickup_geo_label:
                customer_defaults["geo_label"] = pickup_geo_label

            customer, _created = Customer.objects.update_or_create(
                phone=phone,
                defaults=customer_defaults,
            )

            # -------------------------------------------------
            # LOT 22B : rattachement sponsor si code referral entrant
            # -------------------------------------------------
            incoming_ref = _client_ref_code(request)
            existing_profile = (
                ReferralLink.objects
                .select_related("customer", "sponsor")
                .filter(customer=customer)
                .order_by("-id")
                .first()
            )

            if not existing_profile:
                sponsor_profile = None
                if incoming_ref:
                    sponsor_profile = (
                        ReferralLink.objects
                        .select_related("customer")
                        .filter(referral_code=incoming_ref)
                        .exclude(customer=customer)
                        .order_by("-id")
                        .first()
                    )

                def _clean_phone_for_code(value: str) -> str:
                    raw = re.sub(r"\D+", "", str(value or ""))
                    if raw:
                        return raw[-8:]
                    return f"{customer.id:04d}"

                def _build_customer_referral_code() -> str:
                    base = f"FAGNI-{_clean_phone_for_code(getattr(customer, 'phone', ''))}"
                    code_ = base
                    i = 2
                    while ReferralLink.objects.exclude(customer=customer).filter(referral_code=code_).exists():
                        code_ = f"{base}-{i}"
                        i += 1
                    return code_

                ReferralLink.objects.create(
                    customer=customer,
                    referral_code=_build_customer_referral_code(),
                    actor_type="client",
                    sponsor=sponsor_profile,
                )

            code = uuid.uuid4().hex[:8].upper()
            while Order.objects.filter(code=code).exists():
                code = uuid.uuid4().hex[:8].upper()

            order_kwargs = {
                "is_draft": True,
                "customer": customer,
                "status": "pending",
                "payment_status": "unpaid",
                "amount_paid": 0,
                "code": code,
                "notes": notes or None,
                "pickup_address": address,
                "delivery_address": address,
            }

            # Compat champs coordonnées / géoloc côté Order
            if "pickup_lat" in order_field_names and pickup_lat is not None:
                order_kwargs["pickup_lat"] = pickup_lat
            if "pickup_lng" in order_field_names and pickup_lng is not None:
                order_kwargs["pickup_lng"] = pickup_lng
            if "pickup_place_id" in order_field_names and pickup_place_id:
                order_kwargs["pickup_place_id"] = pickup_place_id
            if "pickup_geo_source" in order_field_names and pickup_geo_source:
                order_kwargs["pickup_geo_source"] = pickup_geo_source
            if "pickup_geo_label" in order_field_names and pickup_geo_label:
                order_kwargs["pickup_geo_label"] = pickup_geo_label

            if "delivery_lat" in order_field_names and pickup_lat is not None:
                order_kwargs["delivery_lat"] = pickup_lat
            if "delivery_lng" in order_field_names and pickup_lng is not None:
                order_kwargs["delivery_lng"] = pickup_lng
            if "delivery_place_id" in order_field_names and pickup_place_id:
                order_kwargs["delivery_place_id"] = pickup_place_id

            order = Order.objects.create(**order_kwargs)

            from orders.models import DeliveryLeg

            DeliveryLeg.objects.get_or_create(
                order=order,
                leg_type="pickup",
                defaults={"status": "pending"},
            )
            DeliveryLeg.objects.get_or_create(
                order=order,
                leg_type="return",
                defaults={"status": "pending"},
            )

            try:
                order.update_financials(save=True)
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=5448")

            try:
                _ensure_pickup_mission_for_order(order)
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=5453")


            upsell, _ = OrderUpsell.objects.get_or_create(order=order)
            upsell.express_24h = bool(request.POST.get("express_24h"))
            upsell.premium_ironing = bool(request.POST.get("premium_ironing"))
            upsell.fragrance = bool(request.POST.get("fragrance"))
            upsell.delicate_care = bool(request.POST.get("delicate_care"))
            upsell.save()

            try:
                upsell_total = upsell.total
            except Exception:
                upsell_total = Decimal("0.00")

            try:
                base_total = Decimal(str(getattr(order, "total_client_ttc", 0) or 0))
            except Exception:
                base_total = Decimal("0.00")

            order.total_client_ttc = (base_total + upsell_total).quantize(Decimal("0.01"))
            try:
                order.save(update_fields=["total_client_ttc"])
            except Exception:
                order.save()

            
            request.session[f"upsell_data_{order.id}"] = {
                "express_24h": bool(request.POST.get("express_24h")),
                "premium_ironing": bool(request.POST.get("premium_ironing")),
                "fragrance": bool(request.POST.get("fragrance")),
                "delicate_care": bool(request.POST.get("delicate_care")),
            }

            return redirect("orders:client_new_order_step2", order_id=order.id)

    return render(request, "orders/client_new_order.html", {
        "phone": phone,
        "customer": customer,
        "error": error,
        "name_init": name_init,
        "address_init": address_init,
        "google_maps_api_key": getattr(settings, "GOOGLE_MAPS_API_KEY", ""),
    })



def _ensure_pickup_mission_for_order(order):
    """
    Crée une mission logistique de collecte si elle n'existe pas déjà.
    Retourne la mission ou None.
    """
    try:
        from logistics.models import Mission
        existing = Mission.objects.filter(
            order=order,
            mission_type="pickup_from_customer",
        ).first()
        if existing:
            return existing

        from logistics.adapters import (
            get_order_source_address_v2,
            get_order_destination_address_v2,
            get_order_contact_name,
            get_order_contact_phone,
        )
        from logistics.services import create_mission_for_order

        source_address = get_order_source_address_v2(order)
        destination_address = get_order_destination_address_v2(order)

        if source_address is None or destination_address is None:
            return None

        mission = create_mission_for_order(
            order=order,
            mission_type="pickup_from_customer",
            source_address=source_address,
            destination_address=destination_address,
            contact_name=get_order_contact_name(order),
            contact_phone=get_order_contact_phone(order),
            priority="normal",
            instructions="Mission de collecte créée automatiquement depuis la commande",
            sequence_index=1,
        )
        return mission
    except Exception:
        return None


# -------------------------------------------------------------------
# Helpers: montants & timeline (client)
# -------------------------------------------------------------------

def _to_dec(v) -> Decimal:
    try:
        return Decimal(str(v or "0"))
    except Exception:
        return Decimal("0")


def _client_order_amounts(order: Order) -> dict:
    """
    Bridge client legacy -> pricing engine canonique.
    """
    finance = build_order_finance_summary(order)
    pricing = _compute_order_pricing(order)
    adjusted_total = pricing.get("total_client_ttc", finance.get("total_client_ttc", Decimal("0")))
    child_referral_discount = pricing.get("child_referral_discount", Decimal("0"))

    try:
        amount_paid = Decimal(str(finance.get("amount_paid", Decimal("0")) or 0))
        amount_remaining = Decimal(str(adjusted_total or 0)) - amount_paid
        if amount_remaining < 0:
            amount_remaining = Decimal("0")
    except Exception:
        amount_remaining = finance.get("amount_remaining", Decimal("0"))

    return {
        "prestation_total": finance.get("prestation_total", Decimal("0")),
        "delivery_fee": finance.get("delivery_fee_client", Decimal("0")),
        "delivery_fee_client": finance.get("delivery_fee_client", Decimal("0")),
        "service_fee_ht": finance.get("service_fee_ht", Decimal("0")),
        "vat_fagni": finance.get("vat_fagni", Decimal("0")),
        "child_referral_discount": child_referral_discount,
        "total_ttc": adjusted_total,
        "total_client_ttc": adjusted_total,
        "amount_paid": amount_paid,
        "amount_remaining": amount_remaining,
    }

# --- LOT_2_32_PAYMENT_UI_JSON_VIEW_OK ---
@client_login_required
def client_order_payment_ui_json(request, order_id):
    """
    Endpoint JSON (GET) pour polling UI paiement côté client.
    - ne modifie PAS la commande
    - renvoie un statut canonique basé sur total_ttc + amount_paid
    """
    from decimal import Decimal
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404

    order = get_object_or_404(Order, id=order_id)

    
    
    # Sécurité alignée sur client_order_detail (phone session)
    phone = _client_phone(request)
    if not phone:
        return JsonResponse({"ok": False, "error": "not_authenticated"}, status=401)

    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)


    snapshot = build_order_canonical_snapshot(order)

    try:
        total = Decimal(str(snapshot.get("total_client_ttc", 0) or 0))
    except Exception:
        total = Decimal("0")
    try:
        paid = Decimal(str(snapshot.get("amount_paid", 0) or 0))
    except Exception:
        paid = Decimal("0")
    try:
        remain = Decimal(str(snapshot.get("amount_due", 0) or 0))
    except Exception:
        remain = Decimal("0")

    if paid < 0:
        paid = Decimal("0")
    if total < 0:
        total = Decimal("0")
    if remain < 0:
        remain = Decimal("0")

    if total <= 0:
        payment_ui = "waiting_amount"
        payment_status_ui = "unpaid"
    elif not snapshot.get("can_pay", False) or paid >= total:
        payment_ui = "paid"
        payment_status_ui = "paid"
    elif paid <= 0:
        payment_ui = "unpaid"
        payment_status_ui = "unpaid"
    else:
        payment_ui = "partial"
        payment_status_ui = "partial"

    wave_declared = bool(request.session.get(f"wave_declared_{order.id}"))

    # LOT_2_34_WAVE_DECLARED_AUTOCLEAR_OK
    # Si la commande est payée, on nettoie le flag "declared" pour éviter un état UI incohérent
    if payment_ui == "paid" and wave_declared:
        try:
            del request.session[f"wave_declared_{order.id}"]
        except KeyError:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=5659")
        wave_declared = False


    resp = JsonResponse({
        "ok": True,
        "order_id": order.id,
        "payment_ui": payment_ui,
        "payment_status_ui": payment_status_ui,
        "total_ttc": float(total),
        "amount_paid": float(paid),
        "remain": float(remain),
        "wave_declared": wave_declared,
    })
    resp["Cache-Control"] = "no-store"
    return resp


def _client_order_timeline(order: Order) -> list:
    """
    Timeline client pilotée prioritairement par les DeliveryLegs,
    avec fallback doux sur quelques timestamps / status bruts.
    """
    created = getattr(order, "created_at", None)
    t_pickup = getattr(order, "pickup_time", None)
    t_wash = getattr(order, "wash_complete_time", None)
    t_deliv = getattr(order, "delivered_time", None)

    st = (getattr(order, "status", "") or "").strip().lower()

    try:
        legs = list(order.legs.all())
    except Exception:
        legs = []

    pickup_leg = None
    return_leg = None
    for leg in legs:
        lt = (getattr(leg, "leg_type", None) or "").strip().lower()
        if lt == "pickup" and pickup_leg is None:
            pickup_leg = leg
        elif lt == "return" and return_leg is None:
            return_leg = leg

    pickup_status = (getattr(pickup_leg, "status", None) or "").strip().lower()
    return_status = (getattr(return_leg, "status", None) or "").strip().lower()

    pickup_done = bool(t_pickup) or pickup_status == "done"
    pickup_active = pickup_status in {"assigned", "in_progress"}

    return_done = bool(t_deliv) or return_status == "done"
    return_active = return_status in {"assigned", "in_progress"}

    wash_done = bool(t_wash)
    wash_active = False

    # Si collecte terminée et retour pas encore lancé/fini => traitement
    if pickup_done and not return_done and not return_active:
        wash_active = True

    # Fallback sur statut brut si besoin
    if st == "done":
        pickup_done = True
        wash_done = True
        return_done = True
        pickup_active = False
        wash_active = False
        return_active = False
    elif st in {"in_progress", "processing"}:
        if not pickup_done and not return_active and not return_done:
            pickup_active = True

    steps = [
        {
            "key": "created",
            "label": "Commande créée",
            "done": bool(created),
            "active": False,
            "ts": created,
        },
        {
            "key": "pickup",
            "label": "Collecte effectuée" if pickup_done else "Collecte en cours",
            "done": pickup_done,
            "active": (not pickup_done and pickup_active),
            "ts": t_pickup,
        },
        {
            "key": "wash",
            "label": "Lavage terminé" if wash_done else "Traitement en cours",
            "done": wash_done,
            "active": (not wash_done and wash_active),
            "ts": t_wash,
        },
        {
            "key": "delivered",
            "label": "Livrée" if return_done else ("Livraison en cours" if return_active else "En attente de livraison"),
            "done": return_done,
            "active": (not return_done and return_active),
            "ts": t_deliv,
        },
    ]

    # S'il n'y a aucune étape active explicite, active la première non terminée
    if not any(step.get("active") for step in steps):
        for step in steps:
            if not step.get("done"):
                step["active"] = True
                break

    return steps


# -------------------------------------------------------------------
# Détail client ✅ FIX: sécurisation par phone, pas customer_id
# -------------------------------------------------------------------

@ensure_csrf_cookie
@client_login_required
def client_order_detail(request, order_id: int):
    phone = _client_phone(request)
    if not phone:
        return redirect("orders:client_login")

    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        return redirect("orders:client_home")

    customer = getattr(order, "customer", None)

    amounts = _client_order_amounts(order)
    snapshot = build_order_canonical_snapshot(order)
    legs = []
    try:
        legs = list(
            order.delivery_legs.all().values(
                "id", "leg_type", "status", "driver_amount", "driver_id"
            )
        )
    except Exception:
        legs = []

    legs_ui = []
    try:
        driver_ids = [x.get("driver_id") for x in legs if x.get("driver_id")]
        driver_name_by_id = {}
        if driver_ids:
            from partners.models import DeliveryPartner
            for d in DeliveryPartner.objects.filter(id__in=driver_ids):
                driver_name_by_id[d.id] = getattr(d, "name", "") or f"Livreur #{d.id}"

        def _leg_type_label(v):
            return {
                "pickup": "Collecte",
                "laundry_in": "Arrivée blanchisserie",
                "laundry_out": "Sortie blanchisserie",
                "return": "Livraison",
            }.get(v, "Logistique")

        def _client_leg_status_label(v):
            return {
                "pending": ("En attente", "pending"),
                "assigned": ("Affectée", "progress"),
                "accepted": ("Acceptée", "progress"),
                "picked_up": ("Collectée", "done"),
                "in_progress": ("En cours", "progress"),
                "done": ("Terminée", "done"),
                "completed": ("Terminée", "done"),
                "canceled": ("Annulée", "danger"),
            }.get(v, ("En attente", "pending"))

        for leg in legs:
            st_label, st_class = _client_leg_status_label(leg.get("status"))
            legs_ui.append({
                **leg,
                "leg_type_label": _leg_type_label(leg.get("leg_type")),
                "status_label": st_label,
                "status_class": st_class,
                "driver_name": driver_name_by_id.get(leg.get("driver_id"), "Livreur partenaire"),
            })
    except Exception:
        legs_ui = []

    items = []
    try:
        for it in order.items.all():
            items.append({
                "label": (
                getattr(it, "designation", None)
                or getattr(getattr(it, "service", None), "name", None)
                or getattr(it, "service_type", None)
                or "Article"
            ),
                "qty": getattr(it, "quantity", 0) or 0,
                "price": getattr(it, "unit_price", 0) or 0,
            })
    except Exception:
        items = []

    timeline = [
        {
            "key": "created",
            "label": "Commande créée",
            "done": True,
            "active": False,
            "ts": getattr(order, "created_at", None),
        },
        {
            "key": "pickup",
            "label": "Collecte en cours" if snapshot.get("status_canonical") in ["submitted", "in_processing"] else "Collecte",
            "done": False,
            "active": snapshot.get("status_canonical") in ["submitted", "in_processing"],
            "ts": None,
        },
        {
            "key": "wash",
            "label": "Traitement en cours",
            "done": snapshot.get("status_canonical") == "ready",
            "active": snapshot.get("status_canonical") == "in_processing",
            "ts": None,
        },
        {
            "key": "delivered",
            "label": "En attente de livraison",
            "done": snapshot.get("status_canonical") == "done",
            "active": snapshot.get("status_canonical") == "ready",
            "ts": None,
        },
    ]

    hero_title = "Suivi de commande"
    hero_text = snapshot.get("display_address") or "Commande en cours"

    live_tracking_title = "Suivi en cours"
    live_tracking_text = "Ta commande est en cours de traitement."

    try:
        if snapshot.get("status_canonical") == "in_processing":
            live_tracking_title = "Collecte en cours"
            live_tracking_text = "Le livreur affecté gère actuellement la collecte de ta commande."
            hero_title = "Collecte planifiée"
            hero_text = "Un livreur va passer récupérer ton linge."
        elif snapshot.get("status_canonical") == "ready":
            live_tracking_title = "Commande prête"
            live_tracking_text = "La commande est prête pour la livraison."
        elif snapshot.get("status_canonical") == "done":
            live_tracking_title = "Commande terminée"
            live_tracking_text = "Ta commande a été finalisée."
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=5913")

    partner_tracking = {
        "pickup": {
            "title": "Collecte",
            "icon": "🚚",
            "partner": {"name": "Livreur partenaire", "phone": "", "city": "", "address": "", "is_assigned": False},
            "status": "En attente",
            "status_raw": "pending",
        },
        "laundry": {
            "title": "Blanchisserie",
            "icon": "🧺",
            "partner": {"name": "Blanchisserie partenaire FAGNI", "phone": "", "city": "", "address": "", "is_assigned": False},
            "status": "En attente",
            "status_raw": "pending",
        },
        "return": {
            "title": "Livraison",
            "icon": "📦",
            "partner": {"name": "Livreur partenaire", "phone": "", "city": "", "address": "", "is_assigned": False},
            "status": "En attente",
            "status_raw": "pending",
        },
    }

    map_summary = {
        "pickup_address": snapshot.get("display_address") or "Adresse non renseignée",
        "driver_name": "Livreur partenaire",
        "laundry_name": "Blanchisserie partenaire FAGNI",
    }

    try:
        paid = Decimal(str(snapshot.get("amount_paid", 0) or 0))
    except Exception:
        paid = DECIMAL_ZERO

    try:
        total_client_ttc = Decimal(str(snapshot.get("total_client_ttc", 0) or 0))
    except Exception:
        total_client_ttc = DECIMAL_ZERO

    wallet_used = paid
    if wallet_used < DECIMAL_ZERO:
        wallet_used = DECIMAL_ZERO
    if total_client_ttc > DECIMAL_ZERO and wallet_used > total_client_ttc:
        wallet_used = total_client_ttc

    try:
        display_summary
    except NameError:
        display_summary = build_order_display_summary(order)

    try:
        finance_summary
    except NameError:
        finance_summary = build_order_finance_summary(order)

    ctx = {
        "phone": phone,
        "customer": customer,
        "display_address": snapshot.get("display_address") or "Adresse non renseignée",
        "order": order,
        "legs": legs,
        "legs_ui": legs_ui,
        "items": items,
        "amounts": amounts,
        "display_summary": display_summary,
        "finance_summary": finance_summary,
        "payment_status": snapshot.get("payment_status_raw"),
        "payment_ui": snapshot.get("payment_status_canonical"),
        "payment_label": snapshot.get("payment_label"),
        "pay_pill_class": "pay-ok" if snapshot.get("is_paid") else "pay-wait",
        "paid_amount": paid,
        "wallet_used": wallet_used,
        "total_amount": snapshot.get("total_client_ttc", DECIMAL_ZERO),
        "remain_amount": snapshot.get("amount_due", DECIMAL_ZERO),
        "timeline": timeline,
        "hero_title": hero_title,
        "hero_text": hero_text,
        "snapshot": snapshot,
        "partner_tracking": partner_tracking,
        "live_tracking_title": live_tracking_title,
        "live_tracking_text": live_tracking_text,
        "tracking_title": live_tracking_title,
        "tracking_text": live_tracking_text,
        "map_summary": map_summary,
    }

    try:
        ctx.update(build_order_display_context(order, customer, amounts, paid))
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=6005")

    resp = render(request, "orders/client_order_detail.html", ctx)
    resp["Cache-Control"] = "no-store"
    return resp


def client_order_detail_json(request, order_id: int):
    """
    JSON léger pour rafraîchir l'état paiement / résumé commande côté page détail client.
    """
    phone = _client_phone(request)
    if not phone:
        resp = JsonResponse({"ok": False, "error": "not_authenticated"}, status=401)
        resp["Cache-Control"] = "no-store"
        return resp

    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        resp = JsonResponse({"ok": False, "error": "order_not_found"}, status=404)
        resp["Cache-Control"] = "no-store"
        return resp

    snapshot = build_order_canonical_snapshot(order)
    finance_summary = build_order_finance_summary(order)

    try:
        total = Decimal(str(finance_summary.get("total_client_ttc", 0) or 0))
    except Exception:
        total = DECIMAL_ZERO

    try:
        paid = Decimal(str(finance_summary.get("amount_paid", 0) or 0))
    except Exception:
        paid = DECIMAL_ZERO

    try:
        remain = Decimal(str(finance_summary.get("amount_remaining", 0) or 0))
    except Exception:
        remain = DECIMAL_ZERO

    if total < DECIMAL_ZERO:
        total = DECIMAL_ZERO
    if paid < DECIMAL_ZERO:
        paid = DECIMAL_ZERO
    if remain < DECIMAL_ZERO:
        remain = DECIMAL_ZERO

    if total <= DECIMAL_ZERO:
        payment_ui = "waiting_amount"
        pay_pill_class = "pay-wait"
        payment_label = "En attente"
    elif not snapshot.get("can_pay", False) and paid >= total:
        payment_ui = "paid"
        pay_pill_class = "pay-ok"
        payment_label = "Payé"
    elif paid <= DECIMAL_ZERO:
        payment_ui = "unpaid"
        pay_pill_class = "pay-wait"
        payment_label = "Paiement en attente"
    else:
        payment_ui = "partial"
        pay_pill_class = "pay-wait"
        payment_label = "Paiement partiel"

    payload = {
        "ok": True,
        "order_id": order.id,
        "code": snapshot.get("code"),
        "status": {
            "raw": snapshot.get("status_raw"),
            "canonical": snapshot.get("status_canonical"),
            "label": snapshot.get("status_label"),
        },
        "payment": {
            "raw": snapshot.get("payment_status_raw"),
            "canonical": snapshot.get("payment_status_canonical"),
            "label": payment_label,
            "ui": payment_ui,
            "pill_class": pay_pill_class,
            "can_pay": bool(snapshot.get("can_pay", False)),
            "is_paid": bool(snapshot.get("is_paid", False)),
        },
        "amounts": {
            "total_ttc": float(total),
            "amount_paid": float(paid),
            "amount_due": float(remain),
        },
    }

    resp = JsonResponse(payload)
    resp["Cache-Control"] = "no-store"
    return resp



@require_POST
@client_login_required
def client_order_pay_simulate(request, order_id: int):

    """
    Paiement simulé strictement réservé aux environnements de développement.

    Cette route ne doit jamais pouvoir créer un paiement lorsque DEBUG=False.
    """
    if not getattr(settings, "DEBUG", False):
        resp = JsonResponse(
            {
                "ok": False,
                "error": "simulate_disabled",
                "message": "Le paiement simulé est désactivé sur cet environnement.",
            },
            status=403,
        )
        resp["Cache-Control"] = "no-store"
        return resp

    phone = _client_phone(request)
    if not phone:
        resp = JsonResponse({"ok": False, "error": "not_authenticated"}, status=401)
        resp["Cache-Control"] = "no-store"
        return resp

    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        resp = JsonResponse({"ok": False, "error": "order_not_found"}, status=404)
        resp["Cache-Control"] = "no-store"
        return resp

    if getattr(order, "status", None) == "canceled":
        resp = JsonResponse(
            {
                "ok": False,
                "error": "order_canceled",
                "message": "Une commande annulée ne peut pas être payée.",
            },
            status=409,
        )
        resp["Cache-Control"] = "no-store"
        return resp

    snapshot = build_order_canonical_snapshot(order)
    finance_summary = build_order_finance_summary(order)

    try:
        total = Decimal(str(finance_summary.get("total_client_ttc", 0) or 0))
    except Exception:
        total = DECIMAL_ZERO

    if total <= DECIMAL_ZERO:
        resp = JsonResponse({"ok": False, "error": "no_total_amount"}, status=400)
        resp["Cache-Control"] = "no-store"
        return resp

    try:
        paid_now = Decimal(str(finance_summary.get("amount_paid", 0) or 0))
        if paid_now < DECIMAL_ZERO:
            paid_now = DECIMAL_ZERO
    except Exception:
        paid_now = DECIMAL_ZERO

    if not snapshot.get("can_pay", False) or paid_now >= total:
        resp = JsonResponse({"ok": False, "error": "already_paid"}, status=409)
        resp["Cache-Control"] = "no-store"
        return resp

    resp = JsonResponse(
        {
            "ok": False,
            "error": "payment_simulation_disabled",
            "message": (
                "La simulation de paiement est désactivée. "
                "Aucun paiement fictif ne peut être enregistré."
            ),
        },
        status=403,
    )
    resp["Cache-Control"] = "no-store"
    return resp

    try:
        order.refresh_from_db()
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=6168")

    try:
        order.refresh_from_db()
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=6173")

    snap = build_order_canonical_snapshot(order)
    finance_summary = build_order_finance_summary(order)

    resp = JsonResponse({
        "ok": True,
        "payment_ui": "paid" if snap.get("payment_status_canonical") == "paid" else "partial",
        "payment_status": snap.get("payment_status_canonical"),
        "amounts": {
            "total_ttc": float(Decimal(str(finance_summary.get("total_client_ttc", 0) or 0))),
            "amount_paid": float(Decimal(str(finance_summary.get("amount_paid", 0) or 0))),
            "amount_remaining": float(Decimal(str(finance_summary.get("amount_remaining", 0) or 0))),
        },
    })
    resp["Cache-Control"] = "no-store"
    return resp

def _parse_money_amount(raw: str) -> Decimal:
    raw = (raw or "").strip()
    if not raw:
        return DECIMAL_ZERO
    raw = raw.replace(" ", "").replace("\u00A0", "")
    raw = raw.replace(",", ".")
    # garder uniquement chiffres + point
    raw = re.sub(r"[^0-9\.]", "", raw)
    if raw.count(".") > 1:
        # si plusieurs points, garder le 1er
        first = raw.find(".")
        raw = raw[:first+1] + raw[first+1:].replace(".", "")
    try:
        return Decimal(raw)
    except Exception:
        return DECIMAL_ZERO


@require_POST
@client_login_required
def client_order_pay_cash(request, order_id: int):
    """
    Ancien point d'entrée cash côté client.

    Le client ne peut jamais comptabiliser lui-même un paiement cash.
    La confirmation doit provenir d'un acteur habilité ou des opérations
    FAGNI, via un parcours vérifié.
    """
    phone = _client_phone(request)
    if not phone:
        resp = JsonResponse(
            {"ok": False, "error": "not_authenticated"},
            status=401,
        )
        resp["Cache-Control"] = "no-store"
        return resp

    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )

    if not order:
        resp = JsonResponse(
            {"ok": False, "error": "order_not_found"},
            status=404,
        )
        resp["Cache-Control"] = "no-store"
        return resp

    if getattr(order, "status", None) == "canceled":
        resp = JsonResponse(
            {
                "ok": False,
                "error": "order_canceled",
                "message": "Une commande annulée ne peut pas être payée.",
            },
            status=409,
        )
        resp["Cache-Control"] = "no-store"
        return resp

    try:
        total = Decimal(
            str(getattr(order, "total_client_ttc", 0) or 0)
        )
    except Exception:
        total = DECIMAL_ZERO

    try:
        paid_now = Decimal(
            str(getattr(order, "amount_paid", 0) or 0)
        )
    except Exception:
        paid_now = DECIMAL_ZERO

    if (
        getattr(order, "payment_status", None) == "paid"
        or (total > DECIMAL_ZERO and paid_now >= total)
    ):
        resp = JsonResponse(
            {"ok": False, "error": "already_paid"},
            status=409,
        )
        resp["Cache-Control"] = "no-store"
        return resp

    # Sécurité P0 : une déclaration provenant uniquement du client
    # ne doit jamais alimenter Payment, source de vérité comptable.
    resp = JsonResponse(
        {
            "ok": False,
            "error": "cash_requires_verification",
            "message": (
                "Le paiement cash doit être confirmé par un livreur "
                "habilité ou par les opérations FAGNI."
            ),
        },
        status=403,
    )
    resp["Cache-Control"] = "no-store"
    return resp


def apply_order_payment(
    order,
    add_amount,
    *,
    channel="manual",
    reference="",
    note="",
):
    """
    Orchestrateur canonique des paiements FAGNI.

    Payment est la source de vérité comptable.
    Order.amount_paid et Order.payment_status sont des projections
    synchronisées à partir des Payment.

    Retourne :
    {
        "applied": Decimal,
        "amount_paid": Decimal,
        "remaining": Decimal,
        "payment_status": str,
        "became_paid": bool,
        "already_settled": bool,
        "already_applied": bool,
    }
    """
    from decimal import Decimal

    from django.core.exceptions import ValidationError
    from django.db import transaction
    from django.db.models import Sum

    from orders.models import OrderPaymentEvent, Payment

    try:
        requested_amount = Decimal(str(add_amount or 0))
    except Exception as exc:
        raise ValidationError(
            "Paiement refusé : montant invalide."
        ) from exc

    reference = (reference or "").strip()
    channel = (channel or "manual").strip() or "manual"

    allowed_channels = {
        "wallet",
        "wallet_auto",
        "wave_webhook",
        "wave_ops",
        "wave_manual_verified",
        "cash",
        "manual",
    }

    if channel not in allowed_channels:
        raise ValidationError(
            f"Canal de paiement non autorisé : {channel}."
        )

    if channel in {
        "wave_webhook",
        "wave_ops",
        "wave_manual_verified",
    } and not reference:
        raise ValidationError(
            "Une référence Wave vérifiée est obligatoire."
        )

    if requested_amount <= DECIMAL_ZERO:
        return {
            "applied": DECIMAL_ZERO,
            "amount_paid": Decimal(
                str(getattr(order, "amount_paid", 0) or 0)
            ),
            "remaining": Decimal(
                str(getattr(order, "amount_remaining", 0) or 0)
            ),
            "payment_status": (
                getattr(order, "payment_status", "") or ""
            ),
            "became_paid": False,
            "already_settled": False,
            "already_applied": False,
        }

    with transaction.atomic():
        order = (
            Order.objects
            .select_for_update()
            .select_related("customer")
            .get(pk=order.pk)
        )

        if getattr(order, "status", None) == "canceled":
            raise ValidationError(
                "Impossible d'enregistrer un paiement sur une commande annulée."
            )

        try:
            finance_summary = build_order_finance_summary(order)
        except Exception:
            finance_summary = {}

        try:
            total = Decimal(
                str(
                    finance_summary.get(
                        "total_client_ttc",
                        getattr(order, "total_client_ttc", 0),
                    )
                    or 0
                )
            )
        except Exception:
            total = DECIMAL_ZERO

        if total <= DECIMAL_ZERO:
            raise ValidationError(
                "Impossible d'enregistrer un paiement : "
                "total_client_ttc est à 0."
            )

        # --------------------------------------------------------
        # Idempotence stricte avant toute opération wallet ou DB
        # --------------------------------------------------------
        if reference:
            if channel in {
                "wave_webhook",
                "wave_ops",
                "wave_manual_verified",
            }:
                conflicting_payment = (
                    Payment.objects
                    .filter(
                        reference=reference,
                    )
                    .exclude(order=order)
                    .order_by("id")
                    .first()
                )

                if conflicting_payment is not None:
                    raise ValidationError(
                        "Cette référence Wave est déjà rattachée "
                        "à une autre commande."
                    )

            existing_payment = (
                Payment.objects
                .filter(
                    order=order,
                    reference=reference,
                    status=Payment.ACCOUNTING_STATUS_CONFIRMED,
                )
                .order_by("id")
                .first()
            )

            if existing_payment is not None:
                existing_amount = Decimal(
                    str(existing_payment.amount or 0)
                )

                if existing_amount != requested_amount:
                    raise ValidationError(
                        "Cette référence de paiement existe déjà avec "
                        "un montant différent."
                    )

                order.refresh_from_db(
                    fields=[
                        "amount_paid",
                        "payment_status",
                        "payment_date",
                    ]
                )

                amount_paid = Decimal(
                    str(getattr(order, "amount_paid", 0) or 0)
                )
                remaining = total - amount_paid

                if remaining < DECIMAL_ZERO:
                    remaining = DECIMAL_ZERO

                return {
                    "applied": DECIMAL_ZERO,
                    "amount_paid": amount_paid,
                    "remaining": remaining,
                    "payment_status": (
                        getattr(order, "payment_status", "") or ""
                    ),
                    "became_paid": False,
                    "already_settled": remaining <= DECIMAL_ZERO,
                    "already_applied": True,
                }

        paid_sum = (
            Payment.objects
            .filter(
                order=order,
                status=Payment.ACCOUNTING_STATUS_CONFIRMED,
            )
            .aggregate(total=Sum("amount"))
            .get("total")
            or 0
        )
        already_paid = Decimal(str(paid_sum))

        if already_paid < DECIMAL_ZERO:
            already_paid = DECIMAL_ZERO

        remaining = total - already_paid

        if remaining < DECIMAL_ZERO:
            remaining = DECIMAL_ZERO

        if remaining <= DECIMAL_ZERO:
            order.refresh_from_db(
                fields=["amount_paid", "payment_status", "payment_date"]
            )

            return {
                "applied": DECIMAL_ZERO,
                "amount_paid": Decimal(
                    str(getattr(order, "amount_paid", 0) or 0)
                ),
                "remaining": DECIMAL_ZERO,
                "payment_status": (
                    getattr(order, "payment_status", "") or ""
                ),
                "became_paid": False,
                "already_settled": True,
                "already_applied": False,
            }

        to_apply = min(requested_amount, remaining)
        to_apply = Decimal(int(to_apply))

        if to_apply <= DECIMAL_ZERO:
            return {
                "applied": DECIMAL_ZERO,
                "amount_paid": already_paid,
                "remaining": remaining,
                "payment_status": (
                    getattr(order, "payment_status", "") or ""
                ),
                "became_paid": False,
                "already_settled": False,
                "already_applied": False,
            }

        previous_payment_status = (
            getattr(order, "payment_status", None) or "pending"
        )
        previous_amount_paid = already_paid

        # wallet_auto est le seul canal pour lequel cette fonction effectue
        # elle-même le débit. Le canal "wallet" est déjà débité explicitement
        # par client_order_pay_wave_page avant cet appel.
        if channel == "wallet_auto":
            from wallets.services import (
                debit_wallet,
                get_or_create_wallet_for_customer,
            )

            wallet = get_or_create_wallet_for_customer(order.customer)

            wallet_tx = debit_wallet(
                wallet,
                to_apply,
                description=f"Paiement commande {order.code}",
                order=order,
                idempotency_key=(
                    reference
                    or f"WALLET-AUTO-{order.id}"
                ),
                tx_type="debit",
            )

            if wallet_tx is None:
                raise ValidationError(
                    "Le débit du wallet n'a pas pu être enregistré."
                )

        payment = order.add_payment(
            amount=to_apply,
            channel=channel,
            reference=reference,
            source="system",
            confirmed_by=None,
            save=True,
        )

        if payment is None:
            order.refresh_from_db(
                fields=["amount_paid", "payment_status", "payment_date"]
            )

            current_paid = Decimal(
                str(getattr(order, "amount_paid", 0) or 0)
            )
            current_remaining = total - current_paid

            if current_remaining < DECIMAL_ZERO:
                current_remaining = DECIMAL_ZERO

            return {
                "applied": DECIMAL_ZERO,
                "amount_paid": current_paid,
                "remaining": current_remaining,
                "payment_status": (
                    getattr(order, "payment_status", "") or ""
                ),
                "became_paid": False,
                "already_settled": current_remaining <= DECIMAL_ZERO,
                "already_applied": False,
            }

        created = bool(getattr(payment, "_fagni_created", True))

        order.refresh_from_db()

        amount_paid_after = Decimal(
            str(getattr(order, "amount_paid", 0) or 0)
        )
        payment_status_after = (
            getattr(order, "payment_status", "") or ""
        )

        applied_amount = (
            Decimal(str(payment.amount or 0))
            if created
            else DECIMAL_ZERO
        )

        if created:
            OrderPaymentEvent.objects.create(
                order=order,
                channel=channel,
                reference=reference,
                amount=applied_amount,
                amount_paid_before=previous_amount_paid,
                amount_paid_after=amount_paid_after,
                status_before=previous_payment_status,
                status_after=payment_status_after,
                note=note or "",
            )

    became_paid = (
        previous_payment_status != "paid"
        and payment_status_after == "paid"
    )

    # Les effets complémentaires restent hors du verrou principal.
    # Les services appelés doivent conserver leur propre idempotence.
    if created:
        try:
            if became_paid:
                generate_mlm_commissions_for_order(order)

                from orders.services import (
                    completer_parrainage_client_si_applicable,
                    completer_parrainage_livreur_si_applicable,
                    completer_parrainage_pressing_si_applicable,
                )

                completer_parrainage_client_si_applicable(order)
                completer_parrainage_livreur_si_applicable(order)
                completer_parrainage_pressing_si_applicable(order)

            apply_fagni_monetization(order)

        except Exception:
            logging.getLogger("fagni.orders.views").exception(
                "Échec des effets post-paiement | order_id=%s",
                getattr(order, "id", None),
            )

    try:
        finance_summary = build_order_finance_summary(order)
        new_remaining = Decimal(
            str(finance_summary.get("amount_remaining", 0) or 0)
        )
    except Exception:
        new_remaining = total - amount_paid_after

    if new_remaining < DECIMAL_ZERO:
        new_remaining = DECIMAL_ZERO

    return {
        "applied": applied_amount,
        "amount_paid": amount_paid_after,
        "remaining": new_remaining,
        "payment_status": payment_status_after,
        "became_paid": became_paid,
        "already_settled": new_remaining <= DECIMAL_ZERO,
        "already_applied": not created,
    }


@client_login_required
@ensure_csrf_cookie
def client_order_pay_wave_page(request, order_id: int):
    """
    Page HTML simple qui affiche un QR code basé sur le numéro Wave (téléphone),
    + montant conseillé (reste à payer).
    MVP: QR = données texte (tel) via Google Chart.
    """
    phone = _client_phone(request)
    if not phone:
        return redirect("orders:client_login")

    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        return redirect("orders:client_home")

    amounts = _client_order_amounts(order)

    display_summary = build_order_display_summary(order)
    finance_summary = build_order_finance_summary(order)

    total = finance_summary.get("total_client_ttc", DECIMAL_ZERO) or DECIMAL_ZERO
    paid = finance_summary.get("amount_paid", DECIMAL_ZERO) or DECIMAL_ZERO
    remain = finance_summary.get("amount_remaining", DECIMAL_ZERO) or DECIMAL_ZERO

    try:
        total = Decimal(str(total))
    except Exception:
        total = DECIMAL_ZERO
    try:
        paid = Decimal(str(paid))
    except Exception:
        paid = DECIMAL_ZERO
    try:
        remain = Decimal(str(remain))
    except Exception:
        remain = DECIMAL_ZERO

    if remain < DECIMAL_ZERO:
        remain = DECIMAL_ZERO

    # ✅ HYBRIDE WALLET + WAVE
    try:
        wallet_obj = get_or_create_wallet_for_customer(order.customer)
    except Exception:
        wallet_obj = None

    try:
        wallet_balance = Decimal(str(getattr(wallet_obj, "balance", 0) or 0)) if wallet_obj else DECIMAL_ZERO
    except Exception:
        wallet_balance = DECIMAL_ZERO

    wallet_usable = wallet_balance if wallet_balance <= remain else remain
    if wallet_usable < DECIMAL_ZERO:
        wallet_usable = DECIMAL_ZERO

    wave_remaining = remain - wallet_usable
    if wave_remaining < DECIMAL_ZERO:
        wave_remaining = DECIMAL_ZERO

    pricing_mode = display_summary.get("pricing_mode", "item")
    bag_label = display_summary.get("bag_label", "")
    finance_breakdown = finance_summary
    service_fee_client_ttc = finance_summary.get("service_fee_client_ttc", DECIMAL_ZERO) or DECIMAL_ZERO


    # ---------------------------------------------------------
    # ✅ Déclaration manuelle "J'ai payé" (Wave) — persistée en base
    # - NE crée PAS de Payment
    # - NE modifie PAS amount_paid
    # - NE déclenche PAS mark_paid / wallets / payouts
    # ---------------------------------------------------------
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        # ✅ Utilisation explicite du wallet avant paiement Wave
        if action == "use_wallet":
            try:
                with transaction.atomic():
                    order.refresh_from_db()
                    if wallet_obj:
                        wallet_obj.refresh_from_db()

                    previous_payment_status = getattr(order, "payment_status", None)

                    paid_now = getattr(order, "amount_paid", DECIMAL_ZERO) or DECIMAL_ZERO
                    try:
                        paid_now = Decimal(str(paid_now))
                    except Exception:
                        paid_now = DECIMAL_ZERO

                    total_now = total
                    try:
                        total_now = Decimal(str(total_now))
                    except Exception:
                        total_now = DECIMAL_ZERO

                    remaining_now = total_now - paid_now
                    if remaining_now < DECIMAL_ZERO:
                        remaining_now = DECIMAL_ZERO

                    if remaining_now <= DECIMAL_ZERO:
                        return redirect("orders:client_order_detail", order_id=order.id)

                    fresh_wallet_balance = DECIMAL_ZERO
                    if wallet_obj:
                        try:
                            fresh_wallet_balance = Decimal(str(getattr(wallet_obj, "balance", 0) or 0))
                        except Exception:
                            fresh_wallet_balance = DECIMAL_ZERO

                    usable_now = fresh_wallet_balance if fresh_wallet_balance <= remaining_now else remaining_now
                    if usable_now <= DECIMAL_ZERO:
                        return redirect(reverse("orders:client_order_pay_wave_page", args=[order.id]))

                    wallet_tx = debit_wallet(
                        wallet_obj,
                        usable_now,
                        description=f"Paiement wallet commande {order.code}",
                        order=order,
                        leg=None,
                        idempotency_key=f"wallet_order_payment_{order.id}",
                        tx_type="debit",
                    )

                    if wallet_tx is None:
                        return redirect(reverse("orders:client_order_pay_wave_page", args=[order.id]) + "?wallet=error")

                    payment_result = apply_order_payment(
                        order,
                        usable_now,
                        channel="wallet",
                        reference=f"wallet_order_payment_{order.id}",
                        note=f"Paiement wallet commande {order.code}",
                    )

                if payment_result["payment_status"] == "paid":
                    return redirect("orders:client_order_detail", order_id=order.id)

                return redirect(reverse("orders:client_order_pay_wave_page", args=[order.id]) + "?wallet=applied")

            except Exception:
                return redirect(reverse("orders:client_order_pay_wave_page", args=[order.id]) + "?wallet=error")

        key = f"wave_declared_{order.id}"

        payment_reference = (request.POST.get("payment_reference") or "").strip()
        payment_proof = request.FILES.get("payment_proof")

        if action == "declare_wave_paid" and not payment_reference:
            return redirect(reverse("orders:client_order_pay_wave_page", args=[order.id]) + "?wave=missing_reference")

        if action == "declare_wave_paid" and not payment_proof and not getattr(order, "payment_proof", None):
            return redirect(reverse("orders:client_order_pay_wave_page", args=[order.id]) + "?wave=missing_proof")

        update_fields = []

        if getattr(order, "payment_status", None) != "paid":
            if getattr(order, "payment_status", None) != "declared":
                order.payment_status = "declared"
                update_fields.append("payment_status")

            if getattr(order, "payment_verification_status", None) != "pending_review":
                order.payment_verification_status = "pending_review"
                update_fields.append("payment_verification_status")

            if not getattr(order, "payment_declared_at", None):
                order.payment_declared_at = timezone.now()
                update_fields.append("payment_declared_at")

            if getattr(order, "payment_declared_channel", "") != "wave":
                order.payment_declared_channel = "wave"
                update_fields.append("payment_declared_channel")

            # 🔒 Sauvegarde référence Wave
            if payment_reference:
                order.payment_declared_reference = payment_reference
                update_fields.append("payment_declared_reference")


            # 🔒 FIX — sauvegarde stricte de la référence
            if payment_reference:
                order.payment_declared_reference = payment_reference
                update_fields.append("payment_declared_reference")

            if payment_proof and hasattr(order, "payment_proof"):
                order.payment_proof = payment_proof
                update_fields.append("payment_proof")

            current_ref = (getattr(order, "payment_declared_reference", "") or "").strip()
            if not current_ref:
                ref_candidate = ""
                try:
                    ref_candidate = (request.POST.get("checkout_id") or "").strip()
                except Exception:
                    ref_candidate = ""
                if ref_candidate:
                    order.payment_declared_reference = ref_candidate
                    update_fields.append("payment_declared_reference")

            if update_fields:
                order.save(update_fields=update_fields)

        if not request.session.get(key):
            request.session[key] = True
            request.session.modified = True
        return redirect(reverse("orders:client_order_detail", args=[order.id]) + "?wave=declared")
    # Numéro Wave: priorise settings.WAVE_RECEIVER_PHONE, sinon fallback sur phone client
    wave_phone = (getattr(settings, "WAVE_RECEIVER_PHONE", "") or "").strip() or phone

    # Session Wave Checkout canonique et partagée avec l'API client.
    # Le helper central assure :
    # - le verrouillage transactionnel ;
    # - la réutilisation pendant le TTL ;
    # - le stockage dans wave_checkout_id / wave_checkout_url ;
    # - la préservation de payment_declared_reference.
    pay_link = ""
    checkout_id = ""

    try:
        amount_xof = int(wave_remaining)
    except (TypeError, ValueError, ArithmeticError):
        amount_xof = 0

    order_status = (
        getattr(order, "status", "") or ""
    ).strip().lower()
    payment_status = (
        getattr(order, "payment_status", "") or ""
    ).strip().lower()

    if (
        order_status != "canceled"
        and payment_status != "paid"
        and amount_xof > 0
    ):
        try:
            from orders.client_api import _get_or_create_wave_checkout

            pay_link, checkout_id = _get_or_create_wave_checkout(
                order,
                amount_xof,
                request,
            )
            pay_link = (pay_link or "").strip()
            checkout_id = (checkout_id or "").strip()
        except Exception:
            # Une indisponibilité Wave ne doit pas casser la page :
            # le lien marchand statique reste disponible en repli.
            pay_link = ""
            checkout_id = ""

    # Fallback: lien marchand Wave (si pas de pay_link API)
    base_link = (getattr(settings, "WAVE_MERCHANT_LINK_BASE", "") or "").strip()
    if (not pay_link) and base_link:
        # Toujours fournir un lien scannable
        try:
            ax = str(amount_xof).strip()
        except Exception:
            ax = ""
        if ax and ax != "0":
            sep = "&" if "?" in base_link else "?"
            pay_link = f"{base_link}{sep}amount={ax}"
        else:
            pay_link = base_link
    # - attention: éviter border-radius sur l'image côté template
    import qrcode
    from io import BytesIO
    import base64
    # QR: encode toujours un URL Wave scannable
    pl = (pay_link or "").strip()
    qr_data = pl if pl else "https://pay.wave.com/"
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    buf = BytesIO()
    img.save(buf, format="PNG")
    qr_b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    qr_data_uri = "data:image/png;base64," + qr_b64

    try:
        wallet_used = Decimal(str(getattr(order, "amount_paid", 0) or 0))
    except Exception:
        wallet_used = DECIMAL_ZERO

    if wallet_used > total:
        wallet_used = total

    return render(request, "orders/client_pay_wave.html", {

        "wallet_balance": wallet_balance,
        "wallet_usable": wallet_usable,
        "wallet_used": wallet_used,
        "wave_remaining": wave_remaining,
        "wallet_applied": request.GET.get("wallet") == "applied",

        
          "wave_declared": bool(request.session.get(f"wave_declared_{order.id}")),
          # LOT_2_15_WAVE_DECLARED_CTX_OK
"order": order,
        "amounts": amounts,
        "total": total,
        "paid": paid,
        "remain": remain,
        "display_summary": display_summary,
        "finance_summary": finance_summary,
        "pricing_mode": pricing_mode,
        "bag_label": bag_label,
        "finance_breakdown": finance_breakdown,
        "service_fee_client_ttc": service_fee_client_ttc,
        "wave_phone": wave_phone,
        "qr_data_uri": qr_data_uri,
              "pay_link": pay_link,
          "checkout_id": checkout_id,
})

def _ops_redirect_back(request, fallback="orders:list"):
    next_url = (
        request.POST.get("next")
        or request.GET.get("next")
        or request.META.get("HTTP_REFERER")
        or ""
    )
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    try:
        return redirect(fallback)
    except Exception:
        return redirect("/orders/")


def _order_is_declared_or_pending_review(order) -> bool:
    return (
        getattr(order, "payment_status", "") == "declared"
        or getattr(order, "payment_verification_status", "") == "pending_review"
    )


@staff_member_required
@require_POST
def admin_confirm_declared_payment(request, order_id: int):
    """
    Confirme un paiement déclaré manuellement par le client.
    Effet recherché :
    - payment_status -> paid
    - payment_verification_status -> verified (si champ présent)
    - amount_paid -> total_client_ttc
    - payment_date -> now() si vide
    - payment_method -> wave si vide
    """
    order = get_object_or_404(Order, pk=order_id)

    declared_like = _order_is_declared_or_pending_review(order)
    already_paid = getattr(order, "payment_status", "") == "paid"

    if not declared_like and not already_paid:
        messages.warning(request, "Cette commande n'est pas en paiement déclaré.")
        return _ops_redirect_back(request)

    finance_summary = build_order_finance_summary(order)
    try:
        total_ttc = Decimal(str(finance_summary.get("total_client_ttc", 0) or 0))
    except Exception:
        total_ttc = DECIMAL_ZERO

    if total_ttc <= DECIMAL_ZERO:
        messages.error(request, "Impossible de confirmer : total TTC nul ou invalide.")
        return _ops_redirect_back(request)

    verified_reference = (
        request.POST.get("verified_wave_reference") or ""
    ).strip()

    human_confirmation = (
        request.POST.get("wave_human_verified") or ""
    ).strip().lower()

    if human_confirmation not in {"1", "true", "on", "yes"}:
        messages.error(
            request,
            "Confirmation refusée : vous devez confirmer avoir "
            "vérifié personnellement la transaction dans Wave.",
        )
        return _ops_redirect_back(request)

    if not verified_reference:
        messages.error(
            request,
            "Confirmation refusée : la référence réellement constatée "
            "dans Wave est obligatoire.",
        )
        return _ops_redirect_back(request)

    now_dt = timezone.now()
    update_fields = []

    try:
        current_amount_paid = Decimal(
            str(getattr(order, "amount_paid", 0) or 0)
        )
    except Exception:
        current_amount_paid = DECIMAL_ZERO

    remaining = total_ttc - current_amount_paid
    if remaining < DECIMAL_ZERO:
        remaining = DECIMAL_ZERO

    if remaining > DECIMAL_ZERO:
        apply_order_payment(
            order,
            remaining,
            channel="wave_manual_verified",
            reference=verified_reference,
            note=(
                "Paiement Wave vérifié manuellement dans l'application "
                f"Wave par {getattr(request.user, 'username', '')}"
            ),
        )

        try:
            order.refresh_from_db()
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception(
                "Erreur refresh après validation Wave manuelle | "
                "order_id=%s",
                getattr(order, "id", None),
            )
        try:
            order.refresh_from_db()
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception(
                "Exception silencieuse après confirmation Wave OPS | "
                "order_id=%s",
                getattr(order, "id", None),
            )

    if hasattr(order, "payment_verification_status"):
        if getattr(order, "payment_verification_status", "") != "verified":
            order.payment_verification_status = "verified"
            update_fields.append("payment_verification_status")

    if hasattr(order, "payment_date") and not getattr(order, "payment_date", None):
        order.payment_date = now_dt
        update_fields.append("payment_date")

    if hasattr(order, "payment_method") and not getattr(order, "payment_method", None):
        order.payment_method = "wave"
        update_fields.append("payment_method")

    if hasattr(order, "payment_declared_at") and not getattr(order, "payment_declared_at", None):
        order.payment_declared_at = now_dt
        update_fields.append("payment_declared_at")

    # Champs optionnels si tu les ajoutes plus tard
    if hasattr(order, "payment_verified_at"):
        order.payment_verified_at = now_dt
        update_fields.append("payment_verified_at")

    if hasattr(order, "payment_verified_by"):
        try:
            if getattr(request, "user", None) and request.user.is_authenticated:
                order.payment_verified_by = request.user
                update_fields.append("payment_verified_by")
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=6974")

    try:
        if update_fields:
            order.save(update_fields=list(dict.fromkeys(update_fields)))
        else:
            order.save()
        messages.success(
            request,
            f"Paiement confirmé pour la commande {getattr(order, 'code', order.id)}."
        )
    except ValidationError as e:
        messages.error(request, f"Confirmation refusée : {e}")
    except Exception as e:
        messages.error(request, f"Erreur lors de la confirmation : {e}")

    return _ops_redirect_back(request)


@staff_member_required
@require_POST
def admin_reject_declared_payment(request, order_id: int):
    """
    Rejette une déclaration de paiement.
    Effet recherché :
    - payment_status -> pending
    - payment_verification_status -> rejected (si champ présent)
    - amount_paid -> 0
    - payment_date -> None
    Ne doit PAS pouvoir rétrograder une commande déjà paid.
    """
    order = get_object_or_404(Order, pk=order_id)

    if getattr(order, "payment_status", "") == "paid":
        messages.error(request, "Impossible de rejeter : la commande est déjà confirmée payée.")
        return _ops_redirect_back(request)

    if not _order_is_declared_or_pending_review(order):
        messages.warning(request, "Cette commande n'est pas en attente de vérification.")
        return _ops_redirect_back(request)

    reason = (request.POST.get("reason") or "").strip()
    update_fields = []

    if getattr(order, "payment_status", "") != "pending":
        order.payment_status = "pending"
        update_fields.append("payment_status")

    if hasattr(order, "payment_verification_status"):
        if getattr(order, "payment_verification_status", "") != "rejected":
            order.payment_verification_status = "rejected"
            update_fields.append("payment_verification_status")

    # 🔒 Ne jamais effacer amount_paid/payment_date lors d'un rejet.
    # Le rejet concerne la déclaration, pas l'historique financier.

    if hasattr(order, "payment_rejected_at"):
        order.payment_rejected_at = timezone.now()
        update_fields.append("payment_rejected_at")

    if hasattr(order, "payment_rejected_by"):
        try:
            if getattr(request, "user", None) and request.user.is_authenticated:
                order.payment_rejected_by = request.user
                update_fields.append("payment_rejected_by")
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=7040")

    if hasattr(order, "payment_rejection_reason") and reason:
        order.payment_rejection_reason = reason
        update_fields.append("payment_rejection_reason")

    try:
        if update_fields:
            order.save(update_fields=list(dict.fromkeys(update_fields)))
        else:
            order.save()
        messages.success(
            request,
            f"Déclaration de paiement rejetée pour la commande {getattr(order, 'code', order.id)}."
        )
    except ValidationError as e:
        messages.error(request, f"Rejet refusé : {e}")
    except Exception as e:
        messages.error(request, f"Erreur lors du rejet : {e}")

    return _ops_redirect_back(request)



def client_order_item_new(request, order_id):
    phone = _client_phone(request)

    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        return redirect("orders:client_home")

    locked_response = _client_order_locked_response(request, order)
    if locked_response is not None:
        return locked_response

    error = None

    if request.method == "POST":
        from decimal import Decimal

        designation = (request.POST.get("designation") or "").strip()
        quantity_raw = (request.POST.get("quantity") or "").strip()
        unit_price_raw = (request.POST.get("unit_price") or "").strip()

        if not designation:
            error = "Merci de renseigner la désignation."
        else:
            try:
                qty = Decimal(str(quantity_raw).replace(",", "."))
            except Exception:
                qty = Decimal("0")
            try:
                up = Decimal(str(unit_price_raw).replace(",", "."))
            except Exception:
                up = Decimal("0")

            if qty <= 0:
                error = "Quantité invalide."
            elif up < 0:
                error = "Prix unitaire invalide."
            else:
                line_total = qty * up

                # service nullable => None OK

                OrderItem.objects.create(
                    order=order,
                    service=None,
                    designation=designation,
                    quantity=qty,
                    unit_price=up,
                    total=line_total,
                )

                try:
                    order.update_financials(save=True)
                except Exception:
                    import logging
                    logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=7118")

                
        upsell_data = request.session.get(f"upsell_data_{order.id}", {})

        if upsell_data:
            upsell, _ = OrderUpsell.objects.get_or_create(order=order)

            upsell.express_24h = upsell_data.get("express_24h", False)
            upsell.premium_ironing = upsell_data.get("premium_ironing", False)
            upsell.fragrance = upsell_data.get("fragrance", False)
            upsell.delicate_care = upsell_data.get("delicate_care", False)
            upsell.save()

            try:
                upsell_total = upsell.total
                base_total = Decimal(str(getattr(order, "total_client_ttc", 0) or 0))
                order.total_client_ttc = (base_total + upsell_total).quantize(Decimal("0.01"))
                order.save(update_fields=["total_client_ttc"])
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=7138")

            request.session.pop(f"upsell_data_{order.id}", None)

        return redirect("orders:client_order_detail", order_id=order.id)

    return render(request, "orders/client_order_item_form.html", {
        "order": order,
        "mode": "new",
        "error": error,
        "designation": "",
        "quantity": "",
        "unit_price": "",
    })


@require_http_methods(["GET", "POST"])
@client_required
def client_order_item_edit(request, order_id, item_id):
    phone = _client_phone(request)

    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        return redirect("orders:client_home")

    locked_response = _client_order_locked_response(request, order)
    if locked_response is not None:
        return locked_response

    item = (
        OrderItem.objects
        .filter(pk=item_id, order=order)
        .first()
    )
    if not item:
        return redirect("orders:client_order_detail", order_id=order.id)

    error = None

    if request.method == "POST":
        from decimal import Decimal

        designation = (request.POST.get("designation") or "").strip()
        quantity_raw = (request.POST.get("quantity") or "").strip()
        unit_price_raw = (request.POST.get("unit_price") or "").strip()

        if not designation:
            error = "Merci de renseigner la désignation."
        else:
            try:
                qty = Decimal(str(quantity_raw).replace(",", "."))
            except Exception:
                qty = Decimal("0")
            try:
                up = Decimal(str(unit_price_raw).replace(",", "."))
            except Exception:
                up = Decimal("0")

            if qty <= 0:
                error = "Quantité invalide."
            elif up < 0:
                error = "Prix unitaire invalide."
            else:
                item.designation = designation
                item.quantity = qty
                item.unit_price = up
                item.total = qty * up
                item.save(update_fields=["designation", "quantity", "unit_price", "total"])

                try:
                    order.update_financials(save=True)
                except Exception:
                    import logging
                    logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=7211")

                return redirect("orders:client_order_detail", order_id=order.id)

    return render(request, "orders/client_order_item_form.html", {
        "order": order,
        "mode": "edit",
        "item": item,
        "error": error,
        "designation": item.designation or "",
        "quantity": item.quantity if item.quantity is not None else "",
        "unit_price": item.unit_price if item.unit_price is not None else "",
    })


@require_http_methods(["GET", "POST"])
@client_required
def client_order_item_delete(request, order_id, item_id):
    phone = _client_phone(request)

    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        return redirect("orders:client_home")

    locked_response = _client_order_locked_response(request, order)
    if locked_response is not None:
        return locked_response

    item = OrderItem.objects.filter(pk=item_id, order=order).first()
    if not item:
        return redirect("orders:client_order_detail", order_id=order.id)

    if request.method == "POST":
        item.delete()
        try:
            order.update_financials(save=True)
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=7249")
        return redirect("orders:client_order_detail", order_id=order.id)

    return render(request, "orders/client_order_item_form.html", {
        "order": order,
        "mode": "delete",
        "item": item,
    })


# -------------------------------------------------------------------
# Lookup client (inchangé)
# -------------------------------------------------------------------

@require_GET
def client_lookup(request):
    phone = _normalize_phone(request.GET.get("phone") or "")
    if not phone:
        return JsonResponse({"ok": False, "error": "missing_phone"}, status=400)

    c = Customer.objects.filter(phone=phone).order_by("-id").first()
    if not c:
        return JsonResponse({"ok": True, "found": False})

    return JsonResponse({
        "ok": True,
        "found": True,
        "name": c.name or "",
        "address": getattr(c, "address", "") or "",
        "phone": c.phone or phone,
    })


# ============================================================
#  PLACEHOLDERS ÉDITION / SUPPRESSION (NON UTILISÉS)
# ============================================================

def edit(request):
    return HttpResponse("edit - placeholder", content_type="text/plain; charset=utf-8")


def delete(request):
    return HttpResponse("delete - placeholder", content_type="text/plain; charset=utf-8")



def _build_order_public_url(request, order, viewname="orders:detail"):
    """
    Construit une URL absolue vers une vue donnée pour une commande.
    """
    relative = reverse(viewname, args=[order.id])
    return request.build_absolute_uri(relative)


def _q(amount: Decimal) -> Decimal:
    # arrondi FCFA à l'unité (pas de centimes)
    return Decimal(amount or 0).quantize(Decimal("1"), rounding=ROUND_HALF_UP)


def _get_vat_rate_percent(cfg) -> Decimal:
    # On essaie de lire un éventuel champ admin.
    # Sinon fallback à 18% (CI).
    for attr in ("vat_percent", "vat_rate", "vat_fagni_percent", "tva_percent"):
        v = getattr(cfg, attr, None)
        if v is not None:
            try:
                return Decimal(str(v))
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=7350")
    return Decimal("18")


def _build_invoice_context(order):
    """
    Contexte canonique facturation/tickets.
    Source de vérité = _compute_order_pricing(order) (inclut Express).
    On garde compute_order_amounts(order) pour compat / infos secondaires.
    """
    cfg = get_pricing_settings()

    # Compat / autres besoins (peut servir ailleurs)
    amounts = compute_order_amounts(order)

    # Source de vérité montants (verrouillage)
    pricing = _compute_order_pricing(order)

    vat_rate = _get_vat_rate_percent(cfg)

    prestations_amount = _q(pricing.get("items_total", 0))
    delivery_amount    = _q(pricing.get("delivery_fee", 0))
    service_amount     = _q(pricing.get("service_fee", 0))
    express_amount     = _q(pricing.get("express_extra_fee", 0))
    vat_amount         = _q(pricing.get("vat_fagni", 0))

    # HT client = prestations + livraison + service + express (sans TVA)
    total_ht_client = _q(prestations_amount + delivery_amount + service_amount + express_amount)

    # TTC client = total canonique (inclut TVA)
    total_ttc_client = _q(pricing.get("total_client", 0))

    # Compat anciennes variables
    base_ht = total_ht_client
    grand_total = total_ttc_client

    # ✅ exposer "express_client" pour les templates existants
    express_client = express_amount

    # --- LOT 5C : couche d'affichage client TTC unifiée ---
    display_delivery_amount = delivery_amount
    display_express_amount = express_amount

    display_prestations_amount = _q(total_ttc_client - display_delivery_amount - display_express_amount)
    if display_prestations_amount < Decimal("0"):
        display_prestations_amount = Decimal("0")

    # (Optionnel mais utile partout) : lignes standardisées
    invoice_lines = {
        # interne / canon
        "prestations": prestations_amount,
        "delivery": delivery_amount,
        "service": service_amount,
        "express": express_amount,
        "vat": vat_amount,
        "total_ht": total_ht_client,
        "total_ttc": total_ttc_client,

        # affichage client
        "display_prestations": display_prestations_amount,
        "display_delivery": display_delivery_amount,
        "display_express": display_express_amount,
        "display_total_ttc": total_ttc_client,
    }

    return {
        "cfg": cfg,
        "amounts": amounts,          # compat existant
        "pricing": pricing,          # debug/ops si besoin
        "invoice_lines": invoice_lines,

        # Champs canon
        "prestations_amount": prestations_amount,
        "delivery_amount": delivery_amount,
        "service_amount": service_amount,
        "express_amount": express_amount,
        "vat_rate": vat_rate,
        "vat_amount": vat_amount,
        "total_ht_client": total_ht_client,
        "total_ttc_client": total_ttc_client,

        # affichage client unifié
        "display_prestations_amount": display_prestations_amount,
        "display_delivery_amount": display_delivery_amount,
        "display_express_amount": display_express_amount,
        "display_total_ttc": total_ttc_client,

        # Compat historique
        "express_client": express_client,
        "base_ht": base_ht,
        "tva_amount": vat_amount,
        "grand_total": grand_total,
    }


@login_required
def order_invoice_pdf(request, order_id):
    """
    Génère une FACTURE PDF A4 (WeasyPrint) :
    - accessible uniquement si commande payée
    - montants alignés sur le ticket (source unique = _build_invoice_context)
    - header/footer pilotés par InvoiceSettings (admin)
    """
    order = get_object_or_404(
        Order.objects.select_related("customer", "laundry_partner", "delivery_partner"),
        pk=order_id
    )

    # Sécurité : facture uniquement si payé
    if order.payment_status != "paid":
        return HttpResponseForbidden("Facture indisponible : commande non payée.")

    # S'assurer que les montants + invoice_number sont à jour (et persistés)
    try:
        order.update_financials(save=True)
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=7466")

    # 🔒 NE PAS écraser items (déjà filtré + photos)
    # items = _build_client_display_items(order)

    # ✅ Source unique (comme ticket_pdf / ticket_thermal)
    ctx = _build_invoice_context(order)

    # ✅ Header/Footer admin (nettoyé)
    invoice_settings = get_invoice_settings_clean()
    if invoice_settings:
        ctx["invoice_settings"] = invoice_settings

    context = {
        "order": order,
        "items": items,
        "now": timezone.now(),
        **ctx,
    }

    html_string = render_to_string("orders/invoice_pdf.html", context=context, request=request)
    try:
        pdf = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()
    except Exception as e:
        return HttpResponse(
            f"<h1>Erreur génération PDF (WeasyPrint)</h1><pre>{e}</pre><hr>{html_string}",
            content_type="text/html",
            status=500,
        )

    filename = f"FACTURE-{order.invoice_number or order.code or order.id}.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{smart_str(filename)}"'
    return response


from decimal import Decimal as _Decimal
from django.contrib import messages
from django.shortcuts import redirect, render

# ... (le reste de tes imports déjà présents)

def _dec(v):
    try:
        if v in (None, "", False):
            return Decimal("0")
        if isinstance(v, Decimal):
            return v
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def _enrich_items_for_ticket(items_qs):
    """
    Ajoute sur chaque item (sans toucher aux champs DB ni aux @property) :
      - ticket_total : total sûr (item.total si >0 sinon quantity*unit_price sinon line_total)
      - ticket_name  : designation sûre (designation > service.name > "—")
    """
    items = list(items_qs)

    for it in items:
        # --- Total robuste ---
        total_db = _dec(getattr(it, "total", None))

        if total_db > 0:
            total = total_db
        else:
            q = _dec(getattr(it, "quantity", 0))
            pu = _dec(getattr(it, "unit_price", 0))
            total_calc = q * pu

            # fallback ultime : la @property line_total (lecture seule)
            if total_calc <= 0:
                try:
                    total_calc = _dec(getattr(it, "line_total", 0))
                except Exception:
                    total_calc = Decimal("0")

            total = total_calc

        # ⚠️ Ne pas faire it.line_total = ...
        setattr(it, "ticket_total", total)

        # --- Nom robuste ---
        name = getattr(it, "designation", None)
        if not name:
            svc = getattr(it, "service", None)
            name = getattr(svc, "name", None) if svc else None

        setattr(it, "ticket_name", name or "—")

    return items


@login_required
def order_ticket_pdf(request, order_id):
    order = (
        Order.objects
        .select_related("customer", "laundry_partner", "delivery_partner")
        .prefetch_related("items__service", "items__photos")
        .filter(pk=order_id)
        .first()
    )
    if not order:
        messages.error(request, f"Commande introuvable (ID={order_id}).")
        return redirect("orders:list")

    items = _enrich_items_for_ticket(order.items.all())

    detail_url = request.build_absolute_uri(reverse("orders:detail", args=[order.id]))
    qr_b64 = _qr_png_base64(detail_url)

    # Montants harmonisés pour template
    ctx = _build_invoice_context(order)
    amounts = ctx.get("amounts") or {}
    prestations_amount = _dec(amounts.get("subtotal", 0))
    delivery_amount = _dec(amounts.get("delivery_fee_client", 0))
    service_amount = _dec(amounts.get("service_fee_ht", 0))
    express_amount = _dec(ctx.get("express_client", 0))  # déjà normalisé
    vat_amount = _dec(ctx.get("vat_amount", 0))
    total_ttc_client = _dec(ctx.get("total_ttc_client", 0))

    context = {
        "order": order,
        "items": items,

        # QR
        "qr_data": detail_url,
        "qr_b64": qr_b64,

        # Contexte facture/pricing existant
        **ctx,
        "invoice_settings": get_invoice_settings_clean(),

        # ✅ variables harmonisées (A4 + thermique)
        "prestations_amount": prestations_amount,
        "delivery_amount": delivery_amount,
        "service_amount": service_amount,
        "express_amount": express_amount,
        "vat_amount": vat_amount,
        "total_ttc_client": total_ttc_client,
    }
    return render(request, "orders/ticket_pdf.html", context)


@login_required
def order_ticket_thermal_pdf(request, order_id):
    """
    Ticket thermique (80mm).
    Source de vérité montants = _build_invoice_context(order) (compute_order_amounts)
    Lignes = line_total calculé ici pour éviter tout écart template/model.
    """
    order = (
        Order.objects
        .filter(pk=order_id)
        .select_related("customer", "laundry_partner", "delivery_partner")
        .prefetch_related("items__service", "items__photos")
        .first()
    )
    if not order:
        return redirect("orders:list")

    items = _enrich_items_for_ticket(order.items.all())

    ctx = _build_invoice_context(order)
    amounts = ctx.get("amounts") or {}

    invoice_settings = get_invoice_settings_clean()
    if invoice_settings:
        ctx["invoice_settings"] = invoice_settings

    prestations_amount = _dec(amounts.get("subtotal", 0))
    delivery_amount = _dec(amounts.get("delivery_fee_client", 0))
    service_amount = _dec(amounts.get("service_fee_ht", 0))
    express_amount = _dec(ctx.get("express_client", 0))
    vat_amount = _dec(ctx.get("vat_amount", 0))
    total_ttc_client = _dec(ctx.get("total_ttc_client", 0))

    # QR total
    qr_total = total_ttc_client
    customer_phone = ""
    if getattr(order, "customer", None) and getattr(order.customer, "phone", None):
        customer_phone = order.customer.phone

    try:
        qr_total_int = int(_dec(qr_total))
    except Exception:
        qr_total_int = 0

    qr_data = f"CMD:{order.code}|TEL:{customer_phone}|TOTAL:{qr_total_int}"
    qr_b64 = _qr_png_base64(qr_data)

    context = {
        "order": order,
        "items": items,
        "qr_data": qr_data,
        "qr_b64": qr_b64,

        **ctx,

        # ✅ variables harmonisées
        "prestations_amount": prestations_amount,
        "delivery_amount": delivery_amount,
        "service_amount": service_amount,
        "express_amount": express_amount,
        "vat_amount": vat_amount,
        "total_ttc_client": total_ttc_client,
    }
    return render(request, "orders/ticket_thermal_pdf.html", context)


def safe_decimal(value, default=Decimal("0")):
    try:
        if value in (None, ""):
            return default
        return Decimal(str(value))
    except Exception:
        return default


@login_required
def update(request, order_id):
    """
    Édition legacy d'une commande EXISTANTE.

    Contrat A5-E3 :
    - cette route reste disponible uniquement pour les commandes encore
      pré-matérialisation ;
    - dès qu'au moins un OrderItem appartient à une ServiceExecution,
      le contrat commercial de la commande est considéré comme matérialisé
      et l'éditeur legacy devient interdit en GET comme en POST ;
    - le verrou modèle OrderItem reste la dernière ligne de défense contre
      toute mutation commerciale post-matérialisation.
    """
    order = get_object_or_404(
        Order.objects.select_related(
            "customer",
            "laundry_partner",
            "delivery_partner",
        ),
        pk=order_id,
    )

    # Frontière d'autorité V2 :
    # le verrou historique basé sur OrderItem reste utile,
    # mais il ne couvre pas toutes les matérialisations possibles.
    #
    # La présence d'une ServiceExecution est désormais le critère
    # canonique d'entrée sous autorité V2.
    from services.services import order_uses_canonical_service_executions
    if order_uses_canonical_service_executions(order=order):
        messages.error(
            request,
            (
                "Cette commande est pilotée par le moteur "
                "ServiceExecution. L'éditeur legacy est verrouillé."
            ),
        )
        return redirect(
            "orders:detail",
            order_id=order.id,
        )

    # 🔒 A5-E3 — HARD LOCK POST-MATERIALISATION
    #
    # Une fois qu'une ligne commerciale est rattachée à une
    # ServiceExecution, l'ancien éditeur ne doit plus pouvoir :
    # - modifier les OrderItem ;
    # - ajouter/supprimer des OrderItem ;
    # - modifier les métadonnées de commande dans le même POST ;
    # - recalculer les montants via son ancien workflow.
    #
    # Le test est fait AVANT toute mutation de l'objet Order.
    is_materialized = order.items.filter(
        service_execution_link__isnull=False,
    ).exists()

    if is_materialized:
        messages.error(
            request,
            (
                "Cette commande est déjà matérialisée dans le moteur "
                "d'exécution FAGNI. L'éditeur legacy est verrouillé."
            ),
        )
        return redirect("orders:detail", order_id=order.id)

    # 🔒 HARD LOCK historique : une commande payée ne doit plus être
    # modifiable, même lorsqu'elle n'a pas encore de matérialisation V2.
    if getattr(order, "payment_status", None) == "paid":
        messages.error(
            request,
            "Commande payée : modification interdite.",
        )
        return redirect("orders:detail", order_id=order.id)

    # Pour l'affichage (GET) : catalogue de services
    service_categories = ServiceCategory.objects.all().order_by("name")
    service_items = (
        ServiceItem.objects.select_related("category")
        .all()
        .order_by("category__name", "name")
    )

    # Paramètres transport pour le front (mêmes noms que dans create.html)
    logistics = getattr(settings, "FAGNI_LOGISTICS", {})
    delivery_min_fee = logistics.get("client_min_fee", 0)
    delivery_price_per_km = logistics.get("client_price_per_km", 0)
    delivery_fixed_fee = logistics.get("client_fixed_fee", 0)

    if request.method == "POST":
        # --- 1) Mise à jour infos ordre (hors items) ---
        status = request.POST.get("status") or order.status
        order.status = status

        laundry_id = request.POST.get("laundry_partner") or None
        delivery_id = request.POST.get("delivery_partner") or None

        if laundry_id:
            try:
                order.laundry_partner_id = int(laundry_id)
            except (TypeError, ValueError):
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=7735")

        if delivery_id:
            try:
                order.delivery_partner_id = int(delivery_id)
            except (TypeError, ValueError):
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=7741")

        # notes / instructions
        notes_raw = request.POST.get("order_notes", None)
        if notes_raw is not None:
            order.notes = notes_raw.strip()

        # --- 2) Gestion des lignes (update / ajout / suppression) ---
        item_indexes = request.POST.getlist("item_index[]")
        item_ids = request.POST.getlist("item_id[]")
        service_ids = request.POST.getlist("service_id[]")
        designations = request.POST.getlist("designation[]")
        quantities = request.POST.getlist("quantity[]")
        unit_prices = request.POST.getlist("unit_price[]")

        existing_items_qs = (
            order.items.all()
            .select_related("service")
            .prefetch_related("photos")
            .order_by("id")
        )
        existing_by_id = {str(it.id): it for it in existing_items_qs}
        kept_ids = set()

        nb_rows = len(service_ids)

        for idx in range(nb_rows):
            raw_service_id = (service_ids[idx] or "").strip()
            raw_designation = (designations[idx] or "").strip()
            raw_item_id = (item_ids[idx] or "").strip()
            raw_quantity = (quantities[idx] or "").strip()
            raw_unit_price = (unit_prices[idx] or "").strip()
            raw_index = (item_indexes[idx] or "").strip()

            existing_item = existing_by_id.get(raw_item_id) if raw_item_id else None

            # Ancienne ligne sans service_id -> on garde (compat vieilles commandes)
            if not raw_service_id and existing_item:
                kept_ids.add(existing_item.id)
                if raw_index != "":
                    files = request.FILES.getlist(f"photos_{raw_index}")
                    for f in files:
                        OrderItemPhoto.objects.create(order_item=existing_item, image=f)
                continue

            if not raw_service_id:
                continue

            # quantité
            try:
                quantity = int(raw_quantity)
            except (TypeError, ValueError):
                quantity = 0

            # prix unitaire Decimal
            clean_unit_price = (raw_unit_price or "").replace(" ", "").replace("\u00a0", "")
            try:
                unit_price = Decimal(clean_unit_price)
            except (TypeError, ValueError, InvalidOperation):
                unit_price = Decimal("0")

            # Ligne invalide (nouvelle) : on n'en crée pas
            if quantity <= 0 or unit_price <= 0:
                if existing_item:
                    kept_ids.add(existing_item.id)
                    if raw_index != "":
                        files = request.FILES.getlist(f"photos_{raw_index}")
                        for f in files:
                            OrderItemPhoto.objects.create(order_item=existing_item, image=f)
                continue

            designation = raw_designation

            # Ligne existante ou nouvelle ?
            if existing_item:
                existing_item.service_id = int(raw_service_id)
                if designation:
                    existing_item.designation = designation
                existing_item.quantity = quantity
                existing_item.unit_price = unit_price
                existing_item.save()
                item = existing_item
            else:
                item = OrderItem.objects.create(
                    order=order,
                    service_id=int(raw_service_id),
                    designation=designation,
                    quantity=quantity,
                    unit_price=unit_price,
                )

            kept_ids.add(item.id)

            # Ajout de nouvelles photos
            if raw_index != "":
                files = request.FILES.getlist(f"photos_{raw_index}")
                for f in files:
                    OrderItemPhoto.objects.create(order_item=item, image=f)

        # --- 3) Suppression des lignes qui ne sont plus envoyées ---
        for it in existing_items_qs:
            if it.id not in kept_ids:
                it.delete()

        # --- 4) Recalcul des totaux ---
        if hasattr(order, "recalculate_totals"):
            order.recalculate_totals()

        order.recompute_distances_from_positions()
        order.compute_totals(save=True)
        order.save()

        messages.success(request, "Commande mise à jour avec succès.")
        return redirect("orders:detail", order_id=order.id)

    # --- GET : affichage du formulaire ---
    context = {
        "order": order,
        "service_categories": service_categories,
        "service_items": service_items,
        "delivery_min_fee": delivery_min_fee,
        "delivery_price_per_km": delivery_price_per_km,
        "delivery_fixed_fee": delivery_fixed_fee,
    }
    return render(request, "orders/update.html", context)


# ============================================================
#  TABLEAU DE BORD LIVREURS
# ============================================================
@login_required
def driver_dashboard(request):
    """
    Tableau 'Commandes avec livreurs' :
    - Filtre par livreur (delivery_partner)
    - KPIs : nombre de commandes, statuts, distance totale, coût livreur, marge logistique
    """

    # 1) Récupération du filtre livreur (GET ?driver_id=...)
    driver_id = request.GET.get("driver_id") or ""

    # 2) Liste des livreurs disponibles pour le select
    drivers = DeliveryPartner.objects.all().order_by("name")

    # 3) Base queryset : uniquement les commandes avec un livreur associé
    base_qs = (
        Order.objects
        .select_related("customer", "delivery_partner", "laundry_partner")
        .filter(delivery_partner__isnull=False)
    )

    # 4) Application éventuelle du filtre livreur
    if driver_id:
        base_qs = base_qs.filter(legs__driver_id=driver_id).distinct()

    # 5) Ordre d'affichage : plus récentes d'abord
    orders_qs = base_qs.order_by("-created_at")

    # 6) KPIs statistiques sur les commandes filtrées
    total_orders = orders_qs.count()

    pending = orders_qs.filter(status="pending").count()
    in_progress = orders_qs.filter(status="in_progress").count()
    done = orders_qs.filter(status="done").count()
    canceled = orders_qs.filter(status="canceled").count()

    agg = orders_qs.aggregate(
        dist=Sum("distance_km"),
        margin=Sum("logistic_margin"),
    )

    total_distance = agg.get("dist") or 0
    total_logistic_margin = agg.get("margin") or 0

    # ✅ Revenu livreur NET (source de vérité) = payout in - adjustment out
    tx_qs = WalletTransaction.objects.filter(
        order__in=orders_qs,
        leg__isnull=False,
        type__in=["payout", "adjustment"],
    )

    # si filtre livreur appliqué → on restreint à son wallet
    if driver_id:
        tx_qs = tx_qs.filter(wallet__delivery_partner_id=driver_id)

    total_driver_cost = tx_qs.aggregate(net=_wallet_net_expr()).get("net") or Decimal("0")

    # 7) Contexte pour le template
    try:
        selected_driver_id = int(driver_id) if driver_id else None
    except ValueError:
        selected_driver_id = None

    context = {
        "drivers": drivers,
        "selected_driver_id": selected_driver_id,
        "orders": orders_qs,
        # KPIs
        "total_orders": total_orders,
        "pending": pending,
        "in_progress": in_progress,
        "done": done,
        "canceled": canceled,
        "total_distance": total_distance,
        "total_driver_cost": total_driver_cost,
        "total_logistic_margin": total_logistic_margin,
    }

    return render(request, "orders/driver_dashboard.html", context)


# ------------------------------------------------------------
# APP LIVREUR – "Mes courses du jour"
# ------------------------------------------------------------
def _gmaps_url(lat, lng):
    """
    Construit une URL Google Maps simple à partir d'une latitude/longitude.

    - Si lat/lng sont vides ou invalides → retourne ""
    - Sinon → https://www.google.com/maps?q=lat,lng
    """
    try:
        if lat in (None, "") or lng in (None, ""):
            return ""
        lat_f = float(lat)
        lng_f = float(lng)
        return f"https://www.google.com/maps?q={lat_f},{lng_f}"
    except Exception:
        return ""


def _get_driver_app_context(request):
    """
    Logique commune pour la vue HTML et la vue JSON (AJAX) de l'app livreur.
    Gère :
    - mode livreur (user non staff, lié à un DeliveryPartner par email)
    - mode staff (user staff, choisit un livreur dans la liste)
    - filtres de statut
    """
    user = request.user

    # Tous les livreurs (utile en mode staff)
    drivers = DeliveryPartner.objects.all().order_by("name")

    # Filtres venant de l'URL
    status_filter = request.GET.get("status", "active")
    selected_driver_id = request.GET.get("driver_id") or None

    driver_mode = False
    current_driver = None

    # 🔒 Mode LIVREUR : user connecté NON staff → on le mappe à un DeliveryPartner via son email
    if user.is_authenticated and not user.is_staff:
        email = (user.email or "").strip()
        if email:
            try:
                current_driver = DeliveryPartner.objects.get(email__iexact=email)
                driver_mode = True
                selected_driver_id = str(current_driver.id)
            except DeliveryPartner.DoesNotExist:
                driver_mode = False

    # Base queryset
    orders_qs = Order.objects.select_related(
        "customer",
        "laundry_partner",
        "delivery_partner",
    )

    # Filtre par livreur
    if driver_mode and current_driver:
        # Le livreur ne voit QUE ses propres courses
        orders_qs = orders_qs.filter(legs__driver=current_driver).distinct()
    elif selected_driver_id:
        orders_qs = orders_qs.filter(legs__driver_id=selected_driver_id).distinct()

    # Filtre par statut
    if status_filter == "active":
        orders_qs = orders_qs.filter(status__in=["pending", "in_progress"])
    elif status_filter == "done":
        orders_qs = orders_qs.filter(status="done")
    elif status_filter == "canceled":
        orders_qs = orders_qs.filter(status="canceled")
    # "all" → pas de filtre de statut

    orders_qs = orders_qs.order_by("-created_at")

    # Aujourd'hui
    today = timezone.localdate()
    today_qs = orders_qs.filter(created_at__date=today)

    # Agrégats
    aggregates = orders_qs.aggregate(
        total_distance_km=Sum("distance_km"),
    )

    # ✅ revenu livreur net basé sur WalletTransaction
    tx_qs = WalletTransaction.objects.filter(
        order__in=orders_qs,
        leg__isnull=False,
        type__in=["payout", "adjustment"],
    )

    # driver_mode => uniquement le wallet du livreur connecté
    if driver_mode and current_driver:
        tx_qs = tx_qs.filter(wallet__delivery_partner=current_driver)
    elif selected_driver_id:
        tx_qs = tx_qs.filter(wallet__delivery_partner_id=selected_driver_id)

    aggregates["total_driver_income"] = tx_qs.aggregate(net=_wallet_net_expr()).get("net") or Decimal("0")

    context = {
        # données principales
        "orders": orders_qs,
        "drivers": drivers,
        "selected_driver_id": int(selected_driver_id) if selected_driver_id else None,
        "status_filter": status_filter,

        # mode / identité livreur
        "driver_mode": driver_mode,
        "current_driver": current_driver,

        # KPIs
        "filtered_orders_count": orders_qs.count(),
        "total_orders": orders_qs.count(),     # ici, on considère "total" = après filtres livreur+statut
        "today_orders": today_qs.count(),
        "pending": orders_qs.filter(status="pending").count(),
        "in_progress": orders_qs.filter(status="in_progress").count(),
        "done": orders_qs.filter(status="done").count(),
        "canceled": orders_qs.filter(status="canceled").count(),
        "total_distance_km": aggregates["total_distance_km"] or 0,
        "total_driver_income": aggregates["total_driver_income"] or 0,
    }
    return context


@login_required
def order_scan_redirect(request, order_code):
    """
    Lors d'un scan QR, on retrouve la commande via son code unique.
    Puis redirection vers la vue livreur, si le livreur est bien assigné.
    """
    # 1) Retrouver la commande
    order = get_object_or_404(
        Order,
        code__iexact=order_code
    )

    
    # 2) Vérifier le livreur connecté
    user_email = (request.user.email or "").strip().lower()

    try:
        delivery_partner = DeliveryPartner.objects.get(email__iexact=user_email)
    except DeliveryPartner.DoesNotExist:
        return HttpResponseForbidden(
            "Aucun profil livreur associé à cet email (%s)." % user_email
        )

    # 3) Vérifier que c'est bien SA mission via DeliveryLeg
    if not DeliveryLeg.objects.filter(order=order, driver=delivery_partner).exists():
        return HttpResponseForbidden("Vous n'êtes pas assigné à cette course.")

    # 4) Redirection vers sa page
    return redirect("orders:driver_order_detail", order_id=order.id)


@login_required
def driver_order_timeline_action(request, order_id, action):
    """
    App livreur – actions chronométrées sur la timeline :
    - pickup_done       -> pickup_time
    - dropoff_done      -> dropoff_time
    - wash_done         -> wash_complete_time
    - return_done       -> return_time
    - delivered_done    -> delivered_time

    Sécurisé :
    - staff : peut tout faire
    - livreur : uniquement sur SES courses

    ✅ En plus :
    - Quand wash_done est validé (linge prêt), on ASSIGNE automatiquement
      le 2e tronçon (delivery/return) s'il était en pending, pour ce même livreur.
    """
    from partners.models import DeliveryPartner  # import local pour éviter les cycles

    if request.method != "POST":
        return redirect("orders:driver_order_detail", order_id=order_id)

    order = get_object_or_404(
        Order.objects.select_related("delivery_partner"),
        pk=order_id,
    )

    # --- Sécurité accès ---
    user = request.user

    # Staff : OK sur tout
    if not user.is_staff:
        user_email = (user.email or "").strip()
        if not user_email:
            return HttpResponseForbidden("Votre compte n'a pas d'email défini.")

        try:
            delivery_partner = DeliveryPartner.objects.get(email__iexact=user_email)
        except DeliveryPartner.DoesNotExist:
            return HttpResponseForbidden("Aucun profil livreur associé à cet email.")

        if not DeliveryLeg.objects.filter(order=order, driver=delivery_partner).exists():
            return HttpResponseForbidden("Vous n'êtes pas assigné à cette course.")

    # --- Mapping action -> champ datetime ---
    action_map = {
        "pickup_done": "pickup_time",
        "dropoff_done": "dropoff_time",
        "wash_done": "wash_complete_time",
        "return_done": "return_time",
        "delivered_done": "delivered_time",
    }

    field_name = action_map.get(action)
    if not field_name:
        # action inconnue : on retourne au détail
        return redirect("orders:driver_order_detail", order_id=order_id)

    # On ne modifie que si le champ est encore vide (idempotent)
    current_val = getattr(order, field_name, None)
    if not current_val:
        now = timezone.now()
        setattr(order, field_name, now)

        # On essaye d'update updated_at s'il existe
        update_fields = [field_name]
        if hasattr(order, "updated_at"):
            order.updated_at = now
            update_fields.append("updated_at")

        order.save(update_fields=update_fields)

        # ✅ Déclencheur : linge prêt → assigne le 2e tronçon (prestataire → client)
        # On ne touche QUE les legs pending du même livreur assigné à la commande.
        if action == "wash_done":
            pass

    next_url = request.POST.get("next") or reverse(
        "orders:driver_order_detail",
        args=[order_id],
    )
    return redirect(next_url)


@login_required
def driver_performance_me(request):
    """
    Redirige le livreur connecté vers sa page de performance,
    en se basant sur son email (DeliveryPartner.email).
    """
    user = request.user
    email = (user.email or "").strip()

    from partners.models import DeliveryPartner  # import local pour éviter les cycles

    driver = None
    if email:
        try:
            driver = DeliveryPartner.objects.get(email__iexact=email)
        except DeliveryPartner.DoesNotExist:
            driver = None

    if not driver:
        # Pas de profil livreur lié → retour au hub
        return redirect("orders:driver_hub")

    # Redirection vers la vue existante qui prend un driver_id
    return redirect("orders:driver_performance", driver_id=driver.id)


@login_required
def driver_performance(request, driver_id):
    """
    Dashboard avancé de performance pour un livreur :
    - KPIs (commandes, distance, revenu, taux de complétion, etc.)
    - Séries journalières pour graphiques (Chart.js)
    - Classement parmi les autres livreurs
    """
    # 1) Récupération du livreur
    driver = get_object_or_404(DeliveryPartner, pk=driver_id)

    # 2) Filtre de période (7 / 30 / 90 / 180 jours)
    period = request.GET.get("period", "30")
    try:
        days = int(period)
    except (TypeError, ValueError):
        days = 30

    if days not in (7, 30, 90, 180):
        days = 30

    end = timezone.now()
    start = end - timedelta(days=days)

    # 3) Query de base : toutes les commandes du livreur sur la période
    orders_qs = (
        Order.objects.filter(
            delivery_partner=driver,
            created_at__gte=start,
            created_at__lte=end,
        )
        .select_related("customer")
        .order_by("-created_at")
    )

    # 4) KPIs globaux
    total_orders = orders_qs.count()
    done_orders = orders_qs.filter(status="done").count()
    in_progress_orders = orders_qs.filter(status="in_progress").count()
    pending_orders = orders_qs.filter(status="pending").count()
    canceled_orders = orders_qs.filter(status="canceled").count()

    aggregates = orders_qs.aggregate(
        total_distance_km=Sum("distance_km"),
        total_income=Sum("driver_logistic_cost"),
    )
    total_distance_km = aggregates["total_distance_km"] or 0
    total_income = aggregates["total_income"] or 0

    # Taux de complétion / annulation
    completion_rate = 0
    cancel_rate = 0
    if total_orders > 0:
        completion_rate = round(done_orders * 100 / total_orders, 1)
        cancel_rate = round(canceled_orders * 100 / total_orders, 1)

    # Moyennes
    avg_income_per_order = 0
    avg_income_per_km = 0
    if done_orders > 0:
        avg_income_per_order = round(total_income / done_orders, 2)
    if total_distance_km:
        avg_income_per_km = round(total_income / total_distance_km, 2)

    # 5) Statistiques journalières (séries pour graphiques)
    daily_stats_qs = (
        orders_qs
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            orders_count=Count("id"),
            day_income=Sum("driver_logistic_cost"),
            day_distance=Sum("distance_km"),
        )
        .order_by("day")
    )

    daily_labels = []
    daily_orders_series = []
    daily_income_series = []
    daily_distance_series = []

    for row in daily_stats_qs:
        day = row["day"]
        daily_labels.append(day.strftime("%d/%m"))
        daily_orders_series.append(row["orders_count"] or 0)
        daily_income_series.append(float(row["day_income"] or 0))
        daily_distance_series.append(float(row["day_distance"] or 0))

    # 6) Classement des livreurs (leaderboard)
    leaderboard_qs = (
        Order.objects.filter(
            created_at__gte=start,
            created_at__lte=end,
            status="done",
            delivery_partner__isnull=False,
        )
        .values("delivery_partner_id", "delivery_partner__name")
        .annotate(
            total_income=Sum("driver_logistic_cost"),
            total_orders=Count("id"),
            total_distance=Sum("distance_km"),
        )
        .order_by("-total_income")
    )

    leaderboard = list(leaderboard_qs[:10])
    current_rank = None
    for idx, row in enumerate(leaderboard_qs, start=1):
        if row["delivery_partner_id"] == driver.id:
            current_rank = idx
            break

    # On ne garde que les 10 premiers pour l'affichage
    leaderboard = leaderboard[:10]

    # 7) Quelques dernières commandes pour le bas de page
    last_orders = orders_qs[:10]

    context = {
        "driver": driver,
        "period_days": days,
        "start_date": start,
        "end_date": end,

        # KPIs
        "total_orders": total_orders,
        "done_orders": done_orders,
        "in_progress_orders": in_progress_orders,
        "pending_orders": pending_orders,
        "canceled_orders": canceled_orders,
        "total_distance_km": total_distance_km,
        "total_income": total_income,
        "completion_rate": completion_rate,
        "cancel_rate": cancel_rate,
        "avg_income_per_order": avg_income_per_order,
        "avg_income_per_km": avg_income_per_km,

        # Séries pour les graphiques
        "daily_labels": daily_labels,
        "daily_orders_series": daily_orders_series,
        "daily_income_series": daily_income_series,
        "daily_distance_series": daily_distance_series,

        # Classement
        "leaderboard": leaderboard,
        "current_rank": current_rank,

        # Dernières commandes
        "last_orders": last_orders,
    }

    return render(request, "orders/driver_performance.html", context)


@login_required
def driver_wallet(request):
    """LEGACY: redirige vers le nouveau wallet livreur (app wallets)."""
    driver_id = (request.GET.get("driver_id") or "").strip()
    url = "/wallets/driver/wallet/"
    if driver_id:
        url = f"{url}?driver_id={driver_id}"
    return redirect(url)
def _get_connected_driver(request, order=None):
    """
    Résout le livreur "connecté" (profil DeliveryPartner) de façon robuste.

    Priorité :
    1) ?driver_id=xx (autorisé si staff OU en DEBUG pour faciliter les tests)
    2) mapping par email : request.user.email == DeliveryPartner.email
    3) fallback : order.delivery_partner (si order fourni)
    4) fallback DEV : premier livreur actif (DEBUG uniquement)
    """
    user = getattr(request, "user", None)
    if not user or not user.is_authenticated:
        return None

    is_debug = bool(getattr(settings, "DEBUG", False))
    is_staff = bool(getattr(user, "is_staff", False))

    # 1) driver_id via URL (staff OU debug)
    driver_id = (request.GET.get("driver_id") or "").strip()
    if driver_id and (is_staff or is_debug):
        d = DeliveryPartner.objects.filter(pk=driver_id, is_active=True).first()
        if d:
            return d

    # 2) mapping par email (prod)
    email = (getattr(user, "email", "") or "").strip()
    if email:
        d = DeliveryPartner.objects.filter(email__iexact=email, is_active=True).first()
        if d:
            return d

    # 3) fallback order
    if order is not None:
        dp = getattr(order, "delivery_partner", None)
        if dp and getattr(dp, "is_active", True):
            return dp

    # 4) fallback DEV
    if is_debug:
        return DeliveryPartner.objects.filter(is_active=True).order_by("id").first()

    return None


def normalize_order_legs(order, driver=None):
    """
    Verrouille la cohérence legs (idempotent) pour une commande.

    Règles :
    - 1 seul driver "actif" = order.delivery_partner (si défini)
    - Pour ce driver actif : au plus 1 pickup actif + 1 return actif
    - Les legs d'autres drivers => canceled (sauf done/canceled)
    - wash_complete_time :
        * NULL  => return actif ne peut pas être assigned/in_progress -> pending
        * OK    => return actif pending -> assigned
    - Ne touche JAMAIS aux legs done/canceled (historique)
    """

    STATUS_RANK = {"pending": 1, "assigned": 2, "in_progress": 3, "done": 4, "canceled": 0}

    def _advance_status_only(current: str, target: str) -> str:
        c = (current or "pending").lower()
        t = (target or "pending").lower()
        # canceled/done ne bougent plus
        if c in ("canceled", "done"):
            return c
        # on n'accepte que si ça avance
        return t if STATUS_RANK.get(t, 1) > STATUS_RANK.get(c, 1) else c

    from django.db import transaction

    from orders.models import DeliveryLeg

    if not order:
        return

    assigned = getattr(order, "delivery_partner", None)

    # driver cible (si driver passé, on normalise sur l'assignation quand elle existe)
    target_driver = assigned or driver

    with transaction.atomic():
        # 1) Annuler les legs "actifs" des autres drivers (si commande assignée)
        if assigned:
            other_active = (
                DeliveryLeg.objects
                .select_for_update()
                .filter(order=order)
                .exclude(driver=assigned)
                .exclude(status__in=["done", "canceled"])
            )
            # 🔒 Ne jamais annuler un leg déjà payé (payout existe)
            try:
                from wallets.models import WalletTransaction
                paid_leg_ids = set(
                    WalletTransaction.objects.filter(
                        order_id=order.id,
                        wallet__owner_type="driver",
                        type="payout",
                        direction="in",
                    ).exclude(leg_id__isnull=True).values_list("leg_id", flat=True)
                )
            except Exception:
                paid_leg_ids = set()

            other_active.exclude(id__in=paid_leg_ids).update(status="canceled")

        # 2) Pour le driver cible : garder 1 pickup + 1 return actifs (le plus récent)
        if target_driver:
            for lt in ["pickup", "return"]:
                qs = (
                    DeliveryLeg.objects
                    .select_for_update()
                    .filter(order=order, driver=target_driver, leg_type=lt)
                    .exclude(status__in=["done", "canceled"])
                    .order_by("-id")
                )
                keep = qs.first()
                if keep:
                    qs_to_cancel = qs.exclude(id=keep.id)

                    # 🔒 Ne jamais annuler un leg déjà payé
                    try:
                        from wallets.models import WalletTransaction
                        paid_leg_ids = set(
                            WalletTransaction.objects.filter(
                                order_id=order.id,
                                wallet__owner_type="driver",
                                type="payout",
                                direction="in",
                            ).exclude(leg_id__isnull=True).values_list("leg_id", flat=True)
                        )
                    except Exception:
                        paid_leg_ids = set()

                    qs_to_cancel.exclude(id__in=paid_leg_ids).update(status="canceled")

        # 3) Sync statut return selon wash_complete_time (sur le return actif)
        wash_ready = bool(getattr(order, "wash_complete_time", None))
        qs_return = (
            DeliveryLeg.objects
            .select_for_update()
            .filter(order=order, driver=target_driver, leg_type="return")
            .exclude(status__in=["done", "canceled"])
        )
        # ✅ IMPORTANT : on n'auto-upgrade plus le return.
        # Le return reste 'pending' tant que le livreur n'a pas "accept".
        # (Le blocage métier wash_complete_time est déjà géré dans update_leg_status.)
        pass


def ensure_default_driver_legs(order, driver):
    """
    SAFE helper pour la vue livreur :
    - normalise les legs (anti-doublons, 1 driver actif)
    - ne fait un sync "models" QUE si aucun leg actif n'existe (legacy)
    - si un type manque (pickup/return), crée UNIQUEMENT le leg manquant
      sans toucher aux statuts déjà en cours.
    """
    from decimal import Decimal, ROUND_HALF_UP
    from orders.models import DeliveryLeg

    if not order or not driver:
        return DeliveryLeg.objects.none()

    # 🔒 0) Normalisation globale (évite doublons et multi-drivers actifs)
    try:
        normalize_order_legs(order, driver=driver)
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=8551")

    # 🔒 Garde-fou : si commande assignée, on ne gère que le driver assigné
    assigned = getattr(order, "delivery_partner", None)
    if assigned and str(getattr(driver, "id", "")) != str(getattr(assigned, "id", "")):
        return DeliveryLeg.objects.filter(order=order, driver=driver).exclude(status="canceled").order_by("id")

    # Legs actuels (actifs)
    qs = DeliveryLeg.objects.filter(order=order, driver=driver).exclude(status="canceled").order_by("id")

    # ✅ Legacy : si aucun leg actif du tout, on peut resync via models
    if not qs.exists():
        try:
            from orders.models import sync_delivery_legs_for_order
            sync_delivery_legs_for_order(order)
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=8567")
        qs = DeliveryLeg.objects.filter(order=order, driver=driver).exclude(status="canceled").order_by("id")

    # Helper arrondi FCFA
    def _round_fcfa(v):
        if v is None:
            return Decimal("0")
        if not isinstance(v, Decimal):
            v = Decimal(str(v))
        return v.quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    # Si un type manque, on crée uniquement le leg manquant (sans resync global)
    have_pickup = qs.filter(leg_type="pickup").exists()
    have_return = qs.filter(leg_type="return").exists()

    # Base status
    if getattr(order, "status", None) == "done":
        base_pickup = "done"
        base_return = "done"
    elif getattr(order, "status", None) == "in_progress":
        base_pickup = "assigned"
        # 🔒 IMPORTANT : return reste toujours pending jusqu'au "accept" du livreur
        base_return = "pending"
    else:
        base_pickup = "pending"
        base_return = "pending"

    # Données financières (split 50/50) — fallback 0
    delivery_fee = Decimal(str(getattr(order, "delivery_fee", 0) or 0))
    driver_total = Decimal(str(getattr(order, "amount_driver_partner", 0) or 0))
    margin_total = Decimal(str(getattr(order, "logistic_margin", 0) or 0))

    client_share_1 = _round_fcfa(delivery_fee / 2)
    client_share_2 = _round_fcfa(delivery_fee - client_share_1)

    driver_share_1 = _round_fcfa(driver_total / 2)
    driver_share_2 = _round_fcfa(driver_total - driver_share_1)

    margin_share_1 = _round_fcfa(margin_total / 2)
    margin_share_2 = _round_fcfa(margin_total - margin_share_1)

    # Distance (si connue)
    distance_total = getattr(order, "distance_km", None) or getattr(order, "distance_km_total", None) or 0
    try:
        distance_total = Decimal(str(distance_total or 0))
    except Exception:
        distance_total = Decimal("0")
    distance_one_way = None
    if distance_total > 0:
        distance_one_way = (distance_total / Decimal("2")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    # Création du leg manquant seulement
    try:
        if not have_pickup:
            DeliveryLeg.objects.create(
                order=order,
                driver=driver,
                leg_type="pickup",
                status=base_pickup,
                distance_km=float(distance_one_way) if distance_one_way is not None else None,
                client_fee_share=client_share_1,
                driver_amount=driver_share_1,
                fagni_margin=margin_share_1,
            )

        if not have_return:
            DeliveryLeg.objects.create(
                order=order,
                driver=driver,
                leg_type="return",
                status=base_return,
                distance_km=float(distance_one_way) if distance_one_way is not None else None,
                client_fee_share=client_share_2,
                driver_amount=driver_share_2,
                fagni_margin=margin_share_2,
            )
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=8644")

    # Re-load + ajuste return selon wash_complete_time (sans toucher done/canceled)
    qs = DeliveryLeg.objects.filter(order=order, driver=driver).exclude(status="canceled").order_by("id")

    try:
        wash_ready = bool(getattr(order, "wash_complete_time", None))
        r = qs.filter(leg_type="return").exclude(status__in=["done", "canceled"]).order_by("-id").first()
        if r:
            # ✅ IMPORTANT : pas d'auto-upgrade du return.
            # Il reste pending jusqu'au "accept" du livreur.
            pass
        qs = DeliveryLeg.objects.filter(order=order, driver=driver).exclude(status="canceled").order_by("id")
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=8658")

    return qs


@login_required
def driver_order_detail(request, order_id):
    """
    Vue détail course côté LIVREUR.

    - Montre uniquement les infos utiles au chauffeur
    - Calcule la distance & le montant qui lui reviennent à partir des DeliveryLeg
    - Crée des DeliveryLeg par défaut s'il n'y en a pas encore
    - Affiche des montants cohérents (moteur compute_totals + fallbacks)
    """

    from decimal import Decimal
    from django.db.models import Sum
    from django.shortcuts import get_object_or_404, redirect, render

    order = get_object_or_404(
        Order.objects.select_related("customer", "delivery_partner", "laundry_partner"),
        pk=order_id,
    )

    display_summary = build_order_display_summary(order)
    finance_summary = build_order_finance_summary(order)

    # ------------------------------------------------------------
    # Driver à utiliser pour la vue "détail course"
    # Règle :
    # - côté livreur (non-staff) : accès UNIQUEMENT si driver_id == order.delivery_partner_id
    # - staff : peut consulter; si order.delivery_partner est vide, peut fallback via ?driver_id
    # ------------------------------------------------------------
    selected_driver_id = (request.GET.get("driver_id") or "").strip()

    assigned_driver = getattr(order, "delivery_partner", None)  # peut être None (legacy)
    driver = assigned_driver  # par défaut on se cale sur l'assignation

    if not request.user.is_staff:
        # Sans assignation : on refuse (sinon on ne sait pas "qui" a le droit)
        if not assigned_driver:
            return redirect("orders:driver_hub")

        # driver_id obligatoire et doit matcher l'assignation
        # 🔓 fallback automatique
        if not selected_driver_id:
            selected_driver_id = str(assigned_driver.id)

        if str(assigned_driver.id) != str(selected_driver_id):
            return redirect("orders:driver_hub")

    else:
        # Staff : si la commande n'a pas de livreur assigné, on peut fallback sur ?driver_id (optionnel)
        if driver is None and selected_driver_id:
            driver = DeliveryPartner.objects.filter(pk=selected_driver_id).first()

    # -----------------------------
    # Defaults
    # -----------------------------
    driver_legs_qs = DeliveryLeg.objects.none()
    driver_leg_distance = Decimal("0")
    driver_leg_amount_done = Decimal("0")   # ✅ payé = legs done uniquement
    driver_leg_amount_all = Decimal("0")    # (optionnel) potentiel = tous les legs
    driver_mission_type_label = "Mission unique (A/R ou globale)"
    driver_wallet = None
    driver_wallet_url = ""

    if driver is not None:

        # ✅ SAFE : ne jamais resync pendant une course si des legs existent déjà
        # (sinon ça peut réécrire assigned/in_progress -> pending)
        try:
            if not DeliveryLeg.objects.filter(order=order).exclude(status="canceled").exists():
                from orders.models import sync_delivery_legs_for_order
                sync_delivery_legs_for_order(order)
        except Exception:
            import logging
            logging.getLogger("fagni.views.legs").exception("Echec silencieux: resync legs pendant course en cours | order_id=%s", getattr(order, "id", None) if "order" in dir() else None)

        # 🔒 Normalisation légère (anti-doublons / return pending si wash pas prêt)
        try:
            normalize_order_legs(order, driver=driver)
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=8742")

        # ==========================================
        # DRIVER_LEG_POST_JSON_V1 (AJAX actions)
        # ==========================================
        if request.method == "POST":
            leg_type = (request.POST.get("leg") or "").strip()
            action = (request.POST.get("action") or "").strip()

            # Détection AJAX (le JS envoie X-Requested-With + Accept: application/json)
            try:
                is_ajax_req = is_ajax(request)
            except Exception:
                acc = (request.headers.get("Accept") or "").lower()
                is_ajax_req = (request.headers.get("X-Requested-With") == "XMLHttpRequest") or ("application/json" in acc)

            if not leg_type or not action:
                if is_ajax_req:
                    from django.http import JsonResponse
                    return JsonResponse({"ok": False, "error": "Paramètres manquants (leg/action)."}, status=400)

            # Sécurité : il faut un driver résolu (non legacy)
            if driver is None:
                if is_ajax_req:
                    from django.http import JsonResponse
                    return JsonResponse({"ok": False, "error": "Accès refusé : driver introuvable."}, status=403)

            # Charger la jambe concernée
            leg = (
                DeliveryLeg.objects
                .filter(order=order, driver=driver, leg_type=leg_type)
                .exclude(status="canceled")
                .order_by("id")
                .first()
            )

            if not leg:
                if is_ajax_req:
                    from django.http import JsonResponse
                    return JsonResponse({"ok": False, "error": f"Leg introuvable: {leg_type}."}, status=404)

            changed, msg = update_leg_status(leg, action, user=request.user)

            if is_ajax_req:
                from django.http import JsonResponse
                return JsonResponse(
                    {
                        "ok": bool(changed),
                        "message": msg or ("OK" if changed else "Aucune modification"),
                        "leg_type": leg_type,
                        "action": action,
                        "new_status": getattr(leg, "status", None),
                    },
                    status=(200 if changed else 400),
                )

        # 1) On charge les legs (pour le driver de la vue)
        driver_legs_qs = DeliveryLeg.objects.filter(order=order, driver=driver).exclude(status="canceled").order_by("id")

        # 2) Agrégats distance / montant (propre)
        #    - distance : somme de toutes les jambes (info logistique)
        #    - montant payé : uniquement les jambes DONE (logique métier Cas B)
        legs_agg_all = driver_legs_qs.aggregate(
            total_distance_km=Sum("distance_km"),
            total_amount=Sum("driver_amount"),
        )
        driver_leg_distance = legs_agg_all["total_distance_km"] or Decimal("0")
        driver_leg_amount_all = legs_agg_all["total_amount"] or Decimal("0")

        legs_agg_done = driver_legs_qs.filter(status="done").aggregate(
            total_amount=Sum("driver_amount"),
        )
        driver_leg_amount_done = legs_agg_done["total_amount"] or Decimal("0")

        # 3) Fallbacks legacy (très anciennes commandes sans legs)
        if (driver_leg_distance == 0) and getattr(order, "distance_km_total", None):
            try:
                driver_leg_distance = Decimal(str(order.distance_km_total))
            except Exception:
                driver_leg_distance = Decimal("0")

        # Si aucune jambe done, on ne force pas le montant "payé".
        # (On peut garder le potentiel via driver_leg_amount_all)
        # MAIS si l'ordre est DONE et qu'il n'y a pas de legs done (legacy), on fallback.
        if (driver_leg_amount_done == 0) and (getattr(order, "status", "") == "done"):
            if getattr(order, "amount_driver_partner", None):
                try:
                    driver_leg_amount_done = Decimal(str(order.amount_driver_partner))
                except Exception:
                    driver_leg_amount_done = Decimal("0")

        # 4) Libellé type de mission à partir des leg_type
        leg_types = list(driver_legs_qs.values_list("leg_type", flat=True).distinct())
        if not leg_types:
            driver_mission_type_label = "Mission unique (A/R ou globale)"
        elif len(leg_types) == 1:
            lt = leg_types[0]
            if lt == "pickup":
                driver_mission_type_label = "Collecte (client → blanchisserie)"
            elif lt == "delivery":
                driver_mission_type_label = "Livraison (blanchisserie → client)"
            elif lt == "round_trip":
                driver_mission_type_label = "Collecte + livraison (A/R complet)"
            else:
                driver_mission_type_label = "Mission particulière"
        else:
            driver_mission_type_label = "Collecte + livraison (tronçons multiples)"

        # 5) Wallet livreur
        try:
            driver_wallet = get_or_create_wallet_for_delivery_partner(driver)

            # Force driver_id dans le querystring
            qd = request.GET.copy()
            qd["driver_id"] = str(driver.id)
            qs = qd.urlencode()

            driver_wallet_url = reverse("wallets:driver_wallet_dashboard")
            if qs:
                driver_wallet_url = f"{driver_wallet_url}?{qs}"

        except Exception:
            driver_wallet = None
            driver_wallet_url = ""

    # ------------------------------------------------------------
    # Montants FAGNI cohérents (moteur compute_totals + fallbacks)
    # ------------------------------------------------------------
    try:
        amounts = order.compute_totals(save=False) or {}
    except Exception:
        amounts = {}

    # Total client TTC : moteur central -> champs DB -> fallback calculé
    total_client_ttc = (
        amounts.get("total_ttc_client")
        or amounts.get("total_client_ttc")
        or getattr(order, "total_client_ttc", None)
    )

    try:
        total_client_ttc = Decimal(str(total_client_ttc)) if total_client_ttc is not None else Decimal("0")
    except Exception:
        total_client_ttc = Decimal("0")

    if total_client_ttc <= 0:
        try:
            # ✅ même logique que driver_app_data : prestation_total sinon somme(items)
            from django.db.models import F as _F, Value as _Value, DecimalField as _DecimalField, Sum as _Sum
            from django.db.models.functions import Coalesce as _Coalesce, Cast as _Cast

            DEC = _DecimalField(max_digits=12, decimal_places=2)

            items_total = order.items.aggregate(
                s=_Coalesce(
                    _Sum(_Cast(_F("quantity"), DEC) * _Cast(_F("unit_price"), DEC)),
                    _Value(0, output_field=DEC),
                )
            ).get("s") or Decimal("0")

            prestation_raw = getattr(order, "prestation_total", None)
            prestation = Decimal(str(prestation_raw)) if prestation_raw not in (None, "", 0) else Decimal(str(items_total))

            if prestation <= 0:
                # fallback ultime (legacy)
                prestation = Decimal(str(getattr(order, "total", None) or 0))

            service = Decimal(str(getattr(order, "service_fee", None) or 0))
            delivery = Decimal(str(getattr(order, "delivery_fee", None) or 0))
            express = Decimal(str(getattr(order, "express_extra_fee", None) or 0))
            vat = Decimal(str(getattr(order, "vat_fagni", None) or 0))

            total_client_ttc = prestation + service + delivery + express + vat
        except Exception:
            total_client_ttc = Decimal("0")


    delivery_fee_client = (
        amounts.get("delivery_fee_client")
        or getattr(order, "delivery_fee", None)
        or Decimal("0")
    )
    try:
        delivery_fee_client = Decimal(str(delivery_fee_client))
    except Exception:
        delivery_fee_client = Decimal("0")

    service_fee_ht = amounts.get("service_fee_ht") or getattr(order, "service_fee", None) or Decimal("0")
    vat_fagni = amounts.get("vat_fagni") or getattr(order, "vat_fagni", None) or Decimal("0")
    express_surcharge = amounts.get("express_surcharge") or getattr(order, "express_extra_fee", None) or Decimal("0")

    try:
        service_fee_ht = Decimal(str(service_fee_ht))
    except Exception:
        service_fee_ht = Decimal("0")
    try:
        vat_fagni = Decimal(str(vat_fagni))
    except Exception:
        vat_fagni = Decimal("0")
    try:
        express_surcharge = Decimal(str(express_surcharge))
    except Exception:
        express_surcharge = Decimal("0")

    # Revenu livreur affiché :
    # ✅ source-of-truth = net wallet (payout/adjustment) sur cette commande
    # (aligné avec driver_app_data)
    # ✅ Revenu livreur affiché : net des tx payout/adjustment liées aux legs du driver
    driver_income = Decimal("0")
    if driver is not None:
        try:
            # ✅ anti-wallet parasite : uniquement les tx du wallet de CE driver
            qs = WalletTransaction.objects.filter(
                order=order,
                type__in=["payout", "adjustment"],
                wallet__delivery_partner=driver,
            )

            # priorité: tx liées à une leg du driver (logique par tronçon)
            net = qs.filter(leg__isnull=False, leg__driver=driver).aggregate(net=_wallet_net_expr()).get("net")

            # fallback: tx order-level sur le wallet du driver (ex: corrections/legacy)
            if net is None:
                net = qs.aggregate(net=_wallet_net_expr()).get("net")

            driver_income = net or Decimal("0")
        except Exception:
            driver_income = Decimal("0")

    else:
        # legacy : pas de driver résolu
        driver_income = (
            amounts.get("amount_driver_partner")
            or getattr(order, "amount_driver_partner_resolved", None)
            or getattr(order, "amount_driver_partner", None)
            or getattr(order, "driver_logistic_cost", None)
            or Decimal("0")
        )
        try:
            driver_income = Decimal(str(driver_income))
        except Exception:
            driver_income = Decimal("0")

    # ------------------------------------------------------------
    # KPI "reste à gagner" + % (potentiel - payé wallet)
    # ------------------------------------------------------------
    driver_income_remaining = Decimal("0")
    driver_income_progress_pct = 0

    try:
        if driver_leg_amount_all and driver_leg_amount_all > 0:
            driver_income_remaining = (driver_leg_amount_all - driver_income) if driver_leg_amount_all > driver_income else Decimal("0")
            driver_income_progress_pct = int((driver_income / driver_leg_amount_all) * 100) if driver_leg_amount_all else 0
            driver_income_progress_pct = max(0, min(100, driver_income_progress_pct))
    except Exception:
        driver_income_remaining = Decimal("0")
        driver_income_progress_pct = 0


    # ------------------------------------------------------------
    # Coordonnées (maps)
    # ------------------------------------------------------------
    pickup_coords = resolve_pickup_coords(order)
    delivery_coords = resolve_delivery_coords(order)
    provider_coords = resolve_provider_coords(order)

    pickup_lat, pickup_lng = pickup_coords if pickup_coords else (None, None)
    delivery_lat, delivery_lng = delivery_coords if delivery_coords else (None, None)
    provider_lat, provider_lng = provider_coords if provider_coords else (None, None)

    context = {
        "order": order,
        "driver": driver,
        "driver_legs_qs": driver_legs_qs,
        "driver_leg_distance": driver_leg_distance,

        # ✅ affichage métier
        "driver_leg_amount": driver_leg_amount_done,   # ce que la page utilisait déjà
        "driver_leg_amount_done": driver_leg_amount_done,
        "driver_leg_amount_all": driver_leg_amount_all,  # potentiel (si tu veux l'afficher dans le template)

        "driver_mission_type_label": driver_mission_type_label,
        "driver_wallet": driver_wallet,
        "driver_wallet_url": driver_wallet_url,

        "amounts": amounts,
        "total_client_ttc": finance_summary.get("total_client_ttc", total_client_ttc),
        "delivery_fee_client": finance_summary.get("delivery_fee_client", delivery_fee_client),
        "driver_income": driver_income,
        "driver_income_remaining": driver_income_remaining,
        "driver_income_progress_pct": driver_income_progress_pct,

        "pickup_coords": pickup_coords,
        "delivery_coords": delivery_coords,
        "provider_coords": provider_coords,
        "pickup_lat": pickup_lat,
        "pickup_lng": pickup_lng,
        "delivery_lat": delivery_lat,
        "delivery_lng": delivery_lng,
        "provider_lat": provider_lat,
        "provider_lng": provider_lng,

        "service_fee_ht": finance_summary.get("service_fee_ht", service_fee_ht),
        "vat_fagni": vat_fagni,
        "express_surcharge": express_surcharge,

    }

    return render(request, "orders/driver_order_detail_v2.html", context)


# ===============================================
#  APP LIVREUR – PAGE
# ===============================================
@login_required
def driver_app(request):
    """
    Page App livreur : UI shell + driver connecté.

    IMPORTANT:
    - Les listes + KPIs sont alimentés par /orders/driver-app/data/ (source-of-truth),
      afin d'éviter les incohérences entre order.status et legs.status.
    - Livreur non-staff: voit uniquement ses courses (filtrage appliqué dans driver_app_data)
    - Staff: peut filtrer via ?driver_id=7 (filtrage appliqué dans driver_app_data)
    """
    from django.shortcuts import render

    connected_driver = _get_connected_driver(request)
    selected_driver_id = (request.GET.get("driver_id") or "").strip()

    context = {
        "connected_driver": connected_driver,
        "selected_driver_id": selected_driver_id,
    }
    return render(request, "orders/driver_app.html", context)


# ===============================================
#  APP LIVREUR – DATA JSON pour refresh KPIs
# ===============================================
# ===============================================
@login_required
def driver_app_data(request):
    """
    Version JSON live de l'app livreur :
    - KPIs (counts, distance, income)
    - listes de commandes (pending/in_progress/done) pour refresh UI
    IMPORTANT:
      - filtrage aligné avec driver_app()
      - grouping basé sur les DeliveryLeg (source de vérité UI), pas sur order.status

    INCOME:
      - paid = WalletTransaction net (payout/adjustment)
      - potential = somme des montants potentiels des legs (fallback)
      - display = paid si >0 sinon potential
    """
    from decimal import Decimal
    from django.http import JsonResponse
    from django.db.models import Sum, F, Value, DecimalField, Case, When
    from django.db.models.functions import Coalesce, Cast
    from django.utils.timezone import localtime

    user = request.user
    connected_driver = _get_connected_driver(request)
    selected_driver_id = (request.GET.get("driver_id") or "").strip()

    # ==========================
    # BASE QUERY (ALIGNÉE driver_app)
    # ==========================
    qs = Order.objects.select_related("customer", "laundry_partner", "delivery_partner")

    if not user.is_staff:
        if connected_driver:
            qs = qs.filter(legs__driver=connected_driver).distinct()
        else:
            qs = qs.none()
    else:
        if selected_driver_id:
            qs = qs.filter(legs__driver_id=selected_driver_id).distinct()
        else:
            qs = qs  # staff voit tout par défaut

    qs = qs.order_by("-created_at")

    # ==========================
    # ANNOTATIONS (montants + distance)
    # ==========================
    DEC = DecimalField(max_digits=12, decimal_places=2)

    items_total_expr = Coalesce(
        Sum(Cast(F("items__quantity"), DEC) * Cast(F("items__unit_price"), DEC)),
        Value(0, output_field=DEC),
    )

    prestation_expr = Coalesce(F("prestation_total"), items_total_expr, output_field=DEC)
    service_expr = Coalesce(F("service_fee"), Value(0, output_field=DEC))
    delivery_expr = Coalesce(F("delivery_fee"), Value(0, output_field=DEC))
    vat_expr = Coalesce(F("vat_fagni"), Value(0, output_field=DEC))
    express_expr = Coalesce(F("express_extra_fee"), Value(0, output_field=DEC))

    total_client_fallback_expr = prestation_expr + service_expr + delivery_expr + express_expr + vat_expr

    # ✅ IMPORTANT: si total_client_ttc == 0 => fallback (pas Coalesce)
    total_client_expr = Case(
        When(total_client_ttc__gt=0, then=F("total_client_ttc")),
        default=total_client_fallback_expr,
        output_field=DEC,
    )

    distance_expr = Coalesce(F("distance_km_total"), Value(0, output_field=DEC), output_field=DEC)

    qs = qs.annotate(
        items_total=items_total_expr,
        total_client_display=total_client_expr,
        driver_distance_display=distance_expr,
    )

    # IMPORTANT: on récupère les legs pour calculer un statut UI fiable + potentiel revenu
    qs = qs.prefetch_related("legs")

    # ==========================================================
    # ✅ INCOME SOURCE-OF-TRUTH: WalletTransaction net par commande
    #    priorité: tx liées à un leg
    #    fallback: tx order-level (leg NULL) sur wallet driver
    # ==========================================================
    order_ids = list(qs.values_list("id", flat=True))

    base_tx = WalletTransaction.objects.filter(
        order_id__in=order_ids,
        type__in=["payout", "adjustment"],
        wallet__owner_type="driver",                 # ✅ IMPORTANT: uniquement wallets livreurs
    )
    base_tx = base_tx.filter(wallet__delivery_partner__isnull=False)

    driver_target_id = None
    if not user.is_staff and connected_driver:
        driver_target_id = connected_driver.id
    elif user.is_staff and selected_driver_id:
        try:
            driver_target_id = int(selected_driver_id)
        except Exception:
            driver_target_id = None

    if driver_target_id:
        base_tx = base_tx.filter(wallet__delivery_partner_id=driver_target_id)

    # 1) priorité: legs-level
    income_rows_legs = (
        base_tx.filter(leg__isnull=False)
        .values("order_id")
        .annotate(net=_wallet_net_expr())
    )
    income_by_order_id = {r["order_id"]: (r["net"] or Decimal("0")) for r in income_rows_legs}

    # 2) fallback: order-level (leg NULL) uniquement si pas déjà présent
    missing_ids = [oid for oid in order_ids if oid not in income_by_order_id]
    if missing_ids:
        income_rows_order = (
            base_tx.filter(order_id__in=missing_ids, leg__isnull=True)
            .values("order_id")
            .annotate(net=_wallet_net_expr())
        )
        for r in income_rows_order:
            income_by_order_id[r["order_id"]] = (r["net"] or Decimal("0"))

    # ✅ SAFETY: ne jamais exposer de 'paid' négatif côté UI (adjustments peuvent être < 0)
    try:
        for _oid, _net in list(income_by_order_id.items()):
            if _net is None or _net < 0:
                income_by_order_id[_oid] = Decimal("0")
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=9213")

    def _safe_float(x):
        try:
            return float(x or 0)
        except Exception:
            return 0.0

    def compute_status_from_legs(order):
        """
        Source de vérité UI:
          - si un leg in_progress OU assigned => in_progress
          - sinon si un leg pending => pending
          - sinon si tous done => done
          - sinon fallback: order.status
        """
        legs = list(order.legs.all())
        if not legs:
            return order.status

        st = [str(l.status or "").lower() for l in legs]

        # ✅ assigned = course démarrée côté UI
        if any(s in ("in_progress", "assigned") for s in st):
            return "in_progress"

        if any(s == "pending" for s in st):
            return "pending"

        # si tous les legs actifs sont done (on tolère canceled)
        active = [s for s in st if s != "canceled"]
        if active and all(s == "done" for s in active):
            return "done"

        return order.status

    def _dec0(x):
        try:
            return Decimal(str(x or "0"))
        except Exception:
            return Decimal("0")

    def _leg_driver_potential(leg) -> Decimal:
        """
        Retourne la part livreur potentielle d'un leg.
        On essaie plusieurs noms de champs possibles pour rester robuste.
        """
        # IMPORTANT: certains champs peuvent être "niveau commande" et se répéter sur chaque jambe.
        # => on protège pour éviter un potentiel multiplié par nb de legs.

        for attr in (
            "amount_driver",
            "driver_amount",
            "driver_share",
            "driver_fee",
            "driver_income",
            "amount_driver_potential",
            "driver_logistic_cost",
            "amount_driver_partner",
        ):
            if hasattr(leg, attr):
                v = getattr(leg, attr, None)
                d = _dec0(v)
                if d > 0:
                    # 🚫 Guard anti-multiplication: amount_driver_partner est souvent un champ "commande"
                    if attr == "amount_driver_partner":
                        try:
                            order = getattr(leg, "order", None)
                            if order:
                                nb = order.legs.exclude(status="canceled").count()
                                if nb > 1:
                                    continue
                        except Exception:
                            continue
                    return d
        return Decimal("0")

    def _order_driver_potential_from_legs(order) -> Decimal:
        try:
            legs = list(order.legs.all())
        except Exception:
            legs = []
        total = Decimal("0")
        for leg in legs:
            # 🚫 Exclure les jambes annulées du potentiel
            try:
                st = (getattr(leg, "status", "") or "").lower().strip()
                if st == "canceled":
                    continue
                if getattr(leg, "is_canceled", False):
                    continue
                if getattr(leg, "canceled_at", None):
                    continue
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=9307")
            total += _leg_driver_potential(leg)
        return total

    def serialize(order, computed_status):
        customer = getattr(order, "customer", None)

        total_client_num = _safe_float(getattr(order, "total_client_display", 0) or 0)

        # ✅ montant déjà payé (DB)
        paid_num = _safe_float(getattr(order, "amount_paid", 0) or 0)

        # ✅ fallback robustesse: si payment_status dit "paid" mais amount_paid=0
        payment_status = str(getattr(order, "payment_status", "") or "").lower().strip()
        if total_client_num > 0 and paid_num <= 0 and payment_status in ("paid", "completed", "succeeded"):
            paid_num = total_client_num

        # ✅ dû = max(0, total - payé)
        due_num = total_client_num - paid_num
        if due_num < 0:
            due_num = 0.0

        # ✅ Label paiement
        if total_client_num <= 0:
            pay_label = "À CALCULER"
            total_client_out = None   # JS affiche "—"
            due_out = 0.0
        else:
            total_client_out = total_client_num
            due_out = due_num

            if due_num <= 0:
                pay_label = "PAYÉ"
            elif 0 < due_num < total_client_num:
                pay_label = "PARTIEL"
            else:
                pay_label = "À ENCAISSER"

        # si c'est "à calculer", on masque le revenu potentiel (cohérence UI)
        income_out = 0.0
        income_paid = Decimal("0")
        income_potential = Decimal("0")

        # ✅ revenu livreur: payé si dispo, sinon potentiel
        income_paid = income_by_order_id.get(order.id, Decimal("0")) or Decimal("0")
        income_potential = _order_driver_potential_from_legs(order)

        # ✅ revenu livreur affiché: projection = max(paid, potential)
        # - paid = net wallet déjà crédité
        # - potential = somme estimée (legs)
        # => on affiche la projection la plus haute (cohérence KPI + UI)
        # ✅ afficher TOUJOURS le payé s'il existe, même si < potentiel
        if income_paid > 0:
            income_out = _safe_float(income_paid)
        else:
            income_out = _safe_float(income_potential)

        # ✅ distance
        dist_out = _safe_float(getattr(order, "driver_distance_display", 0) or 0)

        detail_url = reverse("orders:driver_order_detail", args=[order.id]) + "?back=" + quote(
            reverse("orders:driver_hub")
        )
        dp_id = getattr(order, "delivery_partner_id", None)
        if dp_id:
            detail_url = f"{detail_url}&driver_id={dp_id}"
        return {
            "id": order.id,
            "code": order.code,
            "status": computed_status,
            "raw_status": order.status,

            "pay_label": pay_label,
            "created_at": localtime(order.created_at).strftime("%d/%m/%Y %H:%M") if order.created_at else None,
            "customer_name": getattr(customer, "name", None) or "Client",
            "customer_phone": getattr(customer, "phone", "") if customer else "",
            "customer_address": getattr(customer, "address", "") if customer else "",

            "total_client": total_client_out,
            "due_amount": due_out,

            "driver_income": income_out,
            "driver_distance": dist_out,

            "detail_url": detail_url,

            "driver_income_paid": _safe_float(income_paid),
            "driver_income_potential": _safe_float(income_potential),
    }

    # ==========================
    # BUILD DATA (grouping by computed_status)
    # ==========================
    orders_all = list(qs[:300])

    orders_pending_list = []
    orders_in_progress_list = []
    orders_done_list = []

    pending_count = 0
    in_progress_count = 0
    done_count = 0

    # KPI totals
    total_distance = Decimal("0")
    total_income_paid = Decimal("0")
    total_income_potential = Decimal("0")
    total_income_display = Decimal("0")  # ✅ cohérent avec serialize(): paid si dispo sinon potential

    for o in orders_all:
        cs = compute_status_from_legs(o)

        total_distance += Decimal(str(_safe_float(getattr(o, "driver_distance_display", 0) or 0)))

        paid = income_by_order_id.get(o.id, Decimal("0")) or Decimal("0")
        total_income_paid += paid

        pot = _order_driver_potential_from_legs(o)
        total_income_potential += pot

        # ✅ règle identique à serialize(): paid si dispo sinon potential
        total_income_display += (paid if paid > 0 else pot)

        if cs == "pending":
            pending_count += 1
            if len(orders_pending_list) < 60:
                orders_pending_list.append(serialize(o, cs))
        elif cs == "in_progress":
            in_progress_count += 1
            if len(orders_in_progress_list) < 60:
                orders_in_progress_list.append(serialize(o, cs))
        elif cs == "done":
            done_count += 1
            if len(orders_done_list) < 60:
                orders_done_list.append(serialize(o, cs))

    return JsonResponse({
        "pending": pending_count,
        "in_progress": in_progress_count,
        "done": done_count,

        "total_distance_km": float(total_distance),

        # NEW: paid/potential/display
        "total_driver_income_paid": float(total_income_paid),
        "total_driver_income_potential": float(total_income_potential),
        "total_driver_income_display": float(total_income_display),

        # keep backward compatibility (old key)
        "total_driver_income": float(total_income_paid),

        "source_distance": "distance_km_total",
        "source_income": "total_driver_income_display=sum(paid_else_potential) • cards=paid_else_potential",
        "server_time": localtime().strftime("%H:%M:%S"),

        "orders_pending": orders_pending_list,
        "orders_in_progress": orders_in_progress_list,
        "orders_done": orders_done_list,
    })


# ===============================================
#  APP LIVREUR – EXPORT CSV / XLSX
# ===============================================
@login_required
def driver_app_export_csv(request):
    """
    Export CSV des courses visibles dans l'App livreur.

    - Si user livreur : uniquement SES courses
    - Si user staff : possibilité de filtrer par driver_id
    - Montants harmonisés FAGNI :
        * total_client_display = total_client_ttc (si dispo) OU fallback
        * driver_income_display = amount_driver_partner (fallback driver_logistic_cost)
    """
    user = request.user
    connected_driver = _get_connected_driver(request)

    qs = Order.objects.select_related(
        "customer",
        "delivery_partner",
        "laundry_partner",
    )

    driver_id = (request.GET.get("driver_id") or "").strip()
    status_filter = (request.GET.get("status") or "active").strip()

    # --- Filtre livreur ---
    if connected_driver and not user.is_staff:
        # Mode LIVREUR : uniquement ses propres courses
        qs = qs.filter(legs__driver=connected_driver).distinct()
    elif driver_id:
        qs = qs.filter(legs__driver_id=driver_id).distinct()

    # --- Filtre statut (même logique que driver_app_data) ---
    if status_filter == "active":
        qs = qs.filter(status__in=["pending", "in_progress"])
    elif status_filter == "done":
        qs = qs.filter(status="done")
    elif status_filter == "canceled":
        qs = qs.filter(status="canceled")
    # "all" => pas de filtre supplémentaire

    qs = qs.order_by("-created_at")

    # --- Expressions de montants harmonisés FAGNI ---
    DEC = DecimalField(max_digits=12, decimal_places=2)

    items_total_expr = Coalesce(
        Sum(Cast(F("items__quantity"), DEC) * Cast(F("items__unit_price"), DEC)),
        Value(0, output_field=DEC),
    )

    # ✅ Prestation = prestation_total si dispo, sinon somme des items
    prestation_expr = Coalesce(F("prestation_total"), items_total_expr, output_field=DEC)

    service_expr = Coalesce(F("service_fee"), Value(0, output_field=DEC))
    delivery_expr = Coalesce(F("delivery_fee"), Value(0, output_field=DEC))
    express_expr = Coalesce(F("express_extra_fee"), Value(0, output_field=DEC))
    vat_expr = Coalesce(F("vat_fagni"), Value(0, output_field=DEC))

    total_client_fallback_expr = prestation_expr + service_expr + delivery_expr + express_expr + vat_expr

    # ✅ IMPORTANT : si total_client_ttc == 0 => fallback (pas Coalesce)
    total_client_expr = Case(
        When(total_client_ttc__gt=0, then=Cast(F("total_client_ttc"), DEC)),
        default=total_client_fallback_expr,
        output_field=DEC,
    )

    income_expr = Coalesce(
        F("amount_driver_partner"),
        Coalesce(F("driver_logistic_cost"), Value(0, output_field=DEC)),
        output_field=DEC,
    )

    distance_expr = Coalesce(F("distance_km_total"), Value(0, output_field=DEC), output_field=DEC)

    qs = qs.annotate(
        items_total=items_total_expr,
        total_client_display=total_client_expr,
        driver_income_display=income_expr,
        driver_distance_display=distance_expr,
    )

    # --- Réponse CSV ---
    response = HttpResponse(content_type="text/csv")

    filename_parts = ["driver_app"]
    if connected_driver and not user.is_staff:
        filename_parts.append(f"driver_{connected_driver.id}")
    elif driver_id:
        filename_parts.append(f"driver_{driver_id}")

    filename = "_".join(filename_parts) + ".csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response, delimiter=";")

    def _has_coords(obj) -> bool:
        if not obj:
            return False
        lat = getattr(obj, "latitude", None)
        lng = getattr(obj, "longitude", None)
        return bool(lat and lng)

    def _is_calc_order(order) -> bool:
        customer = getattr(order, "customer", None)
        laundry = getattr(order, "laundry_partner", None)
        has_laundry = bool(getattr(order, "laundry_partner_id", None))
        customer_ok = _has_coords(customer)
        laundry_ok = _has_coords(laundry)

        try:
            dist_total_num = float(getattr(order, "driver_distance_display", 0) or 0)
        except Exception:
            dist_total_num = 0.0

        return (not has_laundry) or (not customer_ok) or (not laundry_ok) or (dist_total_num <= 0)

    # En-têtes
    writer.writerow([
        "Code commande",
        "Date création",
        "Client",
        "Téléphone",
        "Adresse",
        "Statut",
        "Distance_km",
        "Montant client TTC",
        "Revenu livreur (FCFA)",
    ])

    # Lignes
    for order in qs:
        customer = getattr(order, "customer", None)

        calc = _is_calc_order(order)

        dist_val = float(getattr(order, "driver_distance_display", 0) or 0) if not calc else 0.0
        total_val = float(order.total_client_display or 0) if not calc else 0.0
        income_val = float(order.driver_income_display or 0) if not calc else 0.0

        writer.writerow([
            order.code or order.id,
            order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
            getattr(customer, "name", "") if customer else "",
            getattr(customer, "phone", "") if customer else "",
            getattr(customer, "address", "") if customer else "",
            order.get_status_display(),
            dist_val,
            total_val,
            income_val,
        ])

    return response


#@login_required
#def driver_app_export_xlsx(request):
#    """
#    Alias simple vers l'export CSV pour compatibilité.
 #   (On garde le même contenu, tu pourras l'ouvrir dans Excel et
#    mettre en forme avec ton template premium.)
 #   """
   # return driver_app_export_csv(request)


@login_required
def driver_app_export_xlsx(request):
    """
    Export XLSX premium des courses livreur, avec :
    - filtres par livreur + statut (même logique que driver_app / driver_app_export_csv)
    - montants alignés sur le modèle FAGNI :
        * total_client_ttc (fallback : prestation_total + service_fee + delivery_fee + vat_fagni)
        * revenu livreur = amount_driver_partner (fallback driver_logistic_cost)
    - garde-fou “À CALCULER” :
        * si commande non calculable => distance = 0, total = 0, revenu = 0
    - mise en forme Excel : en-têtes stylées, auto-filter, freeze pane, totaux.
    """
    from io import BytesIO
    from datetime import datetime

    from django.http import HttpResponse
    from django.db.models import Sum, Value, F, Case, When
    from django.db.models.functions import Coalesce, Cast
    from django.db.models import DecimalField

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    user = request.user
    connected_driver = _get_connected_driver(request)
    driver_id = None

    # --- Identification du livreur (livreur connecté ou driver_id en GET pour un staff) ---
    if connected_driver:
        driver_id = connected_driver.id
    elif user.is_staff:
        driver_id_param = (request.GET.get("driver_id") or "").strip()
        if driver_id_param:
            try:
                driver_id = int(driver_id_param)
            except ValueError:
                driver_id = None

    status_filter = (request.GET.get("status") or "active").strip()

    # On force un DecimalField commun
    DEC = DecimalField(max_digits=12, decimal_places=2)

    # Base queryset (mêmes relations que l'app livreur)
    qs = (
        Order.objects
        .select_related("customer", "delivery_partner", "laundry_partner")
        .prefetch_related("items")
        .order_by("-created_at")
    )

    if driver_id:
        qs = qs.filter(legs__driver_id=driver_id).distinct()

    # Filtre statut (même logique que driver_app / driver_app_data / driver_orders_csv)
    if status_filter == "active":
        qs = qs.filter(status__in=["pending", "in_progress"])
    elif status_filter == "done":
        qs = qs.filter(status="done")
    elif status_filter == "canceled":
        qs = qs.filter(status="canceled")
    # status_filter == "all" => pas de filtre supplémentaire

    # ---------- ANNOTATIONS FINANCIÈRES ALIGNÉES FAGNI ----------
    items_total_expr = Coalesce(
        Sum(
            Cast(F("items__quantity"), DEC) * Cast(F("items__unit_price"), DEC)
        ),
        Value(0, output_field=DEC),
    )

    prestation_expr = Coalesce(F("prestation_total"), items_total_expr, output_field=DEC)
    service_expr = Coalesce(F("service_fee"), Value(0, output_field=DEC))
    delivery_expr = Coalesce(F("delivery_fee"), Value(0, output_field=DEC))
    vat_expr = Coalesce(F("vat_fagni"), Value(0, output_field=DEC))


    express_expr = Coalesce(F("express_extra_fee"), Value(0, output_field=DEC))
    base_ht_expr = prestation_expr + service_expr + delivery_expr
    total_client_fallback = base_ht_expr + express_expr + vat_expr

    driver_income_expr = Coalesce(
        F("amount_driver_partner"),
        Coalesce(F("driver_logistic_cost"), Value(0, output_field=DEC)),
        output_field=DEC,
    )

    distance_expr = Coalesce(
        F("distance_km_total"),
        Value(0, output_field=DEC),
        output_field=DEC,
    )

    qs = qs.annotate(
        items_total=items_total_expr,
        total_client_display=Case(
            When(total_client_ttc__gt=0, then=Cast(F("total_client_ttc"), DEC)),
            default=total_client_fallback,
            output_field=DEC,
        ),
        driver_income_display=driver_income_expr,
        driver_distance_display=distance_expr,
    )

    # ---------- GARDE-FOU “À CALCULER” ----------
    def _has_coords(obj) -> bool:
        if not obj:
            return False
        lat = getattr(obj, "latitude", None)
        lng = getattr(obj, "longitude", None)
        return bool(lat and lng)

    def _is_calc_order(order) -> bool:
        customer = getattr(order, "customer", None)
        laundry = getattr(order, "laundry_partner", None)

        has_laundry = bool(getattr(order, "laundry_partner_id", None))
        customer_ok = _has_coords(customer)
        laundry_ok = _has_coords(laundry)

        try:
            dist_total_num = float(getattr(order, "distance_km_total", 0) or 0)
        except Exception:
            dist_total_num = 0.0

        return (not has_laundry) or (not customer_ok) or (not laundry_ok) or (dist_total_num <= 0)

    # ---------- CRÉATION DU FICHIER EXCEL ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "Courses livreur"

    # Styles de base
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")  # bleu
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    money_format = "#,##0"
    distance_format = "0.00"

    headers = [
        "Code commande",
        "Date création",
        "Client",
        "Téléphone",
        "Adresse",
        "Statut",
        "Distance (km)",
        "Total client TTC (FCFA)",
        "Revenu livreur (FCFA)",
        "Paiement",  # (optionnel mais utile)
    ]

    ws.append(headers)

    # Header style
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze + autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Lignes
    row_index = 1
    for order in qs:
        row_index += 1
        customer = getattr(order, "customer", None)

        calc = _is_calc_order(order)

        code = order.code or str(order.id)
        created = order.created_at.strftime("%d/%m/%Y %H:%M") if order.created_at else ""
        client_name = getattr(customer, "name", "") if customer else ""
        phone = getattr(customer, "phone", "") if customer else ""
        address = getattr(customer, "address", "") if customer else ""
        status_display = order.get_status_display()

        # ✅ alignement “À CALCULER”
        if calc:
            distance_km = 0.0
            total_client = 0.0
            driver_income = 0.0
            pay_label = "À CALCULER"
        else:
            distance_km = float(getattr(order, "driver_distance_display", 0) or 0)
            total_client = float(getattr(order, "total_client_display", 0) or 0)
            driver_income = float(getattr(order, "driver_income_display", 0) or 0)

            # label paiement similaire à l'API (simple)
            paid_num = 0.0
            try:
                paid_num = float(getattr(order, "amount_paid", 0) or 0)
            except Exception:
                paid_num = 0.0

            due_num = total_client - paid_num
            if due_num < 0:
                due_num = 0.0

            if total_client <= 0:
                pay_label = "À CALCULER"
            elif due_num <= 0:
                pay_label = "PAYÉ"
            elif 0 < due_num < total_client:
                pay_label = "PARTIEL"
            else:
                pay_label = "À ENCAISSER"

        ws.append([
            code,
            created,
            client_name,
            phone,
            address,
            status_display,
            distance_km,
            total_client,
            driver_income,
            pay_label,
        ])

        # Bordures + formats
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_index, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Distance (km)
            if col_idx == 7:
                cell.number_format = distance_format

            # Montants (FCFA)
            if col_idx in (8, 9):
                cell.number_format = money_format

    # Totaux (ligne de synthèse)
    total_row = row_index + 1
    ws.cell(row=total_row, column=1, value="TOTAUX").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = thin_border

    # Sommes sur distance / total client / revenu
    ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{row_index})").number_format = distance_format
    ws.cell(row=total_row, column=8, value=f"=SUM(H2:H{row_index})").number_format = money_format
    ws.cell(row=total_row, column=9, value=f"=SUM(I2:I{row_index})").number_format = money_format

    for col_idx in (7, 8, 9):
        c = ws.cell(row=total_row, column=col_idx)
        c.font = Font(bold=True)
        c.border = thin_border

    # Largeurs de colonnes (premium lisible)
    widths = {
        1: 16,  # code
        2: 18,  # date
        3: 22,  # client
        4: 16,  # phone
        5: 38,  # address
        6: 14,  # statut
        7: 14,  # distance
        8: 20,  # total client
        9: 20,  # revenu livreur
        10: 14, # paiement
    }
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ---------- RÉPONSE HTTP ----------
    now = datetime.now().strftime("%Y%m%d_%H%M")
    filename_parts = ["driver_app", "export", now]
    if driver_id:
        filename_parts.insert(1, f"driver_{driver_id}")
    filename = "_".join(filename_parts) + ".xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def compute_driver_week_stats(driver):
    """
    Calcule les stats SEMAINE (lundi -> aujourd'hui) pour un livreur donné.
    Utilisé par :
    - driver_me_app
    - driver_leaderboard

    On s'aligne sur une logique simple :
    - Périmètre : toutes les commandes de la semaine pour ce livreur
    - Revenus estimés = somme de driver_logistic_cost sur la semaine
    - Distance = somme de distance_km sur la semaine
    - Prime = 500 FCFA * nb de courses TERMINÉES dans la semaine
    - Taux de succès = done / (done + canceled) sur la semaine
    """
    today = timezone.localdate()
    start_week = today - timedelta(days=today.weekday())

    # Toutes les commandes de la semaine pour ce livreur
    week_qs = Order.objects.filter(
        delivery_partner=driver,
        created_at__date__gte=start_week,
        created_at__date__lte=today,
    )

    # Sous-ensembles
    done_qs = week_qs.filter(status="done")
    canceled_qs = week_qs.filter(status="canceled")

    # 1) Distance et revenus estimés sur TOUTES les commandes de la semaine
    weekly_distance_km = (
        week_qs.aggregate(s=Sum("distance_km"))["s"] or Decimal("0.0")
    )
    weekly_earnings = (
        week_qs.aggregate(s=Sum("driver_logistic_cost"))["s"] or Decimal("0.0")
    )

    # 2) Nombre de courses terminées (pour la prime)
    weekly_orders = done_qs.count()

    # 3) Objectifs
    weekly_target_orders = 40
    weekly_target_earnings = 80000  # FCFA

    def pct(part, total):
        if not total:
            return 0
        return int(Decimal(part) * Decimal(100) / Decimal(total))

    weekly_orders_progress = pct(weekly_orders, weekly_target_orders)
    weekly_earnings_progress = pct(weekly_earnings, weekly_target_earnings)

    # 4) Prime : 500 FCFA par course terminée dans la semaine
    weekly_bonus_amount = weekly_orders * 500

    # 5) Taux de succès = done / (done + canceled)
    total_finished_or_canceled = done_qs.count() + canceled_qs.count()
    if total_finished_or_canceled > 0:
        weekly_success_rate = int(
            Decimal(done_qs.count())
            * Decimal(100)
            / Decimal(total_finished_or_canceled)
        )
    else:
        weekly_success_rate = 0

    # 6) Heures de pointe (à affiner plus tard)
    weekly_peak_rides = 0

    return {
        "start_week": start_week,
        "end_week": today,
        "weekly_distance_km": weekly_distance_km,
        "weekly_earnings": weekly_earnings,
        "weekly_orders": weekly_orders,
        "weekly_target_orders": weekly_target_orders,
        "weekly_target_earnings": weekly_target_earnings,
        "weekly_orders_progress": weekly_orders_progress,
        "weekly_earnings_progress": weekly_earnings_progress,
        "weekly_bonus_amount": weekly_bonus_amount,
        "weekly_success_rate": weekly_success_rate,
        "weekly_peak_rides": weekly_peak_rides,
    }


@login_required
def driver_me_app(request):
    """
    Profil livreur (app mobile) :
    - identité du livreur connecté
    - stats globales simples
    - dernières courses
    """
    connected_driver = _get_connected_driver(request)
    if not connected_driver:
        return redirect("orders:driver_hub")

    driver = connected_driver
    today = timezone.localdate()

    qs = (
        Order.objects
        .filter(legs__driver=driver).distinct()
        .select_related("customer", "laundry_partner", "delivery_partner")
        .order_by("-created_at")
    )

    month_qs = qs.filter(
        created_at__year=today.year,
        created_at__month=today.month,
    )

    raw_stats = qs.aggregate(
        total_orders=Count("id", distinct=True),
        total_driver_income=Coalesce(Sum("driver_logistic_cost"), Decimal("0.0")),
    )

    legs_qs = DeliveryLeg.objects.filter(
        order__in=qs,
        distance_km__isnull=False,
    ).values_list("distance_km", flat=True)

    total_distance_km = Decimal("0.0")
    for d in legs_qs:
        if d is not None:
            total_distance_km += Decimal(str(d))
    total_distance_km = total_distance_km.quantize(Decimal("0.1"))

    total_orders = raw_stats["total_orders"] or 0
    pending_orders = qs.filter(status="pending").count()
    in_progress_orders = qs.filter(status="in_progress").count()
    done_orders = qs.filter(status="done").count()
    canceled_orders = qs.filter(status="canceled").count()

    avg_distance_km = Decimal("0.0")
    if total_orders:
        avg_distance_km = (total_distance_km / Decimal(str(total_orders))).quantize(Decimal("0.1"))

    stats = {
        "total_orders": total_orders,
        "month_orders": month_qs.count(),
        "pending_orders": pending_orders,
        "in_progress_orders": in_progress_orders,
        "done_orders": done_orders,
        "canceled_orders": canceled_orders,
        "total_distance_km": total_distance_km,
        "avg_distance_km": avg_distance_km,
        "total_driver_income": raw_stats["total_driver_income"] or Decimal("0.0"),
    }

    context = {
        "driver": driver,
        "connected_driver": driver,
        "stats": stats,
        "today": today,
        "orders": qs[:10],
    }
    return render(request, "orders/driver_me.html", context)


# ===============================================
#  HUB LIVREUR – POINT D'ENTRÉE / TABLEAU DE BORD
# ===============================================
@login_required
def driver_hub(request):
    """
    Hub livreur mobile, strictement côté livreur :
    - KPI simples
    - actions utiles
    - dernières courses
    """
    from django.db.models import Case, When, Value, IntegerField
    from django.core.paginator import Paginator

    connected_driver = _get_connected_driver(request)

    context = {
        "connected_driver": connected_driver,
        "today": timezone.localdate(),
        "period_display": "Aujourd'hui",
        "stats": {
            "total_orders": 0,
            "pending_orders": 0,
            "in_progress_orders": 0,
            "done_orders": 0,
            "canceled_orders": 0,
            "total_distance_km": Decimal("0.0"),
            "avg_distance_km": Decimal("0.0"),
            "total_driver_income": Decimal("0.0"),
        },
        "last_orders": [],
        "current_leg": None,
        "current_order": None,
        "current_mission_label": "Aucune mission active",
        "current_mission_hint": "Aucune action terrain en cours pour le moment.",
        "current_mission_cta": None,
    }

    if not connected_driver:
        return render(request, "orders/driver_hub.html", context)

    qs = (
        Order.objects
        .filter(legs__driver=connected_driver).distinct()
        .select_related("customer", "laundry_partner", "delivery_partner")
        .prefetch_related("items")
        .order_by("-created_at")
    )

    legs_qs = DeliveryLeg.objects.filter(
        order__in=qs,
        distance_km__isnull=False,
    ).values_list("distance_km", flat=True)

    total_distance_km = Decimal("0.0")
    for d in legs_qs:
        if d is not None:
            total_distance_km += Decimal(str(d))
    total_distance_km = total_distance_km.quantize(Decimal("0.1"))

    total_orders = qs.count()
    pending_orders = qs.filter(status="pending").count()
    in_progress_orders = qs.filter(status="in_progress").count()
    done_orders = qs.filter(status="done").count()
    canceled_orders = qs.filter(status="canceled").count()

    total_driver_income = (
        qs.aggregate(s=Coalesce(Sum("driver_logistic_cost"), Decimal("0.0"))).get("s")
        or Decimal("0.0")
    )

    avg_distance_km = Decimal("0.0")
    if total_orders:
        avg_distance_km = (total_distance_km / Decimal(str(total_orders))).quantize(Decimal("0.1"))

    stats = {
        "total_orders": total_orders,
        "pending_orders": pending_orders,
        "in_progress_orders": in_progress_orders,
        "done_orders": done_orders,
        "canceled_orders": canceled_orders,
        "total_distance_km": total_distance_km,
        "avg_distance_km": avg_distance_km,
        "total_driver_income": total_driver_income,
    }

    page_number = request.GET.get("page") or 1
    paginator = Paginator(qs, 5)
    last_orders_page = paginator.get_page(page_number)
    last_orders = list(last_orders_page.object_list)

    mission_ctx = _build_driver_mission_context(request, connected_driver)
    current_leg = mission_ctx.get("leg")
    current_order = mission_ctx.get("order")
    current_mission_label = mission_ctx.get("mission_state_label")
    current_mission_hint = mission_ctx.get("mission_hint")
    current_mission_cta = mission_ctx.get("mission_cta")

    context.update({
        "stats": stats,
        "last_orders": last_orders,
        "last_orders_page": last_orders_page,
        "current_leg": current_leg,
        "current_order": current_order,
        "current_mission_label": current_mission_label,
        "current_mission_hint": current_mission_hint,
        "current_mission_cta": current_mission_cta,
    })

    return render(request, "orders/driver_hub.html", context)

# ===============================================
#  APP LIVREUR – DATA JSON pour refresh KPIs
# ===============================================
@login_required
def driver_me_data(request):
    """
    Endpoint JSON pour l'auto-refresh des KPI de driver_me_app.

    Montants harmonisés FAGNI :
    - Revenu livreur = amount_driver_partner (fallback driver_logistic_cost)
    """
    user_email = (request.user.email or "").strip()
    if not user_email:
        return JsonResponse({"error": "no_email"}, status=403)

    try:
        driver = DeliveryPartner.objects.get(email__iexact=user_email)
    except DeliveryPartner.DoesNotExist:
        return JsonResponse({"error": "no_driver_profile"}, status=403)

    today = timezone.localdate()

    # Toutes les courses du jour pour ce livreur
    qs_today = Order.objects.filter(
        delivery_partner=driver,
        created_at__date=today,
    )

    # Comptages par statut
    total_orders_today = qs_today.count()
    pending_today = qs_today.filter(status="pending").count()
    in_progress_today = qs_today.filter(status="in_progress").count()
    done_today = qs_today.filter(status="done").count()
    canceled_today = qs_today.filter(status="canceled").count()

    # Agrégats distance + revenu livreur (harmonisé)
    DEC = DecimalField(max_digits=12, decimal_places=2)

    income_expr = Coalesce(
        F("amount_driver_partner"),
        Coalesce(F("driver_logistic_cost"), Value(0, output_field=DEC)),
        output_field=DEC,
    )

    aggregates = qs_today.aggregate(
        total_distance_km=Sum("distance_km"),
        total_driver_income=Sum(income_expr),
    )

    total_distance_km_today = float(aggregates["total_distance_km"] or 0)
    total_driver_income_today = float(aggregates["total_driver_income"] or 0)

    data = {
        "date": str(today),
        "total_orders_today": total_orders_today,
        "pending_today": pending_today,
        "in_progress_today": in_progress_today,
        "done_today": done_today,
        "canceled_today": canceled_today,
        "total_distance_km_today": round(total_distance_km_today, 2),
        "total_driver_income_today": round(total_driver_income_today, 2),
    }

    return JsonResponse(data)


# ===============================================
#  DÉTAIL COURSE – APP LIVREUR
# ===============================================
@login_required
def driver_kpi(request):
    """
    Vue KPI livreur :
    - Si user staff : choix du livreur + filtres
    - Si user livreur : KPI uniquement sur ses propres courses

    Montants harmonisés avec FAGNI :
    - Revenu livreur = amount_driver_partner (fallback driver_logistic_cost)
    """
    user = request.user
    connected_driver = None
    selected_driver_id = ""

    # Liste des livreurs (pour les filtres OPS)
    drivers = DeliveryPartner.objects.all().order_by("name")

    # --- Identification du livreur connecté si NON staff ---
    if not user.is_staff:
        user_email = (user.email or "").strip()
        if not user_email:
            return HttpResponseForbidden(
                "Votre compte utilisateur n'a pas d'email défini."
            )
        try:
            connected_driver = DeliveryPartner.objects.get(email__iexact=user_email)
        except DeliveryPartner.DoesNotExist:
            return HttpResponseForbidden(
                "Aucun profil livreur associé à cet email."
            )
        selected_driver_id = connected_driver.id
    else:
        # Mode OPS : on lit le driver_id dans les filtres
        driver_id_param = request.GET.get("driver_id") or ""
        if driver_id_param:
            try:
                connected_driver = DeliveryPartner.objects.get(pk=driver_id_param)
                selected_driver_id = connected_driver.id
            except DeliveryPartner.DoesNotExist:
                selected_driver_id = ""
                connected_driver = None

    # --- Filtre période ---
    period = (request.GET.get("period") or "30d").strip()
    dt_start = None

    if period == "today":
        dt_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "7d":
        dt_start = timezone.now() - timedelta(days=7)
    elif period == "30d":
        dt_start = timezone.now() - timedelta(days=30)
    else:
        period = "all"
        dt_start = None

    # --- Base queryset ---
    orders_qs = Order.objects.select_related(
        "customer", "delivery_partner", "laundry_partner"
    ).all()

    if selected_driver_id:
        orders_qs = orders_qs.filter(legs__driver_id=selected_driver_id).distinct()

    if dt_start:
        orders_qs = orders_qs.filter(created_at__gte=dt_start)

    # --- Agrégats globaux harmonisés FAGNI ---
    DEC = DecimalField(max_digits=12, decimal_places=2)

    driver_income_expr = Coalesce(
        F("amount_driver_partner"),
        Coalesce(F("driver_logistic_cost"), Value(0, output_field=DEC)),
        output_field=DEC,
    )

    aggregates = orders_qs.aggregate(
        total_orders=Count("id"),
        total_distance_km=Sum("distance_km"),
        total_driver_income=Sum(driver_income_expr),
    )

    total_orders = aggregates.get("total_orders") or 0
    total_distance_km = aggregates.get("total_distance_km") or 0
    total_driver_income = aggregates.get("total_driver_income") or 0

    avg_distance_km = total_distance_km / total_orders if total_orders else 0
    avg_driver_income = total_driver_income / total_orders if total_orders else 0

    # --- Répartition par statut ---
    status_counts = {"pending": 0, "in_progress": 0, "done": 0, "canceled": 0}
    for row in orders_qs.values("status").annotate(c=Count("id")):
        status = row["status"]
        c = row["c"] or 0
        if status in status_counts:
            status_counts[status] = c

    # --- Liste des dernières courses (limite 50) ---
    latest_orders = orders_qs.order_by("-created_at")[:50]

    context = {
        "drivers": drivers,
        "connected_driver": connected_driver,
        "selected_driver_id": selected_driver_id,
        "period": period,
        "total_orders": total_orders,
        "total_distance_km": total_distance_km,
        "total_driver_income": total_driver_income,
        "avg_distance_km": avg_distance_km,
        "avg_driver_income": avg_driver_income,
        "pending": status_counts["pending"],
        "in_progress": status_counts["in_progress"],
        "done": status_counts["done"],
        "canceled": status_counts["canceled"],
        "orders": latest_orders,
    }

    return render(request, "orders/driver_kpi.html", context)


@login_required
def driver_orders_csv(request):
    """
    Export CSV des courses livreur selon filtres :
    - Si user staff : peut filtrer par livreur + statut
    - Si user livreur : uniquement ses propres courses

    ⚠️ Revenu livreur harmonisé avec le modèle FAGNI :
    - on prend en priorité amount_driver_partner
    - sinon driver_logistic_cost
    """
    user = request.user
    connected_driver = None
    driver_id = None

    # --- Identification du livreur ---
    if not user.is_staff:
        # Mode LIVREUR : on retrouve le DeliveryPartner via l'email
        user_email = (user.email or "").strip()
        if not user_email:
            return HttpResponseForbidden(
                "Votre compte utilisateur n'a pas d'email défini."
            )
        try:
            connected_driver = DeliveryPartner.objects.get(email__iexact=user_email)
        except DeliveryPartner.DoesNotExist:
            return HttpResponseForbidden(
                "Aucun profil livreur associé à cet email."
            )
        driver_id = connected_driver.id
    else:
        # Mode OPS : possibilité de filtrer par livreur via ?driver_id=...
        driver_id_param = request.GET.get("driver_id") or ""
        if driver_id_param:
            try:
                driver_id = int(driver_id_param)
            except ValueError:
                driver_id = None

    # --- Filtre statut (logique similaire à driver_app) ---
    status_filter = (request.GET.get("status") or "active").strip()

    orders_qs = Order.objects.select_related(
        "customer",
        "delivery_partner",
        "laundry_partner",
    ).all()

    if driver_id:
        orders_qs = orders_qs.filter(legs__driver_id=driver_id).distinct()

    if status_filter == "active":
        orders_qs = orders_qs.filter(status__in=["pending", "in_progress"])
    elif status_filter == "done":
        orders_qs = orders_qs.filter(status="done")
    elif status_filter == "canceled":
        orders_qs = orders_qs.filter(status="canceled")
    # status_filter == "all" => pas de filtre supplémentaire

    orders_qs = orders_qs.order_by("-created_at")

    # --- Construction de la réponse CSV ---
    response = HttpResponse(content_type="text/csv")

    filename_parts = ["courses"]
    if driver_id:
        filename_parts.append(f"driver_{driver_id}")
    filename = "_".join(filename_parts) + ".csv"

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response, delimiter=";")

    # En-tête
    writer.writerow([
        "Code commande",
        "Date création",
        "Client",
        "Téléphone",
        "Adresse",
        "Statut",
        "Distance_km",
        "Revenu_livreur_FCFA",
    ])

    # Lignes
    for order in orders_qs:
        customer = getattr(order, "customer", None)

        # Revenu livreur harmonisé FAGNI :
        # 1) amount_driver_partner (montant dû au livreur)
        # 2) fallback : driver_logistic_cost
        tx_qs = WalletTransaction.objects.filter(
            order=order,
            leg__isnull=False,
            type__in=["payout", "adjustment"],
        )
        if driver_id:
            tx_qs = tx_qs.filter(wallet__delivery_partner_id=driver_id)

        driver_income = tx_qs.aggregate(net=_wallet_net_expr()).get("net") or 0

        writer.writerow([
            order.code or order.id,
            order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
            getattr(customer, "name", "") if customer else "",
            getattr(customer, "phone", "") if customer else "",
            getattr(customer, "address", "") if customer else "",
            order.get_status_display(),
            float(order.distance_km) if order.distance_km is not None else 0,
            float(driver_income),
        ])

    return response


@login_required
def driver_history_me(request):
    """
    Historique des courses du livreur connecté :
    - filtre par période (7j, 30j, mois en cours)
    - filtre par statut
    - bilan : nb courses, distance, revenus, cash à remettre

    Montants harmonisés avec FAGNI :
    - Montant client = total_client_ttc (compute_totals) fallback total_client_ttc / total
    - Revenu livreur = priorité aux legs (DeliveryLeg.driver_amount), sinon amount_driver_partner,
      sinon driver_logistic_cost, sinon estimation distance * 75
    - Cash à remettre (approche cash simplifiée) = montant client - revenu livreur (>= 0)
    """
    user_email = (request.user.email or "").strip()
    if not user_email:
        return HttpResponseForbidden("Votre compte utilisateur n'a pas d'email défini.")

    try:
        delivery_partner = DeliveryPartner.objects.get(email__iexact=user_email)
    except DeliveryPartner.DoesNotExist:
        return HttpResponseForbidden(
            "Aucun profil livreur associé à cet email : %s" % user_email
        )

    today = timezone.localdate()

    # Filtres
    period = (request.GET.get("period") or "7d").strip()
    status_filter = (request.GET.get("status") or "all").strip()

    # Période
    if period == "30d":
        start_date = today - timedelta(days=30)
    elif period == "month":
        start_date = today.replace(day=1)
    else:  # "7d"
        period = "7d"
        start_date = today - timedelta(days=7)

    qs = (
        Order.objects
        .filter(
            delivery_partner=delivery_partner,
            created_at__date__gte=start_date,
            created_at__date__lte=today,
        )
        .select_related("customer", "laundry_partner")
        .prefetch_related("items", "items__photos")
    )

    # Statut
    if status_filter == "active":
        qs = qs.filter(status__in=["pending", "in_progress"])
    elif status_filter in ["pending", "in_progress", "done", "canceled"]:
        qs = qs.filter(status=status_filter)
    else:
        status_filter = "all"

    orders = list(qs.order_by("-created_at"))

    # Agrégats / calculs par ligne
    total_orders = len(orders)
    done_count = sum(1 for o in orders if o.status == "done")
    total_distance_km = qs.aggregate(total=Sum("distance_km"))["total"] or 0

    total_driver_earnings = Decimal("0.00")
    client_total = Decimal("0.00")

    # Tarif de secours par km si aucun revenu renseigné
    fallback_price_per_km = Decimal("75")

    DEC = DecimalField(max_digits=12, decimal_places=2)

    for o in orders:
        # 1) Montants client (source de vérité = compute_totals)
        data = {}
        try:
            data = o.compute_totals(save=False) or {}
        except Exception:
            data = {}

        def _d(v):
            try:
                return Decimal(str(v)) if v is not None else Decimal("0.00")
            except Exception:
                return Decimal("0.00")

        tc = _d(data.get("total_client_ttc")) or _d(getattr(o, "total_client_ttc", None))

        # ✅ IMPORTANT : si tc == 0 => fallback calculé (PAS order.total)
        if tc <= 0:
            items_total = _d(data.get("prestation_total")) or _d(getattr(o, "prestation_total", None))

            service_fee = _d(data.get("service_fee_ht")) or _d(getattr(o, "service_fee", None))
            delivery_fee = _d(data.get("delivery_fee_client")) or _d(getattr(o, "delivery_fee", None))
            vat_fagni = _d(data.get("vat_fagni")) or _d(getattr(o, "vat_fagni", None))
            express_for_client = _d(data.get("express_for_client"))  # si tu veux l'inclure dans tc fallback

            tc = items_total + service_fee + delivery_fee + express_for_client + vat_fagni


        total_client = tc
        o.total_client = total_client
        client_total += total_client

        # 2) Revenu livreur : priorité aux legs (si la mission est découpée)
        legs_agg = DeliveryLeg.objects.filter(
            order=o,
            driver=delivery_partner,
        ).aggregate(
            total_amount=Sum("driver_amount", output_field=DEC),
            total_distance=Sum("distance_km", output_field=DEC),
        )

        legs_amount = legs_agg.get("total_amount") or Decimal("0.00")
        if legs_amount:
            earning = legs_amount
        else:
            earning = (
                data.get("amount_driver_partner")
                or getattr(o, "amount_driver_partner", None)
                or getattr(o, "driver_logistic_cost", None)
                or Decimal("0.00")
            )

        try:
            earning = Decimal(str(earning))
        except Exception:
            earning = Decimal("0.00")

        # 3) Dernier filet : distance * 75
        if (earning is None or earning == 0) and o.distance_km:
            try:
                earning = Decimal(str(o.distance_km)) * fallback_price_per_km
            except Exception:
                earning = Decimal("0.00")

        o.driver_earnings_display = earning
        total_driver_earnings += earning

    # Cash à remettre = montant client - revenus livreur (simplifié)
    cash_to_remit = client_total - total_driver_earnings
    if cash_to_remit < 0:
        cash_to_remit = Decimal("0.00")

    context = {
        "delivery_partner": delivery_partner,
        "orders": orders,

        "period": period,
        "status_filter": status_filter,

        "start_date": start_date,
        "end_date": today,

        "total_orders": total_orders,
        "done_count": done_count,
        "total_distance_km": total_distance_km,
        "total_driver_earnings": total_driver_earnings,
        "client_total": client_total,
        "cash_to_remit": cash_to_remit,
    }
    return render(request, "orders/driver_history.html", context)


@login_required
def driver_missions_history(request):
    """
    V2.2 — Historique des missions (lecture seule) + filtres + export (CSV/XLSX)

    Source de vérité :
      - Potentiel = somme des driver_amount sur les legs du driver (par commande)
      - Payé (net) = WalletTransaction (payout/adjustment) du wallet driver, par commande
      - Reste = max(potentiel - payé, 0)

    Filtres (GET) :
      - status: all|pending|in_progress|done
      - q: recherche (code, nom client, téléphone, adresse)
      - from/to: plage dates (created_at)
      - min_remaining=1 : affiche seulement reste > 0
      - sort: newest|oldest|remaining_desc|paid_desc
      - export: csv|xlsx (mêmes filtres, sans pagination)
    """
    from decimal import Decimal
    from urllib.parse import urlencode
    import csv
    from datetime import datetime

    from django.core.paginator import Paginator
    from django.db.models import Sum, Q
    from django.http import (
        HttpResponse,
        HttpResponseForbidden,
    )
    from django.utils.timezone import localtime, make_aware

    # openpyxl (déjà utilisé ailleurs chez toi)
    try:
        from openpyxl import Workbook
    except Exception:
        Workbook = None

    user = request.user
    connected_driver = _get_connected_driver(request)
    selected_driver_id = (request.GET.get("driver_id") or "").strip()

    # -----------------------------
    # 0) Lire filtres GET
    # -----------------------------
    status_filter = (request.GET.get("status") or "all").strip()
    q = (request.GET.get("q") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()
    sort = (request.GET.get("sort") or "newest").strip()
    min_remaining = (request.GET.get("min_remaining") or "").strip() in ("1", "true", "on", "yes")
    export = (request.GET.get("export") or "").strip().lower()

    # -----------------------------
    # 1) Déterminer le driver cible
    # -----------------------------
    driver = None
    if user.is_staff or user.is_superuser:
        if selected_driver_id:
            driver = DeliveryPartner.objects.filter(pk=selected_driver_id).first()
        if not driver and connected_driver:
            driver = connected_driver
    else:
        # non-staff => uniquement le livreur connecté
        driver = connected_driver
        if not driver:
            return HttpResponseForbidden("Accès refusé : aucun profil livreur connecté.")

    # -----------------------------
    # Helper parse date (input type=date => YYYY-MM-DD)
    # -----------------------------
    def _parse_date_ymd(s: str):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    d_from = _parse_date_ymd(date_from)
    d_to = _parse_date_ymd(date_to)

    # -----------------------------
    # 2) Orders concernés = orders où ce driver a au moins 1 leg
    # -----------------------------
    order_ids = (
        DeliveryLeg.objects
        .filter(driver=driver)
        .values_list("order_id", flat=True)
        .distinct()
    )

    qs = (
        Order.objects
        .filter(id__in=order_ids)
        .select_related("customer", "laundry_partner", "delivery_partner")
        .prefetch_related("legs")
        .order_by("-created_at")
    )

    # Filtre dates (created_at)
    # On filtre au niveau datetime pour rester robuste
    if d_from:
        # début de journée
        dt = datetime(d_from.year, d_from.month, d_from.day, 0, 0, 0)
        try:
            dt = make_aware(dt)
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=10781")
        qs = qs.filter(created_at__gte=dt)

    if d_to:
        # fin de journée
        dt = datetime(d_to.year, d_to.month, d_to.day, 23, 59, 59)
        try:
            dt = make_aware(dt)
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=10790")
        qs = qs.filter(created_at__lte=dt)

    # Recherche DB (réduit la charge)
    if q:
        qs = qs.filter(
            Q(code__icontains=q)
            | Q(customer__name__icontains=q)
            | Q(customer__phone__icontains=q)
            | Q(customer__address__icontains=q)
        )

    # -----------------------------
    # 3) Potentiel par commande (somme driver_amount sur legs du driver)
    # -----------------------------
    potential_rows = (
        DeliveryLeg.objects
        .filter(order_id__in=order_ids, driver=driver)
        .values("order_id")
        .annotate(potential=Sum("driver_amount"))
    )
    potential_by_order = {r["order_id"]: (r["potential"] or 0) for r in potential_rows}

    # -----------------------------
    # 4) Payé net par commande (wallet tx)
    #    payout + adjustment, leg NULL ou pas (legacy inclus)
    # -----------------------------
    tx_qs = WalletTransaction.objects.filter(
        order_id__in=order_ids,
        type__in=["payout", "adjustment"],
        wallet__delivery_partner=driver,
    )

    income_rows = (
        tx_qs.values("order_id")
        .annotate(net=_wallet_net_expr())
    )
    paid_by_order = {r["order_id"]: (r["net"] or Decimal("0")) for r in income_rows}

    # -----------------------------
    # 5) Construire items (UI-ready)
    # -----------------------------
    def fmt_dt(dt):
        if not dt:
            return ""
        try:
            return localtime(dt).strftime("%d/%m/%Y %H:%M")
        except Exception:
            return str(dt)

    items = []
    for o in qs:
        legs_for_driver = [
            l for l in getattr(o, "legs", []).all()
            if str(getattr(l, "driver_id", "")) == str(driver.id)
        ]

        statuses = [getattr(l, "status", None) for l in legs_for_driver]

        # statut mission (UI) basé sur legs
        if any(s == "in_progress" for s in statuses):
            mission_status = "in_progress"
        elif any(s in ("pending", "assigned") for s in statuses):
            mission_status = "pending"
        elif statuses and all(s in ("done", "canceled") for s in statuses):
            mission_status = "done"
        else:
            mission_status = (getattr(o, "status", None) or "pending")

        potential = Decimal(str(potential_by_order.get(o.id, 0) or 0))
        paid = Decimal(str(paid_by_order.get(o.id, Decimal("0")) or 0))
        remaining = potential - paid
        if remaining < 0:
            remaining = Decimal("0")

        # total km (legs)
        total_km = Decimal("0")
        for l in legs_for_driver:
            try:
                total_km += Decimal(str(getattr(l, "distance_km", 0) or 0))
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=10871")

        items.append({
            "order": o,
            "code": getattr(o, "code", "") or f"#{o.id}",
            "created_at": fmt_dt(getattr(o, "created_at", None)),
            "created_at_raw": getattr(o, "created_at", None),
            "mission_status": mission_status,
            "legs": [{
                "id": l.id,
                "leg_type": getattr(l, "leg_type", ""),
                "status": getattr(l, "status", ""),
                "distance_km": float(getattr(l, "distance_km", 0) or 0),
                "driver_amount": float(getattr(l, "driver_amount", 0) or 0),
                "started_at": fmt_dt(getattr(l, "started_at", None)),
                "finished_at": fmt_dt(getattr(l, "finished_at", None)),
            } for l in legs_for_driver],
            "potential": float(potential),
            "paid": float(paid),
            "remaining": float(remaining),
            "total_km": float(total_km),
            "customer_name": getattr(getattr(o, "customer", None), "name", "") or "",
            "customer_phone": getattr(getattr(o, "customer", None), "phone", "") or "",
            "customer_address": getattr(getattr(o, "customer", None), "address", "") or "",
            "laundry_name": getattr(getattr(o, "laundry_partner", None), "name", "") or "",
            "detail_url": reverse("orders:driver_order_detail", args=[o.id]) + "?back=" + quote(reverse('orders:driver_hub')),
        })

    # 6) Filtres Python (mission_status + remaining)
    if status_filter and status_filter != "all":
        if status_filter == "canceled":
            items = [it for it in items if it.get("mission_status") in ("canceled", "cancelled")]
        else:
            items = [it for it in items if it.get("mission_status") == status_filter]

    if min_remaining:
        items = [it for it in items if (it.get("remaining") or 0) > 0]

    # -----------------------------
    # 7) Tri
    # -----------------------------
    if sort == "oldest":
        items.sort(key=lambda x: (x.get("created_at_raw") is None, x.get("created_at_raw")))
    elif sort == "remaining_desc":
        items.sort(key=lambda x: (x.get("remaining") or 0), reverse=True)
    elif sort == "paid_desc":
        items.sort(key=lambda x: (x.get("paid") or 0), reverse=True)
    else:
        # newest
        items.sort(key=lambda x: (x.get("created_at_raw") is None, x.get("created_at_raw")), reverse=True)


    # -----------------------------
    # 8) Export CSV/XLSX (sans pagination)
    # -----------------------------
    def _export_filename(ext: str) -> str:
        base = f"missions_driver_{driver.id}"
        return f"{base}.{ext}"

    def _status_label(s: str) -> str:
        m = {
            "pending": "En attente",
            "assigned": "En attente",
            "in_progress": "En cours",
            "done": "Terminée",
            "canceled": "Annulée",
            "cancelled": "Annulée",
        }
        return m.get((s or "").strip(), (s or "").strip())

    if export == "csv":
        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{_export_filename("csv")}"'
        w = csv.writer(resp)
        w.writerow([
            "Date", "Code", "Statut",
            "Client", "Téléphone", "Adresse",
            "Blanchisserie",
            "Potentiel", "Payé", "Reste",
            "Km (total)", "Nb legs",
            "Legs (détails)",
            "Lien détail",
        ])
        for it in items:
            legs_txt = " | ".join(
                f"{(l.get('leg_type') or '')}:{(l.get('status') or '')}:{int(float(l.get('driver_amount') or 0))}FCFA:{float(l.get('distance_km') or 0):.2f}km"
                for l in (it.get("legs") or [])
            )
            w.writerow([
                it.get("created_at") or "",
                it.get("code") or "",
                _status_label(it.get("mission_status") or ""),
                it.get("customer_name") or "",
                it.get("customer_phone") or "",
                it.get("customer_address") or "",
                it.get("laundry_name") or "",
                int(float(it.get("potential") or 0)),
                int(float(it.get("paid") or 0)),
                int(float(it.get("remaining") or 0)),
                float(it.get("total_km") or 0),
                len(it.get("legs") or []),
                legs_txt,
                it.get("detail_url") or "",
            ])
        return resp

    if export == "xlsx":
        if Workbook is None:
            return HttpResponse("Export XLSX indisponible (openpyxl manquant).", status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = "Missions"

        ws.append([
            "Date", "Code", "Statut",
            "Client", "Téléphone", "Adresse",
            "Blanchisserie",
            "Potentiel", "Payé", "Reste",
            "Km (total)", "Nb legs",
            "Legs (détails)",
            "Lien détail",
        ])

        for it in items:
            legs_txt = " | ".join(
                f"{(l.get('leg_type') or '')}:{(l.get('status') or '')}:{int(float(l.get('driver_amount') or 0))}FCFA:{float(l.get('distance_km') or 0):.2f}km"
                for l in (it.get("legs") or [])
            )
            ws.append([
                it.get("created_at") or "",
                it.get("code") or "",
                _status_label(it.get("mission_status") or ""),
                it.get("customer_name") or "",
                it.get("customer_phone") or "",
                it.get("customer_address") or "",
                it.get("laundry_name") or "",
                int(float(it.get("potential") or 0)),
                int(float(it.get("paid") or 0)),
                int(float(it.get("remaining") or 0)),
                float(it.get("total_km") or 0),
                len(it.get("legs") or []),
                legs_txt,
                it.get("detail_url") or "",
            ])

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{_export_filename("xlsx")}"'
        wb.save(resp)
        return resp

    # -----------------------------
    # 9) KPI (sur items filtrés)
    # -----------------------------
    kpi_total = len(items)
    kpi_potential = sum(float(it.get("potential") or 0) for it in items)
    kpi_paid = sum(float(it.get("paid") or 0) for it in items)
    kpi_remaining = sum(float(it.get("remaining") or 0) for it in items)
    kpi_km = sum(float(it.get("total_km") or 0) for it in items)

    # -----------------------------
    # 10) Pagination
    # -----------------------------
    paginator = Paginator(items, 12)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    # conserver les params (sans page/export) pour pagination & links
    preserved = {}
    for k, v in request.GET.items():
        if k in ("page", "export"):
            continue
        if v is None:
            continue
        preserved[k] = v
    qs_params = urlencode(preserved)

    context = {
        "connected_driver": connected_driver,
        "driver": driver,
        "selected_driver_id": selected_driver_id,
        "page_obj": page_obj,
        "qs_params": qs_params,
        # pour pré-remplir le form
        "status_filter": status_filter,
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
        "sort": sort,
        "min_remaining": min_remaining,

        # KPI
        "kpi_total": kpi_total,
        "kpi_potential": kpi_potential,
        "kpi_paid": kpi_paid,
        "kpi_remaining": kpi_remaining,
        "kpi_km": kpi_km,
    }
    return render(request, "orders/driver_missions_history.html", context)


@login_required
def driver_missions_export_csv(request):
    params = request.GET.copy()
    params["export"] = "csv"
    return redirect(f"/orders/driver/missions/?{params.urlencode()}")

@login_required
def driver_missions_export_xlsx(request):
    params = request.GET.copy()
    params["export"] = "xlsx"
    return redirect(f"/orders/driver/missions/?{params.urlencode()}")


@login_required
def driver_leaderboard(request):
    """
    Classement des livreurs.
    Utilise compute_driver_week_stats pour :
    - revenus semaine
    - distance semaine
    - prime semaine
    - taux de succès semaine
    afin d'être 100% cohérent avec /orders/driver/me/.
    """
    today = timezone.localdate()
    start_week = today - timedelta(days=today.weekday())

    drivers = DeliveryPartner.objects.filter(is_active=True).order_by("name")

    leaderboard = []

    for d in drivers:
        week_stats = compute_driver_week_stats(d)

        leaderboard.append({
            "id": d.id,
            "name": d.name,
            "city": d.city,
            # total commandes semaine terminées
            "orders_count": week_stats["weekly_orders"],
            # pour affichage des courses terminées dans le tableau
            "done_count": week_stats["weekly_orders"],
            # taux de succès semaine
            "success_rate": week_stats["weekly_success_rate"],
            # distance semaine
            "total_distance_km": week_stats["weekly_distance_km"],
            # revenus estimés semaine
            "driver_earnings": week_stats["weekly_earnings"],
            # prime semaine (même calcul que /driver/me/)
            "bonus_amount": week_stats["weekly_bonus_amount"],
        })

    context = {
        "leaderboard": leaderboard,
        "start_week": start_week,
        "end_week": today,
    }

    return render(request, "orders/driver_leaderboard.html", context)


@transaction.atomic
def ensure_delivery_legs_for_order(order):
    """
    Option B — 2 legs dès le départ (niveau commande) :

    - Si des legs existent déjà -> on ne recrée rien (idempotent)
    - Sinon, on crée :
        * pickup (assigned)
        * return (pending si wash pas prêt, sinon assigned)
    - Distance et montant répartis entre les jambes.
    - Garde-fou : pas de legs si commande non chiffrée (items + total_client_ttc > 0)
    """
    from decimal import Decimal, InvalidOperation

    existing_legs = DeliveryLeg.objects.filter(order=order)

    if existing_legs.exists():
        wash_ready = bool(getattr(order, "wash_complete_time", None))

        qs_return = existing_legs.filter(leg_type="return").exclude(status__in=["done", "canceled"])

        if wash_ready:
            # pending -> assigned
            # 🔒 auto-upgrade désactivé: le livreur doit "accept"
            pass
        else:
            # assigned/in_progress -> pending (linge pas prêt)
            qs_return.filter(status__in=["assigned", "in_progress"]).update(status="pending")

        return existing_legs

    driver = getattr(order, "delivery_partner", None)
    if not driver:
        return existing_legs

    # ----------------------------
    # Garde-fou : commande chiffrée
    # ----------------------------
    has_items = False
    try:
        has_items = order.items.exists()
    except Exception:
        has_items = False

    total_client = getattr(order, "total_client_ttc", None)
    if total_client is None:
        total_client = getattr(order, "total", None)

    try:
        from orders.models import DeliveryLeg, sync_delivery_legs_for_order
        if not DeliveryLeg.objects.filter(order=order).exclude(status="canceled").exists():
            sync_delivery_legs_for_order(order)
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11187")

    if (not has_items) or total_client_dec <= 0:
        return existing_legs

    # ----------------------------
    # Distances
    # ----------------------------
    total_distance = (
        getattr(order, "distance_km_total", None)
        or getattr(order, "distance_km", None)
        or 0
    )
    pickup_distance = getattr(order, "distance_km_pickup", None)
    delivery_distance = getattr(order, "distance_km_delivery", None)

    if pickup_distance is not None and delivery_distance is not None:
        pass
    elif pickup_distance is not None and delivery_distance is None:
        if total_distance and pickup_distance <= total_distance:
            delivery_distance = max(total_distance - pickup_distance, 0)
        else:
            delivery_distance = pickup_distance
    elif delivery_distance is not None and pickup_distance is None:
        if total_distance and delivery_distance <= total_distance:
            pickup_distance = max(total_distance - delivery_distance, 0)
        else:
            pickup_distance = delivery_distance
    else:
        if total_distance:
            pickup_distance = total_distance / 2
            delivery_distance = total_distance / 2
        else:
            pickup_distance = 0
            delivery_distance = 0

    try:
        pickup_distance_dec = Decimal(str(pickup_distance or 0))
        delivery_distance_dec = Decimal(str(delivery_distance or 0))
    except (InvalidOperation, TypeError, ValueError):
        pickup_distance_dec = Decimal("0")
        delivery_distance_dec = Decimal("0")

    # ----------------------------
    # Montant livreur total
    # ----------------------------
    total_amount = getattr(order, "amount_driver_partner", None)
    if not total_amount:
        total_amount = (
            getattr(order, "driver_logistic_cost", None)
            or getattr(order, "delivery_fee", None)
            or getattr(order, "delivery_cost_driver", None)
            or 0
        )

    if (not total_amount or float(total_amount) <= 0) and hasattr(order, "compute_totals"):
        try:
            totals = order.compute_totals(save=False) or {}
            total_amount = (
                totals.get("amount_driver_partner")
                or totals.get("driver_amount")
                or totals.get("driver_logistic_cost")
                or totals.get("delivery_cost_driver")   # ✅ TON CAS
                or 0
            )
        except Exception:
            total_amount = 0

    try:
        total_amount_dec = Decimal(str(total_amount or 0))
    except (InvalidOperation, TypeError, ValueError):
        total_amount_dec = Decimal("0")

    # Répartition montant
    dist_sum = pickup_distance_dec + delivery_distance_dec
    if dist_sum > 0 and total_amount_dec > 0:
        ratio_pickup = pickup_distance_dec / dist_sum
        pickup_amount_dec = (total_amount_dec * ratio_pickup)
        pickup_amount = int(pickup_amount_dec.quantize(Decimal("1")))
        delivery_amount = int(total_amount_dec) - pickup_amount
    else:
        pickup_amount = int(total_amount_dec / 2) if total_amount_dec > 0 else 0
        delivery_amount = int(total_amount_dec) - pickup_amount if total_amount_dec > 0 else 0

    # ✅ Source unique : création/normalisation legs dans models (anti-doublons)
    try:
        from orders.models import sync_delivery_legs_for_order
        sync_delivery_legs_for_order(order)
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11276")

    return (
        DeliveryLeg.objects
        .filter(order=order, driver=driver)
        .exclude(status="canceled")
        .order_by("id")
    )


def update_leg_status(leg, action, user=None):
    """
    Met à jour proprement le statut d'une DeliveryLeg selon l'action demandée.

    Actions autorisées :
    - accept  : pending → assigned
    - start   : assigned / pending → in_progress
    - finish  : assigned / in_progress → done
    - cancel  : * → canceled (sauf done)

    Retourne (changed: bool, message: str)
    """
    from django.utils import timezone

    old_status = leg.status
    # ✅ Auto-fix: si le leg n'a pas de driver, on prend le driver assigné à la commande
    try:
        if getattr(leg, "driver_id", None) is None:
            assigned = getattr(getattr(leg, "order", None), "delivery_partner", None)
            if assigned:
                leg.driver = assigned
                leg.save(update_fields=["driver"])
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11309")

    action = (action or "").lower().strip()

    # 🔒 Garde-fou métier : le "return/delivery" ne démarre pas tant que le linge n'est pas prêt
    if getattr(leg, "leg_type", None) == "return":
        if getattr(leg, "status", None) not in {"done", "canceled"}:
            if not getattr(getattr(leg, "order", None), "wash_complete_time", None):
                if action in {"start", "finish"}:
                    return False, "Retour impossible : le linge n\'est pas encore prêt."
    # 🔒 Verrouillage transitions LIVREUR :
    #    pickup doit être DONE avant d'autoriser le RETURN (accept/start/finish)
    try:
        from orders.models import DeliveryLeg

        order = getattr(leg, "order", None)
        driver = getattr(leg, "driver", None)

        pickup_leg = (
            DeliveryLeg.objects
            .filter(order=order, driver=driver, leg_type="pickup")
            .exclude(status="canceled")
            .order_by("-id")
            .first()
        )

        return_leg = (
            DeliveryLeg.objects
            .filter(order=order, driver=driver, leg_type="return")
            .exclude(status="canceled")
            .order_by("-id")
            .first()
        )

        pickup_done = bool(pickup_leg and pickup_leg.status == "done")
        return_started = bool(return_leg and return_leg.status in {"assigned", "in_progress", "done"})

        # 1) Interdire toute avancée du RETURN si le PICKUP n'est pas terminé
        if getattr(leg, "leg_type", None) == "return":
            if action in {"accept", "start", "finish"} and not pickup_done:
                return False, "Impossible : le pickup doit être terminé avant le retour."

        # 2) Interdire d'annuler le pickup si le return a déjà démarré/terminé
        if getattr(leg, "leg_type", None) == "pickup":
            if action == "cancel" and return_started:
                return False, "Annulation pickup impossible : le retour a déjà démarré."
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11356")

    # 🔒 Normalisation anti-doublons avant transition
    try:
        normalize_order_legs(getattr(leg, "order", None), driver=getattr(leg, "driver", None))
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11362")

    if action == "accept":

        # ✅ Cas normal : pending → assigned
        if leg.status == "pending":
            leg.status = "assigned"

            # 1) Fixer delivery_partner sur la commande si absent
            try:
                order = getattr(leg, "order", None)
                driver = getattr(leg, "driver", None)
                if order and driver and not getattr(order, "delivery_partner_id", None):
                    order.delivery_partner = driver
                    order.save(update_fields=["delivery_partner"])
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11378")

            # 2) Aligner driver sur pickup+return
            try:
                from orders.models import DeliveryLeg
                order = getattr(leg, "order", None)
                driver_id = getattr(getattr(leg, "driver", None), "id", None)
                if order and driver_id:
                    DeliveryLeg.objects.filter(
                        order=order,
                        leg_type__in=["pickup", "return"],
                    ).update(driver_id=driver_id)
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11391")

            # 3) Réactiver return canceled non payé
            try:
                from orders.models import DeliveryLeg
                from wallets.models import WalletTransaction

                order = getattr(leg, "order", None)
                driver_id = getattr(getattr(leg, "driver", None), "id", None)

                if order and driver_id:
                    r = DeliveryLeg.objects.filter(
                        order=order,
                        leg_type="return"
                    ).order_by("-id").first()

                    if r and (r.status or "").lower() == "canceled":
                        has_payout = WalletTransaction.objects.filter(
                            order_id=getattr(order, "id", None),
                            leg_id=getattr(r, "id", None),
                            wallet__owner_type="driver",
                            type="payout",
                            direction="in",
                        ).exists()

                        if not has_payout:
                            DeliveryLeg.objects.filter(
                                pk=r.pk,
                                status="canceled"
                            ).update(status="pending")
            except Exception:
                import logging
                logging.getLogger("fagni.views.legs").exception("Echec silencieux: reactivation leg return si non paye | order_id=%s", getattr(order, "id", None) if "order" in dir() else None)

        # ✅ Déjà accepté
        elif leg.status == "assigned":
            return True, "Mission déjà acceptée."

        # ✅ Déjà démarré
        elif leg.status == "in_progress":
            return True, "Mission déjà en cours."

        # ✅ Déjà terminé
        elif leg.status == "done":
            return True, "Mission déjà terminée."

    elif action == "start":
        # 🔒 Le livreur doit d'abord ACCEPT (pending → assigned)
        if leg.status == "assigned":
            leg.status = "in_progress"
            if not getattr(leg, "started_at", None):
                leg.started_at = timezone.now()
        elif leg.status == "in_progress":
            return True, "Mission déjà démarrée."
        elif leg.status == "pending":
            return False, "Vous devez d'abord accepter la mission."

    elif action == "finish":
        if leg.status == "done":
            return True, "Mission déjà terminée."
        if leg.status in {"assigned", "in_progress"}:
            # 🔒 Verrou logique : on ne finit pas un pickup si un return est déjà actif
            if getattr(leg, "leg_type", None) == "pickup":
                try:
                    r = (
                        DeliveryLeg.objects
                        .filter(order=getattr(leg, "order", None), leg_type="return")
                        .exclude(status__in=("canceled", "pending"))
                        .exists()
                    )
                except Exception:
                    r = False
                if r:
                    return False, "Impossible : le retour est déjà actif, terminez la séquence proprement."

            leg.status = "done"
            if not getattr(leg, "finished_at", None):
                leg.finished_at = timezone.now()

    elif action == "cancel":
        # ✅ Idempotent: si déjà annulé, on valide sans erreur
        if leg.status == "canceled":
            return True, "Mission déjà annulée."

        # 🔒 Verrouillage backend: si payout existe, on interdit toute annulation (même staff)
        try:
            from wallets.models import WalletTransaction
            has_payout = WalletTransaction.objects.filter(
                order_id=getattr(leg, "order_id", None),
                leg_id=getattr(leg, "id", None),
                wallet__owner_type="driver",
                type="payout",
                direction="in",
            ).exists()
        except Exception:
            has_payout = False

        if has_payout:
            return False, "Annulation impossible : payout livreur déjà effectué."

        # ✅ Autoriser annulation si done NON payé, mais uniquement staff
        if leg.status == "done":
            if not user or not getattr(user, "is_staff", False):
                return False, "Annulation impossible : leg déjà terminé (réservé au staff)."

        if leg.status != "canceled":
            leg.status = "canceled"
            # ✅ Cohérence: une annulation ne doit pas garder des timestamps de progression
            if hasattr(leg, "started_at"):
                leg.started_at = None
            if hasattr(leg, "finished_at"):
                leg.finished_at = None

            # ✅ Neutraliser montants côté view (sécurité + cohérence immédiate UI)
            try:
                from decimal import Decimal
                leg.client_fee_share = Decimal("0")
                leg.driver_amount = Decimal("0")
                leg.fagni_margin = Decimal("0")
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11511")
    else:
        return False, "Action inconnue"

    if leg.status == old_status:
        return False, f"Aucune transition valide depuis le statut '{old_status}'"

    update_fields = ["status"]
    # ✅ Si annulation: persister aussi la neutralisation des montants
    if leg.status == "canceled":
        if hasattr(leg, "client_fee_share"):
            update_fields.append("client_fee_share")
        if hasattr(leg, "driver_amount"):
            update_fields.append("driver_amount")
        if hasattr(leg, "fagni_margin"):
            update_fields.append("fagni_margin")
    if hasattr(leg, "started_at"):
        update_fields.append("started_at")
    if hasattr(leg, "finished_at"):
        update_fields.append("finished_at")

    # ✅ Sauvegarde leg
    try:
        # évite doublons dans update_fields
        update_fields = list(dict.fromkeys(update_fields))
        leg.save(update_fields=update_fields)
    except Exception:
        # fallback safe
        try:
            leg.save()
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11542")


    # ✅ Auto-sync après transition
    try:
        order = getattr(leg, "order", None)
        if order:
            from orders.models import sync_order_status_from_legs
            sync_order_status_from_legs(order, save=True)

            # 1) Si return terminé, s'assurer que wash_complete_time existe
            if getattr(leg, "leg_type", None) == "return" and leg.status == "done":
                if not getattr(order, "wash_complete_time", None):
                    order.wash_complete_time = leg.finished_at or timezone.now()
                    order.save(update_fields=["wash_complete_time"])

    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11559")

    # 🔒 Normalisation après transition (ex: cancel d'un leg ancien)
    try:
        normalize_order_legs(getattr(leg, "order", None), driver=getattr(leg, "driver", None))
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11565")

    return True, f"Statut mis à jour : {old_status} → {leg.status}"


def _redirect_back(request, fallback_name, **kwargs):
    """
    Utilise ?back=… si présent et sûr, sinon redirige vers un named URL (fallback).
    """
    back = (request.GET.get("back") or request.POST.get("back") or "").strip()
    if back and url_has_allowed_host_and_scheme(back, allowed_hosts={request.get_host()}):
        return redirect(back)
    return redirect(fallback_name, **kwargs)


@require_POST
@login_required
def driver_leg_action(request, leg_id, action):
    leg = get_object_or_404(
        DeliveryLeg.objects.select_related("order", "driver"),
        pk=leg_id
    )
    order = leg.order
    driver = leg.driver

    # 🔒 Parcours livreur verrouillé : toujours revenir vers driver_app
    requested_driver_id = (
        (request.GET.get("driver_id") or "").strip()
        or (request.POST.get("driver_id") or "").strip()
        or str(getattr(driver, "id", "") or "")
    )

    flow_back = reverse("orders:driver_app")
    if requested_driver_id:
        flow_back = f"{flow_back}?driver_id={requested_driver_id}"

    def _driver_flow_redirect():
        return redirect(flow_back)

    # 🔒 Permission : soit staff, soit le livreur assigné connecté (via DeliveryPartner.user)
    if not request.user.is_staff:
        dp_user = getattr(driver, "user", None)
        if dp_user is not None and dp_user != request.user:
            messages.error(request, "Accès refusé.")
            return _driver_flow_redirect()

    # 🔐 Actions autorisées (garde-fou param)
    action = (action or "").lower().strip()
    if action not in {"accept", "start", "finish", "cancel"}:
        messages.error(request, "Action inconnue.")
        return _driver_flow_redirect()

    # 🔒 Verrouillage transitions : pickup → return → done
    # Règles:
    # - Retour (return) interdit tant que pickup n'est pas DONE
    # - Retour (return) interdit tant que linge pas "Prêt" (wash_complete_time)
    # - Une mission canceled ne peut pas être relancée via l'app livreur
    try:
        # Si commande déjà terminée : aucune action
        if getattr(order, "status", None) == "done":
            messages.info(request, "Commande déjà terminée.")
            return _driver_flow_redirect()

        # Si la jambe est annulée, on bloque accept/start/finish (freeze côté model)
        if (leg.status or "").strip() in ("canceled", "cancelled") and action in {"accept", "start", "finish"}:
            messages.error(request, "Cette mission est annulée.")
            return _driver_flow_redirect()

        # Verrou sur RETURN
        if (leg.leg_type or "").strip() == "return" and action in {"accept", "start", "finish"}:
            if not getattr(order, "wash_complete_time", None):
                messages.error(request, "Retour impossible : linge pas encore marqué prêt par la blanchisserie.")
                return _driver_flow_redirect()

            pickup_done = DeliveryLeg.objects.filter(
                order=order,
                leg_type="pickup",
                status="done",
            ).exists()
            if not pickup_done:
                messages.error(request, "Retour impossible : la collecte (pickup) n'est pas terminée.")
                return _driver_flow_redirect()

    except Exception:
        logger.exception("driver_leg_action: transition guard failed (leg_id=%s action=%s)", leg_id, action)

    # ⚙️ Transition (payout géré par orders/signals.py)
    try:
        # 📸 Verrou métier: FINISH interdit sans preuve photo
        # Règle:
        # - preuve liée au leg (idéal) OU fallback preuve au niveau commande (leg=None)
        # - on refuse 'issue' (litige) comme preuve de finish
        if action == "finish" and (leg.status or "").strip() != "done":
            try:
                from orders.models import OrderEvidencePhoto
                qs = OrderEvidencePhoto.objects.filter(order=order, actor_type="driver").exclude(kind="issue")
                has_leg_proof = qs.filter(leg=leg).exists()
                has_order_proof = qs.filter(leg__isnull=True).exists()

                if not (has_leg_proof or has_order_proof):
                    from django.contrib import messages
                    messages.error(request, "Impossible de terminer : ajoute au moins une photo preuve (non-litige) avant de valider.")

                    from urllib.parse import quote

                    back = quote(flow_back, safe="")
                    url = reverse("orders:driver_weighing", kwargs={"order_id": order.id})

                    full_url = f"{url}?leg_id={leg.id}&driver_id={requested_driver_id}&back={back}"

                    # Si submit via fetch/AJAX, renvoyer JSON pour que le front navigue
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        try:
                            from django.http import JsonResponse
                            return JsonResponse({"ok": False, "redirect_url": full_url, "message": "Preuve photo requise avant de terminer."}, status=409)
                        except Exception:
                            import logging
                            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11681")

                    full_url = f"{url}?leg_id={leg.id}&back={back}"
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        try:
                            from django.http import JsonResponse
                            return JsonResponse({"ok": False, "redirect_url": full_url, "message": "Preuve photo requise avant de terminer."}, status=409)
                        except Exception:
                            import logging
                            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11689")
                    return redirect(full_url)
            except Exception:
                # permissif si modèle/migration pas dispo (évite de casser)
                pass

        changed, msg = update_leg_status(leg=leg, action=action, user=request.user)

        # Après chaque action livreur, le statut commande doit être recalculé
        # depuis les DeliveryLeg, jamais forcé manuellement.
        try:
            from orders.models import sync_order_status_from_legs
            sync_order_status_from_legs(order, save=True)
        except Exception:
            logger.exception("driver_leg_action: sync_order_status_from_legs failed (order_id=%s)", getattr(order, "id", None))
    except Exception:
        logger.exception("driver_leg_action: update_leg_status failed (leg_id=%s action=%s)", leg_id, action)
        messages.error(request, "Erreur interne : action impossible pour le moment.")
        return _driver_flow_redirect()

    from django.contrib import messages
    if changed:
        messages.success(request, msg)
    else:
        messages.info(request, msg)

    # ↩️ Redirection : priorité à ?back=…, sinon détail de la course
    return _driver_flow_redirect()


@login_required
def driver_order_live_status(request, order_id):
    """
    Lot 4.4 — Endpoint JSON pour rafraîchir la vue livreur sans reload.

    Règles:
    - staff/superuser: peut demander ?driver_id=... sinon fallback sur order.delivery_partner
    - non-staff: doit correspondre au driver assigné à la commande
      et (si DeliveryPartner.user existe) l'user doit matcher
    - payload legs: uniquement ceux du driver cible
    """
    from decimal import Decimal
    from django.db.models import Sum
    from django.utils import timezone

    order = get_object_or_404(
        Order.objects.select_related("laundry_partner", "customer", "delivery_partner"),
        pk=order_id,
    )

    def dt_iso(v):
        return v.isoformat() if v else None

    def dt_fr(v):
        if not v:
            return None
        try:
            v = timezone.localtime(v)
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11748")
        return v.strftime("%d/%m/%Y %H:%M")

    # ---------------------------------------
    # 1) Déterminer le driver cible
    # ---------------------------------------
    driver_id = (request.GET.get("driver_id") or "").strip()
    driver = None

    if request.user.is_staff or request.user.is_superuser:
        if driver_id:
            driver = DeliveryPartner.objects.filter(pk=driver_id).first()
        if not driver:
            first_leg = DeliveryLeg.objects.filter(order=order, driver__isnull=False).select_related("driver").first()
            driver = first_leg.driver if first_leg else getattr(order, "delivery_partner", None)
    else:
        # Source de vérité : le livreur doit être rattaché à au moins une DeliveryLeg de cette commande.
        possible_driver = None
        try:
            possible_driver = _get_connected_driver(request)
        except Exception:
            possible_driver = None

        if not possible_driver:
            return HttpResponseForbidden("Livreur non connecté.")

        has_leg = DeliveryLeg.objects.filter(order=order, driver=possible_driver).exists()
        if not has_leg:
            return HttpResponseForbidden("Accès refusé : aucune mission assignée à ce livreur.")

        if driver_id and str(possible_driver.id) != str(driver_id):
            return HttpResponseForbidden("Accès refusé : driver_id invalide.")

        driver = possible_driver

    # ---------------------------------------
    # 2) Legs : UNIQUEMENT ceux du driver cible
    # ---------------------------------------
    if driver:
        legs_qs = (
            DeliveryLeg.objects
            .filter(order=order, driver=driver)
            .select_related("driver")
            .order_by("id")
        )
    else:
        legs_qs = DeliveryLeg.objects.none()

    legs = []
    for leg in legs_qs:
        legs.append({
            "id": leg.id,
            "leg_type": leg.leg_type,
            "status": leg.status,
            "distance_km": float(leg.distance_km or 0),
            "driver_amount": float(getattr(leg, "driver_amount", 0) or 0),
            "started_at": dt_iso(getattr(leg, "started_at", None)),
            "finished_at": dt_iso(getattr(leg, "finished_at", None)),
        })

    # ---------------------------------------
    # 3) KPI driver (wallet net + potentiel legs)
    #    ⚠️ UI: jamais négatif côté livreur
    # ---------------------------------------
    driver_leg_amount_all = Decimal("0")
    driver_leg_amount_done = Decimal("0")
    driver_income = Decimal("0")
    driver_income_remaining = Decimal("0")
    driver_income_progress_pct = 0

    try:
        if driver:
            agg_all = legs_qs.aggregate(total=Sum("driver_amount"))
            driver_leg_amount_all = Decimal(str(agg_all["total"] or 0))

            agg_done = legs_qs.filter(status="done").aggregate(total=Sum("driver_amount"))
            driver_leg_amount_done = Decimal(str(agg_done["total"] or 0))

            qs_tx = WalletTransaction.objects.filter(
                order=order,
                type__in=["payout", "adjustment"],
                wallet__delivery_partner=driver,
            )

            net = qs_tx.filter(leg__isnull=False, leg__driver=driver).aggregate(net=_wallet_net_expr()).get("net")
            if net is None:
                net = qs_tx.aggregate(net=_wallet_net_expr()).get("net")

            driver_income = Decimal(str(net or 0))
            if driver_income < 0:
                driver_income = Decimal("0")

            if driver_leg_amount_all > 0:
                driver_income_remaining = driver_leg_amount_all - driver_income
                if driver_income_remaining < 0:
                    driver_income_remaining = Decimal("0")

                driver_income_progress_pct = int((driver_income * 100) / driver_leg_amount_all)
                if driver_income_progress_pct < 0:
                    driver_income_progress_pct = 0
                if driver_income_progress_pct > 100:
                    driver_income_progress_pct = 100
    except Exception:
        import logging
        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=11851")

    payload = {
        "ok": True,
        "driver_id": getattr(driver, "id", None),

        "order": {
            "id": order.id,
            "code": getattr(order, "code", None),
            "status": order.status,
            "payment_status": getattr(order, "payment_status", None),
            "amount_paid": float(getattr(order, "amount_paid", 0) or 0),
            "total_ttc": float(getattr(order, "total_client_ttc", None) or getattr(order, "total", 0) or 0),
            "amount_remaining": float(max((getattr(order, "total_client_ttc", None) or getattr(order, "total", 0) or 0) - (getattr(order, "amount_paid", 0) or 0), 0)),
            "created_at": dt_iso(getattr(order, "created_at", None)),
            "pickup_time": dt_iso(getattr(order, "pickup_time", None)),
            "dropoff_time": dt_iso(getattr(order, "dropoff_time", None)),
            "wash_complete_time": dt_iso(getattr(order, "wash_complete_time", None)),
            "return_time": dt_iso(getattr(order, "return_time", None)),
            "delivered_time": dt_iso(getattr(order, "delivered_time", None)),
        },

        # ✅ timeline live (format UI)
        "order_times": {
            "pickup_time": dt_fr(getattr(order, "pickup_time", None)),
            "dropoff_time": dt_fr(getattr(order, "dropoff_time", None)),
            "wash_complete_time": dt_fr(getattr(order, "wash_complete_time", None)),
            "return_time": dt_fr(getattr(order, "return_time", None)),
            "delivered_time": dt_fr(getattr(order, "delivered_time", None)),
        },

        "legs": legs,

        "kpi": {
            "driver_income": float(driver_income or 0),
            "driver_leg_amount_all": float(driver_leg_amount_all or 0),
            "driver_leg_amount_done": float(driver_leg_amount_done or 0),
            "driver_income_remaining": float(driver_income_remaining or 0),
            "driver_income_progress_pct": int(driver_income_progress_pct or 0),
        },
    }

    return JsonResponse(payload)


# ============================================================
#  TOP CLIENTS CSV
# ============================================================
def export_top_clients_csv(request):
    """
    Export CSV des 100 meilleurs clients par montant cumulé.
    On agrège sur Order.total et, en fallback, sur quantity * unit_price.
    """
    items_total = Sum(
        ExpressionWrapper(
            F("items__quantity") * F("items__unit_price"),
            output_field=DEC,
        ),
        output_field=DEC,
    )

    qs = (
        Order.objects
        .select_related("customer")
        .values("customer__name")
        .annotate(
            nb_cmd=Count("id", distinct=True),
            montant_total=Coalesce(
                Sum("total"),
                items_total,
                output_field=DEC,
            ),
        )
        .order_by("-montant_total")[:100]
    )

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="top_clients.csv"'
    resp.write("Client,Nb Commandes,Montant Total\n")
    for row in qs:
        name = (
            (row.get("customer__name") or "")
            .replace('"', "")
            .replace("\n", " ")
            .replace("\r", " ")
        )
        nb = int(row.get("nb_cmd") or 0)
        total = row.get("montant_total") or 0
        resp.write(f"{name},{nb},{total}\n")
    return resp


def export_top_clients_xlsx(request):
    """
    Export Excel des 100 meilleurs clients par montant cumulé.
    Même logique que export_top_clients_csv mais avec un design Excel.
    """
    items_total = Sum(
        ExpressionWrapper(
            F("items__quantity") * F("items__unit_price"),
            output_field=DEC,
        ),
        output_field=DEC,
    )

    qs = (
        Order.objects
        .select_related("customer")
        .values("customer__name")
        .annotate(
            nb_cmd=Count("id", distinct=True),
            montant_total=Coalesce(
                Sum("total"),
                items_total,
                output_field=DEC,
            ),
        )
        .order_by("-montant_total")[:100]
    )

    wb = Workbook()
    title_font = Font(size=16, bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0056B3")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws = wb.active
    ws.title = "Top clients"

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "FAGNI – Top 100 clients"
    t.font = title_font
    t.fill = header_fill
    t.alignment = center

    ws["A3"] = "Généré le"
    ws["B3"] = timezone.localtime().strftime("%d/%m/%Y %H:%M")

    headers = ["Rang", "Client", "Nb commandes", "Montant total (FCFA)"]
    for col_idx, head in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col_idx, value=head)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    row_idx = 6
    rank = 1
    for row in qs:
        name = (
            (row.get("customer__name") or "")
            .replace('"', "")
            .replace("\n", " ")
            .replace("\r", " ")
        )
        nb = int(row.get("nb_cmd") or 0)
        total = row.get("montant_total") or 0

        vals = [rank, name, nb, float(total)]
        for col_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin_border
            if col_idx in (1, 3, 4):
                c.alignment = right
                if col_idx == 4:
                    c.number_format = "#,##0"
            else:
                c.alignment = left

        rank += 1
        row_idx += 1

    widths = [8, 32, 14, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.auto_filter.ref = f"A5:D{row_idx - 1}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="fagni_top_clients.xlsx"'
    return resp



@login_required
def driver_update_location(request):
    """
    Reçoit la position GPS du livreur connecté et met à jour
    les champs latitude / longitude de son DeliveryPartner.

    Supporte:
    - JSON: {"lat": ..., "lng": ...}
    - Form: lat=...&lng=...
    - QueryString: ?lat=...&lng=...
    - Fallback keys: latitude/longitude

    En cas d'erreur lat/lng manquants, renvoie debug (content_type, body_len, preview).
    """
    from decimal import Decimal
    import json
    from urllib.parse import parse_qs

    if request.method != "POST":
        return JsonResponse({"error": "Méthode non autorisée"}, status=405)

    user_email = (getattr(request.user, "email", "") or "").strip()
    if not user_email:
        return JsonResponse(
            {"error": "Utilisateur sans email, impossible de lier un livreur"},
            status=400,
        )

    try:
        dp = DeliveryPartner.objects.get(email__iexact=user_email)
    except DeliveryPartner.DoesNotExist:
        return JsonResponse(
            {"error": f"Aucun livreur associé à {user_email}"},
            status=404,
        )

    ct = (request.content_type or "").lower()
    raw_bytes = request.body or b""
    raw_text = raw_bytes.decode("utf-8", errors="replace").strip()

    payload = {}

    # 1) JSON (même si ct a un charset)
    if raw_text and ("application/json" in ct or raw_text[:1] in "{["):
        try:
            payload = json.loads(raw_text)
            if not isinstance(payload, dict):
                payload = {}
        except Exception:
            payload = {}

    # 2) urlencoded body
    if raw_text and not payload and ("application/x-www-form-urlencoded" in ct or "=" in raw_text):
        try:
            qs = parse_qs(raw_text, keep_blank_values=True)
            payload = {k: (v[-1] if isinstance(v, list) and v else v) for k, v in qs.items()}
        except Exception:
            payload = {}

    def pick(*vals):
        for v in vals:
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            return v
        return None

    lat = pick(
        payload.get("lat"), payload.get("latitude"),
        request.POST.get("lat"), request.POST.get("latitude"),
        request.GET.get("lat"), request.GET.get("latitude"),
    )
    lng = pick(
        payload.get("lng"), payload.get("longitude"),
        request.POST.get("lng"), request.POST.get("longitude"),
        request.GET.get("lng"), request.GET.get("longitude"),
    )

    if lat is None or lng is None:
        return JsonResponse(
            {
                "error": "lat et lng sont requis",
                "debug": {
                    "content_type": ct,
                    "body_len": len(raw_bytes),
                    "body_preview": raw_text[:200],
                    "post_keys": list(request.POST.keys()),
                    "get_keys": list(request.GET.keys()),
                },
            },
            status=400,
        )

    try:
        dp.latitude = Decimal(str(lat))
        dp.longitude = Decimal(str(lng))
        dp.save(update_fields=["latitude", "longitude", "updated_at"])
    except Exception as e:
        return JsonResponse({"error": f"Erreur lors de la sauvegarde : {e}"}, status=400)

    return JsonResponse({"ok": True, "lat": str(dp.latitude), "lng": str(dp.longitude)})
@login_required
def driver_map(request):
    """
    Carte des livreurs FAGNI :
    - Affiche les livreurs actifs avec coordonnées.
    - Envoie les données JSON pour Leaflet.
    """

    from decimal import Decimal
    from django.db.models import Sum

    # Période de la semaine courante
    today = timezone.localdate()
    start_week = today - timezone.timedelta(days=today.weekday())

    # On ne garde que les livreurs actifs avec latitude & longitude renseignées
    drivers_qs = DeliveryPartner.objects.filter(
        is_active=True,
        latitude__isnull=False,
        longitude__isnull=False,
    ).order_by("name")

    drivers = []

    for d in drivers_qs:
        # Commandes de la semaine pour ce livreur
        week_orders = Order.objects.filter(
            legs__driver=d,
            created_at__date__gte=start_week,
            created_at__date__lte=today,
        )

        week_orders_count = week_orders.count()
        week_earnings = week_orders.aggregate(
            total=Sum("amount_driver_partner")
        )["total"] or Decimal("0")

        # Conversion des coordonnées en float pour Leaflet
        try:
            lat = float(d.latitude)
            lng = float(d.longitude)
        except (TypeError, ValueError):
            # Si on n'arrive pas à convertir, on skip ce livreur
            continue

        drivers.append({
            "id": d.id,
            "name": d.name,
            "city": getattr(d, "city", "") or "",
            "latitude": lat,
            "longitude": lng,
            "week_orders": week_orders_count,
            # on convertit en int pour éviter les Decimals en JSON
            "week_earnings": int(week_earnings),
        })

    # JSON pour le JS (Leaflet)
    drivers_json = json.dumps(drivers, ensure_ascii=False)

    context = {
        "drivers": drivers,  # utilisé dans la colonne de droite
        "drivers_json": drivers_json,  # utilisé par le <script id="drivers-json">
        "connected_driver": _get_connected_driver(request),
    }
    return render(request, "orders/driver_map.html", context)


@login_required
@require_GET
def driver_map_data(request):
    """
    Endpoint JSON pour rafraîchir la carte des livreurs (Leaflet).
    Utilisé par driver_map.html en auto-refresh.
    Ajouts OPS :
    - status : available | busy | stale
    - age_seconds / age_label
    """
    from decimal import Decimal
    from django.db.models import Sum

    def _age_label(seconds: int) -> str:
        try:
            s = int(seconds)
        except Exception:
            return "—"
        if s < 60:
            return f"{s}s"
        m = s // 60
        if m < 60:
            return f"{m} min"
        h = m // 60
        if h < 24:
            return f"{h} h"
        d = h // 24
        return f"{d} j"

    now = timezone.now()
    today = timezone.localdate()
    start_week = today - timezone.timedelta(days=today.weekday())

    drivers_qs = DeliveryPartner.objects.filter(
        is_active=True,
        latitude__isnull=False,
        longitude__isnull=False,
    ).order_by("name")

    drivers = []

    for d in drivers_qs:
        week_orders = Order.objects.filter(
            legs__driver=d,
            created_at__date__gte=start_week,
            created_at__date__lte=today,
        )

        week_orders_count = week_orders.count()
        week_earnings = week_orders.aggregate(total=Sum("amount_driver_partner"))["total"] or Decimal("0")

        # Dernière commande active (optionnel, utile OPS)
        active_order = (
            Order.objects.filter(legs__driver=d, status__in=["pending", "in_progress"]).distinct()
            .order_by("-created_at")
            .first()
        )

        try:
            lat = float(d.latitude)
            lng = float(d.longitude)
        except (TypeError, ValueError):
            continue

        updated_at = getattr(d, "updated_at", None)
        age_seconds = None
        if updated_at:
            try:
                age_seconds = int((now - updated_at).total_seconds())
            except Exception:
                age_seconds = None

        # status
        status = "available"
        if active_order:
            status = "busy"
        # stale si pas d'update récente (15 min)
        if (age_seconds is not None) and (age_seconds > 15 * 60) and (not active_order):
            status = "stale"

        drivers.append({
            "id": d.id,
            "name": d.name,
            "city": getattr(d, "city", "") or "",
            "latitude": lat,
            "longitude": lng,
            "week_orders": int(week_orders_count),
            "week_earnings": int(week_earnings),
            "active_order_id": active_order.id if active_order else None,
            "active_order_code": getattr(active_order, "code", None) if active_order else None,
            "updated_at": updated_at.isoformat() if updated_at else None,
            "age_seconds": age_seconds,
            "age_label": _age_label(age_seconds) if age_seconds is not None else "—",
            "status": status,
        })

    return JsonResponse({"drivers": drivers})


# ============================================================
#  WEIGHING (driver / laundry)
# ============================================================


def _can_driver_touch_order(request, order):
    """
    Sécurité livreur:
    - Autorise staff/superuser
    - Autorise uniquement un livreur rattaché à une DeliveryLeg de la commande
    """
    u = getattr(request, "user", None)
    if not u or not getattr(u, "is_authenticated", False):
        return False

    if getattr(u, "is_superuser", False) or getattr(u, "is_staff", False):
        return True

    try:
        from partners.models import DeliveryPartner
        user_email = (getattr(u, "email", "") or "").strip().lower()
        driver = None

        dp_id = getattr(u, "delivery_partner_id", None) or getattr(u, "driver_id", None)
        if dp_id:
            driver = DeliveryPartner.objects.filter(pk=dp_id).first()

        if not driver and user_email:
            driver = DeliveryPartner.objects.filter(email__iexact=user_email).first()

        if not driver:
            return False

        return DeliveryLeg.objects.filter(order=order, driver=driver).exists()
    except Exception:
        return False


def _sanitize_evidence_kind(kind, OrderEvidencePhoto):
    allowed = {k for (k, _lbl) in getattr(OrderEvidencePhoto, "KIND_CHOICES", [])}
    kind = (kind or "").strip() or "pickup_items"
    return kind if kind in allowed else "pickup_items"


@login_required
@require_http_methods(["GET", "POST"])
def driver_weighing(request, order_id):
    """
    Page de pesée côté livreur.
    - Enregistre OrderWeighing.weight_kg + scale_photo + notes
    - Ajoute une preuve photo (OrderEvidencePhoto) si scale_photo uploadée
    Template: orders/driver_weighing.html
    """
    from decimal import Decimal
    from .models import OrderWeighing, OrderEvidencePhoto

    order = get_object_or_404(Order, pk=order_id)

    
    # 🔐 Sécurité: seul le livreur rattaché à une DeliveryLeg de la commande peut accéder
    if not _can_driver_touch_order(request, order):
        user_driver_id = (
            getattr(request.user, "delivery_partner_id", None)
            or getattr(request.user, "driver_id", None)
            or (request.GET.get("driver_id") or "").strip()
            or (request.POST.get("driver_id") or "").strip()
        )
        messages.warning(request, "Cette course ne t'est pas assignée.")
        return redirect(reverse("orders:driver_app") + (f"?driver_id={user_driver_id}" if user_driver_id else ""))


    ow, _created = OrderWeighing.objects.get_or_create(order=order)

    # leg optionnel pour lier la pesée / preuve à la mission active
    leg_obj = None
    raw_leg_id = (request.GET.get("leg_id") or request.POST.get("leg_id") or "").strip()
    if raw_leg_id.isdigit():
        try:
            leg_obj = DeliveryLeg.objects.filter(pk=int(raw_leg_id), order=order).first()
        except Exception:
            leg_obj = None

    back_url = (request.GET.get("back") or request.POST.get("back") or "").strip()

    # Seule une pesée "draft" est modifiable par le livreur.
    # confirmed/disputed/resolved sont en lecture seule.
    is_locked = (ow.status != "draft")
    if is_locked:
        messages.info(request, "Pesée déjà traitée (lecture seule).")

    if request.method == "POST":
        if is_locked:
            messages.warning(request, "Pesée confirmée : modification désactivée.")
            if back_url and url_has_allowed_host_and_scheme(back_url, allowed_hosts={request.get_host()}):
                return redirect(back_url)
            url = reverse("orders:driver_weighing", kwargs={"order_id": order.id})
            qs = []
            if raw_leg_id:
                qs.append(f"leg_id={raw_leg_id}")
            if back_url:
                qs.append(f"back={back_url}")
            if qs:
                url += "?" + "&".join(qs)
            return redirect(url)

        raw_weight = (request.POST.get("weight_kg") or "").strip()
        notes = (request.POST.get("notes") or "").strip()
        scale_photo = request.FILES.get("scale_photo")

        weight = None
        if raw_weight:
            try:
                weight = Decimal(raw_weight.replace(" ", "").replace(",", "."))
            except Exception:
                weight = None

        if weight is not None and weight <= 0:
            weight = None

        if weight is None:
            messages.warning(request, "Poids invalide. Exemple attendu : 3.50 (strictement superieur a 0)")
            if back_url and url_has_allowed_host_and_scheme(back_url, allowed_hosts={request.get_host()}):
                return redirect(back_url)
            url = reverse("orders:driver_weighing", kwargs={"order_id": order.id})
            qs = []
            if raw_leg_id:
                qs.append(f"leg_id={raw_leg_id}")
            if back_url:
                qs.append(f"back={back_url}")
            if qs:
                url += "?" + "&".join(qs)
            return redirect(url)

        # Update weighing
        ow.weight_kg = weight
        if notes:
            ow.notes = notes
        if scale_photo:
            ow.scale_photo = scale_photo

        ow.entered_by_type = "driver"
        ow.entered_by_id = getattr(request.user, "id", None)
        ow.status = "draft"
        ow.save()

        # Evidence scale photo
        if scale_photo:
            OrderEvidencePhoto.objects.create(
                order=order,
                leg=leg_obj,
                actor_type="driver",
                actor_id=getattr(request.user, "id", None),
                kind="weighing_scale",
                image=scale_photo,
                caption="Photo balance (pesée) – livreur",
            )

        messages.success(request, "✅ Pesée enregistrée")

        # 🔁 retour automatique mission
        driver_id = request.GET.get("driver_id") or request.POST.get("driver_id")
        url = reverse("orders:driver_app")
        if driver_id:
            url += f"?driver_id={driver_id}"
        return redirect(url)
        if back_url and url_has_allowed_host_and_scheme(back_url, allowed_hosts={request.get_host()}):
            return redirect(back_url)
        url = reverse("orders:driver_weighing", kwargs={"order_id": order.id})
        qs = []
        if raw_leg_id:
            qs.append(f"leg_id={raw_leg_id}")
        if back_url:
            qs.append(f"back={back_url}")
        if qs:
            url += "?" + "&".join(qs)
        return redirect(url)

    evidence = OrderEvidencePhoto.objects.filter(order=order).order_by("-created_at")[:50]

    latest_issue = (
        OrderEvidencePhoto.objects
        .filter(order=order, kind="issue")
        .order_by("-created_at")
        .first()
    )

    return render(
        request,
        "orders/driver_weighing.html",
        {
            "order": order,
            "weighing": ow,
            "evidence": evidence,
            "driver_id": (
                (request.GET.get("driver_id") or request.POST.get("driver_id") or "").strip()
                or str(getattr(getattr(order, "delivery_partner", None), "id", "") or "")
            ),
            "is_locked": is_locked,
            "latest_issue": latest_issue,
            "leg_obj": leg_obj,
            "back_url": back_url,
        },
    )


@login_required
@require_http_methods(["POST"])
def driver_evidence_upload(request, order_id):
    """
    Upload preuve(s) photo côté livreur.

    Attendus:
    - files: image (single) OU images (multiple)
    - kind: une valeur de OrderEvidencePhoto.KIND_CHOICES (default pickup_items)
    - caption: texte optionnel
    - leg_id: optionnel (lier la preuve à une mission pickup/dropoff/return)
    - back: optionnel (redirige si URL sûre)
    """
    from .models import OrderEvidencePhoto
    from .models import DeliveryLeg

    order = get_object_or_404(Order, pk=order_id)

    # 🔐 Sécurité: preuve autorisée uniquement pour un livreur rattaché à une DeliveryLeg de la commande
    if not _can_driver_touch_order(request, order):
        user_driver_id = (
            getattr(request.user, "delivery_partner_id", None)
            or getattr(request.user, "driver_id", None)
            or (request.POST.get("driver_id") or request.GET.get("driver_id") or "").strip()
        )
        messages.warning(request, "Cette course ne t'est pas assignée.")
        return redirect(reverse("orders:driver_app") + (f"?driver_id={user_driver_id}" if user_driver_id else ""))

    kind = _sanitize_evidence_kind(request.POST.get("kind"), OrderEvidencePhoto)
    caption = (request.POST.get("caption") or "").strip()

    # Optionnel: lier la preuve à un leg (pickup/dropoff/return)
    leg_obj = None
    raw_leg_id = (request.POST.get("leg_id") or "").strip()
    if raw_leg_id.isdigit():
        try:
            leg_obj = DeliveryLeg.objects.filter(pk=int(raw_leg_id), order=order).first()
        except Exception:
            leg_obj = None
    files = []
    if "images" in request.FILES:
        files = request.FILES.getlist("images")
    elif "image" in request.FILES:
        f = request.FILES.get("image")
        if f:
            files = [f]

    if not files:
        messages.warning(request, "Aucune photo reçue.")
        back = (request.GET.get("back") or request.POST.get("back") or "").strip()
        if back and url_has_allowed_host_and_scheme(back, allowed_hosts={request.get_host()}):
            return redirect(back)
        url = reverse("orders:driver_weighing", kwargs={"order_id": order.id})
        qs = []
        raw_leg_id = (request.POST.get("leg_id") or request.GET.get("leg_id") or "").strip()
        if raw_leg_id:
            qs.append(f"leg_id={raw_leg_id}")
        if back:
            qs.append(f"back={back}")
        if qs:
            url += "?" + "&".join(qs)
        return redirect(url)

    for f in files:
        OrderEvidencePhoto.objects.create(
            order=order,
            leg=leg_obj,  # ✅ on lie au leg si fourni
            actor_type="driver",
            actor_id=getattr(request.user, "id", None),
            kind=kind,
            image=f,
            caption=caption,
        )

    messages.success(request, f"{len(files)} photo(s) ajoutée(s).")

    # ↩️ Redirect back si sûr, sinon fallback
    back = (request.GET.get("back") or request.POST.get("back") or "").strip()
    if back and url_has_allowed_host_and_scheme(back, allowed_hosts={request.get_host()}):
        return redirect(back)

    
    # 🔁 retour automatique mission après upload
    driver_id = request.GET.get("driver_id") or request.POST.get("driver_id")
    url = reverse("orders:driver_app")
    if driver_id:
        url += f"?driver_id={driver_id}"
    return redirect(url)


def _safe_decimal(v, default="0"):
    try:
        if v is None:
            return Decimal(default)
        return Decimal(str(v))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


def _item_unit_price_fcfa(item) -> Decimal:
    """
    Essaye de trouver un prix unitaire (FCFA) sur l'item.
    On teste plusieurs champs possibles pour être robuste.
    """
    for attr in ("unit_price", "price", "unit_price_fcfa", "price_fcfa", "amount_fcfa", "unit_amount"):
        if hasattr(item, attr):
            val = getattr(item, attr)
            d = _safe_decimal(val, default="0")
            if d >= 0:
                return d
    return Decimal("0")


def _resolve_laundry_for_user(request):
    """
    Même logique que laundry_app:
    - non-staff: déduction via email/username
    - staff: ?laundry_id=XX
    """
    laundry = None
    error = None

    if not getattr(request.user, "is_staff", False):
        user_email = (getattr(request.user, "email", "") or "").strip().lower()
        user_name = (getattr(request.user, "username", "") or "").strip().lower()

        q = LaundryPartner.objects.all()
        if user_email:
            laundry = q.filter(email__iexact=user_email).first()
        if laundry is None and user_name:
            laundry = q.filter(email__iexact=user_name).first() or q.filter(name__iexact=user_name).first()

        if laundry is None:
            error = "Aucune blanchisserie liée à ce compte. Contacte l'admin pour associer ton compte à une blanchisserie."
    else:
        laundry_id = (request.GET.get("laundry_id") or "").strip()
        if not laundry_id:
            error = "laundry_id manquant. Exemple: /orders/laundry/app/?laundry_id=1"
        else:
            try:
                lid = int(laundry_id)
                laundry = LaundryPartner.objects.filter(id=lid).first()
                if not laundry:
                    error = "Blanchisserie introuvable (laundry_id invalide)."
            except Exception:
                error = "laundry_id invalide (doit être un entier)."

    return laundry, error


@login_required
@require_http_methods(["GET", "POST"])
def laundry_weighing(request, order_id):
    """
    ✅ MVP V1 (prix par article) — Détails commande côté blanchisserie.
    - Zéro infos client
    - Actions: note interne, démarrer lavage, marquer prêt
    (On garde le nom/URL 'weighing' pour compat sans casser les liens.)
    """
    laundry, error = _resolve_laundry_for_user(request)

    # Si pas de laundry, on affiche une page propre
    if not laundry:
        return render(request, "orders/laundry_weighing.html", {
            "laundry": None,
            "order": None,
            "items": [],
            "totals": {"total_fcfa": 0},
            "error": error,
            "laundry_id": (request.GET.get("laundry_id") or "").strip(),
            "evidence": [],
        })

    order = get_object_or_404(Order, id=order_id)

    display_summary = build_order_display_summary(order)
    finance_summary = build_order_finance_summary(order)

    selected_driver_id = (request.GET.get("driver_id") or "").strip()
    assigned_driver = getattr(order, "delivery_partner", None)

    if not request.user.is_staff:
        # fallback automatique côté livreur
        if assigned_driver and not selected_driver_id:
            selected_driver_id = str(assigned_driver.id)

        # 🔥 CAS 1 : commande non assignée → on autorise (mode souple)
        if not assigned_driver:
            pass

        # 🔥 CAS 2 : assignée à quelqu’un d’autre → bloqué
        elif selected_driver_id and str(assigned_driver.id) != str(selected_driver_id):
            return redirect("orders:driver_hub")


    
    # Sécurité: la commande doit appartenir à cette blanchisserie
    if getattr(order, "laundry_partner_id", None) != laundry.id:
        messages.warning(request, "Accès refusé : cette commande n'est pas attribuée à ta blanchisserie.")
        return redirect(f"{reverse('orders:laundry_app')}?laundry_id={laundry.id}")

    # Résumé bag/item
    items = []
    total = Decimal(str(finance_summary.get("prestation_total", 0) or 0))

    if display_summary.get("is_item"):
        items_qs = order.items.all().order_by("id") if hasattr(order, "items") else []
        for it in items_qs:
            qty = getattr(it, "quantity", 1) or 1
            try:
                qty = int(qty)
            except Exception:
                qty = 1

            unit = _item_unit_price_fcfa(it)
            line_total = unit * Decimal(qty)

            items.append({
                "obj": it,
                "designation": getattr(it, "designation", "Article"),
                "quantity": qty,
                "unit_price_fcfa": unit,
                "line_total_fcfa": line_total,
            })

    # Evidence (si modèle dispo)
    evidence = []
    try:
        from orders.models import EvidencePhoto  # adapte si ton modèle s'appelle autrement
        evidence = list(EvidencePhoto.objects.filter(order=order).order_by("-id")[:50])
    except Exception:
        evidence = []

    # POST: note/démarrer
    if request.method == "POST":
        # Frontière d'autorité V2.
        # Le garde est volontairement placé avant toute écriture, y compris
        # les notes, afin qu'un POST refusé soit sans effet de bord.
        from services.services import order_uses_canonical_service_executions
        if order_uses_canonical_service_executions(order=order):
            return JsonResponse({
                "error": "autorite_v2",
                "message": (
                    "Le statut de cette commande est piloté par "
                    "ses ServiceExecution."
                ),
            }, status=409)

        action = (request.POST.get("action") or "").strip()
        note = (request.POST.get("notes") or "").strip()[:250]

        # Note interne (on la met dans un champ "laundry_notes" si existe, sinon on ignore)
        if note:
            if hasattr(order, "laundry_notes"):
                order.laundry_notes = note
                try:
                    order.save(update_fields=["laundry_notes"])
                except Exception:
                    try:
                        order.save()
                    except Exception:
                        import logging
                        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=12841")

        if action == "start" and order.status != "canceled":
            pickup_done = DeliveryLeg.objects.filter(
                order=order, leg_type="pickup", status="done",
            ).exists()
            if pickup_done:
                order.status = "in_progress"
                order.save(update_fields=["status"])

        redirect_url = reverse("orders:laundry_weighing", kwargs={"order_id": order.id})
        return redirect(f"{redirect_url}?laundry_id={laundry.id}")

    return render(request, "orders/laundry_weighing.html", {
        "laundry": laundry,
        "order": order,
        "display_summary": display_summary,
        "finance_summary": finance_summary,
        "items": items,
        "totals": {"total_fcfa": int(total)},
        "error": error,
        "laundry_id": str(laundry.id),
        "evidence": evidence,
    })


@login_required
@require_http_methods(["POST"])
def laundry_weighing_confirm(request, order_id):
    """
    ✅ MVP V1 — Marquer 'Prêt' (wash_complete_time) + créer/activer mission return.
    """
    laundry, _error = _resolve_laundry_for_user(request)
    if not laundry:
        return redirect("orders:laundry_app")

    order = get_object_or_404(Order, id=order_id)

    # --- Validations (aucune écriture avant ce point) ---
    if getattr(order, "laundry_partner_id", None) != laundry.id:
        return redirect(f"{reverse('orders:laundry_app')}?laundry_id={laundry.id}")

    # Frontière d'autorité V2 avant toute mutation métier.
    from services.services import order_uses_canonical_service_executions
    if order_uses_canonical_service_executions(order=order):
        return JsonResponse({
            "error": "autorite_v2",
            "message": (
                "Le statut de cette commande est piloté par "
                "ses ServiceExecution."
            ),
        }, status=409)

    if order.status == "canceled":
        messages.warning(request, "Une commande annulée ne peut plus être marquée prête.")
        return redirect(f"{reverse('orders:laundry_weighing', kwargs={'order_id': order.id})}?laundry_id={laundry.id}")

    pickup_done = DeliveryLeg.objects.filter(
        order=order, leg_type="pickup", status="done",
    ).exists()
    if not pickup_done:
        messages.warning(request, "La collecte n'est pas encore terminée : impossible de marquer la commande prête.")
        return redirect(f"{reverse('orders:laundry_weighing', kwargs={'order_id': order.id})}?laundry_id={laundry.id}")

    # --- Écritures (idempotentes) ---
    if not getattr(order, "wash_complete_time", None):
        order.wash_complete_time = timezone.now()
        order.status = "ready"
        order.save(update_fields=["wash_complete_time", "status", "updated_at"])
    elif order.status != "ready":
        order.status = "ready"
        order.save(update_fields=["status", "updated_at"])

    # ✅ Dès que c'est prêt, on crée/active la mission RETOUR
    try:
        DeliveryLeg.objects.get_or_create(
            order=order,
            leg_type="return",
            defaults={
                "status": "pending",
                "driver": order.delivery_partner if getattr(order, "delivery_partner_id", None) else None,
            },
        )

        # Auto-assign driver sur return sans driver
        if getattr(order, "delivery_partner_id", None):
            DeliveryLeg.objects.filter(
                order=order, leg_type="return", driver__isnull=True
            ).update(driver=order.delivery_partner)

        # Ne réactiver que les returns annulés NON payés
        qs_reactivate = DeliveryLeg.objects.filter(
            order=order,
            leg_type="return",
            status__in=("canceled", "cancelled"),
        )
        try:
            from wallets.models import WalletTransaction
            paid_leg_ids = set(
                WalletTransaction.objects.filter(
                    order_id=order.id,
                    wallet__owner_type="driver",
                    type="payout",
                    direction="in",
                )
                .exclude(leg_id__isnull=True)
                .values_list("leg_id", flat=True)
            )
        except Exception:
            paid_leg_ids = set()

        qs_reactivate.exclude(id__in=paid_leg_ids).update(status="pending")

    except Exception:
        logger.exception("laundry_weighing_confirm: return leg activation failed order=%s", getattr(order, "id", None))

    # redirect propre
    url = reverse("orders:laundry_weighing", kwargs={"order_id": order.id})
    return redirect(f"{url}?laundry_id={laundry.id}")


@login_required
@require_http_methods(["POST"])
def laundry_weighing_dispute(request, order_id):
    """
    ⚠️ MVP V1 — Signaler un problème (raison + photo optionnelle) => OrderEvidencePhoto.
    """
    laundry, _error = _resolve_laundry_for_user(request)
    if not laundry:
        return redirect("orders:laundry_app")

    order = get_object_or_404(Order, id=order_id)

    if getattr(order, "laundry_partner_id", None) != laundry.id:
        return redirect(f"{reverse('orders:laundry_app')}?laundry_id={laundry.id}")

    reason = (request.POST.get("reason") or "").strip()[:250]
    image = request.FILES.get("image")

    if reason or image:
        try:
            from .models import OrderEvidencePhoto, OrderWeighing

            allowed = {k for (k, _lbl) in getattr(OrderEvidencePhoto, "KIND_CHOICES", [])}
            kind = "issue" if "issue" in allowed else ("laundry" if "laundry" in allowed else "pickup_items")

            OrderEvidencePhoto.objects.create(
                order=order,
                leg=None,
                actor_type="laundry",
                actor_id=getattr(request.user, "id", None),
                kind=kind,
                image=image if image else None,
                caption=reason or "Problème signalé par la blanchisserie",
            )

            ow, _created = OrderWeighing.objects.get_or_create(order=order)
            # Une pesee resolue par OPS ne doit jamais repasser en dispute.
            if ow.status != "resolved":
                ow.status = "disputed"
                ow.save(update_fields=["status", "updated_at"])
        except Exception:
            logger.exception("laundry_weighing_dispute: cannot create evidence for order=%s", getattr(order, "id", None))

    url = reverse("orders:laundry_weighing", kwargs={"order_id": order.id})
    return redirect(f"{url}?laundry_id={laundry.id}")


def _client_order_lock_reason(order) -> str:
    """
    Retourne la raison pour laquelle une commande client ne peut plus être
    modifiée, ou une chaîne vide lorsqu'elle reste modifiable.

    Une session Wave déjà créée gèle définitivement le montant présenté au
    prestataire de paiement. Une modification nécessite alors une nouvelle
    commande, et non la réécriture du prix de la commande existante.
    """
    if getattr(order, "status", None) == "canceled":
        return "order_canceled"

    if not bool(getattr(order, "is_draft", True)):
        return "order_confirmed"

    payment_status = (
        getattr(order, "payment_status", "") or ""
    ).strip().lower()

    if payment_status in {"partial", "paid"}:
        return "payment_started"

    try:
        amount_paid = Decimal(
            str(getattr(order, "amount_paid", 0) or 0)
        )
    except Exception:
        amount_paid = DECIMAL_ZERO

    if amount_paid > DECIMAL_ZERO:
        return "payment_started"

    if (getattr(order, "wave_checkout_id", "") or "").strip():
        return "wave_checkout_active"

    return ""


def _client_order_locked_response(request, order):
    reason = _client_order_lock_reason(order)

    if not reason:
        return None

    if request.method == "POST":
        resp = JsonResponse(
            {
                "ok": False,
                "error": "order_locked",
                "reason": reason,
                "message": (
                    "Cette commande est verrouillée et ne peut plus être "
                    "modifiée."
                ),
            },
            status=409,
        )
        resp["Cache-Control"] = "no-store"
        return resp

    return redirect(
        "orders:client_order_detail",
        order_id=order.id,
    )


# ============================================================
# Wizard Client V1 (Step2/3/4) — ajout safe (anti-crash URLs)
# ============================================================

@require_http_methods(["GET", "POST"])
@client_required
def client_new_order_step2(request, order_id: int):
    """
    Step 2/4:
    - choix du mode de commande (bag / item)
    - catégorie uniquement si mode=item (stockée en session)
    - mode livraison (standard/express/scheduled)
    - scheduled_delivery_at accepte datetime-local (YYYY-MM-DDTHH:MM) ou "YYYY-MM-DD HH:MM"
    """
    from orders.models import ServiceCategory
    from django.utils import timezone
    from django.utils.dateparse import parse_datetime

    phone = _client_phone(request)
    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        return redirect("orders:client_new_order")

    locked_response = _client_order_locked_response(request, order)
    if locked_response is not None:
        return locked_response

    display_summary = build_order_display_summary(order)
    finance_summary = build_order_finance_summary(order)

    categories = ServiceCategory.objects.all().order_by("name")
    error = None

    selected_pricing_mode = getattr(order, "pricing_mode", None) or request.session.get(f"client_wizard_pricing_mode_{order.id}") or "bag"

    if request.method == "POST":
        # 1) mode de commande
        pricing_mode = (request.POST.get("pricing_mode") or "").strip().lower() or "bag"
        if pricing_mode not in ("bag", "item"):
            pricing_mode = "bag"

        request.session[f"client_wizard_pricing_mode_{order.id}"] = pricing_mode
        selected_pricing_mode = pricing_mode
        order.pricing_mode = pricing_mode

        # 2) catégorie uniquement si mode=item
        if pricing_mode == "item":
            cat_id = (request.POST.get("category_id") or "").strip()
            if not cat_id.isdigit():
                error = "Merci de choisir une catégorie."
            else:
                request.session[f"client_wizard_category_id_{order.id}"] = int(cat_id)
        else:
            request.session.pop(f"client_wizard_category_id_{order.id}", None)

        # 3) mode livraison
        delivery_mode = (request.POST.get("delivery_mode") or "").strip() or "standard"
        scheduled_raw = (request.POST.get("scheduled_delivery_at") or "").strip()

        if delivery_mode not in ("standard", "express", "scheduled"):
            delivery_mode = "standard"

        if not error:
            order.delivery_mode = delivery_mode

            if delivery_mode == "scheduled":
                dt = parse_datetime(scheduled_raw) if scheduled_raw else None
                if dt is None and scheduled_raw and "T" in scheduled_raw:
                    dt = parse_datetime(scheduled_raw.replace("T", " ", 1))

                if not dt:
                    error = "Merci de choisir une date/heure de livraison programmée."
                else:
                    if timezone.is_naive(dt):
                        dt = timezone.make_aware(dt, timezone.get_current_timezone())
                    order.scheduled_delivery_at = dt
            else:
                order.scheduled_delivery_at = None

        if not error:
            order.save(update_fields=["pricing_mode", "delivery_mode", "scheduled_delivery_at", "updated_at"])
            return redirect("orders:client_new_order_step3", order_id=order.id)

    return render(request, "orders/client_new_order_step2.html", {
        "order": order,
        "categories": categories,
        "selected_category_id": request.session.get(f"client_wizard_category_id_{order.id}"),
        "selected_pricing_mode": selected_pricing_mode,
        "error": error,
    })


@require_http_methods(["GET", "POST"])
@client_required
def client_new_order_step3(request, order_id: int):
    from django.contrib import messages
    from orders.models import ServiceCategory, ServiceItem, OrderItem

    phone = _client_phone(request)
    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        return redirect("orders:client_new_order")

    locked_response = _client_order_locked_response(request, order)
    if locked_response is not None:
        return locked_response

    display_summary = build_order_display_summary(order)
    finance_summary = build_order_finance_summary(order)

    pricing_mode = getattr(order, "pricing_mode", None) or request.session.get(f"client_wizard_pricing_mode_{order.id}") or "bag"

    # Catégorie choisie en step2 (stockée en session)
    categories = ServiceCategory.objects.order_by("name")
    selected_id = request.session.get(f"client_wizard_category_id_{order.id}")
    category = None
    if selected_id:
        category = categories.filter(id=selected_id).first()
    if not category and pricing_mode == "item":
        category = categories.first()

    services = ServiceItem.objects.none()
    if category and pricing_mode == "item":
        services = ServiceItem.objects.filter(category=category, is_active=True).order_by("name")

    existing_items = _build_client_display_items(order)

    # Pré-remplissage des quantités par service (pour l'UI step3 item)
    qty_by_service = {}
    try:
        for it in existing_items:
            sid = getattr(it, "service_id", None)
            qty = getattr(it, "quantity", None)
            if sid:
                qty_by_service[sid] = qty
                continue

            if isinstance(it, dict):
                sid = it.get("service_id")
                qty = it.get("quantity", 0)
                if sid:
                    qty_by_service[sid] = qty
    except Exception:
        qty_by_service = {}

    error = None
    selected_bag_size = getattr(order, "bag_size", None) or "medium"
    bag_rules_confirmed = bool(request.session.get(f"client_bag_rules_confirmed_{order.id}", False))
    cgu_accepted = bool(request.session.get(f"client_cgu_accepted_{order.id}", False))

    if request.method == "POST":
        if pricing_mode == "bag":
            bag_size = (request.POST.get("bag_size") or "").strip().lower()
            confirm_rules = (request.POST.get("confirm_bag_rules") or "").strip()

            if bag_size not in ("small", "medium", "large"):
                error = "Merci de choisir une taille de sac."
            elif confirm_rules != "1":
                error = "Merci de confirmer que ton sac contient uniquement des vêtements."
            else:
                order.bag_size = bag_size
                request.session[f"client_bag_rules_confirmed_{order.id}"] = True

                try:
                    order.update_financials(save=True)
                except Exception:
                    try:
                        order.save(update_fields=["bag_size", "updated_at"])
                    except Exception:
                        import logging
                        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13157")

                return redirect("orders:client_new_order_step4", order_id=order.id)

        else:
            if not category:
                error = "Aucune catégorie de service n'est disponible. Ajoute une catégorie dans l'admin."
            elif not services.exists():
                error = "Aucun article actif dans cette catégorie. Ajoute des articles dans l'admin."
            else:
                changed = False

                current_service_type = (getattr(order, "service_type", None) or "").strip().lower()
                selected_service_type = ((getattr(category, "slug", None) or getattr(category, "name", "") or "").strip().lower())

                if not current_service_type:
                    try:
                        existing_last = (
                            OrderItem.objects
                            .filter(order=order)
                            .select_related("service__category")
                            .order_by("-id")
                            .first()
                        )
                        if existing_last and getattr(existing_last, "service", None) and getattr(existing_last.service, "category", None):
                            current_service_type = (
                                getattr(existing_last.service.category, "slug", None)
                                or getattr(existing_last.service.category, "name", "")
                                or ""
                            ).strip().lower()
                    except Exception:
                        current_service_type = ""

                if current_service_type and selected_service_type and current_service_type != selected_service_type:
                    messages.error(
                        request,
                        "Une commande ne peut contenir qu’un seul type de service. Crée une commande séparée pour un autre service."
                    )
                    return redirect("orders:client_new_order_step3", order_id=order.id)

                if not current_service_type and selected_service_type:
                    try:
                        order.service_type = selected_service_type
                        order.save(update_fields=["service_type"])
                        current_service_type = selected_service_type
                    except Exception:
                        import logging
                        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13203")

                for svc in services:
                    raw = (request.POST.get(f"qty_{svc.id}") or "").strip()
                    try:
                        qty = int(raw) if raw != "" else 0
                    except Exception:
                        qty = 0

                    it = (
                        OrderItem.objects
                        .filter(order=order, service=svc)
                        .order_by("-id")
                        .first()
                    )

                    if qty <= 0:
                        if it:
                            it.delete()
                            changed = True
                        continue

                    if it:
                        it.designation = svc.name
                        it.quantity = qty
                        it.unit_price = svc.default_price
                        it.save()
                        changed = True
                    else:
                        it = OrderItem.objects.create(
                            order=order,
                            service=svc,
                            designation=svc.name,
                            quantity=qty,
                            unit_price=svc.default_price,
                        )
                        changed = True

                    photo_files = request.FILES.getlist(f"photos_{svc.id}")
                    if photo_files:
                        try:
                            it.photos.all().delete()
                        except Exception:
                            import logging
                            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13246")
                        for photo_file in photo_files:
                            if not photo_file:
                                continue
                            OrderItemPhoto.objects.create(
                                order_item=it,
                                image=photo_file,
                            )
                            changed = True

                if changed:
                    try:
                        order.update_financials(save=True)
                    except Exception:
                        import logging
                        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13260")

                action = (request.POST.get("action") or "").strip()
                if action == "add_more":
                    messages.success(request, "Articles ajoutés. Tu peux continuer à compléter la liste.")
                    return redirect("orders:client_new_order_step3", order_id=order.id)

                if not changed:
                    error = "Sélectionne au moins un article pour continuer."
                else:
                    return redirect("orders:client_new_order_step4", order_id=order.id)

    return render(request, "orders/client_new_order_step3.html", {
        "order": order,
        "display_summary": display_summary,
        "finance_summary": finance_summary,
        "pricing_mode": pricing_mode,
        "category": category,
        "services": services,
        "qty_by_service": qty_by_service,
        "selected_bag_size": selected_bag_size,
        "bag_rules_confirmed": bag_rules_confirmed,
        "cgu_accepted": cgu_accepted,
        "error": error,
    })

@require_http_methods(["GET", "POST"])
@client_required
def client_new_order_step4(request, order_id: int):
    phone = _client_phone(request)
    order = (
        Order.objects
        .select_related("customer")
        .filter(pk=order_id, customer__phone=phone)
        .first()
    )
    if not order:
        return redirect("orders:client_new_order")

    if getattr(order, "status", None) == "canceled":
        if request.method == "POST":
            resp = JsonResponse(
                {
                    "ok": False,
                    "error": "order_canceled",
                    "message": (
                        "Une commande annulée ne peut pas être confirmée."
                    ),
                },
                status=409,
            )
            resp["Cache-Control"] = "no-store"
            return resp

        return redirect(
            "orders:client_order_detail",
            order_id=order.id,
        )

    from decimal import Decimal
    from orders.models import OrderItem

    display_summary = build_order_display_summary(order)
    finance_summary = build_order_finance_summary(order)

    pricing_mode = display_summary.get("pricing_mode", "bag") or "bag"
    bag_size = display_summary.get("bag_size", "medium") or "medium"
    bag_label = display_summary.get("bag_label", "Sac moyen") or "Sac moyen"

    bag_price_map = {
        "small": 7000,
        "medium": 10000,
        "large": 14000,
    }

    current_service_type = (getattr(order, "service_type", None) or "").strip().lower()

    items_qs = (
        OrderItem.objects
        .filter(order=order)
        .select_related("service", "service__category")
        .prefetch_related("photos")
        .order_by("id")
    )

    if current_service_type:
        filtered_items = []
        for it in items_qs:
            cat = getattr(getattr(it, "service", None), "category", None)
            item_service_type = (
                getattr(cat, "slug", None)
                or getattr(cat, "name", "")
                or ""
            ).strip().lower()
            if not item_service_type or item_service_type == current_service_type:
                filtered_items.append(it)
        items = filtered_items
    else:
        items = list(items_qs)

    has_items = len(items) > 0

    # 🔒 Si la commande n'est plus draft, on ne repasse pas par confirmation
    if not getattr(order, "is_draft", True):
        amounts_locked = _client_order_amounts(order)

        try:
            total_ttc_locked = Decimal(str((amounts_locked or {}).get("total_ttc", 0) or 0))
        except Exception:
            total_ttc_locked = Decimal("0")

        paid_locked = getattr(order, "amount_paid", Decimal("0")) or Decimal("0")
        remain_locked = total_ttc_locked - paid_locked

        if remain_locked > Decimal("0"):
            return redirect("orders:client_order_pay_wave_page", order_id=order.id)
        return redirect("orders:client_order_detail", order_id=order.id)

    amounts = _client_order_amounts(order)

    finance_breakdown = finance_summary

    try:
        service_fee_client_ttc = Decimal(str(finance_summary.get("service_fee_client_ttc", 0) or 0))
    except Exception:
        service_fee_client_ttc = Decimal("0")

    has_items = len(items) > 0

    is_bag_mode = bool(display_summary.get("is_bag", pricing_mode == "bag"))
    bag_base_price = finance_summary.get("prestation_total", Decimal("0")) or Decimal("0")

    # En mode item, items obligatoires. En mode bag, bag_size obligatoire.
    if not is_bag_mode and not has_items:
        try:
            from django.contrib import messages
            messages.error(request, "Ajoute au moins un article avant de confirmer la commande.")
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13378")
        return redirect("orders:client_new_order_step3", order_id=order.id)

    if is_bag_mode and bag_size not in ("small", "medium", "large"):
        try:
            from django.contrib import messages
            messages.error(request, "Choisis une taille de sac avant de confirmer la commande.")
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13386")
        return redirect("orders:client_new_order_step3", order_id=order.id)

    if request.method == "POST":
        accepted_cgu = (request.POST.get("accepted_cgu") or "").strip() == "1"
        if not accepted_cgu:
            try:
                from django.contrib import messages
                messages.error(request, "Merci d'accepter les Conditions Générales d'Utilisation FAGNI avant de confirmer la commande.")
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13396")
            return redirect("orders:client_new_order_step4", order_id=order.id)

        if not bool(request.session.get(f"client_cgu_accepted_{order.id}", False)):
            try:
                from orders.models import log_event
                log_event(
                    "cgu.accepted",
                    order=order,
                    actor_type="client",
                    actor_id=order.customer_id,
                    cgu_version="1.3",
                    source="client_new_order_step4",
                )
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13411")
        request.session[f"client_cgu_accepted_{order.id}"] = True

        if not is_bag_mode:
            # --- Garde-fou item: refuse confirmation si total prestations = 0 ---
            try:
                items_total = Decimal('0')
                for it in items:
                    if isinstance(it, dict):
                        q_raw = it.get('quantity', None)
                        if q_raw is None:
                            q_raw = it.get('qty', 0)

                        u_raw = it.get('unit_price', None)
                        if u_raw is None:
                            u_raw = it.get('price', 0)

                        line_total = it.get('total', None)

                        q = Decimal(str(q_raw or 0))
                        u = Decimal(str(u_raw or 0))

                        if q < 0:
                            q = Decimal('0')
                        if u < 0:
                            u = Decimal('0')

                        if line_total is not None:
                            lt = Decimal(str(line_total or 0))
                            if lt < 0:
                                lt = Decimal('0')
                            items_total += lt
                        else:
                            items_total += (q * u)
                    else:
                        q = Decimal(str(getattr(it, 'quantity', 0) or 0))
                        u = Decimal(str(getattr(it, 'unit_price', 0) or 0))
                        if q < 0:
                            q = Decimal('0')
                        if u < 0:
                            u = Decimal('0')
                        items_total += (q * u)
            except Exception:
                items_total = Decimal('0')

            if items_total <= Decimal('0'):
                try:
                    from django.contrib import messages
                    messages.error(request, "Montant prestations = 0. Vérifie les quantités et prix de tes articles avant de confirmer.")
                except Exception:
                    import logging
                    logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13461")
                return redirect('orders:client_new_order_step3', order_id=order.id)

        # Finalisation commerciale canonique :
        # les ServiceExecution doivent exister AVANT que la commande
        # puisse sortir du mode draft.
        try:
            from services.resolution import ServiceCatalogResolutionError
            from services.services import finalize_commercial_order

            finalize_commercial_order(order=order)
            order.refresh_from_db()
        except ServiceCatalogResolutionError as exc:
            from django.contrib import messages

            logger.warning(
                "Finalisation commerciale refusee : catalogue V2 incomplet | "
                "order_id=%s | error=%s",
                getattr(order, "id", None),
                exc,
            )
            messages.error(
                request,
                "Cette commande ne peut pas encore être confirmée : "
                "le service sélectionné n'est pas disponible.",
            )
            return redirect(
                "orders:client_new_order_step3",
                order_id=order.id,
            )

        # Les ressources logistiques ne sont créées qu'après
        # finalisation commerciale réussie.
        from orders.models import DeliveryLeg

        DeliveryLeg.objects.get_or_create(
            order=order,
            leg_type="pickup",
            defaults={"status": "pending"},
        )
        DeliveryLeg.objects.get_or_create(
            order=order,
            leg_type="return",
            defaults={"status": "pending"},
        )

        try:
            order.update_financials(save=True)
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13480")

        try:
            ensure_order_geocoded(order, save=True)
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13485")

        try:
            payment_total = Decimal(
                str(getattr(order, "total_client_ttc", 0) or 0)
            )
        except Exception:
            payment_total = Decimal("0")

        try:
            payment_paid = Decimal(
                str(getattr(order, "amount_paid", 0) or 0)
            )
        except Exception:
            payment_paid = Decimal("0")

        payment_status = (
            getattr(order, "payment_status", "") or ""
        ).strip().lower()

        payment_confirmed = bool(
            payment_total > Decimal("0")
            and payment_paid >= payment_total
            and payment_status == "paid"
        )

        # Aucun pressing, livreur ou statut opérationnel actif avant
        # confirmation comptable du paiement.
        if payment_confirmed:
            try:
                from orders.assignment import pick_best_laundry
                laundry, _laundry_reason = pick_best_laundry(order)
            except Exception:
                laundry, _laundry_reason = None, None

            if laundry:
                try:
                    order.laundry_partner = laundry
                except Exception:
                    import logging
                    logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13497")

            try:
                from orders.assignment import pick_best_driver
                driver, _driver_reason = pick_best_driver(order)
            except Exception:
                driver, _driver_reason = None, None

            if driver:
                try:
                    order.delivery_partner = driver
                except Exception:
                    import logging
                    logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13509")

            try:
                update_fields = []
                if hasattr(order, "laundry_partner_id"):
                    update_fields.append("laundry_partner")
                if hasattr(order, "delivery_partner_id"):
                    update_fields.append("delivery_partner")
                if update_fields:
                    order.save(update_fields=update_fields)
                else:
                    order.save()
            except Exception:
                try:
                    order.save()
                except Exception:
                    import logging
                    logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13525")

            try:
                from orders.models import sync_delivery_legs_for_order
                sync_delivery_legs_for_order(order)
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13531")

            try:
                from orders.service_layer.legs import normalize_order_legs
                normalize_order_legs(order, save=True)
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13537")

            try:
                has_laundry = bool(getattr(order, "laundry_partner_id", None))
                has_driver = bool(getattr(order, "delivery_partner_id", None))
                # A5-E5 :
                # une affectation de ressource ne signifie pas
                # qu'une ServiceExecution a démarré.
                if has_laundry or has_driver:
                    try:
                        from orders.models import sync_delivery_legs_for_order, DeliveryLeg
                        from orders.service_layer.legs import normalize_order_legs

                        sync_delivery_legs_for_order(order)
                        normalize_order_legs(order, save=True)

                        if getattr(order, "delivery_partner_id", None):
                            DeliveryLeg.objects.filter(order=order).exclude(status="canceled").update(
                                driver_id=order.delivery_partner_id
                            )
                    except Exception:
                        import logging
                        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13558")
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13560")

        try:
            amounts_after = _client_order_amounts(order)
            total_ttc = Decimal(str((amounts_after or {}).get("total_ttc", 0) or 0))
        except Exception:
            total_ttc = Decimal("0")

        paid = getattr(order, "amount_paid", Decimal("0")) or Decimal("0")
        remain = total_ttc - paid

        # ✅ AUTOPAIEMENT WALLET si le solde couvre le montant restant
        try:
            wallet = get_or_create_wallet_for_customer(order.customer)
        except Exception:
            wallet = None

        try:
            wallet_balance = Decimal(str(getattr(wallet, "balance", 0) or 0)) if wallet else Decimal("0")
        except Exception:
            wallet_balance = Decimal("0")

        if remain > Decimal("0") and wallet and wallet_balance >= remain:
            try:
                apply_order_payment(
                    order,
                    remain,
                    channel="wallet_auto",
                    reference=f"WALLET-AUTO-{order.id}",
                    note="Auto paiement wallet",
                )
            except Exception:
                import logging
                logging.getLogger("fagni.views.wallet").exception("Echec silencieux: paiement auto wallet (wallet_auto) | order_id=%s", getattr(order, "id", None) if "order" in dir() else None)

            update_fields = []
            try:
                order.refresh_from_db()
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13599")

            try:
                if update_fields:
                    order.save(update_fields=update_fields)
                else:
                    order.save()
            except Exception:
                order.save()

            return redirect("orders:client_order_detail", order_id=order.id)

        if remain > Decimal("0"):
            return redirect("orders:client_order_pay_wave_page", order_id=order.id)
        return redirect("orders:client_order_detail", order_id=order.id)

    cgu_accepted = bool(request.session.get(f"client_cgu_accepted_{order.id}", False))

    return render(request, "orders/client_new_order_step4.html", {
        "display_summary": display_summary,
        "finance_summary": finance_summary,
        "order": order,
        "pricing_mode": pricing_mode,
        "is_bag_mode": is_bag_mode,
        "bag_size": bag_size,
        "bag_label": display_summary.get("bag_label", "Sac moyen"),
        "cgu_accepted": cgu_accepted,
        "bag_base_price": bag_price_map.get(bag_size, 10000),
        "amounts": amounts,
        "finance_breakdown": finance_breakdown,
        "service_fee_client_ttc": service_fee_client_ttc,
        "items": items,
        "has_items": has_items,
    })

@require_http_methods(["GET"])
def laundry_app(request):
    """
    App Blanchisseur (MVP V1): liste des commandes attribuées à une blanchisserie.
    Règle: le partenaire ne voit JAMAIS les infos client (nom/tel/adresse).
    Auth: compte Django requis.
    Ciblage:
      - staff: peut passer ?laundry_id=XX
      - non-staff: on essaie de déduire la blanchisserie via email/username
    """
    laundry = None
    error = None

    # 1) Déduire blanchisserie pour utilisateur non-staff (sécurité)
    if not getattr(request.user, "is_staff", False):
        user_email = (getattr(request.user, "email", "") or "").strip().lower()
        user_name = (getattr(request.user, "username", "") or "").strip().lower()

        q = LaundryPartner.objects.all()
        # On tente email puis username si ton modèle a un champ email/name
        if user_email:
            laundry = q.filter(email__iexact=user_email).first()
        if laundry is None and user_name:
            # fallback: match sur email OU name
            laundry = q.filter(email__iexact=user_name).first() or q.filter(name__iexact=user_name).first()

        if laundry is None:
            error = "Aucune blanchisserie liée à ce compte. Contacte l'admin pour associer ton compte à une blanchisserie."
    else:
        # 2) Staff: peut cibler n'importe quelle blanchisserie via querystring
        laundry_id = (request.GET.get("laundry_id") or "").strip()
        if not laundry_id:
            error = "laundry_id manquant. Exemple: /orders/laundry/app/?laundry_id=1"
        else:
            try:
                lid = int(laundry_id)
                laundry = LaundryPartner.objects.filter(id=lid).first()
                if not laundry:
                    error = "Blanchisserie introuvable (laundry_id invalide)."
            except Exception:
                error = "laundry_id invalide (doit être un entier)."

    orders_qs = Order.objects.none()
    if laundry:
        # ✅ IMPORTANT: pas de select_related("customer") ici pour éviter toute tentation côté template
        orders_qs = (
            Order.objects
            .select_related("laundry_partner")
            .prefetch_related("items")
            .filter(laundry_partner_id=laundry.id)
            .order_by("-id")
        )

    orders_list = list(orders_qs[:200])

    stats_todo = 0
    stats_doing = 0
    stats_ready = 0
    stats_done = 0

    for o in orders_list:
        st = (getattr(o, "status", "") or "").lower().strip()
        ready = bool(getattr(o, "wash_complete_time", None))

        if st == "done":
            stats_done += 1
        elif ready:
            stats_ready += 1
        elif st == "in_progress":
            stats_doing += 1
        else:
            stats_todo += 1

    return render(
        request,
        "orders/laundry_app.html",
        {
            "laundry": laundry,
            "orders": orders_list,
            "error": error,
            "stats_todo": stats_todo,
            "stats_doing": stats_doing,
            "stats_ready": stats_ready,
            "stats_done": stats_done,
        },
    )


@login_required
@require_http_methods(["GET","POST"])
def laundry_order_detail(request, order_id):
    """
    Détail commande côté Blanchisserie (MVP V1 = prix par article, zéro pesée).
    Règle: AUCUNE info client (nom/tel/adresse).
    """

    laundry = None
    error = None
    laundry_id = (request.GET.get("laundry_id") or "").strip()

    # 0) Fallback explicite via ?laundry_id=XX même en dev / non-staff
    if laundry_id:
        try:
            lid = int(laundry_id)
            laundry = LaundryPartner.objects.filter(id=lid).first()
            if not laundry:
                error = "Blanchisserie introuvable (laundry_id invalide)."
        except Exception:
            error = "laundry_id invalide (doit être un entier)."

    # 1) Si rien trouvé, déduire via le compte connecté
    if laundry is None:
        user_email = (getattr(request.user, "email", "") or "").strip().lower()
        user_name = (getattr(request.user, "username", "") or "").strip().lower()

        q = LaundryPartner.objects.all()
        if user_email:
            laundry = q.filter(email__iexact=user_email).first()
        if laundry is None and user_name:
            laundry = q.filter(email__iexact=user_name).first() or q.filter(name__iexact=user_name).first()

        if laundry is None:
            error = "Aucune blanchisserie liée à ce compte. Contacte l'admin pour associer ton compte à une blanchisserie."

    if not laundry:
        return render(request, "orders/laundry_weighing.html", {
            "order": None,
            "laundry": None,
            "items": [],
            "error": error or "Blanchisserie non détectée.",
            "laundry_id": laundry_id or "",
        })

    order = (
        Order.objects
        .select_related("laundry_partner")
        .prefetch_related("items")
        .filter(pk=order_id, laundry_partner_id=laundry.id)
        .first()
    )
    if not order:
        raise Http404("Commande introuvable pour cette blanchisserie.")

    # Auto-réparation:
    # - Si status=done et pas prêt => on remplit wash_complete_time
    # - Si prêt (wash_complete_time existe) => on réactive la mission RETOUR si elle est canceled
    try:
        if (order.status == "done") and (not getattr(order, "wash_complete_time", None)):
            order.wash_complete_time = timezone.now()
            order.save(update_fields=["wash_complete_time"])

        if getattr(order, "wash_complete_time", None):
            DeliveryLeg.objects.get_or_create(
                order=order,
                leg_type="return",
                defaults={
                    "status": "pending",
                    "driver": order.delivery_partner if getattr(order, "delivery_partner_id", None) else None,
                },
            )

            # ✅ Auto-assign driver sur les legs return existants sans driver
            try:
                if getattr(order, "delivery_partner_id", None):
                    DeliveryLeg.objects.filter(
                        order=order,
                        leg_type="return",
                        driver__isnull=True,
                    ).update(driver=order.delivery_partner)
            except Exception:
                import logging
                logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13804")

            # 🔒 Ne jamais réactiver un return déjà payé (payout existe)
            qs_reactivate = DeliveryLeg.objects.filter(
                order=order,
                leg_type="return",
                status__in=("canceled", "cancelled"),
            )

            try:
                from wallets.models import WalletTransaction
                paid_leg_ids = set(
                    WalletTransaction.objects.filter(
                        order_id=order.id,
                        wallet__owner_type="driver",
                        type="payout",
                        direction="in",
                    ).exclude(leg_id__isnull=True).values_list("leg_id", flat=True)
                )
            except Exception:
                paid_leg_ids = set()

            qs_reactivate.exclude(id__in=paid_leg_ids).update(status="pending") 

    except Exception:
        logger.exception("laundry_order_detail auto-repair failed for order %s", getattr(order, "id", None))

    items = _build_client_display_items(order)

    # États UI des boutons
    status = getattr(order, "status", "") or ""
    ready = bool(getattr(order, "wash_complete_time", None))

    is_start_disabled = status in ("in_progress", "done")
    is_ready_disabled = ready or status == "done"
    is_done_disabled = (not ready) or status == "done"

    # Actions V1 (bloquées par règles)
    if request.method == "POST":
        # Frontière d'autorité V2 avant toute écriture de statut legacy.
        from services.services import order_uses_canonical_service_executions
        if order_uses_canonical_service_executions(order=order):
            return JsonResponse({
                "error": "autorite_v2",
                "message": (
                    "Le statut de cette commande est piloté par "
                    "ses ServiceExecution."
                ),
            }, status=409)

        action = (request.POST.get("action") or "").strip()

        choices = getattr(Order._meta.get_field("status"), "choices", []) or []
        allowed_values = {c[0] for c in choices if isinstance(c, (list, tuple)) and c}

        cur = (order.status or "").strip()
        already_ready = bool(getattr(order, "wash_complete_time", None))

        def set_status(value: str):
            if value in allowed_values:
                order.status = value

        # 0) Si déjà done => on bloque tout
        if cur == "done":
            return redirect(f"{request.path}?laundry_id={laundry.id}")

        # 1) START : une seule fois (pas prêt, pas done)
        if action == "start":
            if (not already_ready) and ("in_progress" in allowed_values) and (cur not in ("in_progress", "done")):
                try:
                    set_status("in_progress")
                    order.save(update_fields=["status"])
                except Exception:
                    import logging
                    logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13866")

        # 2) READY : une seule fois (piloté par wash_complete_time)
        elif action == "ready":
            if not already_ready:
                try:
                    order.wash_complete_time = timezone.now()

                    # optionnel: si in_progress existe et pas encore démarré
                    if ("in_progress" in allowed_values) and (cur not in ("in_progress", "done")):
                        set_status("in_progress")
                        order.save(update_fields=["status", "wash_complete_time"])
                    else:
                        order.save(update_fields=["wash_complete_time"])

                    try:
                        from orders.models import log_event
                        log_event(
                            "pressing.ready",
                            order=order,
                            actor_type="partner",
                            actor_id=laundry.id,
                            partner_id=laundry.id,
                            partner_name=getattr(laundry, "name", ""),
                        )
                    except Exception:
                        import logging
                        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13892")

                    # ✅ Dès que c'est "Prêt", on crée/active la mission RETOUR
                    DeliveryLeg.objects.get_or_create(
                        order=order,
                        leg_type="return",
                        defaults={
                            "status": "pending",
                            "driver": order.delivery_partner if getattr(order, "delivery_partner_id", None) else None,
                        },
                    )

                    # ✅ Auto-assign driver sur les legs return existants sans driver
                    try:
                        if getattr(order, "delivery_partner_id", None):
                            DeliveryLeg.objects.filter(
                                order=order,
                                leg_type="return",
                                driver__isnull=True,
                            ).update(driver=order.delivery_partner)
                    except Exception:
                        import logging
                        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13913")

                    # 🔒 Réactiver uniquement les returns annulés NON payés
                    qs_reactivate = DeliveryLeg.objects.filter(
                        order=order,
                        leg_type="return",
                        status__in=("canceled", "cancelled"),
                    )

                    try:
                        from wallets.models import WalletTransaction
                        paid_leg_ids = set(
                            WalletTransaction.objects.filter(
                                order_id=order.id,
                                wallet__owner_type="driver",
                                type="payout",
                                direction="in",
                            )
                            .exclude(leg_id__isnull=True)
                            .values_list("leg_id", flat=True)
                        )
                    except Exception:
                        paid_leg_ids = set()

                    qs_reactivate.exclude(id__in=paid_leg_ids).update(status="pending")

                    # Bonus safe: s'assurer qu'un pickup existe
                    DeliveryLeg.objects.get_or_create(
                        order=order,
                        leg_type="pickup",
                        defaults={"status": "done"},
                    )
                except Exception:
                    logger.exception("READY action failed for order %s", getattr(order, "id", None))

        # 3) DONE : autorisé seulement si prêt

        # 3) DONE : autorisé uniquement si prêt
        elif action == "done":
            if getattr(order, "wash_complete_time", None):
                try:
                    set_status("done")
                    order.save(update_fields=["status"])
                except Exception:
                    try:
                        order.status = "done"
                        order.save()
                    except Exception:
                        import logging
                        logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13961")

        # Recharge l'objet après action POST pour éviter tout affichage stale
        try:
            order.refresh_from_db()
        except Exception:
            import logging
            logging.getLogger("fagni.orders.views").exception("Exception silencieuse (auto-log) - fichier=orders/views.py ligne=13967")

        return redirect(f"{request.path}?laundry_id={laundry.id}")


    return render(request, "orders/laundry_order_detail.html", {
        "order": order,
        "laundry": laundry,
        "items": list(items),
        "error": error,
        "laundry_id": str(laundry.id),
        "is_start_disabled": is_start_disabled,
        "is_ready_disabled": is_ready_disabled,
        "is_done_disabled": is_done_disabled,
    })


# -------------------------------------------------------------------
# ✅ Lot 2.18 — OPS: confirmer paiement Wave (déclaration client → validation ops)
# -------------------------------------------------------------------
# LOT_2_18_WAVE_OPS_CONFIRM_OK
@require_http_methods(["POST"])

@staff_member_required
# LOT_2_18B_WAVE_OPS_STAFF_REQUIRED_OK
def ops_order_confirm_wave_paid(request, order_id: int):
    """
    Confirme manuellement un paiement Wave côté Ops/Admin.

    ✅ Règle FAGNI (source de vérité):
    - On crée un Payment (channel="wave") pour le RESTE dû.
    - Payment.save() applique les guards, sync Order.amount_paid/payment_status/payment_date,
      et déclenche mark_as_paid_and_distribute() si la commande devient PAID (idempotent).

    Notes:
    - Idempotent: si déjà soldée => OK sans recréer.
    - On n'essaie pas d'effacer la session "client" (wave_declared_*), car l'ops n'a pas cette session.
      Côté UI, dès que paid, on n'affiche plus "declared".
    """
    from decimal import Decimal

    order = Order.objects.filter(pk=order_id).first()
    if not order:
        return JsonResponse({"ok": False, "error": "order_not_found"}, status=404)

    try:
        finance_summary = build_order_finance_summary(order)
    except Exception:
        finance_summary = {}

    try:
        remaining = Decimal(str(finance_summary.get("amount_remaining", 0) or 0))
    except Exception:
        remaining = Decimal("0")

    if remaining <= 0:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({
                "ok": True,
                "order_id": order.id,
                "already_paid": True,
                "payment_status": getattr(order, "payment_status", ""),
            })
        return redirect("orders:ops_dashboard")

    checkout_id = (
        getattr(order, "wave_checkout_id", "") or ""
    ).strip()

    if not checkout_id:
        return JsonResponse(
            {
                "ok": False,
                "error": "missing_wave_checkout_id",
                "message": (
                    "Aucune session Wave vérifiable n'est "
                    "rattachée à cette commande."
                ),
            },
            status=400,
        )

    from orders.services import verify_wave_checkout_session

    try:
        verified_wave = verify_wave_checkout_session(checkout_id)
    except ValidationError as exc:
        return JsonResponse(
            {
                "ok": False,
                "error": "wave_verification_failed",
                "message": str(exc),
            },
            status=400,
        )

    remote_amount = Decimal(
        str(verified_wave.get("amount", 0) or 0)
    )

    if remote_amount < remaining:
        return JsonResponse(
            {
                "ok": False,
                "error": "wave_amount_insufficient",
                "message": (
                    "Le montant réellement confirmé par Wave "
                    "est inférieur au solde de la commande."
                ),
            },
            status=400,
        )

    ref = checkout_id

    payment_result = apply_order_payment(
        order,
        remaining,
        channel="wave_ops",
        reference=ref,
        note=(
            "Confirmation OPS Wave vérifiée via API Wave "
            f"par {getattr(request.user, 'username', '')}"
        ),
    )

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({
            "ok": True,
            "order_id": order.id,
            "reference": ref,
            "applied_amount": str(payment_result["applied"]),
            "payment_status": payment_result["payment_status"],
            "already_paid": bool(payment_result["already_settled"]),
        })

    return redirect("orders:ops_dashboard")


# --- LOT_3_1_GUARD_DOUBLE_PAY_OK ---


@login_required
@require_POST
def laundry_update_status(request, order_id):
    order = get_object_or_404(Order, id=order_id)

    laundry, _error = _resolve_laundry_for_user(request)
    if not laundry:
        return JsonResponse({
            "error": "blanchisserie_introuvable",
            "message": "Aucune blanchisserie liee a ce compte.",
        }, status=403)

    if getattr(order, "laundry_partner_id", None) != laundry.id:
        return JsonResponse({
            "error": "acces_refuse",
            "message": "Cette commande n'est pas attribuee a ta blanchisserie.",
        }, status=403)

    # Frontière d'autorité V2.
    from services.services import order_uses_canonical_service_executions
    if order_uses_canonical_service_executions(order=order):
        return JsonResponse({
            "error": "autorite_v2",
            "message": (
                "Le statut de cette commande est piloté par "
                "ses ServiceExecution."
            ),
        }, status=409)

    if getattr(order, "status", None) == "canceled":
        return JsonResponse({
            "error": "commande_annulee",
            "message": "Une commande annulee ne peut plus etre modifiee.",
        }, status=400)

    action = (request.POST.get("action") or "").strip()

    # Le pressing ne doit jamais pouvoir finaliser directement la commande :
    # Order.status ne passe a "done" que via sync_order_status_from_legs.
    if action == "done":
        return JsonResponse({
            "error": "statut_interdit",
            "message": "Le pressing ne peut pas terminer directement la commande.",
        }, status=400)

    if action != "start":
        return JsonResponse({
            "error": "action_invalide",
            "message": f"Action inconnue : {action!r}",
        }, status=400)

    pickup_done = DeliveryLeg.objects.filter(
        order=order, leg_type="pickup", status="done",
    ).exists()
    if not pickup_done:
        return JsonResponse({
            "error": "pickup_non_termine",
            "message": "La collecte (DeliveryLeg pickup) n'est pas encore terminee.",
        }, status=409)

    order.status = "in_progress"
    order.save(update_fields=["status", "updated_at"])

    return JsonResponse({
        "success": True,
        "new_status": order.status
    })


# ============================================================
# Dashboard Superviseur Blanchisserie
# ============================================================
@login_required
@require_http_methods(["GET"])
def laundry_supervisor_dashboard(request):
    """
    Dashboard superviseur dédié au pilotage blanchisserie.
    - staff: peut cibler ?laundry_id=XX
    - non-staff: on déduit la blanchisserie via email/username
    - aucune information client sensible affichée
    """
    laundry = None
    error = None

    if not getattr(request.user, "is_staff", False):
        user_email = (getattr(request.user, "email", "") or "").strip().lower()
        user_name = (getattr(request.user, "username", "") or "").strip().lower()

        q = LaundryPartner.objects.all()
        if user_email:
            laundry = q.filter(email__iexact=user_email).first()
        if laundry is None and user_name:
            laundry = q.filter(email__iexact=user_name).first() or q.filter(name__iexact=user_name).first()

        if laundry is None:
            error = "Aucune blanchisserie liée à ce compte. Contacte l'admin pour associer ton compte à une blanchisserie."
    else:
        laundry_id = (request.GET.get("laundry_id") or "").strip()
        if not laundry_id:
            error = "laundry_id manquant. Exemple: /orders/laundry/supervisor/?laundry_id=1"
        else:
            try:
                lid = int(laundry_id)
                laundry = LaundryPartner.objects.filter(id=lid).first()
                if not laundry:
                    error = "Blanchisserie introuvable (laundry_id invalide)."
            except Exception:
                error = "laundry_id invalide (doit être un entier)."

    orders_qs = Order.objects.none()
    orders_list = []

    stats_todo = 0
    stats_doing = 0
    stats_ready = 0
    stats_done = 0

    recent_orders = []
    alerts = []
    recommendation = "Situation normale."

    if laundry:
        orders_qs = (
            Order.objects
            .select_related("laundry_partner")
            .prefetch_related("items")
            .filter(laundry_partner_id=laundry.id)
            .order_by("-id")
        )

        orders_list = list(orders_qs[:300])

        for o in orders_list:
            st = (getattr(o, "status", "") or "").lower().strip()
            ready = bool(getattr(o, "wash_complete_time", None))

            if st == "done":
                stats_done += 1
            elif ready:
                stats_ready += 1
            elif st == "in_progress":
                stats_doing += 1
            else:
                stats_todo += 1

        recent_orders = orders_list[:12]

        if stats_todo > 0 and stats_doing == 0:
            alerts.append("Aucune commande n'est actuellement en cours de traitement.")
            recommendation = "Démarrer rapidement au moins une commande."
        elif stats_doing > 5:
            alerts.append("Beaucoup de commandes sont en cours en même temps.")
            recommendation = "Stabiliser le flux et finaliser les commandes prêtes."
        elif stats_ready > 0:
            alerts.append("Certaines commandes sont prêtes et attendent la fin de traitement.")
            recommendation = "Finaliser les commandes prêtes en priorité."
        elif stats_todo == 0 and stats_doing == 0 and stats_ready == 0:
            recommendation = "Toutes les commandes visibles semblent traitées."
        else:
            recommendation = "Poursuivre le traitement normal des commandes."

    return render(
        request,
        "orders/laundry_supervisor_dashboard.html",
        {
            "laundry": laundry,
            "error": error,
            "orders": recent_orders,
            "stats_todo": stats_todo,
            "stats_doing": stats_doing,
            "stats_ready": stats_ready,
            "stats_done": stats_done,
            "alerts": alerts,
            "recommendation": recommendation,
        },
    )



def _resolve_driver_for_request(request):
    """
    Retourne le DeliveryPartner courant :
    - non staff : driver connecté
    - staff : fallback via ?driver_id=
    """
    from partners.models import DeliveryPartner

    driver = _get_connected_driver(request)
    if not driver and getattr(request.user, "is_staff", False):
        driver_id = (request.GET.get("driver_id") or "").strip()
        if driver_id.isdigit():
            driver = DeliveryPartner.objects.filter(pk=int(driver_id)).first()
    return driver


def _select_driver_active_leg(driver, order=None):
    """
    Source unique de vérité FAGNI pour choisir UNE mission active visible.

    Règles métier :
    - statuts visibles : pending / assigned / in_progress
    - une livraison n'est visible que si wash_complete_time est renseigné
    - priorité de statut : in_progress > assigned > pending
    - priorité métier :
        1) livraison prête / en cours
        2) collecte
    """
    from django.db.models import Case, When, Value, IntegerField, Q

    if not driver:
        return None

    qs = (
        DeliveryLeg.objects
        .select_related("order", "order__customer", "order__laundry_partner", "driver")
        .filter(driver=driver, status__in=["pending", "assigned", "in_progress"])
        .exclude(status="canceled")
    )

    if order is not None:
        qs = qs.filter(order=order)

    # Livraison invisible tant que le linge n'est pas prêt
    qs = qs.exclude(
        Q(leg_type__in=["return", "delivery"]) &
        Q(order__wash_complete_time__isnull=True)
    )

    qs = qs.annotate(
        global_priority=Case(
            # 1. mission en cours
            When(status="in_progress", then=Value(0)),

            # 2. livraison déjà acceptée
            When(
                status="assigned",
                leg_type__in=["return", "delivery"],
                order__wash_complete_time__isnull=False,
                then=Value(1)
            ),

            # 3. livraison prête à accepter
            When(
                status="pending",
                leg_type__in=["return", "delivery"],
                order__wash_complete_time__isnull=False,
                then=Value(2)
            ),

            # 4. collecte déjà acceptée
            When(
                status="assigned",
                leg_type="pickup",
                then=Value(3)
            ),

            # 5. collecte à accepter
            When(
                status="pending",
                leg_type="pickup",
                then=Value(4)
            ),

            default=Value(9),
            output_field=IntegerField(),
        ),
    ).order_by("global_priority", "order__created_at", "id")

    return qs.first()


def _build_driver_mission_context(request, driver, order=None):
    """
    Construit un contexte homogène pour :
    - driver_app_v3
    - driver_mission_v3
    - driver_hub
    """
    from urllib.parse import quote

    base = {
        "driver": driver,
        "order": None,
        "leg": None,
        "next_action": None,
        "next_action_label": None,
        "mission_type_label": None,
        "mission_state_label": "✅ Aucune mission active",
        "mission_state_tone": "gray",
        "mission_hint": "Tu n’as aucune action terrain à faire pour le moment.",
        "address_display": None,
        "maps_url": None,
        "proof_url": None,
        "mission_cta": None,
    }

    if not driver:
        return base

    leg = _select_driver_active_leg(driver, order=order)
    current_order = getattr(leg, "order", None)

    if not leg or not current_order:
        if order is not None:
            base["mission_state_label"] = "⏸️ Aucune mission active sur cette commande"
            if getattr(order, "wash_complete_time", None):
                base["mission_hint"] = "Cette commande ne présente plus d’action livreur en attente."
            else:
                base["mission_hint"] = "Collecte terminée ou aucune action livreur disponible pour le moment."
        return base

    leg_type = (getattr(leg, "leg_type", "") or "").strip().lower()
    st = (getattr(leg, "status", "") or "").strip().lower()
    customer = getattr(current_order, "customer", None)

    pickup_address = getattr(current_order, "pickup_address", None) or getattr(customer, "address", None)
    delivery_address = getattr(current_order, "delivery_address", None) or getattr(customer, "address", None)

    address_display = pickup_address if leg_type == "pickup" else delivery_address
    maps_url = None
    if address_display:
        maps_url = "https://www.google.com/maps/search/?api=1&query=" + quote(str(address_display))

    back = reverse("orders:driver_app")
    if getattr(driver, "id", None):
        back = f"{back}?driver_id={driver.id}"

    proof_url = reverse("orders:driver_weighing", kwargs={"order_id": current_order.id})

    params = []

    if getattr(leg, "id", None):
        params.append(f"leg_id={leg.id}")

    if getattr(driver, "id", None):
        params.append(f"driver_id={driver.id}")

    if back:
        params.append(f"back={quote(back, safe='')}")

    if params:
        proof_url = proof_url + "?" + "&".join(params)

    mission_cta = reverse("orders:driver_mission_v3", kwargs={"order_id": current_order.id})
    if getattr(driver, "id", None):
        mission_cta = f"{mission_cta}?driver_id={driver.id}"

    ACTION_LABELS = {
        "accept": "✅ Accepter mission",
        "start": "🚀 Démarrer",
        "finish": "✅ Terminer mission",
    }

    next_action = None
    next_action_label = None

    if st == "pending":
        next_action = "accept"
        next_action_label = ACTION_LABELS["accept"]

    elif st == "assigned":
        next_action = "start"
        if leg_type == "pickup":
            next_action_label = "🚀 Démarrer collecte"
        else:
            next_action_label = "🚀 Démarrer livraison"

    elif st == "in_progress":
        next_action = "finish"
        if leg_type == "pickup":
            next_action_label = "✅ Collecte terminée"
        else:
            next_action_label = "✅ Livraison terminée"

    if leg_type == "pickup":
        mission_type_label = "Collecte"
        mission_state_tone = "green"

        if st == "pending":
            mission_state_label = "🟢 Collecte à faire"
            mission_hint = "Accepte puis démarre la collecte chez le client."
        elif st == "assigned":
            mission_state_label = "🟢 Collecte à démarrer"
            mission_hint = "Rends-toi chez le client pour récupérer le linge."
        elif st == "in_progress":
            mission_state_label = "🟢 Collecte en cours"
            mission_hint = "Ajoute une preuve puis valide la collecte."
        else:
            mission_state_label = "🟢 Collecte"
            mission_hint = "Mission de collecte."
    else:
        mission_type_label = "Livraison"
        mission_state_tone = "blue"

        if st == "pending":
            mission_state_label = "🔵 Livraison à faire"
            mission_hint = "Le linge est prêt. Accepte puis démarre la livraison."
        elif st == "assigned":
            mission_state_label = "🔵 Livraison à démarrer"
            mission_hint = "Le linge est prêt. Lance maintenant la remise au client."
        elif st == "in_progress":
            mission_state_label = "🔵 Livraison en cours"
            mission_hint = "Ajoute une preuve de remise puis valide la livraison."
        else:
            mission_state_label = "🔵 Livraison"
            mission_hint = "Mission de livraison."

    base.update({
        "order": current_order,
        "leg": leg,
        "next_action": next_action,
        "next_action_label": next_action_label,
        "mission_type_label": mission_type_label,
        "mission_state_label": mission_state_label,
        "mission_state_tone": mission_state_tone,
        "mission_hint": mission_hint,
        "address_display": address_display,
        "maps_url": maps_url,
        "proof_url": proof_url,
        "mission_cta": mission_cta,
    })
    return base



# ============================================================
#  DRIVER V3 — MISSION UNIQUE
# ============================================================
@login_required
@require_http_methods(["GET"])
def driver_app_v3(request):
    """
    App livreur V3 :
    - affiche uniquement la mission active prioritaire
    - logique métier unifiée FAGNI
    """
    driver = _resolve_driver_for_request(request)
    ctx = _build_driver_mission_context(request, driver)
    return render(request, "orders/driver_mission.html", ctx)


@login_required
@require_http_methods(["GET"])
def driver_mission_v3(request, order_id):
    """
    Détail mission V3 centré sur UNE course et UNE mission visible.
    """
    driver = _resolve_driver_for_request(request)

    if not driver:
        return HttpResponseForbidden("Aucun livreur connecté.")

    order = get_object_or_404(
        Order.objects.select_related("customer", "laundry_partner", "delivery_partner"),
        pk=order_id
    )

    if not getattr(request.user, "is_staff", False):
        # Source de vérité livreur : DeliveryLeg.driver, pas Order.delivery_partner.
        has_leg = DeliveryLeg.objects.filter(order=order, driver=driver).exists()
        if not has_leg:
            return HttpResponseForbidden("Accès refusé : aucune mission assignée à ce livreur.")

    ctx = _build_driver_mission_context(request, driver, order=order)
    return render(request, "orders/driver_mission.html", ctx)


# ============================
# MULTI LEVEL REFERRAL ENGINE
# ============================

def compute_referral_gains(profile):
    level1 = 500
    level2 = 200
    level3 = 100

    gains = {
        "level1": 0,
        "level2": 0,
        "level3": 0,
        "total": 0,
    }

    # Niveau 1
    level1_refs = profile.direct_referrals.all() if hasattr(profile, 'direct_referrals') else []
    gains["level1"] = len(level1_refs) * level1

    # Niveau 2
    level2_count = 0
    for ref in level1_refs:
        if hasattr(ref, 'direct_referrals'):
            level2_count += ref.direct_referrals.count()

    gains["level2"] = level2_count * level2

    # Niveau 3 (simplifié)
    gains["level3"] = int(level2_count * 0.5) * level3

    gains["total"] = gains["level1"] + gains["level2"] + gains["level3"]

    return gains


# =========================================
# REFERRAL ENGINE V1 (SAFE + SIMPLE)
# =========================================

def handle_referral_reward(order):
    try:
        from mlm.models import ReferralLink
        from wallets.services import credit_wallet
        from wallets.models import WalletTransaction

        customer = getattr(order, "customer", None)
        if not customer:
            return

        referral_code = getattr(order, "referral_code", None)
        if not referral_code:
            return

        sponsor_profile = ReferralLink.objects.filter(referral_code=referral_code).first()
        if not sponsor_profile:
            return

        sponsor_customer = getattr(sponsor_profile, "customer", None)
        sponsor_phone = getattr(sponsor_customer, "phone", None)
        child_phone = getattr(customer, "phone", None)
        same_phone = bool(sponsor_phone and child_phone and str(sponsor_phone).strip() == str(child_phone).strip())
        if same_phone:
            return

        # Paiement valide : status OU payment_status
        order_status = str(getattr(order, "status", "") or "").lower()
        payment_status = str(getattr(order, "payment_status", "") or "").lower()
        if order_status not in ["paid", "completed", "done"] and payment_status not in ["paid"]:
            return

        # anti-fraude 48h
        from datetime import timedelta
        if order.created_at > timezone.now() - timedelta(hours=48):
            return

        # première commande payée uniquement avec referral_code
        first_paid_order = (
            Order.objects
            .filter(
                customer=customer,
                payment_status="paid",
            )
            .exclude(referral_code__isnull=True)
            .exclude(referral_code__exact="")
            .order_by("created_at", "id")
            .first()
        )
        if not first_paid_order or getattr(first_paid_order, "id", None) != getattr(order, "id", None):
            return

        reward_amount = 1000  # FCFA (Pilot Growth Plan, 9 juillet 2026)

        # idempotence réelle : si déjà une tx MLM liée à cette commande -> stop
        already_rewarded = WalletTransaction.objects.filter(
            wallet__owner_type="customer",
            wallet__customer=sponsor_customer,
            order=order,
            type="mlm_commission",
            direction="in",
        ).exists()
        if already_rewarded:
            return

        # crédit wallet parrain
        sponsor_wallet = get_or_create_wallet_for_customer(sponsor_customer)
        credit_wallet(
            wallet=sponsor_wallet,
            amount=reward_amount,
            description=f"Parrainage commande {getattr(order, 'code', order.id)}",
            order=order,
            tx_type="mlm_commission",
        )

    except Exception as e:
        print("Referral error:", e)


# =========================================
# CHILD INCENTIVE ENGINE V1
# =========================================

def get_child_referral_discount(customer, order=None):
    try:
        if not customer or order is None:
            return 0

        referral_code_used = getattr(order, "referral_code", None)
        if not referral_code_used:
            return 0

        first_order = (
            Order.objects
            .filter(customer=customer)
            .order_by("created_at", "id")
            .first()
        )

        if not first_order:
            return 0

        if getattr(first_order, "id", None) != getattr(order, "id", None):
            return 0

        return 500
    except Exception:
        return 0


def validate_and_get_coupon_discount(customer, prestation_total, coupon_code, exclude_order_pk=None):
    """
    Valide un coupon et calcule la reduction (ADR-025, Pilot Growth Plan, 9 juillet 2026).
    La reduction est toujours calculee sur le montant HORS livraison (prestation seule),
    et absorbee entierement par la marge FAGNI - jamais par les commissions
    partenaire/livreur, deja verrouillees a la creation de la commande (ADR-001).

    Fonction pure lecture : aucune ecriture, aucun CouponUsage cree ici. Peut donc
    etre appelee en apercu (avant creation de commande, cf api_coupon_preview) sans
    consommer un usage reel du coupon - seule la creation de commande reelle
    (api_create_order, plus bas dans ce module) cree un CouponUsage et fait donc
    progresser max_uses_per_customer/max_total_uses.

    `prestation_total` : montant HORS livraison/service (le meme calcul que
    articlesTotal cote frontend - deja soustrait par l'appelant, jamais recalcule
    ici a partir d'une Order). `exclude_order_pk` : optionnel, exclut une commande
    du controle "premiere commande payee" - utilise par le chemin de creation de
    commande reelle, ou l'order courante existe deja en base au moment de l'appel ;
    laisser None pour un apercu avant creation (aucune commande a exclure).

    Retourne (discount_amount: Decimal, error: str|None, coupon: Coupon|None)
    """
    from orders.models import Coupon, CouponUsage

    if not coupon_code:
        return Decimal("0"), None, None

    coupon_code = coupon_code.strip().upper()

    try:
        coupon = Coupon.objects.get(code=coupon_code)
    except Coupon.DoesNotExist:
        return Decimal("0"), "coupon_introuvable", None

    if not coupon.is_currently_valid():
        return Decimal("0"), "coupon_invalide_ou_expire", coupon

    if not customer:
        return Decimal("0"), "client_introuvable", coupon

    if coupon.first_order_only:
        # Une commande technique, abandonnée, non payée ou annulée ne doit
        # pas faire perdre au client son offre de première commande.
        #
        # La première commande commerciale correspond ici à la première
        # commande réellement payée. L'usage du coupon reste également
        # protégé plus bas par CouponUsage.
        previous_paid_order_exists = (
            Order.objects
            .filter(
                customer=customer,
                payment_status="paid",
            )
            .exclude(pk=exclude_order_pk)
            .exists()
        )

        if previous_paid_order_exists:
            return Decimal("0"), "reserve_premiere_commande", coupon

    already_used = CouponUsage.objects.filter(coupon=coupon, customer=customer).count()
    if already_used >= coupon.max_uses_per_customer:
        return Decimal("0"), "deja_utilise", coupon

    prestation_total = _safe_dec(prestation_total)

    if coupon.discount_type == "percent":
        discount = (prestation_total * coupon.discount_value / Decimal("100")).quantize(Decimal("1"))
        if coupon.max_discount_amount is not None and discount > coupon.max_discount_amount:
            discount = coupon.max_discount_amount
    else:
        discount = coupon.discount_value

    if discount > prestation_total:
        discount = prestation_total
    if discount < 0:
        discount = Decimal("0")

    return discount, None, coupon


def get_child_referral_discount_amount(order):
    try:
        if not order:
            return Decimal("0")

        customer = getattr(order, "customer", None)
        if not customer:
            return Decimal("0")

        amount = get_child_referral_discount(customer, order=order) or 0
        amount = Decimal(str(amount))

        if amount <= 0:
            return Decimal("0")

        return amount.quantize(Decimal("1"))
    except Exception:
        return Decimal("0")


# ==========================
# WEBHOOK WAVE
# ==========================
@csrf_exempt
def wave_webhook(request):
    import json
    import hmac
    import hashlib
    import time
    import urllib.request
    import urllib.error
    from decimal import Decimal
    from django.conf import settings
    from django.db import transaction
    from django.http import JsonResponse
    from django.utils import timezone

    from orders.models import Order, WaveEvent, OrderPaymentEvent

    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "method_not_allowed"}, status=405)

    raw_body = request.body or b""

    # DEBUG local : on skip la signature et on parse directement
    if settings.DEBUG:
        try:
            event = json.loads(raw_body.decode("utf-8"))
        except Exception:
            return JsonResponse({"ok": False, "error": "invalid_json"}, status=400)
    else:
        wave_signature = (request.headers.get("Wave-Signature") or "").strip()
        webhook_secret = (getattr(settings, "WAVE_WEBHOOK_SIGNING_SECRET", "") or "").strip()

        if not webhook_secret:
            return JsonResponse({"ok": False, "error": "missing_webhook_secret"}, status=500)

        try:
            parts = {}
            for part in wave_signature.split(","):
                if "=" in part:
                    k, v = part.split("=", 1)
                    parts[k.strip()] = v.strip()

            timestamp = parts.get("t", "")
            sig_v1 = parts.get("v1", "")

            if not timestamp or not sig_v1:
                return JsonResponse({"ok": False, "error": "invalid_signature_header"}, status=401)

            now_ts = int(time.time())
            req_ts = int(timestamp)
            if abs(now_ts - req_ts) > 300:
                return JsonResponse({"ok": False, "error": "stale_signature"}, status=401)

            signed_payload = timestamp.encode("utf-8") + raw_body
            expected_sig = hmac.new(
                webhook_secret.encode("utf-8"),
                signed_payload,
                hashlib.sha256,
            ).hexdigest()

            if not hmac.compare_digest(expected_sig, sig_v1):
                return JsonResponse({"ok": False, "error": "bad_signature"}, status=401)

            event = json.loads(raw_body.decode("utf-8"))
        except Exception:
            return JsonResponse({"ok": False, "error": "signature_verification_failed"}, status=401)

    event_id = (event.get("id") or "").strip()
    event_type = (event.get("type") or "").strip()
    data = event.get("data") or {}

    if event_type != "checkout.session.completed":
        return JsonResponse({"ok": True, "ignored": True})

    checkout_id = (data.get("id") or "").strip()
    payment_status = (data.get("payment_status") or "").strip()
    checkout_status = (data.get("checkout_status") or "").strip()
    currency = (data.get("currency") or "").strip()
    amount_raw = data.get("amount")

    if not checkout_id:
        return JsonResponse({"ok": False, "error": "missing_checkout_id"}, status=400)

    if payment_status != "succeeded" or checkout_status != "complete":
        return JsonResponse({"ok": True, "ignored": True})

    # Sécurité P0 — wave_checkout_id est l'unique clé de rattachement PSP.
    #
    # payment_declared_reference est une donnée déclarative/legacy et ne
    # constitue jamais une preuve qu'une session Wave appartient à cette
    # commande.
    order = (
        Order.objects
        .select_related("customer")
        .filter(wave_checkout_id=checkout_id)
        .first()
    )

    if not order:
        return JsonResponse(
            {"ok": False, "error": "order_not_found"},
            status=404,
        )

    # DEBUG local : on accepte checkout_test_xxx sans appel API Wave
    if settings.DEBUG and checkout_id.startswith("checkout_test_"):
        remote_id = checkout_id
        remote_payment_status = payment_status
        remote_checkout_status = checkout_status
        remote_currency = currency
        try:
            remote_amount = Decimal(str(amount_raw or "0"))
        except Exception:
            remote_amount = Decimal("0")
    else:
        api_key = (getattr(settings, "WAVE_CHECKOUT_API_KEY", "") or "").strip()
        if not api_key:
            return JsonResponse({"ok": False, "error": "missing_api_key"}, status=500)

        try:
            req = urllib.request.Request(
                f"https://api.wave.com/v1/checkout/sessions/{checkout_id}",
                headers={"Authorization": f"Bearer {api_key}"},
                method="GET",
            )
            with urllib.request.urlopen(req, timeout=8) as resp:
                remote = json.loads(resp.read().decode("utf-8"))
        except Exception as e:
            print("[WAVE ERROR]", e)
            return JsonResponse({"ok": False, "error": f"checkout_retrieve_failed:{e}"}, status=502)

        remote_id = (remote.get("id") or "").strip()
        remote_payment_status = (remote.get("payment_status") or "").strip()
        remote_checkout_status = (remote.get("checkout_status") or "").strip()
        remote_currency = (remote.get("currency") or "").strip()

        try:
            remote_amount = Decimal(str(remote.get("amount") or "0"))
        except Exception:
            remote_amount = Decimal("0")

    if remote_id != checkout_id:
        return JsonResponse({"ok": False, "error": "checkout_id_mismatch"}, status=400)

    if remote_payment_status != "succeeded" or remote_checkout_status != "complete":
        return JsonResponse({"ok": False, "error": "checkout_not_confirmed"}, status=400)

    if remote_currency != "XOF" or currency != "XOF":
        return JsonResponse({"ok": False, "error": "currency_mismatch"}, status=400)

    try:
        remote_amount = Decimal(str(remote_amount or "0"))
    except Exception:
        remote_amount = Decimal("0")

    if remote_amount <= Decimal("0"):
        return JsonResponse({"ok": False, "error": "invalid_amount"}, status=400)

    with transaction.atomic():
        # 🔒 Idempotence STRONG : un event_id ne passe qu'une seule fois
        if event_id:
            try:
                obj, created = WaveEvent.objects.get_or_create(event_id=event_id)
            except Exception:
                print(f"[WAVE WEBHOOK] race condition caught event_id={event_id}")
                return JsonResponse({
                    "ok": True,
                    "idempotent": True,
                    "event_id": event_id
                })

            if not created:
                print(f"[WAVE WEBHOOK] idempotent replay ignored event_id={event_id}")
                return JsonResponse({
                    "ok": True,
                    "idempotent": True,
                    "event_id": event_id
                })

        # 🔒 Verrou DB sur la commande pour éviter les doubles écritures concurrentes
        order = (
            Order.objects
            .select_for_update()
            .select_related("customer")
            .get(pk=order.pk)
        )

        try:
            finance_summary = build_order_finance_summary(order)
        except Exception:
            finance_summary = {}

        try:
            total = Decimal(str(finance_summary.get("total_client_ttc", 0) or 0))
        except Exception:
            total = Decimal("0")

        try:
            already_paid = Decimal(str(getattr(order, "amount_paid", 0) or 0))
        except Exception:
            already_paid = Decimal("0")

        if already_paid < Decimal("0"):
            already_paid = Decimal("0")

        remaining = total - already_paid
        if remaining < Decimal("0"):
            remaining = Decimal("0")

        if remaining <= Decimal("0"):
            return JsonResponse({"ok": True, "idempotent": True, "event_id": event_id})

        to_apply = remote_amount if remote_amount <= remaining else remaining
        if to_apply <= Decimal("0"):
            return JsonResponse({"ok": True, "idempotent": True, "event_id": event_id})

        payment_result = apply_order_payment(
            order,
            to_apply,
            channel="wave_webhook",
            reference=checkout_id,
            note=f"Webhook Wave event_id={event_id}",
        )

    print(
        f"[WAVE WEBHOOK] applied event_id={event_id} "
        f"checkout_id={checkout_id} amount={payment_result['applied']} "
        f"order_id={order.id} payment_status={payment_result['payment_status']}"
    )

    return JsonResponse({
        "ok": True,
        "event_id": event_id,
        "checkout_id": checkout_id,
        "applied_amount": str(payment_result["applied"]),
        "payment_status": payment_result["payment_status"],
    })



from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def admin_payment_review(request):
    from orders.models import Order

    orders = Order.objects.filter(
        payment_status="declared",
        payment_verification_status="pending_review"
    ).order_by("-payment_declared_at")

    return render(request, "admin/payment_review.html", {
        "orders": orders
    })


@staff_member_required
@require_POST
def admin_payment_review_confirm(request, order_id: int):
    from decimal import Decimal
    from django.shortcuts import redirect, get_object_or_404
    from django.utils import timezone
    from orders.models import Order
    from orders.views import build_order_finance_summary, apply_order_payment

    order = get_object_or_404(Order, pk=order_id)

    fs = build_order_finance_summary(order)
    total = Decimal(str(fs.get("total_client_ttc", 0) or 0))
    paid = Decimal(str(getattr(order, "amount_paid", 0) or 0))
    remaining = total - paid

    if remaining > 0:
        verified_reference = (
            request.POST.get("verified_wave_reference") or ""
        ).strip()

        human_confirmation = (
            request.POST.get("wave_human_verified") or ""
        ).strip().lower()

        if human_confirmation not in {"1", "true", "on", "yes"}:
            messages.error(
                request,
                "Validation refusée : confirmez d'abord que la "
                "transaction a été contrôlée dans Wave.",
            )
            return redirect("orders:admin_payment_review")

        if not verified_reference:
            messages.error(
                request,
                "Validation refusée : la référence Wave réellement "
                "vérifiée est obligatoire.",
            )
            return redirect("orders:admin_payment_review")

        apply_order_payment(
            order,
            remaining,
            channel="wave_manual_verified",
            reference=verified_reference,
            note=(
                "Paiement Wave vérifié manuellement dans l'application "
                f"Wave par {getattr(request.user, 'username', '')}"
            ),
        )

    order.refresh_from_db()
    update_fields = []

    if getattr(order, "payment_verification_status", None) != "verified":
        order.payment_verification_status = "verified"
        update_fields.append("payment_verification_status")

    order.payment_verified_at = timezone.now()
    update_fields.append("payment_verified_at")

    if getattr(request, "user", None) and request.user.is_authenticated:
        order.payment_verified_by = request.user
        update_fields.append("payment_verified_by")

    if update_fields:
        order.save(update_fields=list(dict.fromkeys(update_fields)))

    return redirect("orders:admin_payment_review")


@staff_member_required
@require_POST
def admin_payment_review_reject(request, order_id: int):
    from django.shortcuts import redirect, get_object_or_404
    from django.utils import timezone
    from orders.models import Order

    order = get_object_or_404(Order, pk=order_id)
    update_fields = []

    if getattr(order, "payment_verification_status", None) != "rejected":
        order.payment_verification_status = "rejected"
        update_fields.append("payment_verification_status")

    if getattr(order, "payment_status", None) != "paid":
        order.payment_status = "pending"
        update_fields.append("payment_status")

    if hasattr(order, "payment_verified_at"):
        order.payment_verified_at = None
        update_fields.append("payment_verified_at")

    if hasattr(order, "payment_verified_by"):
        order.payment_verified_by = None
        update_fields.append("payment_verified_by")

    if update_fields:
        order.save(update_fields=list(dict.fromkeys(update_fields)))

    return redirect("orders:admin_payment_review")
